from __future__ import annotations

from autobot.paths import REPO_ROOT
import io
import gzip
import hmac
import json
import math
import mimetypes
import os
import re
import shutil
import subprocess
import uuid
import sys
import time
import traceback
import threading
import html as html_mod
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import parse_qs, quote, urlencode, urlparse

import pandas as pd

try:
    from dotenv import load_dotenv

    load_dotenv(REPO_ROOT / ".env")
except ImportError:
    pass

from flask import (
    Flask,
    abort,
    redirect,
    jsonify,
    make_response,
    render_template,
    render_template_string,
    request,
    send_file,
    send_from_directory,
    url_for,
)

from autobot.site_public_url import get_report_site_public_base
from autobot.source_documents import (
    build_source_bytes_preview,
    build_source_file_preview,
    format_file_size,
    list_tender_source_files,
    read_archive_member,
    repair_filename,
    resolve_tender_source_file,
)
from autobot.tender_detail import build_tender_detail
from autobot.tender_deletion import delete_tender_data
from autobot.workflow_overview import build_storage_overview, build_workflow_payload

_AUTOBOT_MAIN_FILE = REPO_ROOT / "autobot" / "main.py"
_TOOLS_RUN_MODULE = REPO_ROOT / "tools" / "run_module.py"

try:
    from autobot.report_prompt import (
        BASE_DIR,
        REPORTS_DIR,
        TENDERS_JSON,
        load_tender_metadata,
    )
except ModuleNotFoundError as e:
    if getattr(e, "name", None) == "report_prompt":
        sys.stderr.write("Не удалось загрузить autobot.report_prompt (проверьте установку пакета autobot/).\n")
    raise

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 26 * 1024 * 1024  # загрузка Excel обоснования НМЦК

FAVICON_SVG = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 64 64">
<defs>
  <linearGradient id="g" x1="0" y1="0" x2="1" y2="1">
    <stop offset="0%" stop-color="#5ea2ff"/>
    <stop offset="100%" stop-color="#5ecf8a"/>
  </linearGradient>
</defs>
<rect x="6" y="6" width="52" height="52" rx="14" fill="#0f1830"/>
<path d="M20 16h16l8 8v20a4 4 0 0 1-4 4H20a4 4 0 0 1-4-4V20a4 4 0 0 1 4-4z" fill="url(#g)"/>
<path d="M36 16v8h8" fill="none" stroke="#e8eefc" stroke-width="3" stroke-linejoin="round"/>
<path d="M24 30h12M24 36h10" stroke="#e8eefc" stroke-width="3" stroke-linecap="round"/>
<circle cx="42" cy="42" r="8" fill="#0f1830" stroke="#e8eefc" stroke-width="3"/>
<path d="M47.5 47.5L53 53" stroke="#e8eefc" stroke-width="3" stroke-linecap="round"/>
</svg>"""

# Совпадает с NEEDED_STAGE в main.py — единственная стадия, которую подсвечиваем зелёным.
STAGE_SUBMISSION = "Подача заявок"

parse_state = {
    "running": False,
    "task": "",
    "command": "",
    "started_at": None,
    "ended_at": None,
    "exit_code": None,
    "log_lines": [],
}
parse_lock = threading.Lock()

merge_site_state: dict = {
    "running": False,
    "total": 0,
    "done": 0,
    "current_tid": "",
    "market_done": 0,
    "market_total": 0,
    "last_market_chat_done": 0,
    "started_at": None,
    "ended_at": None,
    "error_ids": [],
    "log_lines": [],
    "chat_events": [],
    "last_ended_at": None,
    "last_summary": "",
    "last_reason_counts": {},
}
merge_site_lock = threading.Lock()

estimate_upload_jobs: dict[str, dict] = {}
estimate_upload_lock = threading.Lock()
estimate_market_jobs: dict[str, dict] = {}
estimate_market_lock = threading.Lock()
tender_delete_lock = threading.Lock()
agent_market_import_lock = threading.Lock()


def _agent_market_token() -> str:
    from autobot.agent_market_queue import get_or_create_worker_token

    try:
        return get_or_create_worker_token()
    except OSError:
        return ""


def _agent_market_authorized() -> bool:
    expected = _agent_market_token()
    if not expected:
        return False
    authorization = str(request.headers.get("Authorization") or "").strip()
    supplied = authorization[7:].strip() if authorization.casefold().startswith("bearer ") else ""
    if not supplied:
        supplied = str(request.headers.get("X-AutoBot-Agent-Token") or "").strip()
    return bool(supplied) and hmac.compare_digest(supplied, expected)


def _require_agent_market_token():
    if not _agent_market_token():
        return jsonify({"ok": False, "message": "MARKET_AGENT_TOKEN не настроен в AutoBot"}), 503
    if not _agent_market_authorized():
        return jsonify({"ok": False, "message": "Неверный токен агента"}), 401
    return None


def _agent_offer_url(value: object) -> str:
    raw = html_mod.unescape(str(value or "")).strip()
    match = re.search(r"https?://[^\s<>\]\[\)\}\"']+", raw, flags=re.IGNORECASE)
    url = (match.group(0) if match else raw).rstrip(".,;:")
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc or parsed.username or parsed.password:
        return ""
    return url[:2000]


def _agent_offer_price(value: object) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
    else:
        raw = str(value or "").replace("\xa0", " ").replace("\u202f", " ")
        match = re.search(r"\d[\d\s]*(?:[,.]\d{1,2})?", raw)
        if not match:
            return None
        normalized = match.group(0).replace(" ", "").replace(",", ".")
        try:
            number = float(normalized)
        except ValueError:
            return None
    return number if math.isfinite(number) and 1 <= number <= 500_000_000 else None


def _validate_agent_market_result(result: object, expected_position_key: str) -> dict:
    if not isinstance(result, dict):
        raise ValueError("Результат должен быть JSON-объектом")
    returned_key = str(result.get("position_key") or expected_position_key).strip()
    if returned_key != expected_position_key:
        raise ValueError("Агент вернул результат для другой позиции")
    offers: list[dict] = []
    seen_urls: set[str] = set()
    for raw_offer in list(result.get("offers") or [])[:20]:
        if not isinstance(raw_offer, dict):
            continue
        currency = str(raw_offer.get("currency") or "RUB").strip().upper()
        if currency not in {"RUB", "RUR", "₽", "РУБ", "РУБ."}:
            continue
        price = _agent_offer_price(raw_offer.get("price"))
        url = _agent_offer_url(raw_offer.get("url"))
        if price is None or not url or url in seen_urls:
            continue
        seen_urls.add(url)
        try:
            confidence = float(raw_offer.get("confidence") or 0.45)
        except (TypeError, ValueError):
            confidence = 0.45
        offers.append(
            {
                "title": re.sub(r"\s+", " ", str(raw_offer.get("title") or "Источник цены")).strip()[:500],
                "price": price,
                "currency": "RUB",
                "unit": re.sub(r"\s+", " ", str(raw_offer.get("unit") or "")).strip()[:80],
                "url": url,
                "evidence": str(raw_offer.get("evidence") or raw_offer.get("snippet") or "").strip()[:1600],
                "observed_at": str(raw_offer.get("observed_at") or "").strip()[:80],
                "published_at": str(raw_offer.get("published_at") or "").strip()[:120],
                "location": re.sub(r"\s+", " ", str(raw_offer.get("location") or "")).strip()[:250],
                "price_scope": re.sub(r"\s+", " ", str(raw_offer.get("price_scope") or "")).strip()[:120],
                "confidence": max(0.0, min(1.0, confidence)),
            }
        )
        if len(offers) >= 10:
            break
    return {
        "schema_version": 1,
        "position_key": expected_position_key,
        "offers": offers,
        "notes": str(result.get("notes") or "").strip()[:2000],
        "observed_at": str(result.get("observed_at") or "").strip()[:80],
    }


def _telegram_cfg() -> tuple[str, str] | None:
    token = (os.environ.get("TELEGRAM_BOT_TOKEN") or "").strip()
    chat = (os.environ.get("TELEGRAM_CHAT_ID") or "").strip()
    if token and chat:
        return token, chat
    return None


def _tg_send(text: str) -> None:
    cfg = _telegram_cfg()
    if not cfg:
        return
    try:
        from autobot.telegram_notify import send_message

        send_message(cfg[0], cfg[1], text, parse_mode="HTML", disable_web_page_preview=False)
    except Exception:
        pass


def _tg_flush_spool() -> None:
    """Догнать outbox Telegram до отправки следующих сообщений (иначе порядок в чате ломается)."""
    cfg = _telegram_cfg()
    if not cfg:
        return
    try:
        from autobot.telegram_notify import flush_spooled_messages

        flush_spooled_messages(cfg[0])
    except Exception:
        pass


def _merge_chat_add(kind: str, text: str, *, tender_id: str = "", seq: int = 0, total: int = 0) -> None:
    event = {
        "ts": datetime.now().isoformat(timespec="seconds"),
        "source": "web",
        "kind": kind,
        "text": (text or "").strip()[:700],
        "tender_id": (tender_id or "").strip(),
        "seq": int(seq or 0),
        "total": int(total or 0),
    }
    if not event["text"]:
        return
    with merge_site_lock:
        events = list(merge_site_state.get("chat_events") or [])
        events.append(event)
        merge_site_state["chat_events"] = events[-120:]


def _market_web_events_path(tender_id: str) -> Path:
    safe = re.sub(r"[^0-9A-Za-z_.-]+", "_", (tender_id or "unknown").strip())[:80] or "unknown"
    return REPO_ROOT / "data" / "logs" / f"market_web_events_{safe}.jsonl"


def _read_market_web_events(tender_id: str, *, limit: int = 80) -> list[dict]:
    paths = [_market_web_events_path(tender_id)]
    raw_lines: list[str] = []
    for path in paths:
        if not path.is_file():
            continue
        try:
            raw_lines.extend(path.read_text(encoding="utf-8", errors="ignore").splitlines()[-limit:])
        except OSError:
            continue
    if not raw_lines:
        return []
    out: list[dict] = []
    for line in raw_lines[-limit * 2 :]:
        line = line.strip()
        if not line:
            continue
        try:
            ev = json.loads(line)
        except Exception:
            continue
        kind = str(ev.get("kind") or "")
        seq = int(ev.get("seq") or 0)
        total = int(ev.get("total") or 0)
        work = str(ev.get("work_name") or "").strip()
        detail = str(ev.get("detail") or "").strip()
        tid = str(ev.get("tender_id") or tender_id or "").strip()
        source = str(ev.get("source") or "").strip().lower()
        if source == "market" and str(ev.get("text") or "").strip():
            text = str(ev.get("text") or "").strip()
        elif kind == "begin":
            text = f"Работа {seq} из {total} началась" + (f": {work}" if work else "")
        elif kind == "done":
            text = f"✅ {seq}/{total} · готово."
        elif kind == "warn":
            text = f"⚠️ {seq}/{total} · пустой ответ" + (f": {work}" if work else "")
        elif kind == "error":
            text = f"⚠️ {seq}/{total} · ошибка" + (f": {detail}" if detail else "")
        else:
            text = str(ev.get("text") or "").strip()
        if not text:
            continue
        out.append(
            {
                "ts": str(ev.get("ts") or ""),
                "source": source or "market",
                "kind": kind,
                "text": text[:700],
                "tender_id": tid,
                "seq": seq,
                "total": total,
            }
        )
    return out


def eis_notice_url(tender_id: str, stored_url: str | None) -> str:
    """Ссылка на карточку закупки: из tenders.json или запасной URL по regNumber."""
    u = (stored_url or "").strip()
    if u.startswith("http://") or u.startswith("https://"):
        return u
    tid = (tender_id or "").strip()
    if not tid:
        return ""
    return (
        "https://zakupki.gov.ru/epz/order/notice/zk20/view/common-info.html"
        f"?regNumber={quote(tid, safe='')}"
    )


INDEX_TEMPLATE = """
<!doctype html>
<html lang="ru">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  {% if embed_mode %}<base target="_top" />{% endif %}
  <link rel="icon" href="/favicon.svg" type="image/svg+xml" />
  <title>Помощник по госзакупкам</title>
  <link rel="icon" type="image/svg+xml" href="data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 64 64'%3E%3Crect x='8' y='10' width='34' height='44' rx='8' fill='%23121a30' stroke='%236db7ff' stroke-width='3'/%3E%3Cpath d='M18 22h14M18 30h14M18 38h10' stroke='%239fd2ff' stroke-width='3' stroke-linecap='round'/%3E%3Ccircle cx='45' cy='42' r='10' fill='none' stroke='%235ecf8a' stroke-width='4'/%3E%3Cpath d='M52 49l6 6' stroke='%235ecf8a' stroke-width='4' stroke-linecap='round'/%3E%3C/svg%3E" />
  <style>
    :root {
      color-scheme: light;
      --bg: #f4f7fb;
      --panel: #ffffff;
      --panel-soft: #f7fafe;
      --border: #d8e2ef;
      --border-soft: #e7edf5;
      --text: #172235;
      --muted: #61748c;
      --muted-soft: #75859a;
      --accent: #1f72dc;
      --accent-2: #195fba;
      --accent-bright: #4d9bff;
      --ok: #2e8b57;
      --danger: #cf5a5a;
      --shadow: 0 16px 42px rgba(28, 49, 84, 0.08);
    }
    html, body { min-height: 100%; margin: 0; box-sizing: border-box; }
    *, *::before, *::after { box-sizing: inherit; }
    body { font-family: "Segoe UI", Arial, sans-serif; background: linear-gradient(180deg, #ffffff 0%, var(--bg) 100%); color: var(--text); }
    .page { max-width: 1220px; margin: 0 auto; padding: 26px 18px 44px; display: flex; flex-direction: column; }
    .page > .hero-title, .page > h1, .page > .sub { order: 0; }
    .page > .action-hub { order: 1; }
    .page > #reportCoverageBanner { order: 2; }
    .page > .tenders-section { order: 3; }
    .page > .help-section { order: 4; }
    .page > .tool-section { order: 5; }
    .hero-title { display: flex; align-items: center; gap: 10px; margin-bottom: 8px; }
    .hero-mark {
      width: 42px; height: 42px; flex: 0 0 42px;
      display: inline-flex; align-items: center; justify-content: center;
      border-radius: 12px;
      background: linear-gradient(180deg, rgba(57, 126, 209, 0.22), rgba(40, 93, 164, 0.3));
      border: 1px solid rgba(109, 183, 255, 0.35);
      box-shadow: inset 0 1px 0 rgba(255,255,255,0.06), 0 10px 24px rgba(0, 0, 0, 0.18);
    }
    .hero-mark svg { width: 26px; height: 26px; display: block; }
    h1 { margin: 0 0 6px 0; font-size: 1.45rem; font-weight: 700; letter-spacing: -0.02em; line-height: 1.15; }
    .section-title { font-size: 1.1rem; font-weight: 700; color: #e0eaff; margin: 0 0 7px 0; letter-spacing: -0.01em; }
    .section-lead { color: var(--muted); font-size: 13px; margin: 0 0 12px 0; line-height: 1.45; max-width: 72ch; }
    .sub { color: var(--muted); font-size: 13px; margin: 0 0 18px 0; line-height: 1.45; max-width: 62ch; }
    .meta { color: var(--muted); font-size: 12px; margin-bottom: 10px; line-height: 1.4; }
    .controls, .filters {
      border: 1px solid var(--border);
      border-radius: 12px;
      background: linear-gradient(180deg, var(--panel), var(--panel-soft));
      box-shadow: var(--shadow);
    }
    .controls { padding: 12px; margin-bottom: 14px; }
    .filters {
      padding: 8px 12px;
      margin-bottom: 10px;
      display: flex;
      flex-wrap: wrap;
      align-items: center;
      gap: 10px 14px;
      font-size: 12px;
      color: #b8c7ea;
    }
    .filters label { display:flex; align-items:center; gap:8px; cursor:pointer; font-weight: 600; }
    .filters input[type="checkbox"] { width: 15px; height: 15px; accent-color: var(--ok); }
    .filters select {
      background:#0b1223;
      border:1px solid var(--border-soft);
      color:var(--text);
      border-radius:8px;
      padding:6px 8px;
      font-size:12px;
      outline:none;
    }
    .filters select:focus {
      border-color: var(--accent);
      box-shadow: 0 0 0 2px rgba(75, 101, 187, 0.2);
    }
    .filters .muted { color: var(--muted-soft); }
    .filters a { color: #87bbff; font-size: 12px; text-decoration: none; }
    .filters a:hover { text-decoration: underline; color: #b8d4ff; }
    .btn {
      border:1px solid var(--accent);
      background: linear-gradient(180deg, #397ed1, #285da4);
      color:#ecf2ff;
      border-radius:8px;
      padding:7px 11px;
      cursor:pointer;
      font-size:12px;
      font-weight: 600;
      transition: transform .14s ease, filter .14s ease;
    }
    .btn:hover { transform: translateY(-1px); filter: brightness(1.08); }
    .btn.secondary { border-color:#4a567e; background: linear-gradient(180deg, #2d3853, #283247); }
    .btn:disabled { opacity:.58; cursor:not-allowed; transform:none; filter:none; }
    .btn-lg { padding: 10px 18px; font-size: 13px; border-radius: 10px; }
    .action-bar { display: flex; flex-wrap: wrap; gap: 10px; align-items: center; margin-top: 4px; }
    .controls-hint { font-size: 12px; color: var(--muted-soft); margin: 10px 0 0 0; line-height: 1.4; }
    details.advanced {
      margin-top: 14px;
      border: 1px solid var(--border-soft);
      border-radius: 11px;
      background: rgba(10, 14, 28, 0.55);
    }
    details.advanced > summary {
      list-style: none;
      cursor: pointer;
      padding: 11px 14px;
      font-size: 12px;
      color: #a8b8e6;
      user-select: none;
    }
    details.advanced > summary::-webkit-details-marker { display: none; }
    details.advanced[open] > summary {
      color: #d2defa;
      border-bottom: 1px solid var(--border-soft);
    }
    details.advanced .advanced-body { padding: 12px 14px 14px; }
    .link-refresh {
      font-size: 12px;
      color: var(--muted-soft);
      margin-left: 6px;
      text-decoration: none;
      align-self: center;
    }
    .link-refresh:hover { color: #c8d8f8; text-decoration: underline; }
    .stat-strip { line-height: 1.5; }
    .page-footer { margin-top: 22px; font-size: 11px; color: var(--muted-soft); text-align: center; }
    .btn-row { display:flex; flex-wrap:wrap; gap:8px; margin-top:8px; align-items:center; }
    .opts { display:grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap:8px; margin-top:8px; font-size:12px; color:#b8c7ea; }
    .opts label { display:flex; flex-direction:column; gap:4px; }
    .opts input, .link-row input, .rebuild-row select {
      background:#0b1223;
      border:1px solid var(--border-soft);
      color:var(--text);
      border-radius:8px;
      padding:7px 9px;
      font-size:12px;
      outline: none;
    }
    .opts input:focus, .link-row input:focus, .rebuild-row select:focus {
      border-color: var(--accent);
      box-shadow: 0 0 0 2px rgba(75, 101, 187, 0.2);
    }
    .link-row, .rebuild-row { display:flex; flex-wrap:wrap; align-items:center; gap:8px; margin-top:10px; font-size:12px; color:#b8c7ea; }
    .link-row input, .rebuild-row select { flex:1; min-width:220px; max-width:100%; }
    .status { margin-top:8px; font-size:12px; color:#b8c7ea; }
    .logs, .tender-grid {
      margin-top:8px;
      max-height:165px;
      overflow:auto;
      border:1px solid var(--border-soft);
      border-radius:10px;
      background:#0b1223;
      padding:8px;
      font-size:12px;
    }
    .logs { font-family: Consolas, monospace; white-space:pre-wrap; }
    .parse-bar-wrap { height: 10px; background: #0b1223; border-radius: 8px; border: 1px solid var(--border-soft); overflow: hidden; margin-top: 6px; }
    .parse-bar-fill { height: 100%; width: 0%; background: linear-gradient(90deg, #3d5290, #5ecf8a); transition: width .28s ease; }
    .parse-bar-fill.running { width: 65%; animation: parseIndeterminate 1.3s ease-in-out infinite; }
    @keyframes parseIndeterminate {
      0% { transform: translateX(-45%); width: 35%; }
      50% { transform: translateX(10%); width: 55%; }
      100% { transform: translateX(120%); width: 30%; }
    }
    .region-block {
      margin-bottom: 12px;
      padding: 10px 12px 12px;
      border: 1px solid var(--border);
      border-radius: 12px;
      background: linear-gradient(180deg, #121a31, #10182c);
      box-shadow: var(--shadow);
    }
    .region-title { font-size: 13px; font-weight: 700; color: #d2defa; margin: 0 0 9px 0; padding-bottom: 6px; border-bottom: 1px solid #243356; }
    .tender-filter-row {
      display: flex;
      flex-wrap: wrap;
      align-items: center;
      gap: 8px 12px;
      margin: 4px 0 14px;
      padding: 10px 12px;
      border: 1px solid var(--border-soft);
      border-radius: 10px;
      background: rgba(8, 12, 24, 0.42);
      color: #b8c7ea;
      font-size: 12px;
    }
    .tender-filter-row label { display: flex; align-items: center; gap: 8px; font-weight: 700; }
    .tender-filter-row select {
      min-width: 230px;
      max-width: 100%;
      background: #0b1223;
      border: 1px solid var(--border-soft);
      color: var(--text);
      border-radius: 8px;
      padding: 7px 9px;
      font-size: 12px;
      outline: none;
    }
    .tender-filter-row select:focus {
      border-color: var(--accent);
      box-shadow: 0 0 0 2px rgba(75, 101, 187, 0.2);
    }
    .tender-filter-row a { color: #87bbff; text-decoration: none; }
    .tender-filter-row a:hover { text-decoration: underline; color: #b8d4ff; }
    .tender-group { margin-top: 14px; }
    .tender-group:first-of-type { margin-top: 8px; }
    .tender-group-title {
      margin: 0 0 10px;
      padding: 0 0 7px;
      border-bottom: 1px solid #243356;
      color: #d2defa;
      font-size: 15px;
      font-weight: 800;
      letter-spacing: -0.01em;
      line-height: 1.3;
    }
    .tender-group-body { margin-top: 0; }
    .tender-grid-main { display: grid; grid-template-columns: repeat(auto-fill, minmax(285px, 1fr)); gap: 9px; }
    .tender-cell { display: flex; flex-direction: column; gap: 5px; position: relative; min-width: 0; }
    .tender-card {
      display: flex;
      flex-direction: column;
      gap: 2px;
      padding: 12px;
      border-radius: 14px;
      border: 1px solid #35508d;
      background:
        linear-gradient(180deg, rgba(109, 183, 255, 0.12), rgba(109, 183, 255, 0) 32%),
        linear-gradient(145deg, #1d294a, #141d34);
      color: var(--text);
      transition: transform .15s ease, border-color .15s, box-shadow .15s, background .15s ease;
      min-height: 0;
      min-width: 0;
      overflow: hidden;
      box-shadow: 0 10px 26px rgba(0,0,0,.22);
    }
    .tender-card:hover {
      transform: translateY(-2px);
      border-color: #81b8ff;
      box-shadow: 0 14px 30px rgba(5, 10, 25, 0.34);
      background:
        linear-gradient(180deg, rgba(109, 183, 255, 0.18), rgba(109, 183, 255, 0.02) 32%),
        linear-gradient(145deg, #212f56, #17213b);
    }
    .tender-card[data-href] { cursor: pointer; }
    .tender-card.no-data { border-left: 3px solid var(--danger); }
    .tender-card-link {
      display: block;
      min-width: 0;
      text-decoration: none;
      color: inherit;
      padding-right: 34px;
    }
    .tender-card-link--more { flex: 1 1 auto; margin-top: 2px; }
    .tender-card .title {
      font-size: 14px;
      font-weight: 750;
      line-height: 1.35;
      max-height: 4.05em;
      overflow: hidden;
      word-break: break-word;
      color: #f4f8ff;
      text-shadow: 0 1px 0 rgba(0,0,0,.18);
    }
    .tender-card-row {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 8px;
      margin-top: 6px;
      min-width: 0;
    }
    .tender-card-sub {
      flex: 1 1 auto;
      min-width: 0;
      text-decoration: none;
      color: inherit;
    }
    .tender-card-sub:hover .tid { color: #c8d8f8; }
    .tender-card .tid { font-size: 11px; color: var(--muted); margin-top: 0; word-break: break-word; line-height: 1.35; }
    .tender-card-meta {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 8px;
      margin-top: 10px;
    }
    .tender-meta-item {
      min-width: 0;
      padding: 9px 10px;
      border-radius: 10px;
      background: linear-gradient(180deg, rgba(17, 28, 53, 0.92), rgba(9, 16, 31, 0.86));
      border: 1px solid rgba(109, 183, 255, 0.2);
      box-shadow: inset 0 1px 0 rgba(255,255,255,0.03);
    }
    .tender-meta-item--wide { grid-column: 1 / -1; }
    .tender-meta-label {
      display: block;
      font-size: 10px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.07em;
      color: #92a7d6;
      margin-bottom: 5px;
    }
    .tender-meta-value {
      display: block;
      font-size: 13px;
      color: #edf3ff;
      line-height: 1.35;
      word-break: break-word;
      font-weight: 650;
    }
    .tender-meta-value--mono { font-variant-numeric: tabular-nums; color: #d9e7ff; }
    .tender-card-pub {
      display: flex;
      flex-wrap: wrap;
      align-items: center;
      gap: 6px 10px;
      margin-top: 11px;
      padding: 8px 10px;
      border-radius: 11px;
      background: rgba(10, 18, 34, 0.46);
      border: 1px solid rgba(109, 183, 255, 0.14);
    }
    .tender-card-pub-label {
      font-size: 9px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.07em;
      color: #7d8fbb;
    }
    .tender-card-pub-date {
      display: inline-block;
      font-size: 12px;
      font-weight: 600;
      font-variant-numeric: tabular-nums;
      color: #f0f7ff;
      background: linear-gradient(180deg, rgba(88, 118, 210, 0.42), rgba(52, 72, 140, 0.55));
      border: 1px solid rgba(130, 160, 230, 0.45);
      border-radius: 999px;
      padding: 4px 11px;
      line-height: 1.2;
      box-shadow: 0 0 0 1px rgba(0,0,0,.12) inset;
    }
    .tender-status-row {
      display: flex;
      flex-wrap: wrap;
      align-items: center;
      gap: 6px 8px;
      margin-top: 10px;
      min-width: 0;
    }
    .tender-status-row .eis-in-card { margin-left: auto; }
    .tender-progress {
      margin-top: 8px;
      padding: 7px 8px;
      border-radius: 9px;
      background: rgba(255,255,255,0.08);
      border: 1px solid rgba(140, 172, 220, 0.2);
    }
    .tender-progress-head {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 8px;
      font-size: 10px;
      color: #c7d6ef;
      margin-bottom: 5px;
    }
    .tender-progress-label { font-weight: 700; letter-spacing: 0.02em; }
    .tender-progress-value { color: #ecf2ff; font-weight: 700; font-variant-numeric: tabular-nums; }
    .tender-progress-track {
      height: 6px;
      border-radius: 999px;
      overflow: hidden;
      background: rgba(255,255,255,0.08);
      border: 1px solid rgba(255,255,255,0.04);
    }
    .tender-progress-fill {
      height: 100%;
      width: 0%;
      background: linear-gradient(90deg, #4b7dff, #5ecf8a);
    }
    .tender-progress-note {
      margin-top: 5px;
      font-size: 10px;
      color: #9fb2d7;
      line-height: 1.35;
    }
    .eis-in-card {
      display: inline-flex;
      align-items: center;
      flex-shrink: 0;
      font-size: 11px;
      font-weight: 600;
      color: #b8d8ff;
      text-decoration: none;
      padding: 3px 8px;
      border-radius: 7px;
      background: rgba(20, 32, 64, 0.65);
      border: 1px solid rgba(100, 140, 220, 0.35);
      white-space: nowrap;
    }
    .eis-in-card:hover { background: rgba(75, 101, 187, 0.35); color: #fff; border-color: rgba(140, 175, 255, 0.55); }
    .tag { font-size: 11px; padding: 4px 8px; border-radius: 999px; font-weight: 700; letter-spacing: .1px; }
    .tag-ok { background: #1e4d35; color: #9df0b8; }
    .tag-nodata { background: #5a1a22; color: #ffc9cc; border: 1px solid #a04048; }
    .tag-stage-open { background: #1e4d35; color: #9df0b8; border: 1px solid #3d8a67; }
    .tag-stage-closed { background: #5a1a22; color: #ffc9cc; border: 1px solid #a04048; }
    .tender-menu-wrap { position: absolute; top: 7px; right: 7px; z-index: 3; }
    .tender-menu-btn {
      width: 30px;
      height: 30px;
      border-radius: 8px;
      border: 1px solid #d7e2ec;
      background: #ffffff;
      color: #1f334d;
      cursor: pointer;
      font-weight: 700;
      font-size: 16px;
      line-height: 1;
      display: flex;
      align-items: center;
      justify-content: center;
      box-shadow: 0 4px 12px rgba(15, 23, 42, 0.08);
    }
    .tender-menu-btn:hover { background: #f7fafc; border-color: #c3d4e6; color: #10263d; }
    .tender-menu-wrap > summary {
      list-style: none;
      display: block;
      cursor: pointer;
      user-select: none;
    }
    .tender-menu-wrap > summary::-webkit-details-marker { display: none; }
    .tender-menu {
      display: none;
      position: absolute;
      top: 34px;
      right: 0;
      min-width: 260px;
      background: #ffffff;
      border: 1px solid #d7e2ec;
      border-radius: 10px;
      padding: 6px;
      box-shadow: 0 14px 30px rgba(15, 23, 42, 0.14);
    }
    .tender-menu-wrap[open] .tender-menu,
    .tender-menu-wrap.menu-open .tender-menu { display: block; }
    .tender-menu button {
      width: 100%;
      text-align: left;
      background: transparent;
      color: #1f334d;
      border: none;
      padding: 8px;
      border-radius: 8px;
      cursor: pointer;
      font-size: 12px;
    }
    .tender-menu button:hover { background: #f4f8fc; }
    .parse-progress-panel {
      margin-top: 18px; padding: 18px 20px; border-radius: 15px;
      background: linear-gradient(135deg, rgba(34, 57, 101, 0.96), rgba(18, 29, 53, 0.98));
      border: 1px solid rgba(109, 183, 255, 0.58);
      box-shadow: 0 14px 34px rgba(0, 0, 0, 0.28);
    }
    .parse-progress-panel[hidden] { display: none !important; }
    .parse-progress-head { display: flex; align-items: center; gap: 12px; margin-bottom: 8px; font-size: 16px; color: #e4edff; }
    .parse-pulse { width: 12px; height: 12px; border-radius: 50%; background: #5ecf8a; flex-shrink: 0; animation: parsePulse 1.2s ease-in-out infinite; box-shadow: 0 0 12px #5ecf8a; }
    @keyframes parsePulse { 0%, 100% { opacity: 1; transform: scale(1); } 50% { opacity: 0.55; transform: scale(0.92); } }
    .parse-progress-time { margin-top: 9px; font-size: 14px; color: #9df0b8; font-variant-numeric: tabular-nums; }
    .parse-progress-hint { font-size: 12px; color: #b4c4e5; margin-top: 8px; line-height: 1.45; }
    .parse-status-line { margin-top: 8px; font-size: 12px; color: #9aabd0; word-break: break-all; }
    .parse-summary {
      margin-top: 14px;
      padding: 14px 15px;
      border-radius: 12px;
      background: rgba(9, 16, 31, 0.42);
      border: 1px solid rgba(109, 183, 255, 0.2);
    }
    .parse-summary-grid { display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 10px; }
    .parse-summary-item {
      padding: 11px 12px;
      border-radius: 10px;
      background: rgba(255,255,255,0.03);
      border: 1px solid rgba(255,255,255,0.05);
      min-width: 0;
    }
    .parse-summary-label {
      font-size: 11px;
      color: var(--muted-soft);
      text-transform: uppercase;
      letter-spacing: 0.04em;
      margin-bottom: 6px;
    }
    .parse-summary-value { font-size: 14px; color: #ecf2ff; line-height: 1.4; word-break: break-word; }
    .parse-summary-value.ok { color: #9df0b8; }
    .parse-summary-value.warn { color: #ffd7a8; }
    .parse-summary-value.bad { color: #ffc9cc; }
    details.compact-details {
      margin-top: 12px;
      border: 1px solid rgba(109, 183, 255, 0.18);
      border-radius: 10px;
      background: rgba(9, 16, 31, 0.24);
    }
    details.compact-details > summary {
      list-style: none;
      cursor: pointer;
      padding: 10px 12px;
      color: #b8c7ea;
      font-size: 13px;
      user-select: none;
    }
    details.compact-details > summary::-webkit-details-marker { display: none; }
    details.compact-details .logs { margin: 0 10px 10px; }
    details.compact-details .parse-status-line { margin: 0 10px 8px; }
    .merge-bar-wrap { height: 12px; background: #0f1324; border-radius: 8px; overflow: hidden; margin-top: 10px; border: 1px solid #2b365e; }
    .merge-bar-fill { height: 100%; background: linear-gradient(90deg, #3d5290, #5ecf8a); transition: width 0.35s ease; border-radius: 8px; }
    .merge-logs { margin-top: 8px; max-height: 140px; overflow: auto; border: 1px solid #2b365e; border-radius: 8px; background: #0f1324; padding: 8px; font-family: Consolas, monospace; font-size: 11px; white-space: pre-wrap; }
    .site-chat-fab {
      position: fixed; right: 18px; bottom: 18px; z-index: 50;
      width: 54px; height: 54px; border-radius: 999px;
      border: 1px solid rgba(109, 183, 255, 0.62);
      background: linear-gradient(180deg, #397ed1, #285da4);
      color: #fff; cursor: pointer; box-shadow: 0 14px 32px rgba(0,0,0,.38);
      display: flex; align-items: center; justify-content: center; font-size: 23px;
    }
    .site-chat-fab.has-new::after {
      content: ""; position: absolute; right: 7px; top: 7px;
      width: 10px; height: 10px; border-radius: 999px; background: #5ecf8a;
      box-shadow: 0 0 0 3px rgba(94,207,138,.22);
    }
    .site-chat-panel {
      position: fixed; right: 18px; bottom: 84px; z-index: 49;
      width: min(390px, calc(100vw - 28px)); max-height: min(560px, calc(100vh - 120px));
      border: 1px solid rgba(109, 183, 255, 0.42);
      border-radius: 15px; overflow: hidden;
      background: linear-gradient(180deg, #121a30, #0e1528);
      box-shadow: 0 18px 46px rgba(0,0,0,.46);
    }
    .site-chat-panel[hidden] { display: none !important; }
    .site-chat-head {
      display: flex; align-items: center; justify-content: space-between; gap: 10px;
      padding: 11px 12px; border-bottom: 1px solid var(--border-soft);
      color: #e4edff; font-size: 13px; font-weight: 750;
    }
    .site-chat-close {
      border: 1px solid #3a4677; background: rgba(15,19,36,.85);
      color: #c8d8f8; border-radius: 8px; cursor: pointer; padding: 4px 8px;
    }
    .site-chat-feed {
      max-height: 430px; overflow: auto; padding: 10px;
      display: flex; flex-direction: column; gap: 8px;
    }
    .site-chat-empty { color: var(--muted-soft); font-size: 12px; line-height: 1.45; padding: 4px 2px 8px; }
    .site-chat-msg {
      border: 1px solid rgba(109, 183, 255, 0.13); border-radius: 11px;
      background: rgba(8, 12, 24, 0.48); padding: 8px 9px;
    }
    .site-chat-msg.is-done { border-color: rgba(94,207,138,.32); }
    .site-chat-msg.is-error, .site-chat-msg.is-warn { border-color: rgba(255,201,204,.28); }
    .site-chat-meta { color: #7d8fbb; font-size: 10px; margin-bottom: 4px; font-variant-numeric: tabular-nums; }
    .site-chat-text { color: #edf3ff; font-size: 12px; line-height: 1.4; white-space: pre-wrap; word-break: break-word; }
    .cov-banner { padding: 13px 16px; border-radius: 12px; margin-bottom: 16px; font-size: 13px; line-height: 1.5; }
    .cov-warn { background: rgba(90, 26, 34, 0.45); border: 1px solid #a04048; color: #ffc9cc; }
    .cov-partial { background: rgba(77, 53, 30, 0.45); border: 1px solid #8a623d; color: #ffd7a8; }
    .cov-ok { background: rgba(30, 77, 53, 0.35); border: 1px solid #3d8a67; color: #9df0b8; }
    .workflow-strip {
      display: none; flex-wrap: wrap; align-items: center; gap: 8px 12px;
      margin-bottom: 20px; padding: 13px 15px; border-radius: 12px;
      background: rgba(10, 14, 28, 0.55); border: 1px solid var(--border-soft);
      font-size: 13px; color: #c4d2ef;
    }
    .wf-step { display: flex; align-items: center; gap: 8px; font-weight: 600; }
    .wf-num {
      display: inline-flex; align-items: center; justify-content: center;
      width: 28px; height: 28px; border-radius: 999px;
      background: linear-gradient(180deg, #397ed1, #285da4); color: #ecf2ff;
      font-size: 12px; font-weight: 800;
    }
    .wf-arrow { color: #607dce; font-weight: 700; }
    .action-hub, .tenders-section, .tool-section {
      margin-bottom: 14px;
      padding: 14px;
      border: 1px solid var(--border);
      border-radius: 12px;
      background: linear-gradient(180deg, var(--panel), var(--panel-soft));
      box-shadow: var(--shadow);
    }
    .action-hub {
      position: relative;
      overflow: hidden;
      border-color: rgba(109, 183, 255, 0.48);
      background:
        linear-gradient(135deg, rgba(25, 42, 77, 0.98), rgba(13, 23, 44, 0.99));
    }
    .action-hub::before {
      content: ""; position: absolute; inset: 0 auto auto 0; width: 100%; height: 3px;
      background: linear-gradient(90deg, var(--accent-bright), var(--ok), transparent 78%);
    }
    .action-grid { display: grid; grid-template-columns: repeat(12, minmax(0, 1fr)); gap: 10px; }
    .action-card {
      grid-column: span 6;
      padding: 12px;
      border-radius: 10px;
      border: 1px solid var(--border-soft);
      background: linear-gradient(145deg, rgba(18, 29, 53, 0.92), rgba(10, 17, 33, 0.9));
      box-shadow: 0 8px 20px rgba(0, 0, 0, 0.18);
    }
    .action-card:nth-child(1) { border-top: 3px solid var(--accent-bright); }
    .action-card:nth-child(2) { border-top: 3px solid var(--ok); }
    .action-card--wide { grid-column: span 8; }
    .action-card:last-child { grid-column: span 4; }
    .action-card-title { margin: 0 0 7px 0; font-size: 14px; font-weight: 700; color: #e2ebff; line-height: 1.35; }
    .action-card-desc { margin: 0 0 10px 0; font-size: 12px; color: #9fb0d6; line-height: 1.45; }
    .action-card .btn-row { margin-top: 0; }
    .action-card .opts { margin-top: 10px; }
    .action-card > .btn.btn-lg { width: 100%; }
    .action-card .btn-row .btn-lg { flex: 1 1 260px; }
    .tender-actions {
      display: flex; flex-direction: column; gap: 8px;
      margin-top: 12px; padding-top: 12px;
      border-top: 1px solid rgba(76, 108, 181, 0.45);
    }
    .tender-act {
      display: inline-flex; align-items: center; justify-content: center;
      border: 1px solid #3a4677; background: rgba(15, 19, 36, 0.85);
      color: #d5e4ff; border-radius: 11px; padding: 9px 11px;
      font-size: 12px; font-weight: 700; cursor: pointer; text-decoration: none;
      line-height: 1.25; text-align: center;
    }
    .tender-act:hover { background: rgba(75, 101, 187, 0.35); color: #fff; border-color: rgba(140, 175, 255, 0.55); }
    .tender-act:disabled { opacity: 0.45; cursor: not-allowed; transform: none; }
    .tender-act--primary { border-color: var(--accent); background: linear-gradient(180deg, #334b93, #2a3f82); color: #ecf2ff; }
    .tender-act--main { width: 100%; min-height: 43px; font-size: 13px; padding: 10px 12px; }
    .tender-act--crm {
      width: 100%;
      min-height: 46px;
      font-size: 14px;
      border-color: #ffb64d;
      background: linear-gradient(180deg, #ffb84f, #ea8e1f);
      color: #241300;
      box-shadow: 0 10px 22px rgba(234, 142, 31, 0.28);
    }
    .tender-act--crm:hover {
      background: linear-gradient(180deg, #ffc364, #f39b27);
      border-color: #ffd089;
      color: #1b0f00;
      box-shadow: 0 12px 24px rgba(243, 155, 39, 0.34);
    }
    .tender-more-actions {
      display: grid; grid-template-columns: 1fr; gap: 5px; padding: 7px;
    }
    .tender-more-actions .tender-act { width: 100%; justify-content: flex-start; text-align: left; }
    .tender-menu .tender-act,
    .tender-menu .tender-act--primary,
    .tender-menu .tender-act--crm,
    .tender-menu .tender-act--main,
    .tender-menu .tender-act-btn {
      width: 100%;
      min-height: 0;
      justify-content: flex-start;
      text-align: left;
      padding: 9px 10px;
      border-radius: 8px;
      border: 1px solid #d7e2ec;
      background: #ffffff;
      color: #1f334d;
      box-shadow: none;
      font-size: 12px;
      font-weight: 600;
    }
    .tender-menu .tender-act:hover,
    .tender-menu .tender-act--primary:hover,
    .tender-menu .tender-act--crm:hover,
    .tender-menu .tender-act--main:hover,
    .tender-menu .tender-act-btn:hover {
      background: #f4f8fc;
      border-color: #c3d4e6;
      color: #10263d;
    }
    .tender-menu .tender-act:focus,
    .tender-menu .tender-act-btn:focus,
    .tender-menu-btn:focus {
      outline: 2px solid #c9ddff;
      outline-offset: 1px;
    }
    .tender-next { display: none; }
    details.tender-more { display: none; }
    .tender-card-link--disabled { cursor: default; }
    .tag-merge { background: #2a3a6e; color: #b8d4ff; border: 1px solid #4a67b8; }
    .tag-nomerge { background: #3a3048; color: #d0c4e8; border: 1px solid #5a4a72; }
    .help-section { display: none; }
    .help-section .section-title { color: #e8f0ff; margin-bottom: 8px; }
    .help-steps { margin: 0 0 14px 0; padding-left: 22px; color: #c8d8f8; font-size: 14px; line-height: 1.6; }
    .help-steps li { margin-bottom: 6px; }
    .help-steps strong { color: #fff; }
    details.help-glossary {
      border: 1px solid var(--border-soft); border-radius: 10px;
      background: rgba(8, 12, 24, 0.5); font-size: 12px; color: #b8c7ea;
    }
    details.help-glossary > summary {
      cursor: pointer; padding: 10px 12px; font-weight: 600; color: #a8c4ff;
      list-style: none;
    }
    details.help-glossary > summary::-webkit-details-marker { display: none; }
    details.help-glossary[open] > summary { border-bottom: 1px solid var(--border-soft); }
    .glossary-grid { display: grid; grid-template-columns: 110px 1fr; gap: 6px 12px; padding: 10px 12px 12px; line-height: 1.45; }
    .glossary-term { font-weight: 700; color: #d2defa; }
    .btn-effect {
      margin: 8px 0 0 0; padding: 8px 10px; border-radius: 8px;
      background: rgba(15, 22, 44, 0.85); border: 1px dashed #3a4677;
      font-size: 11px; color: #9fb0d6; line-height: 1.45;
    }
    .btn-effect strong { color: #c8e0ff; font-weight: 600; }
    .card-legend {
      margin: 0 0 12px 0; padding: 10px 12px; border-radius: 10px;
      border: 1px solid var(--border-soft); background: rgba(8, 12, 24, 0.45);
      font-size: 11px; color: #9fb0d6; line-height: 1.5;
    }
    .card-legend strong { color: #d2defa; }
    .tool-section--optional { border-style: dashed; opacity: 0.95; }
    .optional-badge {
      display: inline-block; font-size: 10px; font-weight: 700; text-transform: uppercase;
      letter-spacing: 0.06em; color: #ffd7a8; background: rgba(77, 53, 30, 0.5);
      border: 1px solid #8a623d; border-radius: 999px; padding: 2px 8px; margin-left: 8px;
    }
    .main-tabs {
      display:flex; flex-wrap:wrap; gap:8px; margin: 14px 0 18px;
    }
    .main-tab {
      display:inline-flex; align-items:center; gap:7px;
      padding:9px 12px; border-radius:999px; text-decoration:none;
      color:#c8d8f8; background:rgba(15, 22, 44, .72);
      border:1px solid var(--border-soft); font-size:13px; font-weight:700;
    }
    .main-tab:hover { color:#fff; border-color:#6d8fe8; background:rgba(49, 78, 145, .45); }
    .main-tab.is-active { color:#fff; background:linear-gradient(180deg, #345095, #263d78); border-color:#6d8fe8; }
    details.action-options {
      margin-top: 13px;
      border: 1px solid var(--border-soft);
      border-radius: 9px;
      background: rgba(8, 12, 24, 0.42);
    }
    details.action-options > summary {
      list-style: none; cursor: pointer; padding: 11px 13px;
      font-size: 12px; font-weight: 600; color: #a8b8e6; user-select: none;
    }
    details.action-options > summary::-webkit-details-marker { display: none; }
    details.action-options > summary::before { content: "Показать: "; color: #7891cc; }
    details.action-options[open] > summary {
      border-bottom: 1px solid var(--border-soft); color: #d2defa;
    }
    details.action-options .opts,
    details.action-options .rebuild-row { margin: 0; padding: 10px; }
    .hero-mark {
      background: linear-gradient(180deg, #eef5ff, #f8fbff);
      border-color: #bfd4ef;
      box-shadow: inset 0 1px 0 rgba(255,255,255,0.9), 0 10px 24px rgba(35, 74, 135, 0.08);
    }
    .section-title,
    .help-section .section-title,
    .glossary-term,
    .region-title,
    .tender-card .title,
    .tender-meta-value,
    .parse-progress-head,
    .parse-summary-value,
    .site-chat-text,
    .action-card-title,
    .help-steps strong,
    .tender-meta-value--mono {
      color: #1b2a41;
    }
    .controls,
    .filters,
    .action-card,
    .parse-progress,
    .region-card,
    .tender-card,
    .tender-meta-item,
    .tender-card-pub,
    .tender-progress,
    .merge-status-card,
    .site-chat,
    .workflow-note,
    .help-note,
    .glossary-card,
    .btn-effect,
    .card-legend,
    details.advanced,
    details.action-options,
    .tender-menu,
    .merge-logs,
    .status,
    .logs {
      background: #ffffff;
      border-color: var(--border);
      box-shadow: var(--shadow);
      color: var(--text);
    }
    .filters,
    .meta,
    .sub,
    .section-lead,
    .controls-hint,
    .page-footer,
    .tender-next,
    .parse-progress-hint,
    .parse-status-line,
    .parse-summary-label,
    .tender-card .tid,
    .tender-card-pub-label,
    .tender-meta-label,
    .site-chat-meta,
    .help-steps,
    .card-legend,
    .btn-effect,
    .status-line,
    .market-links-note {
      color: var(--muted);
    }
    .filters select,
    .opts input,
    .link-row input,
    .rebuild-row select,
    .tender-filter-row select,
    .tender-filter-row input,
    .type-picker,
    textarea,
    input[type="text"],
    input[type="file"],
    select {
      background: #ffffff;
      border-color: #cfd9e8;
      color: var(--text);
    }
    .btn,
    .main-tab.is-active,
    .tender-act--primary,
    .market-link-chip,
    .chip.is-active {
      background: linear-gradient(180deg, #2e80e8, #1d6fdc);
      border-color: #2e80e8;
      color: #ffffff;
    }
    .btn.secondary,
    .main-tab,
    .tender-act,
    .chip,
    .eis-in-card,
    .offer-source,
    .workflow-pill,
    .upload-step,
    .tag,
    .tag-merge,
    .tag-nomerge {
      background: #f4f8fd;
      border-color: #cfd9e8;
      color: #35506f;
    }
    .main-tab:hover,
    .tender-act:hover,
    .market-link-chip:hover,
    .chip:hover,
    .eis-in-card:hover {
      background: #eaf2fd;
      border-color: #9ec0ef;
      color: #173a65;
    }
    .tender-card,
    .tender-meta-item,
    .tender-card-pub,
    .parse-summary-item,
    .action-card {
      background: linear-gradient(180deg, #ffffff, #f8fbff);
      border-color: #d9e4f1;
    }
    .tender-card:hover {
      background: linear-gradient(180deg, #ffffff, #f2f7fd);
      border-color: #9ec0ef;
      box-shadow: 0 18px 34px rgba(43, 78, 131, 0.12);
    }
    .tag-ok,
    .tag-stage-open,
    .cov-ok {
      background: #e9f8ef;
      color: #257347;
      border-color: #bfe5cc;
    }
    .tag-nodata,
    .tag-stage-closed,
    .cov-warn {
      background: #fff1f1;
      color: #a94444;
      border-color: #f0c5c5;
    }
    .cov-partial {
      background: #fff8e8;
      color: #91621c;
      border-color: #f0deb1;
    }
    .parse-bar-wrap,
    .merge-bar-wrap {
      background: #edf3fa;
      border-color: #d6e0ee;
    }
    .merge-logs,
    .logs,
    .site-chat-msg,
    .parse-summary-item,
    .status-box {
      background: #f8fbff;
      border-color: #dfe7f1;
      color: var(--text);
    }
    .opts,
    .link-row,
    .rebuild-row,
    .status {
      color: var(--muted);
    }
    details.advanced > summary,
    details.action-options > summary {
      color: #35506f;
    }
    details.action-options > summary::before,
    details.advanced[open] > summary,
    .wf-arrow,
    .where {
      color: #5f7ca5;
    }
    @media (max-width: 980px) {
      .action-grid { grid-template-columns: 1fr; }
      .action-card, .action-card--wide, .action-card:last-child { grid-column: 1 / -1; }
      .opts { grid-template-columns: 1fr 1fr; }
    }
    @media (max-width: 720px) {
      .page { padding: 22px 12px 32px; }
      .hero-title { gap: 10px; }
      .hero-mark { width: 46px; height: 46px; flex-basis: 46px; border-radius: 14px; }
      .hero-mark svg { width: 28px; height: 28px; }
      h1 { font-size: 2rem; }
      .action-hub, .tenders-section, .tool-section { padding: 15px; border-radius: 14px; }
      .action-card { padding: 15px; }
      .parse-summary-grid { grid-template-columns: 1fr; }
      .opts { grid-template-columns: 1fr; }
      .tender-grid-main { grid-template-columns: 1fr; }
      .btn-row .btn { width: 100%; }
      .link-row, .rebuild-row { align-items: stretch; flex-direction: column; }
      .link-row input, .rebuild-row select { width: 100%; min-width: 0; }
    }
  </style>
</head>
<body>
  <div class="page">
    <div class="hero-title">
      <span class="hero-mark" aria-hidden="true">
        <svg viewBox="0 0 64 64" fill="none" xmlns="http://www.w3.org/2000/svg">
          <rect x="10" y="9" width="30" height="42" rx="8" fill="#121a30" stroke="#6db7ff" stroke-width="3"/>
          <path d="M19 22H31" stroke="#9fd2ff" stroke-width="3" stroke-linecap="round"/>
          <path d="M19 30H31" stroke="#9fd2ff" stroke-width="3" stroke-linecap="round"/>
          <path d="M19 38H27" stroke="#9fd2ff" stroke-width="3" stroke-linecap="round"/>
          <circle cx="45" cy="42" r="10" stroke="#5ecf8a" stroke-width="4"/>
          <path d="M52 49L58 55" stroke="#5ecf8a" stroke-width="4" stroke-linecap="round"/>
        </svg>
      </span>
      <h1>Помощник по госзакупкам</h1>
    </div>
    <p class="sub" style="max-width:none;">Программа ищет закупки на <strong>zakupki.gov.ru</strong>, вытаскивает из документов <strong>смету</strong> (список работ и цен), ищет <strong>рыночные источники</strong> в интернете и показывает, где заказчик завысил или занизил.</p>
    <nav class="main-tabs" aria-label="Разделы сайта">
      <a class="main-tab is-active" href="/tenders">📋 Тендеры</a>
      <a class="main-tab" href="/estimates">📊 Сметы</a>
      <a class="main-tab" href="/research">🔎 Поиск по позиции</a>
    </nav>

    <section class="help-section" aria-labelledby="helpTitle">
      <h2 class="section-title" id="helpTitle">Как пользоваться — три шага</h2>
      <ol class="help-steps">
        <li><strong>Шаг 1.</strong> Нажмите «Найти новые закупки» — программа скачает документы и извлечёт сметы.</li>
        <li><strong>Шаг 2.</strong> Нажмите «Подготовить недостающие сравнения» — программа найдёт рыночные цены и ссылки на источники. <strong>Это долгий этап</strong>, он может идти часами.</li>
        <li><strong>Шаг 3.</strong> В готовой карточке нажмите «Посмотреть сравнение цен».</li>
      </ol>
      <details class="help-glossary">
        <summary>Словарь: что значат непонятные слова</summary>
        <div class="glossary-grid">
          <div class="glossary-term">Тендер</div>
          <div>Государственная закупка: кто дешевле выполнит работы — тот выиграет контракт.</div>
          <div class="glossary-term">ЕИС</div>
          <div>Официальный портал <strong>zakupki.gov.ru</strong>. «Скачать с ЕИС» = скачать с этого сайта.</div>
          <div class="glossary-term">Смета</div>
          <div>Таблица из документов: какие работы, объёмы и цены заложил заказчик.</div>
          <div class="glossary-term">Рыночные источники</div>
          <div>Объявления и страницы в интернете, откуда берём примерные цены по позициям сметы.</div>
          <div class="glossary-term">Сравнение цен</div>
          <div>Готовая страница: цена заказчика рядом с найденными рыночными ценами. Это главный результат работы.</div>
          <div class="glossary-term">НМЦК</div>
          <div>Максимальная цена контракта — сколько заказчик готов заплатить. Блок внизу страницы — <strong>отдельный инструмент</strong>, к шагам 1–3 не относится.</div>
          <div class="glossary-term">Telegram</div>
          <div>Кнопка на карточке шлёт краткий вывод «выгодно / невыгодно» в ваш чат (если настроен бот).</div>
        </div>
      </details>
    </section>

    <div id="reportCoverageBanner" class="cov-banner stat-strip {% if coverage.tender_count == 0 %}cov-warn{% elif coverage.tenders_missing_merge_html > 0 %}{% if coverage.merge_html_among_tenders == 0 and coverage.svodka_xlsx_count == 0 %}cov-warn{% else %}cov-partial{% endif %}{% else %}cov-ok{% endif %}">
      {% if coverage.tender_count == 0 %}
      Закупок в базе пока нет. Нажмите «Найти новые закупки» в верхней панели.
      {% else %}
      {% if coverage.merge_html_among_tenders >= coverage.tender_count %}
      Все {{ coverage.tender_count }} закупок имеют готовую страницу сравнения «смета vs рынок».
      {% else %}
      В базе <strong>{{ coverage.tender_count }}</strong> закупок · готовых страниц сравнения: <strong>{{ coverage.merge_html_among_tenders }}</strong>
      {% if coverage.tenders_missing_merge_html > 0 %}
      · ещё <strong>{{ coverage.tenders_missing_merge_html }}</strong> ждут шага 2 («Подготовить недостающие сравнения»)
      {% if coverage.missing_no_svodka > 0 and coverage.missing_no_estimate == 0 %}
      — у {{ coverage.missing_no_svodka }} смета уже есть, но рыночные источники ещё не собирались
      {% endif %}
      {% endif %}
      {% endif %}
      {% endif %}
    </div>

    <section class="tenders-section" aria-labelledby="tendersTitle">
      <h2 class="section-title" id="tendersTitle">Список тендеров</h2>
      <p class="section-lead">В каждой карточке показан один рекомендуемый следующий шаг. Повторные и служебные операции находятся в «Дополнительных действиях».</p>
      <form class="tender-filter-row" method="get" action="/tenders">
        {% if show_all %}<input type="hidden" name="all" value="1">{% endif %}
        <input type="hidden" name="sort" value="{{ sort_mode }}">
        <label>
          Регион
          <select name="region" onchange="this.form.submit()">
            <option value="" {% if not selected_region %}selected{% endif %}>Все регионы</option>
            {% for region in region_options %}
            <option value="{{ region }}" {% if selected_region == region %}selected{% endif %}>{{ region }}</option>
            {% endfor %}
          </select>
        </label>
        {% if selected_region %}
        <a href="/tenders?sort={{ sort_mode }}{% if show_all %}&all=1{% endif %}">сбросить регион</a>
        {% endif %}
      </form>
      <p class="section-lead" style="margin-top:-4px;">
        Показано <strong>{{ visible_count }}</strong> из <strong>{{ tender_count }}</strong> тендеров
        {% if selected_region %}
        · регион: <strong>{{ selected_region }}</strong>
        {% endif %}
        {% if show_all %}
        · все этапы
        · <a href="/tenders?sort={{ sort_mode }}{% if selected_region %}&region={{ selected_region|urlencode }}{% endif %}" style="color:#87bbff;">показать только «Подача заявок»</a>
        {% else %}
        · только этап «Подача заявок»
        · <a href="/tenders?all=1&sort={{ sort_mode }}{% if selected_region %}&region={{ selected_region|urlencode }}{% endif %}" style="color:#87bbff;">показать все этапы</a>
        {% endif %}
      </p>

      <div class="tender-grid-main">
        {% for t in items %}
          <div class="tender-cell">
          <div class="tender-card{% if not t.has_estimate %} no-data{% endif %}" data-href="/merge-report/{{ t.tender_id }}/">
            <details class="tender-menu-wrap">
              <summary class="tender-menu-btn" title="Дополнительные действия">&#9776;</summary>
              <div class="tender-menu">
                <div class="tender-more-actions">
                  <button type="button" class="tender-act tender-act-btn" data-tid="{{ t.tender_id }}" onclick="runFullForTender('{{ t.tender_id }}')" title="Продолжить поиск недостающих рыночных цен и заново собрать страницу сравнения.">Продолжить или обновить поиск цен</button>
                  <button type="button" class="tender-act tender-act-btn" data-tid="{{ t.tender_id }}" onclick="exportTenderToCrm('{{ t.tender_id }}')" title="Создать объект в PM.bi CRM и перенести туда строки сметы как материалы.">Добавить в объекты</button>
                  <button type="button" class="tender-act tender-act-btn" data-tid="{{ t.tender_id }}" onclick="rerunMarketForTender('{{ t.tender_id }}')" title="Удалить прогресс поиска цен и опросить Алису по всем позициям заново.">Начать поиск цен заново</button>
                  <button type="button" class="tender-act tender-act-btn" data-tid="{{ t.tender_id }}" onclick="rebuildReportForTender('{{ t.tender_id }}')" title="Повторно прочитать уже скачанные документы. Поиск рыночных цен не запускается.">Повторно извлечь смету из файлов</button>
                  {% if t.has_estimate %}
                  <a class="tender-act" href="/tenders/{{ t.tender_id }}/estimate.xlsx">Скачать Excel сметы</a>
                  {% endif %}
                  {% if t.has_market_partial %}
                  <a class="tender-act" href="/tenders/{{ t.tender_id }}/market-sources.xlsx">Скачать источники рынка</a>
                  {% endif %}
                  {% if t.has_svodka %}
                  <a class="tender-act" href="/tenders/{{ t.tender_id }}/svodka.xlsx">Скачать Excel выгодности</a>
                  <button type="button" class="tender-act tender-act-btn" data-tid="{{ t.tender_id }}" onclick="runViabilityOnly('{{ t.tender_id }}')" title="Обновить вывод о выгодности и отправить его в настроенный Telegram-чат.">Отправить вывод в Telegram</button>
                  {% endif %}
                </div>
              </div>
            </details>
            {% if t.has_merge_report %}
            <a class="tender-card-link" href="/merge-report/{{ t.tender_id }}/">
              <div class="title">{{ t.display_title }}</div>
            </a>
            {% else %}
            <div class="tender-card-link tender-card-link--disabled">
              <div class="title">{{ t.display_title }}</div>
            </div>
            {% endif %}
            <div class="tid">{{ t.tender_id }}</div>

            <div class="tender-card-meta">
              <div class="tender-meta-item">
                <span class="tender-meta-label">&#1062;&#1077;&#1085;&#1072; &#1090;&#1077;&#1085;&#1076;&#1077;&#1088;&#1072;</span>
                <span class="tender-meta-value">{{ t.price_fmt }}</span>
              </div>
              <div class="tender-meta-item">
                <span class="tender-meta-label">&#1056;&#1077;&#1075;&#1080;&#1086;&#1085;</span>
                <span class="tender-meta-value">{{ t.region }}</span>
              </div>
              <div class="tender-meta-item">
                <span class="tender-meta-label">&#1069;&#1090;&#1072;&#1087;</span>
                <span class="tender-meta-value{% if t.stage_open %} tone-good{% elif not t.stage_display or t.stage_display == '—' %} muted{% endif %}">{{ t.stage_display }}</span>
              </div>
            </div>

            {% if t.market_progress_total > 0 %}
            <div class="tender-progress">
              <div class="tender-progress-head">
                <span class="tender-progress-label">&#1055;&#1088;&#1086;&#1072;&#1085;&#1072;&#1083;&#1080;&#1079;&#1080;&#1088;&#1086;&#1074;&#1072;&#1085;&#1086;</span>
                <span class="tender-progress-value">{{ t.market_progress_done }}/{{ t.market_progress_total }}</span>
              </div>
              <div class="tender-progress-track">
                <div class="tender-progress-fill" style="width: {{ t.market_progress_percent }}%;"></div>
              </div>
              <div class="tender-progress-note">
                {% if t.market_progress_done >= t.market_progress_total %}
                &#1042;&#1089;&#1077; &#1089;&#1090;&#1088;&#1086;&#1082;&#1080; &#1089;&#1084;&#1077;&#1090;&#1099; &#1086;&#1073;&#1088;&#1072;&#1073;&#1086;&#1090;&#1072;&#1085;&#1099;.
                {% elif t.has_market_partial %}
                &#1054;&#1073;&#1088;&#1072;&#1073;&#1086;&#1090;&#1072;&#1085;&#1086; {{ t.market_progress_done }} &#1080;&#1079; {{ t.market_progress_total }}, &#1086;&#1089;&#1090;&#1072;&#1083;&#1086;&#1089;&#1100; {{ t.market_progress_left }}.
                {% else %}
                &#1057;&#1084;&#1077;&#1090;&#1072; &#1075;&#1086;&#1090;&#1086;&#1074;&#1072;. &#1055;&#1086;&#1080;&#1089;&#1082; &#1094;&#1077;&#1085; &#1077;&#1097;&#1105; &#1085;&#1077; &#1079;&#1072;&#1087;&#1091;&#1089;&#1082;&#1072;&#1083;&#1089;&#1103;.
                {% endif %}
              </div>
            </div>
            {% endif %}

            <div class="tender-card-pub">
              <span class="tender-card-pub-label">&#1055;&#1091;&#1073;&#1083;&#1080;&#1082;&#1072;&#1094;&#1080;&#1103;</span>
              <span class="tender-card-pub-date">{{ t.publish_date or "&#1044;&#1072;&#1090;&#1072; &#1085;&#1077; &#1091;&#1082;&#1072;&#1079;&#1072;&#1085;&#1072;" }}</span>
              <span class="tender-card-pub-label">&#1054;&#1082;&#1086;&#1085;&#1095;&#1072;&#1085;&#1080;&#1077;</span>
              <span class="tender-card-pub-date">{{ t.deadline_date }}</span>
            </div>

            <div class="tender-status-row">
              {% if t.has_svodka %}
              <span class="tag tag-merge">&#1050;&#1072;&#1088;&#1090;&#1086;&#1095;&#1082;&#1072; &#1075;&#1086;&#1090;&#1086;&#1074;&#1072;</span>
              {% elif t.has_market_partial %}
              <span class="tag tag-merge">&#1045;&#1089;&#1090;&#1100; &#1095;&#1072;&#1089;&#1090;&#1080;&#1095;&#1085;&#1099;&#1077; &#1094;&#1077;&#1085;&#1099;</span>
              {% elif t.has_estimate %}
              <span class="tag tag-ok">&#1057;&#1084;&#1077;&#1090;&#1072; &#1075;&#1086;&#1090;&#1086;&#1074;&#1072;</span>
              {% else %}
              <span class="tag tag-nodata">&#1053;&#1077;&#1090; &#1089;&#1084;&#1077;&#1090;&#1099;</span>
              {% endif %}
              <span class="tag {% if t.stage_open %}tag-stage-open{% else %}tag-stage-closed{% endif %}">{{ t.stage_display }}</span>
              <a class="eis-in-card" href="{{ t.eis_url }}" target="_blank" rel="noopener noreferrer">&#1045;&#1048;&#1057;</a>
            </div>

            <div class="tender-actions">
              {% if t.has_svodka %}
              <a class="tender-act tender-act--primary tender-act--main" href="/merge-report/{{ t.tender_id }}/">Посмотреть сравнение цен</a>
              <button type="button" class="tender-act tender-act--crm tender-act-btn" data-tid="{{ t.tender_id }}" onclick="exportTenderToCrm('{{ t.tender_id }}')" title="Создать объект в PM.bi CRM и перенести туда строки сметы как материалы.">+ Добавить в объекты</button>
              <p class="tender-next">Готово или частично готово: сохранённые строки Алисы будут вверху таблицы.</p>
              {% elif t.has_market_partial %}
              <a class="tender-act tender-act--primary tender-act--main" href="/merge-report/{{ t.tender_id }}/">Посмотреть частичные цены</a>
              <button type="button" class="tender-act tender-act--crm tender-act-btn" data-tid="{{ t.tender_id }}" onclick="exportTenderToCrm('{{ t.tender_id }}')" title="Создать объект в PM.bi CRM и перенести туда строки сметы как материалы.">+ Добавить в объекты</button>
              <p class="tender-next">Есть сохранённый прогресс Алисы. Можно открыть карточку и продолжить поиск.</p>
              {% elif t.has_estimate %}
              <a class="tender-act tender-act--primary tender-act--main" href="/merge-report/{{ t.tender_id }}/">Открыть карточку тендера</a>
              <button type="button" class="tender-act tender-act--crm tender-act-btn" data-tid="{{ t.tender_id }}" onclick="exportTenderToCrm('{{ t.tender_id }}')" title="Создать объект в PM.bi CRM и перенести туда строки сметы как материалы.">+ Добавить в объекты</button>
              <p class="tender-next">Смета готова. В карточке можно запустить поиск цен и смотреть сохранённые строки.</p>
              {% else %}
              <button type="button" class="tender-act tender-act--primary tender-act--main tender-act-btn" data-tid="{{ t.tender_id }}" onclick="runFullForTender('{{ t.tender_id }}')">Скачать документы и подготовить сравнение</button>
              <p class="tender-next">Смета не извлечена. Программа попробует скачать документы повторно.</p>
              {% endif %}
              <details class="tender-more">
                <summary>Дополнительные действия</summary>
                <div class="tender-more-actions">
                  <button type="button" class="tender-act tender-act-btn" data-tid="{{ t.tender_id }}" onclick="runFullForTender('{{ t.tender_id }}')" title="Продолжить поиск недостающих рыночных цен и заново собрать страницу сравнения.">Продолжить или обновить поиск цен</button>
                  <button type="button" class="tender-act tender-act-btn" data-tid="{{ t.tender_id }}" onclick="exportTenderToCrm('{{ t.tender_id }}')" title="Создать объект в PM.bi CRM и перенести туда строки сметы как материалы.">Добавить в объекты</button>
                  <button type="button" class="tender-act tender-act-btn" data-tid="{{ t.tender_id }}" onclick="rerunMarketForTender('{{ t.tender_id }}')" title="Удалить прогресс поиска цен и опросить Алису по всем позициям заново.">Начать поиск цен заново</button>
                  <button type="button" class="tender-act tender-act-btn" data-tid="{{ t.tender_id }}" onclick="rebuildReportForTender('{{ t.tender_id }}')" title="Повторно прочитать уже скачанные документы. Поиск рыночных цен не запускается.">Повторно извлечь смету из файлов</button>
                  {% if t.has_estimate %}
                  <a class="tender-act" href="/tenders/{{ t.tender_id }}/estimate.xlsx">Скачать Excel сметы</a>
                  {% endif %}
                  {% if t.has_market_partial %}
                  <a class="tender-act" href="/tenders/{{ t.tender_id }}/market-sources.xlsx">Скачать источники рынка</a>
                  {% endif %}
                  {% if t.has_svodka %}
                  <a class="tender-act" href="/tenders/{{ t.tender_id }}/svodka.xlsx">Скачать Excel выгодности</a>
                  {% endif %}
                  {% if t.has_svodka %}
                  <button type="button" class="tender-act tender-act-btn" data-tid="{{ t.tender_id }}" onclick="runViabilityOnly('{{ t.tender_id }}')" title="Обновить вывод о выгодности и отправить его в настроенный Telegram-чат.">Отправить вывод в Telegram</button>
                  {% endif %}
                </div>
              </details>
            </div>
          </div>
        </div>
        {% endfor %}
      </div>
    {% if not items %}
    {% if tender_count == 0 %}
    <p class="sub" style="margin:0;">База пуста — нажмите «Найти новые закупки» в верхней панели.</p>
    {% elif not show_all %}
    <p class="sub" style="margin:0;">Нет закупок на этапе «Подача заявок». Включите «Показать все этапы» выше или скачайте новые тендеры.</p>
    {% else %}
    <p class="sub" style="margin:0;">Нет данных для отображения.</p>
    {% endif %}
    {% endif %}

    {% if tender_count %}
    <p class="meta" style="margin:12px 0 0;">В базе {{ tender_count }} тендеров · сметы с таблицей позиций: {{ display_report_count }} / {{ report_count }} · <span style="color:#d89090;">красная полоска слева</span> — в смете нет извлечённых работ.</p>
    {% endif %}
    </section>

    <section class="action-hub controls" aria-labelledby="actionHubTitle">
      <h2 class="section-title" id="actionHubTitle">Главное</h2>
      <p class="section-lead">Сначала нажмите «Найти новые закупки». Если поиск не сработает, ниже появится короткое объяснение причины и следующий шаг.</p>

      <div class="workflow-strip" aria-hidden="true">
        <div class="wf-step"><span class="wf-num">1</span> Найти закупки</div>
        <span class="wf-arrow">→</span>
        <div class="wf-step"><span class="wf-num">2</span> Сравнить цены</div>
        <span class="wf-arrow">→</span>
        <div class="wf-step"><span class="wf-num">3</span> Посмотреть результат</div>
      </div>

      <div class="action-grid">
        <article class="action-card">
          <h3 class="action-card-title">Шаг 1. Найти новые закупки</h3>
          <p class="action-card-desc">Ищет закупки по вашим регионам и ключевым словам, скачивает архивы документов, распаковывает их и извлекает смету в Excel. Результат попадает в список выше.</p>
          <button class="btn btn-lg" type="button" id="startBtn" onclick="startParsing()">Найти новые закупки</button>
          <details class="action-options">
            <summary>параметры поиска</summary>
            <div class="opts">
              <label title="Сколько страниц результатов просматривать на каждую пару регион × ключевое слово">Страниц результатов
                <input type="number" id="optMaxPages" min="1" max="20" value="2" />
              </label>
              <label title="Максимум новых тендеров за один запуск">Закупок за запуск
                <input type="number" id="optMaxTenders" min="1" max="50" value="15" />
              </label>
              <label title="Не брать закупки старше указанного числа дней">Опубликованы за последние, дней
                <input type="number" id="optDaysBack" min="1" max="365" value="60" />
              </label>
            </div>
          </details>
        </article>

        <article class="action-card">
          <h3 class="action-card-title">Шаг 2. Сравнить цены заказчика с рынком</h3>
          <p class="action-card-desc">Программа ищет рыночные цены и реальные ссылки для позиций сметы, а затем собирает готовую страницу сравнения. <strong>Долго</strong> — обработка нескольких закупок может занять часы.</p>
          <div class="btn-row">
            <button class="btn btn-lg" type="button" id="genMergeMissingBtn" onclick="generateMergeSiteMissing()">Подготовить недостающие сравнения</button>
            <button class="btn secondary" type="button" id="genMergeSiteBtn" onclick="generateMergeSiteAll()">Обновить сравнения для всех</button>
          </div>
          <p class="btn-effect"><strong>Рекомендуется первая кнопка:</strong> она пропускает уже готовые результаты. Вторая повторно обрабатывает все доступные сметы.</p>
        </article>

        <article class="action-card action-card--wide">
          <h3 class="action-card-title">Проверить одну закупку по ссылке</h3>
          <p class="action-card-desc">Вставьте ссылку с zakupki.gov.ru или номер закупки. Программа скачает документы, извлечёт смету, найдёт рыночные цены и подготовит сравнение.</p>
          <div class="link-row">
            <span>Ссылка или номер:</span>
            <input id="tenderLinkInput" type="text" placeholder="https://zakupki.gov.ru/... или 19-значный номер" />
            <button class="btn" type="button" id="runByLinkBtn" onclick="runByTenderLink()">Проверить эту закупку</button>
          </div>
          <div id="quickTenderCheck" class="meta" style="margin-top:6px;display:none;">
            Последний запуск: <a id="quickTenderReportLink" href="#" target="_blank" rel="noopener noreferrer">открыть сводку</a>
            · <a id="quickTenderEisLink" href="#" target="_blank" rel="noopener noreferrer">карточка на ЕИС</a>
          </div>
          <details class="action-options">
            <summary>повторное извлечение сметы из уже скачанных файлов</summary>
            <div class="rebuild-row">
              <span>Выберите закупку:</span>
              <select id="rebuildTenderSelect" {% if not rebuild_options %}disabled{% endif %}>
                {% for o in rebuild_options %}
                <option value="{{ o.tender_id }}">{{ o.tender_id }} — {{ o.display_title }}</option>
                {% endfor %}
                {% if not rebuild_options %}
                <option value="">— нет тендеров —</option>
                {% endif %}
              </select>
              <button class="btn secondary" type="button" id="rebuildBtn" onclick="rebuildReport()">Извлечь смету повторно</button>
              <button class="btn secondary" type="button" id="rebuildAllBtn" onclick="rebuildAllReports()" {% if tender_count < 1 %}disabled title="Нет тендеров в базе"{% endif %}>Повторить для всех</button>
            </div>
          </details>
        </article>

        <article class="action-card">
          <h3 class="action-card-title">Уведомления в браузере</h3>
          <p class="action-card-desc">Всплывающее окно, когда закончится поиск закупок или появится новое сравнение цен. Это не Telegram — уведомление работает только в этом браузере.</p>
          <button class="btn secondary" type="button" id="enablePushBtn" onclick="enableWebPush()">Включить уведомления</button>
          <a class="link-refresh" href="#" onclick="location.reload(); return false;" style="display:inline-block;margin-top:10px;">Обновить страницу (F5)</a>
        </article>
      </div>

      <div id="mergeSitePanel" class="parse-progress-panel" role="status" aria-live="polite" hidden>
        <div class="parse-progress-head">
          <span class="parse-pulse" aria-hidden="true"></span>
          <strong id="mergeSiteLabel">Подготавливаем сравнения цен</strong>
        </div>
        <div class="merge-bar-wrap"><div id="mergeBarFill" class="merge-bar-fill" style="width:0%"></div></div>
        <div class="parse-progress-time" id="mergePercentText">0%</div>
        <div class="parse-progress-hint" id="mergeSiteDetail"></div>
        <div class="merge-logs" id="mergeSiteLogs"></div>
      </div>
      <div id="mergeIdleSummary" class="meta" style="margin-top:4px;"></div>
      <div id="mergeMissingReason" class="meta" style="margin-top:4px;"></div>
      <div id="parseProgressPanel" class="parse-progress-panel" role="status" aria-live="polite" hidden>
        <div class="parse-progress-head">
          <span class="parse-pulse" aria-hidden="true"></span>
          <strong id="parseProgressLabel">Выполняется…</strong>
        </div>
        <div class="parse-bar-wrap"><div id="parseBarFill" class="parse-bar-fill"></div></div>
        <div class="parse-progress-time" id="parseProgressTime">Прошло: 0 с</div>
        <div class="status" id="parseStatus"></div>
        <div class="parse-summary">
          <div class="parse-summary-grid">
            <div class="parse-summary-item">
              <div class="parse-summary-label">Итог</div>
              <div class="parse-summary-value" id="parseResultMain">Ждём запуск</div>
            </div>
            <div class="parse-summary-item">
              <div class="parse-summary-label">Причина</div>
              <div class="parse-summary-value" id="parseResultIssue">Пока нет</div>
            </div>
            <div class="parse-summary-item">
              <div class="parse-summary-label">Что делать</div>
              <div class="parse-summary-value" id="parseResultNext">Нажать кнопку поиска</div>
            </div>
          </div>
        </div>
        <div id="parseProgressLogCount" class="parse-progress-hint" style="margin-top:8px;color:#b8c7ea;"></div>
        <details class="compact-details">
          <summary>Технические подробности</summary>
          <div class="parse-status-line" id="parseCommandLine"></div>
          <div class="logs" id="parseLogs"></div>
        </details>
      </div>
    </section>

    <section class="tool-section region-block" id="nmckParseBlock" aria-labelledby="nmckParseTitle">
      <h2 class="section-title" id="nmckParseTitle">Дополнительно: обоснование НМЦК (Приложение №2)</h2>
      <p class="action-card-desc">Загрузите Excel «Приложение №2 к извещению (Обоснование НМЦК)» — получите таблицу и JSON с позициями, количествами, коммерческими предложениями и НМЦК.</p>
      <div class="btn-row" style="margin-top:0;">
        <input type="file" id="nmckFileInput" accept=".xlsx,.xls,.xlsm,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" style="max-width:100%;font-size:12px;color:#b8c7ea;" />
        <button type="button" class="btn secondary" id="nmckParseBtn" onclick="parseNmckJustification()">Разобрать Excel в таблицу и JSON</button>
        <a class="btn secondary" id="nmckPreviewLink" href="#" target="_blank" rel="noopener noreferrer" hidden>Открыть таблицу</a>
        <button type="button" class="btn secondary" id="nmckCopyBtn" onclick="copyNmckJson()" disabled>Скопировать JSON</button>
        <button type="button" class="btn secondary" id="nmckDownloadBtn" onclick="downloadNmckJson()" disabled>Скачать JSON-файл</button>
      </div>
      <p class="status" id="nmckParseStatus" style="margin-top:8px;"></p>
      <textarea id="nmckJsonOut" readonly hidden style="width:100%;min-height:220px;margin-top:10px;font-family:Consolas,monospace;font-size:11px;line-height:1.35;background:#0b1223;border:1px solid var(--border-soft);color:var(--text);border-radius:8px;padding:10px;box-sizing:border-box;resize:vertical;"></textarea>
    </section>
  </div>
  <button type="button" class="site-chat-fab" id="siteChatFab" onclick="toggleSiteChat()" title="Логи и сообщения выполнения">🧾</button>
  <aside class="site-chat-panel" id="siteChatPanel" hidden>
    <div class="site-chat-head">
      <span>🧾 Логи выполнения</span>
      <button type="button" class="site-chat-close" onclick="toggleSiteChat(false)">закрыть</button>
    </div>
    <div class="site-chat-feed" id="siteChatFeed">
      <div class="site-chat-empty">Пока событий нет. Когда запустится сравнение цен, здесь появятся сообщения как в Telegram.</div>
    </div>
  </aside>
  <script>
    function switchMarketSection(key) {
      document.querySelectorAll("[data-market-tab]").forEach(el => el.classList.toggle("is-active", el.getAttribute("data-market-tab") === key));
      document.querySelectorAll("[data-market-pane]").forEach(el => el.classList.toggle("is-active", el.getAttribute("data-market-pane") === key));
    }

    document.addEventListener("click", function(e) {
      const card = e.target && e.target.closest ? e.target.closest(".tender-card[data-href]") : null;
      if (!card) return;
      if (e.target.closest("a, button, summary, details, input, select, textarea, label")) return;
      const href = card.getAttribute("data-href");
      if (href) window.location.href = href;
    });

    (function bindRebuildSelect() {
      const sel = document.getElementById("rebuildTenderSelect");
      if (!sel) return;
      sel.addEventListener("change", function() {
        applyToolbarDisabled(parseRunning, !!window.__mergeRunLive);
      });
    })();

    let lastNmckJson = "";
    async function parseNmckJustification() {
      const inp = document.getElementById("nmckFileInput");
      const st = document.getElementById("nmckParseStatus");
      const ta = document.getElementById("nmckJsonOut");
      const copyB = document.getElementById("nmckCopyBtn");
      const dlB = document.getElementById("nmckDownloadBtn");
      const prevA = document.getElementById("nmckPreviewLink");
      const f = inp && inp.files && inp.files[0];
      if (!f) { alert("Выберите файл Excel (.xlsx)"); return; }
      if (st) st.textContent = "Загрузка и разбор…";
      lastNmckJson = "";
      if (copyB) copyB.disabled = true;
      if (dlB) dlB.disabled = true;
      if (prevA) { prevA.hidden = true; prevA.href = "#"; }
      if (ta) { ta.hidden = true; ta.value = ""; }
      const fd = new FormData();
      fd.append("file", f);
      try {
        const r = await fetch("/api/parse-nmck-justification", { method: "POST", body: fd });
        let data = {};
        try { data = await r.json(); } catch (e) {}
        if (!r.ok || !data.ok) {
          if (st) st.textContent = (data && data.message) ? data.message : ("Ошибка " + r.status);
          return;
        }
        const pack = { columns: data.columns, rows: data.rows, meta: data.meta };
        lastNmckJson = JSON.stringify(pack, null, 2);
        const m = data.meta || {};
        if (st) {
          st.textContent = "Готово: " + (m.row_count != null ? m.row_count : "?") + " поз., колонок "
            + (m.column_count != null ? m.column_count : "?") + ", лист «" + (m.sheet || "") + "»";
        }
        if (data.preview_url && prevA) {
          prevA.href = data.preview_url;
          prevA.hidden = false;
        }
        if (ta) { ta.value = lastNmckJson; ta.hidden = false; }
        if (copyB) copyB.disabled = false;
        if (dlB) dlB.disabled = false;
      } catch (e) {
        if (st) st.textContent = "Запрос не выполнен (сеть или сервер).";
      }
    }
    function copyNmckJson() {
      if (!lastNmckJson) return;
      navigator.clipboard.writeText(lastNmckJson).then(function() {
        const st = document.getElementById("nmckParseStatus");
        if (st) st.textContent += " · JSON в буфере обмена";
      }).catch(function() { alert("Не удалось скопировать"); });
    }
    function downloadNmckJson() {
      if (!lastNmckJson) return;
      const blob = new Blob([lastNmckJson], { type: "application/json;charset=utf-8" });
      const a = document.createElement("a");
      a.href = URL.createObjectURL(blob);
      a.download = "nmck_prilozhenie_2.json";
      a.click();
      URL.revokeObjectURL(a.href);
    }

    function getRebuildTenderId() {
      const s = document.getElementById("rebuildTenderSelect");
      return s && s.value ? String(s.value).trim() : "";
    }
    function setQuickTenderLinks(tid) {
      const t = String(tid || "").trim();
      const box = document.getElementById("quickTenderCheck");
      const rep = document.getElementById("quickTenderReportLink");
      const eis = document.getElementById("quickTenderEisLink");
      if (!box || !rep || !eis) return;
      if (!t) {
        box.style.display = "none";
        return;
      }
      rep.href = "/merge-report/" + encodeURIComponent(t) + "/";
      eis.href = "https://zakupki.gov.ru/epz/order/notice/ea20/view/common-info.html?regNumber=" + encodeURIComponent(t);
      rep.textContent = "сводка " + t;
      box.style.display = "";
      try { localStorage.setItem("lastTenderCheckId", t); } catch (e) {}
    }
    try {
      const lastTid = localStorage.getItem("lastTenderCheckId") || "";
      if (lastTid) setQuickTenderLinks(lastTid);
    } catch (e) {}
    const TENDER_COUNT = {{ tender_count }};

    let parseRunning = false;
    let parseStartMs = null;
    let parsePendingUntil = 0;
    let parseAutoReloadArmed = false;
    let notifyState = {
      enabled: localStorage.getItem("webPushEnabled") === "1",
      prev: null,
    };

    function formatElapsed(sec) {
      const s = Math.max(0, Math.floor(sec));
      const m = Math.floor(s / 60);
      const h = Math.floor(m / 60);
      if (h > 0) return `${h} ч ${m % 60} мин ${s % 60} с`;
      if (m > 0) return `${m} мин ${s % 60} с`;
      return `${s} с`;
    }

    function updateParseElapsed() {
      if (!parseRunning || parseStartMs == null) return;
      const sec = (Date.now() - parseStartMs) / 1000;
      const el = document.getElementById("parseProgressTime");
      if (el) el.textContent = "Прошло: " + formatElapsed(sec);
    }

    setInterval(updateParseElapsed, 1000);

    function showParseLaunchFeedback() {
      parseRunning = true;
      parseStartMs = Date.now();
      parsePendingUntil = Date.now() + 8000;
      parseAutoReloadArmed = true;
      const panel = document.getElementById("parseProgressPanel");
      const label = document.getElementById("parseProgressLabel");
      const bar = document.getElementById("parseBarFill");
      const time = document.getElementById("parseProgressTime");
      const status = document.getElementById("parseStatus");
      const logs = document.getElementById("parseLogs");
      const logCount = document.getElementById("parseProgressLogCount");
      const cmd = document.getElementById("parseCommandLine");
      if (panel) panel.hidden = false;
      if (label) label.textContent = "Запускаем поиск новых закупок…";
      if (bar) {
        bar.classList.add("running");
        bar.style.width = "65%";
      }
      if (time) time.textContent = "Прошло: 0 с";
      if (status) status.textContent = "Статус: передаём задачу серверу";
      if (logs) logs.textContent = "Ожидаем первые сообщения от программы…";
      if (logCount) logCount.textContent = "Поиск запускается.";
      if (cmd) cmd.textContent = "";
      applyToolbarDisabled(true, false);
      window.setTimeout(function() {
        if (panel) panel.scrollIntoView({ behavior: "smooth", block: "nearest" });
      }, 80);
    }

    function showParseLaunchError(message) {
      parseRunning = false;
      parseStartMs = null;
      parsePendingUntil = 0;
      parseAutoReloadArmed = false;
      const panel = document.getElementById("parseProgressPanel");
      const label = document.getElementById("parseProgressLabel");
      const bar = document.getElementById("parseBarFill");
      const time = document.getElementById("parseProgressTime");
      const status = document.getElementById("parseStatus");
      if (panel) panel.hidden = false;
      if (label) label.textContent = "Поиск не запустился";
      if (bar) {
        bar.classList.remove("running");
        bar.style.width = "100%";
      }
      if (time) time.textContent = "";
      if (status) status.textContent = message || "Сервер не смог запустить поиск.";
      applyToolbarDisabled(false, !!window.__mergeRunLive);
    }

    function applyToolbarDisabled(parseRun, mergeRun) {
      const busy = parseRun || mergeRun;
      const startBtn = document.getElementById("startBtn");
      const rebuildBtn = document.getElementById("rebuildBtn");
      const rebuildAllBtn = document.getElementById("rebuildAllBtn");
      const genBtn = document.getElementById("genMergeSiteBtn");
      const genMissingBtn = document.getElementById("genMergeMissingBtn");
      const runByLinkBtn = document.getElementById("runByLinkBtn");
      if (startBtn) {
        startBtn.disabled = busy;
        startBtn.textContent = parseRun ? "Ищем закупки…" : "Найти новые закупки";
      }
      if (rebuildBtn) rebuildBtn.disabled = busy || !getRebuildTenderId();
      if (rebuildAllBtn) rebuildAllBtn.disabled = busy || TENDER_COUNT < 1;
      if (genBtn) genBtn.disabled = busy;
      if (genMissingBtn) genMissingBtn.disabled = busy;
      if (runByLinkBtn) runByLinkBtn.disabled = busy;
      document.querySelectorAll(".tender-act-btn").forEach(function(btn) {
        btn.disabled = busy;
      });
    }

    function updatePushButtonUi() {
      const b = document.getElementById("enablePushBtn");
      if (!b) return;
      if (!("Notification" in window)) {
        b.textContent = "Браузер не поддерживает уведомления";
        b.disabled = true;
        return;
      }
      const perm = Notification.permission;
      if (notifyState.enabled && perm === "granted") {
        b.textContent = "Уведомления включены";
        b.disabled = true;
        return;
      }
      b.textContent = "Включить уведомления";
      b.disabled = false;
    }

    async function enableWebPush() {
      if (!("Notification" in window)) {
        alert("Браузер не поддерживает уведомления.");
        return;
      }
      const permission = await Notification.requestPermission();
      if (permission !== "granted") {
        alert("Разрешение на уведомления не выдано.");
        updatePushButtonUi();
        return;
      }
      notifyState.enabled = true;
      localStorage.setItem("webPushEnabled", "1");
      updatePushButtonUi();
      new Notification("AutoBot", { body: "Уведомления в браузере включены." });
    }

    function safeNotify(title, body) {
      if (!notifyState.enabled) return;
      if (!("Notification" in window)) return;
      if (Notification.permission !== "granted") return;
      try {
        new Notification(title, { body });
      } catch (e) {}
    }

    function handlePushDiff(nextState) {
      const prev = notifyState.prev;
      notifyState.prev = nextState;
      if (!prev) return;

      if (prev.parse_running && !nextState.parse_running) {
        const ok = nextState.parse_exit_code === 0;
        safeNotify(
          ok ? "Поиск закупок завершён" : "Поиск закупок завершён с ошибкой",
          ok ? "Обновите страницу, чтобы увидеть результат." : "Откройте ход работы на странице."
        );
      }

      if (prev.merge_running && !nextState.merge_running) {
        safeNotify("Сравнения цен готовы", nextState.merge_last_summary || "Обработка завершена.");
      }

      if ((nextState.coverage_merge_html || 0) > (prev.coverage_merge_html || 0)) {
        const delta = (nextState.coverage_merge_html || 0) - (prev.coverage_merge_html || 0);
        safeNotify("Появились новые сравнения цен", "Готово новых страниц: " + delta);
      }
    }

    async function refreshPushState() {
      try {
        const r = await fetch("/api/push-state");
        if (!r.ok) return;
        const st = await r.json();
        handlePushDiff(st);
      } catch (e) {}
    }

    async function startParsing() {
      showParseLaunchFeedback();
      try {
        const body = {
          max_pages: parseInt(document.getElementById("optMaxPages").value, 10) || 2,
          max_tenders: parseInt(document.getElementById("optMaxTenders").value, 10) || 15,
          days_back: parseInt(document.getElementById("optDaysBack").value, 10) || 60,
        };
        const r = await fetch("/api/start-parse", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify(body),
        });
        const data = await r.json();
        if (!r.ok || !data.ok) {
          const message = data.message || "Не удалось запустить поиск закупок";
          showParseLaunchError(message);
          alert(message);
          return;
        }
        window.setTimeout(refreshStatus, 200);
      } catch (e) {
        const message = "Не удалось запустить поиск закупок. Проверьте, работает ли сервер.";
        showParseLaunchError(message);
        alert(message);
      }
    }

    async function rebuildReport() {
      const tid = getRebuildTenderId();
      if (!tid) { alert("Выберите закупку."); return; }
      if (!confirm("Повторно извлечь смету для закупки " + tid + " из уже скачанных документов?\\n\\nРыночные цены обновляться не будут.")) return;
      applyToolbarDisabled(true, false);
      try {
        const r = await fetch("/api/rebuild-report", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ tender_id: tid }),
        });
        const data = await r.json();
        if (!data.ok) {
          alert(data.message || "Не удалось запустить повторное извлечение сметы");
          refreshStatus();
        }
      } catch (e) {
        alert("Не удалось отправить запрос на повторное извлечение сметы.");
        refreshStatus();
      }
    }

    async function rebuildReportForTender(tid) {
      const t = String(tid || "").trim();
      if (!t) return;
      if (!confirm("Повторно извлечь смету для закупки " + t + " из уже скачанных документов?\\n\\nРыночные цены обновляться не будут.")) return;
      applyToolbarDisabled(true, false);
      try {
        const r = await fetch("/api/rebuild-report", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ tender_id: t }),
        });
        const data = await r.json();
        if (!data.ok) {
          alert(data.message || "Не удалось запустить повторное извлечение сметы");
          refreshStatus();
        }
      } catch (e) {
        alert("Не удалось отправить запрос на повторное извлечение сметы.");
        refreshStatus();
      }
    }

    async function rebuildAllReports() {
      if (TENDER_COUNT < 1) { alert("В списке пока нет закупок."); return; }
      if (!confirm(
        "Повторно извлечь сметы для всех " + TENDER_COUNT + " закупок?\\n\\n"
        + "Программа перечитает уже скачанные документы. Рыночные цены обновляться не будут."
      )) return;
      applyToolbarDisabled(true, false);
      try {
        const r = await fetch("/api/rebuild-all-reports", { method: "POST" });
        const data = await r.json();
        if (!data.ok) {
          alert(data.message || "Не удалось запустить повторное извлечение смет");
          refreshStatus();
        }
      } catch (e) {
        alert("Ошибка запроса");
        refreshStatus();
      }
    }

    async function generateMergeSiteAll() {
      if (!confirm("Обновить сравнения цен для всех закупок со сметой?\\n\\nПрограмма повторно проверит рыночные источники. Процесс может занять несколько часов.")) return;
      try {
        const r = await fetch("/api/generate-merge-site-all", {
          method: "POST",
          headers: { "Content-Type": "application/json", "Accept": "application/json" },
          body: "{}",
        });
        let data = {};
        try {
          data = await r.json();
        } catch (e) {
          alert("Сервер вернул не JSON (код " + r.status + "). Проверьте консоль web_ui.py.");
          refreshStatus();
          return;
        }
        if (!r.ok || !data.ok) {
          alert(data.message || ("Запрос отклонён (HTTP " + r.status + ")"));
        }
      } catch (e) {
        alert("Сеть или сервер недоступен: " + e);
      }
      refreshStatus();
      refreshCoverage();
    }

    async function generateMergeSiteMissing() {
      if (!confirm("Подготовить сравнения только там, где результата ещё нет или прошлая обработка завершилась с ошибкой?\\n\\nУже готовые страницы будут пропущены.")) return;
      try {
        const r = await fetch("/api/generate-merge-site-missing", {
          method: "POST",
          headers: { "Content-Type": "application/json", "Accept": "application/json" },
          body: "{}",
        });
        let data = {};
        try {
          data = await r.json();
        } catch (e) {
          alert("Сервер вернул не JSON (код " + r.status + "). Проверьте консоль web_ui.py.");
          refreshStatus();
          return;
        }
        if (!r.ok || !data.ok) {
          alert(data.message || ("Запрос отклонён (HTTP " + r.status + ")"));
        }
      } catch (e) {
        alert("Сеть или сервер недоступен: " + e);
      }
      refreshStatus();
      refreshCoverage();
    }

    async function runFullForTender(tid) {
      const t = String(tid || "").trim();
      if (!t) return;
      if (!confirm("Подготовить сравнение цен для закупки " + t + "?\\n\\nПрограмма проверит документы, найдёт рыночные цены и соберёт готовую страницу.")) return;
      setQuickTenderLinks(t);
      try {
        const r = await fetch("/api/generate-merge-site-one", {
          method: "POST",
          headers: { "Content-Type": "application/json", "Accept": "application/json" },
          body: JSON.stringify({ tender_id: t }),
        });
        const data = await r.json();
        if (!r.ok || !data.ok) {
          alert(data.message || ("Запрос отклонён (HTTP " + r.status + ")"));
        }
      } catch (e) {
        alert("Сеть или сервер недоступен: " + e);
      }
      refreshStatus();
      refreshCoverage();
    }

    function navigateCrmProject(url) {
      if (!url) return;
      try {
        if (window.parent && window.parent !== window) {
          window.parent.postMessage({ type: "pmbi:navigate", href: url }, "*");
          return;
        }
      } catch (e) {}
      window.location.href = url;
    }

    async function exportTenderToCrm(tid) {
      const t = String(tid || "").trim();
      if (!t) return;
      if (!confirm("Добавить закупку " + t + " в CRM?\\n\\nБудет создан объект, а строки сметы уйдут в материалы объекта.")) return;
      try {
        const r = await fetch("/api/export-to-crm", {
          method: "POST",
          headers: { "Content-Type": "application/json", "Accept": "application/json" },
          body: JSON.stringify({ tender_id: t }),
        });
        let data = {};
        try { data = await r.json(); } catch (e) {}
        if (!r.ok || !data.ok) {
          alert(data.message || ("CRM-экспорт не прошёл (HTTP " + r.status + ")"));
          return;
        }
        const summary = data.summary || {};
        if (data.already_exists) {
          alert("Эта закупка уже есть в CRM: объект #" + data.project_id + ".");
          if (data.project_url) navigateCrmProject(data.project_url);
          return;
        }
        alert(
          "Готово: объект #" + data.project_id + " создан в CRM.\\n"
          + "Материалов отправлено: " + (data.materials_sent || 0) + ".\\n"
          + "В CRM сейчас материалов: " + (summary.materials == null ? "?" : summary.materials) + "."
        );
        if (data.project_url) navigateCrmProject(data.project_url);
      } catch (e) {
        alert("Не удалось отправить закупку в CRM: " + e);
      }
    }

    async function rerunMarketForTender(tid) {
      const t = String(tid || "").trim();
      if (!t) return;
      if (!confirm("Начать поиск рыночных цен для закупки " + t + " заново?\\n\\nСохранённый прогресс Алисы будет отброшен.")) return;
      setQuickTenderLinks(t);
      try {
        const r = await fetch("/api/generate-merge-site-one-rerun-market", {
          method: "POST",
          headers: { "Content-Type": "application/json", "Accept": "application/json" },
          body: JSON.stringify({ tender_id: t }),
        });
        const data = await r.json();
        if (!r.ok || !data.ok) {
          alert(data.message || ("Запрос отклонён (HTTP " + r.status + ")"));
        }
      } catch (e) {
        alert("Сеть или сервер недоступен: " + e);
      }
      refreshStatus();
      refreshCoverage();
    }

    async function runViabilityOnly(tid) {
      const t = String(tid || "").trim();
      if (!t) return;
      if (!confirm("Обновить вывод о выгодности закупки " + t + " и отправить его в Telegram?")) return;
      try {
        const r = await fetch("/api/tender-viability-refresh", {
          method: "POST",
          headers: { "Content-Type": "application/json", "Accept": "application/json" },
          body: JSON.stringify({ tender_id: t }),
        });
        const data = await r.json();
        if (!r.ok || !data.ok) {
          alert(data.message || ("Запрос отклонён (HTTP " + r.status + ")"));
          return;
        }
        let msg = data.message || "Готово.";
        if (data.report_url) {
          msg += " | Открыть: " + data.report_url;
        }
        if (data.telegram_sent) {
          msg += " | В Telegram отправлен анализ.";
        }
        alert(msg);
      } catch (e) {
        alert("Сеть или сервер недоступен: " + e);
      }
      refreshCoverage();
    }

    async function runByTenderLink() {
      const inp = document.getElementById("tenderLinkInput");
      const raw = inp && inp.value ? String(inp.value).trim() : "";
      if (!raw) { alert("Вставьте ссылку на закупку с zakupki.gov.ru или её номер."); return; }
      if (!confirm("Проверить эту закупку?\\n\\nПрограмма скачает документы, извлечёт смету и найдёт рыночные цены.")) return;
      try {
        const r = await fetch("/api/generate-merge-site-by-link", {
          method: "POST",
          headers: { "Content-Type": "application/json", "Accept": "application/json" },
          body: JSON.stringify({ tender_link: raw }),
        });
        const data = await r.json();
        if (!r.ok || !data.ok) {
          alert(data.message || ("Запрос отклонён (HTTP " + r.status + ")"));
        } else if (inp) {
          inp.value = "";
          setQuickTenderLinks(data.tender_id || "");
        }
      } catch (e) {
        alert("Сеть или сервер недоступен: " + e);
      }
      refreshStatus();
      refreshCoverage();
    }

    async function refreshCoverage() {
      const el = document.getElementById("reportCoverageBanner");
      if (!el) return;
      try {
        const r = await fetch("/api/reports-coverage");
        if (!r.ok) return;
        const c = await r.json();
        const nt = c.tender_count ?? 0;
        const mh = c.merge_html_among_tenders ?? 0;
        const miss = c.tenders_missing_merge_html ?? 0;
        const sx = c.svodka_xlsx_count ?? 0;
        const rs_no_est = c.missing_no_estimate ?? 0;
        const rs_no_svodka = c.missing_no_svodka ?? 0;
        const rs_no_html = c.missing_no_html ?? 0;
        let cls = "cov-banner stat-strip cov-ok";
        if (nt === 0) {
          el.className = "cov-banner stat-strip cov-warn";
          el.innerHTML = "В списке пока нет закупок. Нажмите «Найти новые закупки».";
          return;
        }
        if (miss > 0) cls = mh === 0 && sx === 0 ? "cov-banner stat-strip cov-warn" : "cov-banner stat-strip cov-partial";
        el.className = cls;
        let html = "Готовых сравнений цен: <strong>" + mh + "</strong> из " + nt;
        if (miss > 0) {
          html += " · ждут обработки: <strong>" + miss + "</strong>";
          html += "<br/><span style=\\"opacity:.85;font-size:11px\\">Из них: без извлечённой сметы — " + rs_no_est + ", без найденных рыночных цен — " + rs_no_svodka + ", страница результата не собрана — " + rs_no_html + ".</span>";
        }
        el.innerHTML = html;
      } catch (e) {}
    }

    function parseProgressView(pr, parsePending) {
      const lines = Array.isArray(pr.log_tail) ? pr.log_tail : [];
      const isTenderSearch = String(pr.task || "").includes("поиск новых закупок")
        || lines.some(function(line) { return line === "Поиск тендеров..."; });
      if (parsePending) {
        return { title: "Запускаем поиск новых закупок…", detail: "Передаём задачу серверу", percent: 5, indeterminate: true };
      }
      if (!pr.running) {
        if (pr.exit_code === 0) {
          return {
            title: isTenderSearch ? "Поиск закупок завершён" : "Задание завершено",
            detail: isTenderSearch ? "Готово. Обновите страницу, чтобы увидеть новые закупки." : "Готово.",
            percent: 100,
            indeterminate: false,
          };
        }
        if (pr.exit_code !== null && pr.exit_code !== undefined) {
          return { title: isTenderSearch ? "Поиск завершён с ошибкой" : "Задание завершено с ошибкой", detail: "Подробности — в журнале ниже.", percent: 100, indeterminate: false };
        }
        return { title: "Ожидание", detail: "", percent: 0, indeterminate: false };
      }

      let searchChecks = 0;
      let tenderStep = null;
      let filtersDone = false;
      let finalReport = false;
      for (const line of lines) {
        if (line.startsWith("- ") && line.includes(" найдено")) searchChecks += 1;
        if (line.startsWith("Итого после фильтров:")) filtersDone = true;
        if (line.startsWith("Готово. Общий отчет по сметам:")) finalReport = true;
        const match = line.match(/^\[([0-9]+)\/([0-9]+)\] ([0-9]+):/);
        if (match) tenderStep = { current: Number(match[1]), total: Number(match[2]), id: match[3] };
      }

      if (finalReport) {
        return { title: "Завершаем поиск", detail: "Сохраняем итоговый отчёт и список закупок", percent: 98, indeterminate: false };
      }
      if (tenderStep && tenderStep.total > 0) {
        const completedBefore = Math.max(0, tenderStep.current - 1);
        const percent = 48 + Math.round((completedBefore / tenderStep.total) * 47);
        return {
          title: "Скачиваем документы и извлекаем сметы",
          detail: "Закупка " + tenderStep.current + " из " + tenderStep.total + " · № " + tenderStep.id,
          percent: percent,
          indeterminate: false,
        };
      }
      if (filtersDone) {
        return { title: "Формируем список закупок", detail: "Поиск завершён, применяем фильтры и проверяем ранее найденные закупки", percent: 45, indeterminate: true };
      }
      if (searchChecks > 0 || lines.some(function(line) { return line === "Поиск тендеров..."; })) {
        return {
          title: "Ищем закупки на zakupki.gov.ru",
          detail: searchChecks > 0 ? "Проверено поисковых направлений: " + searchChecks : "Получаем первые результаты…",
          percent: Math.min(40, 10 + searchChecks * 5),
          indeterminate: true,
        };
      }
      return { title: pr.task ? "Сейчас: " + pr.task : "Подготавливаем поиск…", detail: "Процесс запущен, ожидаем первые сообщения", percent: 7, indeterminate: true };
    }

    function parseOutcomeSummary(pr, parsePending) {
      const lines = Array.isArray(pr.log_tail) ? pr.log_tail : [];
      const joined = lines.join("\\n");
      const foundMatch = joined.match(/Итого после фильтров:\s*([0-9]+)/);
      const addedMatch = joined.match(/Добавлено в систему:\s*([0-9]+)/);
      const resultEl = { text: "Поиск ещё не завершён", cls: "" };
      const issueEl = { text: "Идёт выполнение", cls: "" };
      const nextEl = { text: "Дождаться окончания", cls: "" };

      if (parsePending) {
        return {
          result: { text: "Запускаем поиск", cls: "" },
          issue: { text: "Сервер принимает задачу", cls: "" },
          next: { text: "Подождать несколько секунд", cls: "" },
        };
      }
      if (pr.running) {
        return {
          result: { text: "Идёт поиск закупок", cls: "" },
          issue: { text: "Программа проверяет ЕИС и документы", cls: "" },
          next: { text: "Можно просто оставить вкладку открытой", cls: "" },
        };
      }

      if (joined.includes("ERR_CERT_AUTHORITY_INVALID")) {
        resultEl.text = "Новых закупок не получено";
        resultEl.cls = "bad";
        issueEl.text = "Сайт zakupki.gov.ru отклонён из-за проблемы с сертификатом";
        issueEl.cls = "bad";
        nextEl.text = "Проверить сертификаты/антивирус/VPN и повторить поиск";
        nextEl.cls = "warn";
      } else if (joined.includes("ERR_NETWORK_ACCESS_DENIED")) {
        resultEl.text = "Новых закупок не получено";
        resultEl.cls = "bad";
        issueEl.text = "Нет доступа к zakupki.gov.ru из браузера Playwright";
        issueEl.cls = "bad";
        nextEl.text = "Проверить VPN, прокси, фаервол или блокировку сети";
        nextEl.cls = "warn";
      } else if (pr.exit_code === 0) {
        const found = foundMatch ? Number(foundMatch[1]) : null;
        const added = addedMatch ? Number(addedMatch[1]) : null;
        if (found === null) {
          resultEl.text = "Поиск завершён";
          resultEl.cls = "ok";
          issueEl.text = added !== null ? ("Добавлено в базу: " + added) : "Итог поиска сохранён";
          issueEl.cls = "ok";
          nextEl.text = "Проверить список закупок ниже";
        } else if (found === 0) {
          resultEl.text = "Подходящих закупок не найдено";
          resultEl.cls = "warn";
          issueEl.text = "По текущим регионам и ключевым словам результат пустой";
          issueEl.cls = "warn";
          nextEl.text = "Расширить параметры поиска или проверить доступ к ЕИС";
        } else {
          resultEl.text = "Поиск завершён";
          resultEl.cls = "ok";
          issueEl.text = "Найдено: " + found + (added !== null ? " · новых в базе: " + added : "");
          issueEl.cls = "ok";
          nextEl.text = "Проверить список закупок ниже";
        }
      } else if (pr.exit_code !== null && pr.exit_code !== undefined) {
        resultEl.text = "Поиск завершился с ошибкой";
        resultEl.cls = "bad";
        issueEl.text = "Подробности скрыты в технических деталях";
        issueEl.cls = "warn";
        nextEl.text = "Открыть детали и посмотреть последнюю ошибку";
      }

      return { result: resultEl, issue: issueEl, next: nextEl };
    }

    let siteChatOpen = false;
    let siteChatLastKey = "";

    function toggleSiteChat(force) {
      const panel = document.getElementById("siteChatPanel");
      const fab = document.getElementById("siteChatFab");
      if (!panel) return;
      siteChatOpen = typeof force === "boolean" ? force : panel.hidden;
      panel.hidden = !siteChatOpen;
      if (siteChatOpen && fab) fab.classList.remove("has-new");
      const feed = document.getElementById("siteChatFeed");
      if (siteChatOpen && feed) feed.scrollTop = feed.scrollHeight;
    }

    function renderSiteChat(events) {
      const feed = document.getElementById("siteChatFeed");
      const fab = document.getElementById("siteChatFab");
      if (!feed) return;
      const list = Array.isArray(events) ? events.slice(-90) : [];
      const last = list.length ? JSON.stringify(list[list.length - 1]) : "";
      if (last && last !== siteChatLastKey && !siteChatOpen && fab) fab.classList.add("has-new");
      siteChatLastKey = last;
      feed.replaceChildren();
      if (!list.length) {
        const empty = document.createElement("div");
        empty.className = "site-chat-empty";
        empty.textContent = "Пока событий нет. Когда запустится сравнение цен, здесь появятся сообщения как в Telegram.";
        feed.appendChild(empty);
        return;
      }
      for (const ev of list) {
        const msg = document.createElement("div");
        const kind = String(ev.kind || "");
        msg.className = "site-chat-msg" + (kind ? " is-" + kind : "");
        const meta = document.createElement("div");
        meta.className = "site-chat-meta";
        const ts = String(ev.ts || "").replace("T", " ");
        const tid = ev.tender_id ? " · " + ev.tender_id : "";
        meta.textContent = (ts || "сейчас") + tid;
        const text = document.createElement("div");
        text.className = "site-chat-text";
        const icon = kind === "done" ? "✅" : (kind === "error" || kind === "warn") ? "⚠️" : kind === "begin" ? "🔎" : "🧾";
        const rawText = String(ev.text || "");
        text.textContent = rawText.startsWith(icon) ? rawText : (icon + " " + rawText);
        msg.appendChild(meta);
        msg.appendChild(text);
        feed.appendChild(msg);
      }
      if (siteChatOpen) feed.scrollTop = feed.scrollHeight;
    }

    async function refreshStatus() {
      let pr = { running: false };
      let mr = { running: false };
      try {
        const rp = await fetch("/api/parse-status");
        if (rp.ok) pr = await rp.json();
      } catch (e) {}
      try {
        const rm = await fetch("/api/merge-site-status");
        if (rm.ok) mr = await rm.json();
      } catch (e) {}
      try {
        if (pr.running) parsePendingUntil = 0;
        const parsePending = !pr.running && Date.now() < parsePendingUntil;
        parseRunning = !!pr.running || parsePending;
        if (pr.running && pr.started_at) {
          const ms = Date.parse(pr.started_at);
          parseStartMs = Number.isNaN(ms) ? null : ms;
        } else if (!parsePending) {
          parseStartMs = null;
        }

        const hasParseHistory = !!(
          (pr.log_tail && pr.log_tail.length)
          || pr.command
          || pr.ended_at
          || pr.exit_code !== null && pr.exit_code !== undefined
        );
        const panel = document.getElementById("parseProgressPanel");
        if (panel) panel.hidden = !(parseRunning || hasParseHistory);

        const progressView = parseProgressView(pr, parsePending);
        const label = document.getElementById("parseProgressLabel");
        if (label) label.textContent = progressView.title;

        const lc = document.getElementById("parseProgressLogCount");
        if (lc && (parseRunning || hasParseHistory)) {
          const n = pr.log_lines_count ?? 0;
          lc.textContent = parsePending
            ? "Поиск запускается."
            : pr.running
            ? "Строк в логе: " + n + " (растёт, пока идёт вывод)."
            : "Строк в логе: " + n + ".";
        } else if (lc) lc.textContent = "";

        const bar = document.getElementById("parseBarFill");
        if (bar) {
          if (parseRunning && progressView.indeterminate) {
            bar.classList.add("running");
          } else {
            bar.classList.remove("running");
          }
          bar.style.width = Math.min(100, Math.max(0, progressView.percent)) + "%";
        }

        const status = document.getElementById("parseStatus");
        const logs = document.getElementById("parseLogs");
        const cmdLine = document.getElementById("parseCommandLine");
        const summary = parseOutcomeSummary(pr, parsePending);
        const resultMain = document.getElementById("parseResultMain");
        const resultIssue = document.getElementById("parseResultIssue");
        const resultNext = document.getElementById("parseResultNext");
        let st = progressView.detail || (parsePending ? "запускается" : pr.running ? "идёт выполнение" : "ожидание");
        if (!pr.running && pr.exit_code !== null && pr.exit_code !== undefined) {
          st += " · код выхода: " + pr.exit_code;
        }
        if (pr.ended_at && !pr.running) st += " · завершено: " + pr.ended_at;
        status.textContent = st;
        if (resultMain) {
          resultMain.textContent = summary.result.text;
          resultMain.className = "parse-summary-value" + (summary.result.cls ? " " + summary.result.cls : "");
        }
        if (resultIssue) {
          resultIssue.textContent = summary.issue.text;
          resultIssue.className = "parse-summary-value" + (summary.issue.cls ? " " + summary.issue.cls : "");
        }
        if (resultNext) {
          resultNext.textContent = summary.next.text;
          resultNext.className = "parse-summary-value" + (summary.next.cls ? " " + summary.next.cls : "");
        }
        if (cmdLine) {
          cmdLine.textContent = pr.running && pr.command ? "Команда: " + pr.command : "";
        }
        if (logs && (!parsePending || pr.log_tail && pr.log_tail.length)) {
          logs.textContent = (pr.log_tail && pr.log_tail.length ? pr.log_tail.join("\\n") : "");
          logs.scrollTop = logs.scrollHeight;
        }

        if (parseRunning) {
          updateParseElapsed();
        } else {
          const endLine = document.getElementById("parseProgressTime");
          if (endLine) {
            if (pr.started_at && pr.ended_at) {
              const ms1 = Date.parse(pr.started_at);
              const ms2 = Date.parse(pr.ended_at);
              if (!Number.isNaN(ms1) && !Number.isNaN(ms2) && ms2 >= ms1) {
                endLine.textContent = "Длительность: " + formatElapsed((ms2 - ms1) / 1000);
              } else {
                endLine.textContent = pr.ended_at ? ("Завершено: " + pr.ended_at) : "";
              }
            } else {
              endLine.textContent = pr.ended_at ? ("Завершено: " + pr.ended_at) : "";
            }
          }
        }

        const mergeRun = !!mr.running;
        const mp = document.getElementById("mergeSitePanel");
        if (mp) mp.hidden = !mergeRun;

        const pct = typeof mr.percent === "number" ? mr.percent : 0;
        const fill = document.getElementById("mergeBarFill");
        const ptext = document.getElementById("mergePercentText");
        const det = document.getElementById("mergeSiteDetail");
        const mlogs = document.getElementById("mergeSiteLogs");
        const marketDone = Number(mr.market_done || 0);
        const marketTotal = Number(mr.market_total || 0);
        const marketLeft = Number(mr.market_left || Math.max(0, marketTotal - marketDone));
        if (fill) fill.style.width = Math.min(100, Math.max(0, pct)) + "%";
        if (ptext) {
          let text = pct + "% · тендеры " + (mr.done ?? 0) + " / " + (mr.total ?? 0);
          if (marketTotal > 0) text += " · рынок " + marketDone + " / " + marketTotal;
          if (mr.current_tid) text += " · сейчас: " + mr.current_tid;
          ptext.textContent = text;
        }
        if (det) {
          if (mergeRun && marketTotal > 0 && marketDone < marketTotal) {
            det.textContent = "Поиск рынка идёт по строкам сметы: обработано " + marketDone + " из " + marketTotal + ", осталось " + marketLeft + ".";
          } else if (mergeRun && marketTotal > 0 && marketDone >= marketTotal) {
            det.textContent = "Рынок обработал строки сметы, собираем страницу сравнения…";
          } else {
            det.textContent = mergeRun ? "Ищем рыночные цены и собираем страницы сравнения…" : "";
          }
        }
        if (mlogs) {
          mlogs.textContent = (mr.log_tail && mr.log_tail.length ? mr.log_tail.join("\\n") : "");
          mlogs.scrollTop = mlogs.scrollHeight;
        }
        renderSiteChat(mr.chat_events || []);

        const mis = document.getElementById("mergeIdleSummary");
        const mreason = document.getElementById("mergeMissingReason");
        if (mis) {
          if (!mergeRun && mr.last_ended_at) {
            mis.textContent = "Последний прогон сводок: " + mr.last_ended_at + " — " + (mr.last_summary || "");
          } else if (mergeRun) {
            mis.textContent = "";
          }
        }
        if (mreason) {
          const reasons = mr.last_reason_counts || {};
          const txt = "Не удалось обработать: без сметы " + (reasons.no_estimate || 0)
            + ", ошибка поиска цен " + (reasons.market_failed || 0)
            + ", ошибка объединения данных " + (reasons.merge_failed || 0)
            + ", ошибка страницы результата " + (reasons.html_failed || 0);
          mreason.textContent = !mergeRun && mr.last_ended_at ? txt : "";
        }

        window.__mergeRunLive = mergeRun;
        applyToolbarDisabled(parseRunning, mergeRun);
        if (parseAutoReloadArmed && !pr.running && pr.exit_code === 0 && !mergeRun) {
          parseAutoReloadArmed = false;
          window.setTimeout(function() {
            location.reload();
          }, 900);
        }
        if (typeof window._wasMergeRun === "undefined") window._wasMergeRun = false;
        if (window._wasMergeRun && !mergeRun) refreshCoverage();
        window._wasMergeRun = mergeRun;
      } catch (e) {}
    }

    setInterval(refreshStatus, 2000);
    setInterval(refreshCoverage, 5000);
    setInterval(refreshPushState, 5000);
    refreshStatus();
    refreshCoverage();
    refreshPushState();
    updatePushButtonUi();
  </script>
</body>
</html>
"""


TENDERS_SHELL_TEMPLATE = """
<!doctype html>
<html lang="ru">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <link rel="icon" href="/favicon.svg" type="image/svg+xml" />
  <title>Тендеры</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f6f8fb;
      --panel: rgba(255,255,255,.94);
      --line: #dbe5f0;
      --text: #172235;
      --muted: #64748b;
      --accent: #2d6fd2;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: "Segoe UI", Arial, sans-serif;
      background:
        radial-gradient(circle at top left, rgba(92, 149, 224, 0.12), transparent 34%),
        linear-gradient(180deg, #ffffff 0%, var(--bg) 100%);
      color: var(--text);
    }
    .page {
      width: 100%;
      max-width: none;
      margin: 0;
      padding: 18px 18px 0;
    }
    .tabs {
      display: inline-flex;
      gap: 8px;
      flex-wrap: wrap;
      margin-bottom: 18px;
      padding: 6px;
      border-radius: 18px;
      border: 1px solid var(--line);
      background: rgba(255,255,255,.82);
      box-shadow: 0 16px 36px rgba(40, 70, 118, 0.08);
      backdrop-filter: blur(14px);
    }
    .tab {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 38px;
      padding: 0 14px;
      border-radius: 12px;
      color: #35506f;
      text-decoration: none;
      font-size: 14px;
      font-weight: 700;
    }
    .tab.is-active {
      background: linear-gradient(180deg, #ffffff, #eef5ff);
      color: var(--accent);
      box-shadow: inset 0 0 0 1px #cfe0f7;
    }
    .shell {
      position: relative;
      min-height: calc(100vh - 86px);
      border-radius: 0;
      border: 0;
      background: transparent;
      box-shadow: none;
      overflow: visible;
    }
    .loader {
      position: absolute;
      inset: 0;
      display: flex;
      align-items: center;
      justify-content: center;
      padding: 28px;
      background: linear-gradient(180deg, rgba(255,255,255,.92), rgba(246,249,253,.96));
      transition: opacity .32s ease, visibility .32s ease;
      z-index: 3;
    }
    .loader.is-hidden {
      opacity: 0;
      visibility: hidden;
      pointer-events: none;
    }
    .loader-card {
      width: min(780px, 100%);
      padding: 26px;
      border-radius: 24px;
      border: 1px solid #e0e8f3;
      background: var(--panel);
      box-shadow: 0 24px 48px rgba(40, 69, 110, 0.10);
    }
    .loader-card h1 {
      margin: 0 0 8px;
      font-size: 28px;
      line-height: 1.15;
    }
    .loader-card p {
      margin: 0;
      color: var(--muted);
      font-size: 14px;
      line-height: 1.55;
    }
    .bar {
      margin-top: 18px;
      height: 10px;
      border-radius: 999px;
      overflow: hidden;
      background: #e6eef8;
    }
    .bar-fill {
      width: 36%;
      height: 100%;
      border-radius: inherit;
      background: linear-gradient(90deg, #4d8be6, #8abbff);
      animation: loadbar 1.4s ease-in-out infinite;
      transform-origin: left center;
    }
    .skeleton-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 12px;
      margin-top: 18px;
    }
    .sk {
      border-radius: 16px;
      background: linear-gradient(90deg, #eef3fa 20%, #f8fbff 50%, #eef3fa 80%);
      background-size: 220% 100%;
      animation: shimmer 1.35s linear infinite;
    }
    .sk.big { height: 112px; }
    .sk.small { height: 68px; }
    .loader-note {
      margin-top: 14px;
      color: #50657f;
      font-size: 13px;
      line-height: 1.5;
    }
    .loader-note strong { color: #173050; }
    iframe {
      display: block;
      width: 100%;
      min-height: calc(100vh - 86px);
      border: 0;
      background: transparent;
      opacity: 0;
      transition: opacity .28s ease;
    }
    iframe.is-ready { opacity: 1; }
    @keyframes shimmer {
      0% { background-position: 200% 0; }
      100% { background-position: -200% 0; }
    }
    @keyframes loadbar {
      0% { transform: translateX(-105%) scaleX(.85); }
      55% { transform: translateX(150%) scaleX(1.08); }
      100% { transform: translateX(210%) scaleX(.9); }
    }
    @media (max-width: 760px) {
      .page { padding: 16px 12px 0; }
      .shell,
      iframe { min-height: calc(100vh - 78px); }
      .loader-card { padding: 18px; border-radius: 18px; }
      .loader-card h1 { font-size: 22px; }
      .skeleton-grid { grid-template-columns: 1fr; }
    }
  </style>
</head>
<body>
  <div class="page">
    <nav class="tabs">
      <a class="tab is-active" href="/tenders">Тендеры</a>
      <a class="tab" href="/estimates">Сметы</a>
      <a class="tab" href="/research">Поиск по позиции</a>
    </nav>

    <section class="shell">
      <div class="loader" id="tendersLoader">
        <div class="loader-card">
          <h1>Загружаем тендеры</h1>
          <p>Страница тяжёлая: здесь много карточек, статусов, ссылок на Excel и прогресса по рынку. Сначала показываем оболочку, потом подтягиваем содержимое без белого экрана.</p>
          <div class="bar"><div class="bar-fill"></div></div>
          <div class="skeleton-grid" aria-hidden="true">
            <div class="sk big"></div>
            <div class="sk big"></div>
            <div class="sk small"></div>
            <div class="sk small"></div>
          </div>
          <div class="loader-note" id="tendersLoaderNote">Если карточек много, это может занять несколько секунд. <strong>Сметы открываются быстрее</strong> и доступны сразу по умолчанию.</div>
        </div>
      </div>
      <iframe id="tendersFrame" src="{{ iframe_src }}" title="Тендеры" loading="eager"></iframe>
    </section>
  </div>
  <script>
    (function() {
      const frame = document.getElementById("tendersFrame");
      const loader = document.getElementById("tendersLoader");
      const note = document.getElementById("tendersLoaderNote");
      let watchdog = setTimeout(function() {
        if (note) {
          note.innerHTML = 'Загрузка идет дольше обычного. Страница всё еще собирается, это не зависание. Можно подождать или открыть <strong>Сметы</strong>.';
        }
      }, 6000);

      frame.addEventListener("load", function() {
        window.clearTimeout(watchdog);
        frame.classList.add("is-ready");
        loader.classList.add("is-hidden");
      });
    })();
  </script>
</body>
</html>
"""


def _html_reports_by_tender_id() -> dict[str, str]:
    """Номер тендера → имя файла ОТЧЕТ_ПО_СМЕТАМ_<id>.html (без общих сводок)."""
    out: dict[str, str] = {}
    if not REPORTS_DIR.exists():
        return out
    prefix = "ОТЧЕТ_ПО_СМЕТАМ_"
    for p in REPORTS_DIR.iterdir():
        if not p.is_file() or not p.name.startswith(prefix) or not p.name.endswith(".html"):
            continue
        if "ОБЩИЙ" in p.name:
            continue
        tid = p.name[len(prefix) : -len(".html")]
        if tid:
            out[tid] = p.name
    return out


def _smet_report_html_has_position_groups(path: Path) -> bool:
    """True, если в отчёте main.py есть раскрытые блоки сметы (не только «Нет данных для отображения»)."""
    try:
        with path.open("r", encoding="utf-8", errors="ignore") as f:
            chunk = f.read(512_000)
    except OSError:
        return False
    return "<details class=\"group\"" in chunk


def _estimate_rows_by_tender_id() -> dict[str, int]:
    """Количество строк-работ в ОТЧЕТ_ПО_СМЕТАМ_<id>.xlsx."""
    out: dict[str, int] = {}
    prefix = "ОТЧЕТ_ПО_СМЕТАМ_"
    if not REPORTS_DIR.is_dir():
        return out
    for p in REPORTS_DIR.glob(f"{prefix}*.xlsx"):
        if "ОБЩИЙ" in p.name:
            continue
        tid = p.stem[len(prefix) :]
        if not tid:
            continue
        try:
            # Считаем строки в файле сметы; без тяжелых вычислений.
            df = pd.read_excel(p, usecols=[0])
            out[tid] = int(len(df))
        except Exception:
            out[tid] = 0
    return out


def _live_market_progress_by_tender() -> dict[str, tuple[int, int]]:
    """Только живой прогресс активного запуска, без чтения всех Excel при открытии списка тендеров."""
    with merge_site_lock:
        running = bool(merge_site_state.get("running"))
        current_tid = str(merge_site_state.get("current_tid") or "").strip()
        done = int(merge_site_state.get("market_done") or 0)
        total = int(merge_site_state.get("market_total") or 0)
    if not running or not current_tid or total <= 0:
        return {}
    return {current_tid: (max(0, done), max(0, total))}


def _market_progress_for_tender(tid: str) -> tuple[int, int]:
    """
    Прогресс Алисы по строкам сметы: (готово, всего) для тех же строк,
    которые реально идут в real_market_scraper.py
    (без явных дублей и без слишком коротких названий).
    """
    from autobot.market_analytics import COL_DUP, COL_NAME

    tid = (tid or "").strip()
    if not tid:
        return 0, 0
    est_path = REPORTS_DIR / f"ОТЧЕТ_ПО_СМЕТАМ_{tid}.xlsx"
    if not est_path.is_file():
        return 0, 0
    try:
        est = pd.read_excel(est_path)
    except Exception:
        return 0, 0
    if COL_NAME not in est.columns:
        return 0, 0

    total = 0
    for _, row in est.iterrows():
        if COL_DUP in est.columns and str(row.get(COL_DUP, "")).strip() == "Да":
            continue
        name = str(row.get(COL_NAME, "") or "").strip()
        if len(name) < 8:
            continue
        total += 1
    if total <= 0:
        return 0, 0

    market_path = _price_output_path_for_tender(tid)
    if not market_path.is_file():
        return 0, total
    try:
        ali = pd.read_excel(market_path)
    except Exception:
        return 0, total
    if COL_NAME not in ali.columns:
        return 0, total

    ren: dict[str, str] = {}
    if "Цены за ед. (рынок, руб)" not in ali.columns and "Цены (строго, руб)" in ali.columns:
        ren["Цены (строго, руб)"] = "Цены за ед. (рынок, руб)"
    if ren:
        ali = ali.rename(columns=ren)

    done = 0
    for _, row in ali.iterrows():
        name = str(row.get(COL_NAME, "") or "").strip()
        if not name:
            continue
        # Частичный файл с рыночными источниками содержит только уже пройденные строки.
        # Для прогресса считаем строку обработанной даже если цена не найдена
        # или ответ был пустым/ошибочным — Telegram считает этот шаг так же.
        done += 1
    return min(done, total), total


def _tender_deadline_text(meta: dict) -> str:
    for key in (
        "deadline_date",
        "end_date",
        "close_date",
        "finish_date",
        "submission_end",
        "application_end",
        "bidding_end_date",
    ):
        txt = str(meta.get(key) or "").strip()
        if txt:
            return txt
    return ""

def _legacy_price_output_path_for_tender(tid: str) -> Path:
    est_path = REPORTS_DIR / f"ОТЧЕТ_ПО_СМЕТАМ_{tid}.xlsx"
    return REPORTS_DIR / f"РЫНОК_ИСТОЧНИКИ_{est_path.stem}.xlsx"


def _market_output_path_for_tender(tid: str) -> Path:
    est_path = REPORTS_DIR / f"ОТЧЕТ_ПО_СМЕТАМ_{tid}.xlsx"
    return REPORTS_DIR / f"РЫНОК_ИСТОЧНИКИ_{est_path.stem}.xlsx"


def _price_output_path_for_tender(tid: str) -> Path:
    market_path = _market_output_path_for_tender(tid)
    if market_path.is_file():
        return market_path
    return _legacy_price_output_path_for_tender(tid)


def _safe_download_stem(title: str, fallback: str) -> str:
    return re.sub(r"[^0-9A-Za-zА-Яа-яЁё._ -]+", "_", str(title or fallback)).strip(" ._") or str(fallback or "report")


def _crm_base_url() -> str:
    return (
        os.environ.get("PMBI_CRM_URL")
        or os.environ.get("PMBI_CRM_BASE_URL")
        or "http://127.0.0.1:8080"
    ).strip().rstrip("/")


def _crm_public_base_url() -> str:
    return (
        os.environ.get("PMBI_CRM_PUBLIC_URL")
        or os.environ.get("PMBI_PUBLIC_BASE_URL")
        or os.environ.get("PMBI_CRM_URL")
        or "http://127.0.0.1:8080"
    ).strip().rstrip("/")


def _crm_project_url(project_id: int, tab: str | None = None) -> str:
    query = f"openProject={int(project_id)}"
    if tab:
        query += f"&tab={quote(str(tab), safe='')}"
    return f"/app/projects?{query}"


def _crm_credentials() -> tuple[str, str] | None:
    login = (os.environ.get("PMBI_CRM_LOGIN") or "").strip()
    password = os.environ.get("PMBI_CRM_PASSWORD") or ""
    if login and password:
        return login, password
    return None


def _float_or_none(value) -> float | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        out = float(value)
    except (TypeError, ValueError):
        return None
    if out != out:
        return None
    return out


def _tender_estimate_materials_for_crm(tender_id: str) -> list[dict]:
    from autobot.market_analytics import COL_DUP, COL_ITEM, COL_NAME, COL_QTY, COL_SUM, COL_UNIT, COL_UNIT_PRICE
    from autobot.market_strategy import build_search_plan

    tid = (tender_id or "").strip()
    if not tid:
        return []
    path = REPORTS_DIR / f"ОТЧЕТ_ПО_СМЕТАМ_{tid}.xlsx"
    if not path.is_file():
        return []
    try:
        df = pd.read_excel(path)
    except Exception:
        return []
    if COL_NAME not in df.columns:
        return []

    try:
        max_rows = int(os.environ.get("PMBI_CRM_MAX_MATERIALS", "2000") or "2000")
    except ValueError:
        max_rows = 2000
    max_rows = max(1, min(max_rows, 10000))

    meta = load_tender_metadata().get(tid, {}) or {}
    region = str(meta.get("region") or "").strip()
    materials: list[dict] = []
    for row_index, (_, row) in enumerate(df.iterrows(), start=1):
        if COL_DUP in df.columns and str(row.get(COL_DUP, "")).strip().casefold() in {"да", "yes", "true", "1"}:
            continue
        title = str(row.get(COL_NAME, "") or "").strip()
        if len(title) < 4:
            continue
        qty = _float_or_none(row.get(COL_QTY))
        unit_price = _float_or_none(row.get(COL_UNIT_PRICE))
        total = _float_or_none(row.get(COL_SUM))
        if qty is None or qty <= 0:
            qty = 1.0
        if unit_price is None or (unit_price <= 0 and total is not None and total > 0):
            unit_price = (total / qty) if total is not None and qty > 0 else 0.0
        source_file = str(row.get("Файл ЛСР", "") or "").strip()
        if source_file.casefold() in {"nan", "none"}:
            source_file = ""
        file_name = Path(source_file).name if source_file else f"Смета тендера {tid}.xlsx"
        estimate_title = Path(file_name).stem or f"Смета тендера {tid}"
        section = _normalize_section_title(str(row.get("Раздел", "") or ""))
        basis_code = str(
            row.get("basis_code", "")
            or row.get("Шифр расценки", "")
            or row.get("Код", "")
            or ""
        ).strip()
        plan = build_search_plan(title, row.get(COL_UNIT, ""), basis_code, section, region)
        item_kind = plan.position.slug
        if item_kind not in {"work", "material", "service", "product", "other"}:
            item_kind = "other"
        notes = [f"Тендер: {tid}"]
        item_no = str(row.get(COL_ITEM, "") or "").strip()
        if item_no:
            notes.append(f"Позиция: {item_no}")
        if total is not None and total > 0:
            notes.append(f"Сумма по смете: {total:.2f} руб.")
        materials.append(
            {
                "title": title[:500],
                "unit": str(row.get(COL_UNIT, "") or "").strip() or "шт",
                "planned_qty": max(0.000001, float(qty)),
                "planned_price": max(0.0, float(unit_price or 0)),
                "planned_total": max(0.0, float(total or (qty * (unit_price or 0)))),
                "article": basis_code,
                "code": basis_code,
                "basis_code": basis_code,
                "item_kind": item_kind,
                "type": item_kind,
                "type_label": plan.position.label,
                "section_title": section or None,
                "section": section or "",
                "estimate_file_name": file_name,
                "estimate_title": estimate_title,
                "source_item_key": f"{item_no or row_index}:{basis_code}:{title[:160]}",
                "notes": "; ".join(notes),
            }
        )
        if len(materials) >= max_rows:
            break
    return materials


def _estimate_materials_for_crm(estimate_id: str) -> list[dict]:
    estimate_id = re.sub(r"[^0-9a-fA-F-]", "", estimate_id or "")[:40]
    if not estimate_id:
        return []
    rows = _load_estimate_rows(estimate_id)
    if not rows:
        return []

    try:
        max_rows = int(os.environ.get("PMBI_CRM_MAX_MATERIALS", "2000") or "2000")
    except ValueError:
        max_rows = 2000
    max_rows = max(1, min(max_rows, 10000))

    meta = _load_estimate_meta(estimate_id) or {}
    file_name = str(meta.get("original_filename") or f"Смета {estimate_id}.xlsx").strip()
    estimate_title = str(meta.get("title") or Path(file_name).stem or f"Смета {estimate_id}").strip()
    materials: list[dict] = []
    for row_index, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            continue
        title = str(row.get("name") or "").strip()
        if len(title) < 4:
            continue
        qty = _float_or_none(row.get("qty"))
        unit_price = _float_or_none(row.get("unit_price"))
        total = _float_or_none(row.get("total"))
        if qty is None or qty <= 0:
            qty = 1.0
        if unit_price is None or (unit_price <= 0 and total is not None and total > 0):
            unit_price = (total / qty) if total is not None and qty > 0 else 0.0
        notes = [f"Смета: {estimate_id}"]
        item_no = str(row.get("item_no") or "").strip()
        if item_no:
            notes.append(f"Позиция: {item_no}")
        section = _normalize_section_title(str(row.get("section") or ""))
        if section:
            notes.append(f"Раздел: {section}")
        sheet = str(row.get("sheet") or "").strip()
        if sheet:
            notes.append(f"Лист: {sheet}")
        excel_row = row.get("excel_row")
        if excel_row not in (None, ""):
            notes.append(f"Строка Excel: {excel_row}")
        basis_code = str(row.get("basis_code") or row.get("code") or row.get("article") or "").strip()
        if basis_code:
            notes.append(f"Код: {basis_code}")
        type_key = str(row.get("type") or "").strip().lower()
        type_label = str(row.get("type_label") or "").strip()
        code_type = _estimate_code_type(basis_code)
        if code_type:
            type_key, type_label = code_type
        if type_label:
            notes.append(f"Тип: {type_label}")
        if total is not None and total > 0:
            notes.append(f"Сумма по смете: {total:.2f} руб.")
        item_kind = type_key if type_key in {"work", "material", "service", "product", "other"} else (type_label or "")
        materials.append(
            {
                "title": title[:500],
                "unit": str(row.get("unit") or "").strip() or "шт",
                "planned_qty": max(0.000001, float(qty)),
                "planned_price": max(0.0, float(unit_price or 0.0)),
                "planned_total": max(0.0, float(total or (qty * (unit_price or 0.0)))),
                "article": basis_code,
                "code": basis_code,
                "basis_code": basis_code,
                "item_kind": item_kind,
                "type": type_key or item_kind,
                "type_label": type_label,
                "section_title": section or None,
                "section": section or "",
                "estimate_file_name": file_name,
                "estimate_title": estimate_title,
                "source_external_id": estimate_id,
                "source_item_key": f"{excel_row or item_no or row_index}:{basis_code}:{title[:160]}",
                "notes": "; ".join(notes),
            }
        )
        if len(materials) >= max_rows:
            break
    return materials


def _estimate_crm_prefill(estimate_id: str) -> dict:
    estimate_id = re.sub(r"[^0-9a-fA-F-]", "", estimate_id or "")[:40]
    meta = _load_estimate_meta(estimate_id) or {}
    rows = _load_estimate_rows(estimate_id)
    summary = _summarize_estimate_rows(rows)
    type_counts = summary.get("type_counts") or {}
    type_labels = {
        "material": "материалы",
        "work": "работы",
        "service": "услуги",
        "product": "товары",
        "other": "прочее",
    }
    type_bits = [f"{type_labels.get(key, key)}: {int(val)}" for key, val in type_counts.items() if int(val or 0) > 0]
    estimate_title = str(meta.get("title") or f"Смета {estimate_id}").strip()[:240]
    original_name = str(meta.get("original_filename") or "").strip()
    created_at = str(meta.get("created_at") or "").strip()
    budget = _float_or_none(summary.get("total_sum")) or 0.0
    description_lines = [
        "Импортировано из auto_bot по отдельной смете.",
        f"Смета: {estimate_title}",
        f"Файл: {original_name}" if original_name else "",
        f"Дата загрузки: {created_at}" if created_at else "",
        f"Строк в смете: {int(summary.get('row_count') or 0)}",
        f"Состав: {', '.join(type_bits)}" if type_bits else "",
    ]
    return {
        "estimate_id": estimate_id,
        "estimate_title": estimate_title,
        "original_filename": original_name,
        "created_at": created_at,
        "row_count": int(summary.get("row_count") or 0),
        "total_sum": budget if budget > 0 else None,
        "total_sum_fmt": _fmt_money(budget) if budget > 0 else "—",
        "project": {
            "title": estimate_title,
            "client_name": "Объект по смете",
            "address": "Адрес уточнить по смете",
            "region": "",
            "contract_no": f"ESTIMATE-{estimate_id}",
            "budget": budget,
            "description": "\n".join(x for x in description_lines if x),
        },
    }


def _build_estimate_crm_project_payload(estimate_id: str, overrides: dict | None = None) -> tuple[dict, list[dict], dict]:
    prefill = _estimate_crm_prefill(estimate_id)
    base_project = dict(prefill.get("project") or {})
    data = overrides if isinstance(overrides, dict) else {}

    def _txt(key: str, default: str, limit: int) -> str:
        raw = str(data.get(key) if key in data else default).strip()
        return raw[:limit]

    budget_raw = data.get("budget")
    if isinstance(budget_raw, str):
        budget_raw = budget_raw.replace(" ", "").replace("\xa0", "").replace(",", ".")
    budget = _float_or_none(budget_raw)
    if budget is None:
        budget = _float_or_none(base_project.get("budget")) or 0.0

    title = _txt("title", str(base_project.get("title") or f"Смета {estimate_id}"), 240) or f"Смета {estimate_id}"
    client_name = _txt("client_name", str(base_project.get("client_name") or "Объект по смете"), 240) or "Объект по смете"
    address = _txt("address", str(base_project.get("address") or "Адрес уточнить по смете"), 500) or "Адрес уточнить по смете"
    region = _txt("region", str(base_project.get("region") or ""), 160)
    contract_no = _txt("contract_no", str(base_project.get("contract_no") or f"ESTIMATE-{estimate_id}"), 120) or f"ESTIMATE-{estimate_id}"
    description = _txt("description", str(base_project.get("description") or ""), 5000)

    project = {
        "title": title,
        "client_name": client_name,
        "address": address,
        "region": region or None,
        "contract_no": contract_no,
        "budget": max(0.0, float(budget or 0.0)),
        "description": description,
    }
    materials = _estimate_materials_for_crm(estimate_id)
    return project, materials, prefill


def _build_crm_project_payload(tender_id: str) -> tuple[dict, list[dict]]:
    meta = load_tender_metadata().get(tender_id, {}) or {}
    title = str(meta.get("title") or f"Тендер {tender_id}").strip()
    region = str(meta.get("region") or "").strip()
    eis_url = eis_notice_url(tender_id, meta.get("url"))
    stage = str(meta.get("stage") or "").strip()
    publish_date = str(meta.get("publish_date") or "").strip()
    price = _float_or_none(meta.get("price_rub")) or 0.0
    description_lines = [
        f"Импортировано из auto_bot по тендеру {tender_id}.",
        f"ЕИС: {eis_url}" if eis_url else "",
        f"Этап закупки: {stage}" if stage else "",
        f"Дата публикации: {publish_date}" if publish_date else "",
        f"Регион: {region}" if region else "",
    ]
    project = {
        "title": title[:240],
        "client_name": "Заказчик из ЕИС",
        "address": region or f"Адрес уточнить по тендеру {tender_id}",
        "region": region or None,
        "contract_no": tender_id,
        "budget": price,
        "description": "\n".join(x for x in description_lines if x),
    }
    materials = _tender_estimate_materials_for_crm(tender_id)
    return project, materials


def export_tender_to_crm(tender_id: str, project_id: int | None = None) -> dict:
    project_payload, materials = _build_crm_project_payload(tender_id)
    base = _crm_base_url()
    tid = str(tender_id or "").strip()
    meta = load_tender_metadata().get(tid, {}) or {}
    source_reference = eis_notice_url(tid, meta.get("url"))

    import requests

    with requests.Session() as session:
        _crm_login(session, base)
        projects = _crm_projects(session, base)
        requested_project_id = _requested_crm_project_id(project_id)
        target = next((row for row in projects if row["id"] == requested_project_id), None) if requested_project_id else None
        if requested_project_id and not target:
            raise RuntimeError("Выбранный объект не найден или недоступен в CRM.")
        if not target:
            target = next((row for row in projects if row["contract_no"] == tid), None)

        created_new = target is None
        if created_new:
            create_resp = session.post(f"{base}/api/projects", json=project_payload, timeout=30)
            if create_resp.status_code >= 400:
                raise RuntimeError(f"CRM не создала объект: HTTP {create_resp.status_code} {create_resp.text[:300]}")
            project = create_resp.json().get("project") or {}
            target_project_id = int(project.get("id") or 0)
            if target_project_id <= 0:
                raise RuntimeError("CRM создала объект, но не вернула project.id.")
        else:
            target_project_id = int(target["id"])

        import_result = _import_crm_estimate(
            session,
            base,
            target_project_id,
            materials,
            {
                "sourceType": "tender",
                "sourceKey": f"tender:{tid}",
                "externalId": tid,
                "tenderId": tid,
                "title": str(project_payload.get("title") or f"Сметы тендера {tid}"),
                "sourceReference": source_reference,
            },
            source_label=f"Сметы тендера {tid}",
            source_reference=source_reference,
        )

        task_summary = {"tasks": 0, "stages": 0}
        if created_new:
            boot_resp = session.post(
                f"{base}/api/projects/{target_project_id}/bootstrap",
                json={
                    "replace_existing": False,
                    "materials": [],
                    "tasks": [
                        {
                            "title": "Проверить тендер и решение об участии",
                            "description": f"Проверить условия закупки {tid}, сметы, сроки и риски перед дальнейшей работой.",
                            "priority": "high",
                        }
                    ],
                },
                timeout=60,
            )
            if boot_resp.status_code >= 400:
                raise RuntimeError(f"Смета импортирована, но стартовая задача не создана: HTTP {boot_resp.status_code} {boot_resp.text[:300]}")
            task_summary = boot_resp.json().get("summary") or task_summary

        summary = {
            "materials": len(import_result.get("items") or []),
            "tasks": int(task_summary.get("tasks") or 0),
            "stages": int(task_summary.get("stages") or 0),
            "estimate_sources": int(import_result.get("estimateSources") or 0),
        }

    return {
        "project_id": target_project_id,
        "project_url": _crm_project_url(target_project_id, "schedule"),
        "materials_sent": int(import_result.get("imported") or len(materials)),
        "summary": summary,
        "already_exists": not created_new,
        "added_to_existing": not created_new,
    }


def export_estimate_to_crm(
    estimate_id: str,
    overrides: dict | None = None,
    project_id: int | None = None,
) -> dict:
    project_payload, materials, prefill = _build_estimate_crm_project_payload(estimate_id, overrides=overrides)
    base = _crm_base_url()
    source_reference = f"/estimates/{estimate_id}"

    import requests

    with requests.Session() as session:
        _crm_login(session, base)
        projects = _crm_projects(session, base)
        requested_project_id = _requested_crm_project_id(project_id)
        target = next((row for row in projects if row["id"] == requested_project_id), None) if requested_project_id else None
        if requested_project_id and not target:
            raise RuntimeError("Выбранный объект не найден или недоступен в CRM.")
        contract_no = str(project_payload.get("contract_no") or "").strip()
        if not target and contract_no:
            target = next((row for row in projects if row["contract_no"] == contract_no), None)

        created_new = target is None
        if created_new:
            create_resp = session.post(f"{base}/api/projects", json=project_payload, timeout=30)
            if create_resp.status_code >= 400:
                raise RuntimeError(f"CRM не создала объект: HTTP {create_resp.status_code} {create_resp.text[:300]}")
            project = create_resp.json().get("project") or {}
            target_project_id = int(project.get("id") or 0)
            if target_project_id <= 0:
                raise RuntimeError("CRM создала объект, но не вернула project.id.")
        else:
            target_project_id = int(target["id"])

        import_result = _import_crm_estimate(
            session,
            base,
            target_project_id,
            materials,
            {
                "sourceType": "estimate",
                "sourceKey": str(estimate_id),
                "externalId": str(estimate_id),
                "title": str(prefill.get("estimate_title") or f"Смета {estimate_id}"),
                "fileName": str(prefill.get("original_filename") or ""),
                "sourceReference": source_reference,
            },
            source_label=str(prefill.get("estimate_title") or f"Смета {estimate_id}"),
            source_reference=source_reference,
        )

        task_summary = {"tasks": 0, "stages": 0}
        if created_new:
            boot_resp = session.post(
                f"{base}/api/projects/{target_project_id}/bootstrap",
                json={
                    "replace_existing": False,
                    "materials": [],
                    "tasks": [
                        {
                            "title": "Проверить смету и подготовить объект",
                            "description": f"Проверить импортированную смету «{prefill.get('estimate_title') or estimate_id}», уточнить материалы, объёмы и план работ.",
                            "priority": "high",
                        }
                    ],
                },
                timeout=60,
            )
            if boot_resp.status_code >= 400:
                raise RuntimeError(f"Смета импортирована, но стартовая задача не создана: HTTP {boot_resp.status_code} {boot_resp.text[:300]}")
            task_summary = boot_resp.json().get("summary") or task_summary

        summary = {
            "materials": len(import_result.get("items") or []),
            "tasks": int(task_summary.get("tasks") or 0),
            "stages": int(task_summary.get("stages") or 0),
            "estimate_sources": int(import_result.get("estimateSources") or 0),
        }

    return {
        "project_id": target_project_id,
        "project_url": _crm_project_url(target_project_id, "schedule"),
        "materials_sent": int(import_result.get("imported") or len(materials)),
        "summary": summary,
        "already_exists": not created_new,
        "added_to_existing": not created_new,
    }


def delete_estimate(estimate_id: str) -> None:
    estimate_id = re.sub(r"[^0-9a-fA-F-]", "", estimate_id or "")[:40]
    if not estimate_id:
        raise RuntimeError("Нужен estimate_id.")
    meta = _load_estimate_meta(estimate_id)
    if not meta:
        raise RuntimeError("Смета не найдена.")

    with estimate_market_lock:
        cur = estimate_market_jobs.get(estimate_id)
        if cur and cur.get("running"):
            raise RuntimeError("Нельзя удалить смету, пока по ней идёт поиск рынка.")
        estimate_market_jobs.pop(estimate_id, None)

    est_dir = _estimate_dir_path(estimate_id)
    try:
        resolved_root = USER_ESTIMATES_DIR.resolve()
        resolved_dir = est_dir.resolve()
    except Exception:
        resolved_root = USER_ESTIMATES_DIR
        resolved_dir = est_dir
    if resolved_dir == resolved_root or resolved_root not in resolved_dir.parents:
        raise RuntimeError("Небезопасный путь удаления сметы.")

    index_items = [x for x in _read_estimates_index() if str(x.get("id") or "") != estimate_id]
    _write_estimates_index(index_items)
    if est_dir.is_dir():
        shutil.rmtree(est_dir)


def collect_sidebar_tenders() -> tuple[list[dict], int, int, int]:
    """
    Все тендеры из tenders.json + признаки: есть файл отчёта и есть ли в нём блоки позиций
    (иначе внутри отчёта только «Нет данных для отображения»).
    """
    from autobot.merge_estimate_market import OUT_PREFIX

    meta = load_tender_metadata()
    reports_map = _html_reports_by_tender_id()
    estimate_ids = set(_estimate_xlsx_tender_ids())
    live_market_progress = _live_market_progress_by_tender()
    merge_root = REPO_ROOT / "data" / "reports_site"
    items: list[dict] = []
    for tid, tmeta in meta.items():
        report_file = reports_map.get(tid, "")
        has_report = bool(report_file) and (REPORTS_DIR / report_file).is_file()
        if not has_report:
            report_file = ""
        rp = REPORTS_DIR / report_file if report_file else None
        has_display_data = bool(rp) and _smet_report_html_has_position_groups(rp)
        has_estimate = tid in estimate_ids
        market_partial_exists = _price_output_path_for_tender(tid).is_file()
        merge_html_exists = (merge_root / tid / "index.html").is_file()
        svodka_exists = (REPORTS_DIR / f"{OUT_PREFIX}{tid}.xlsx").is_file()
        saved_market_done, saved_market_total = _market_progress_for_tender(tid)
        live_market_done, live_market_total = live_market_progress.get(tid, (0, 0))
        if live_market_total > 0:
            market_done, market_total = live_market_done, live_market_total
        else:
            market_done, market_total = saved_market_done, saved_market_total
        market_left = max(0, market_total - market_done)
        market_pct = int(min(100, max(0, round(100.0 * market_done / market_total)))) if market_total > 0 else 0
        stage_raw = (tmeta.get("stage") or "").strip()
        stage_open = stage_raw == STAGE_SUBMISSION
        stage_display = stage_raw if stage_raw else "—"
        items.append(
            {
                "tender_id": tid,
                "display_title": tmeta.get("title") or f"Тендер {tid}",
                "region": tmeta.get("region") or "Без региона",
                "eis_url": eis_notice_url(tid, tmeta.get("url")),
                "has_report": has_report,
                "has_display_data": has_display_data,
                "has_estimate": has_estimate,
                "has_merge_report": merge_html_exists or svodka_exists or market_partial_exists or has_estimate,
                "has_svodka": svodka_exists,
                "has_market_partial": market_partial_exists,
                "report_file": report_file,
                "stage_open": stage_open,
                "stage_display": stage_display,
                "estimate_rows": None,
                "market_progress_done": market_done,
                "market_progress_total": market_total,
                "market_progress_left": market_left,
                "market_progress_percent": market_pct,
                "publish_date": (tmeta.get("publish_date") or "").strip(),
            }
        )
    items.sort(key=lambda x: (x["region"], x["display_title"], x["tender_id"]))
    n_reports = sum(1 for x in items if x["has_report"])
    n_with_data = sum(1 for x in items if x["has_display_data"])
    return items, len(items), n_reports, n_with_data

def _publish_date_sort_key(raw: str, *, newest_first: bool) -> tuple[int, float]:
    """
    Ключ сортировки для даты публикации:
    - сначала валидные даты, потом пустые/неразобранные;
    - поддержка ISO и привычного формата dd.mm.yyyy (с временем или без).
    """
    txt = (raw or "").strip()
    if not txt:
        return 1, 0.0
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(txt, fmt)
            ts = dt.timestamp()
            return 0, (-ts if newest_first else ts)
        except ValueError:
            continue
    try:
        dt = datetime.fromisoformat(txt.replace("Z", "+00:00"))
        ts = dt.timestamp()
        return 0, (-ts if newest_first else ts)
    except ValueError:
        return 1, 0.0


TENDERS_STATUS_LABELS = {
    "download_documents": "Нужны документы",
    "extract_estimate": "Нужна смета",
    "find_market_prices": "Нужны цены",
    "build_comparison": "Нужно сравнение",
    "review": "Готово",
}

TENDERS_STATUS_DETAILS = {
    "download_documents": "Документы еще не скачаны.",
    "extract_estimate": "Документы есть, смета еще не извлечена.",
    "find_market_prices": "Смета готова, рынок еще не собран.",
    "build_comparison": "Цены есть, итоговая страница еще не готова.",
    "review": "Сравнение готово к просмотру.",
}

TENDERS_STATUS_CLASS = {
    "download_documents": "status-attention",
    "extract_estimate": "status-attention",
    "find_market_prices": "status-work",
    "build_comparison": "status-work",
    "review": "status-ready",
}

TENDERS_STATUS_ORDER = {
    "download_documents": 10,
    "extract_estimate": 20,
    "find_market_prices": 30,
    "build_comparison": 40,
    "review": 90,
}

TENDERS_MAIN_ACTION_LABELS = {
    "download_documents": "Скачать документы",
    "extract_estimate": "Извлечь смету",
    "find_market_prices": "Найти цены",
    "build_comparison": "Собрать сравнение",
    "review": "Открыть результат",
}

TENDERS_RUN_TITLES = {
    "download_documents": "Скачиваем документы",
    "extract_estimate": "Извлекаем смету",
    "find_market_prices": "Ищем цены",
    "build_comparison": "Собираем сравнение",
    "review": "Открываем результат",
}

TENDERS_RUN_DETAILS = {
    "download_documents": "Система скачает документы, извлечет смету и продолжит подготовку результата.",
    "extract_estimate": "Система перечитает документы, извлечет смету и продолжит подготовку результата.",
    "find_market_prices": "Система найдет рыночные цены и соберет сравнение.",
    "build_comparison": "Система обновит итоговую таблицу и страницу результата.",
    "review": "Сравнение уже готово.",
}


def _tender_object_name(raw_title: str, tender_id: str) -> str:
    title = re.sub(r"\s+", " ", str(raw_title or "")).strip()
    if not title:
        return ""
    remainder = title.replace(str(tender_id or ""), "")
    remainder = re.sub(r"[№#\s.:;,_—–-]+", "", remainder).casefold()
    if remainder in {"", "тендер", "закупка", "извещение", "документы"}:
        return ""
    return title


def _tender_law_and_method(tender_id: str, url: str, law: str, method: str) -> tuple[str, str]:
    tid = str(tender_id or "").strip()
    href = str(url or "").casefold()
    law_text = str(law or "").strip()
    method_text = str(method or "").strip()
    if not law_text:
        law_text = "223-ФЗ" if len(tid) == 11 or "notice223" in href or "/223/" in href else "44-ФЗ"
    if not method_text and law_text == "44-ФЗ":
        route_methods = {
            "/ea20/": "Электронный аукцион",
            "/ok20/": "Открытый конкурс",
            "/zk20/": "Запрос котировок",
        }
        method_text = next((label for marker, label in route_methods.items() if marker in href), "")
    return law_text, method_text


def _tenders_items() -> tuple[list[dict], dict[str, int]]:
    payload = build_workflow_payload(include_storage=False)
    meta_by_id = load_tender_metadata()
    items = list(payload.get("tenders") or [])
    for item in items:
        tid = str(item.get("tender_id") or "").strip()
        action = str(item.get("next_action") or "").strip() or "download_documents"
        meta_row = meta_by_id.get(tid, {}) or {}
        raw_title = str(meta_row.get("title") or item.get("title") or "").strip()
        object_name = _tender_object_name(raw_title, tid)
        title = object_name or f"Закупка № {tid}"
        price_rub = _float_or_none(item.get("price_rub") if item.get("price_rub") is not None else meta_row.get("price_rub"))
        eis_url = eis_notice_url(tid, meta_row.get("url"))
        law, purchase_method = _tender_law_and_method(
            tid,
            eis_url,
            str(meta_row.get("law") or ""),
            str(meta_row.get("purchase_method") or ""),
        )
        eis_stage = str(item.get("stage") or meta_row.get("stage") or "").strip()
        if eis_stage.casefold() in {"закупки", "этап", "этап закупки", "статус", "статус закупки"}:
            eis_stage = ""
        item["title"] = title
        item["object_name"] = object_name
        item["customer_name"] = str(meta_row.get("customer_name") or "").strip()
        item["updated_date"] = str(meta_row.get("updated_date") or "").strip()
        item["law"] = law
        item["purchase_method"] = purchase_method
        item["law_method_label"] = " · ".join(x for x in (law, purchase_method) if x)
        item["eis_stage"] = eis_stage
        item["status_label"] = TENDERS_STATUS_LABELS.get(action, item.get("next_action_label") or "Следующий шаг")
        item["status_detail"] = TENDERS_STATUS_DETAILS.get(action, "")
        item["status_class"] = TENDERS_STATUS_CLASS.get(action, "status-work")
        item["sort_weight"] = TENDERS_STATUS_ORDER.get(action, 50)
        item["price_fmt"] = _fmt_money(price_rub) if price_rub else "не указана"
        item["price_value"] = float(price_rub or 0)
        item["eis_url"] = eis_url
        item["result_url"] = f"/tenders/{tid}"
        item["main_button_label"] = TENDERS_MAIN_ACTION_LABELS.get(action, item.get("next_action_label") or "Продолжить")
        item["main_run_title"] = TENDERS_RUN_TITLES.get(action, "Продолжаем закупку")
        item["main_run_detail"] = TENDERS_RUN_DETAILS.get(action, "Система выполнит следующий недостающий шаг.")
        item["can_export_crm"] = bool(item.get("has_estimate"))
    items.sort(
        key=lambda x: (
            int(x.get("sort_weight") or 50),
            _publish_date_sort_key(str(x.get("publish_date") or ""), newest_first=True),
            str(x.get("title") or ""),
            str(x.get("tender_id") or ""),
        )
    )
    counts = {str(k): int(v) for k, v in (payload.get("counts") or {}).items()}
    return items, counts


def _tenders_overview(items: list[dict], counts: dict[str, int]) -> dict:
    ready = int(counts.get("review", 0) or 0)
    needs_work = max(0, len(items) - ready)
    return {
        "total": len(items),
        "ready": ready,
        "needs_work": needs_work,
        "needs_prices": int(counts.get("find_market_prices", 0) or 0),
        "needs_docs": int(counts.get("download_documents", 0) or 0),
        "needs_estimate": int(counts.get("extract_estimate", 0) or 0),
        "needs_comparison": int(counts.get("build_comparison", 0) or 0),
    }


@app.route("/dashboard")
def dashboard_redirect():
    return redirect(url_for("index"))


def _render_tenders_board():
    items, counts = _tenders_items()
    selected_action = (request.args.get("action") or "").strip()
    if selected_action == "needs_work":
        visible_items = [x for x in items if str(x.get("next_action") or "") != "review"]
    elif selected_action and selected_action != "all":
        visible_items = [x for x in items if str(x.get("next_action") or "") == selected_action]
    else:
        selected_action = "all"
        visible_items = items
    filters = [
        {"key": "all", "label": "Все", "count": len(items)},
        {"key": "needs_work", "label": "В работе", "count": max(0, len(items) - int(counts.get("review", 0) or 0))},
        {"key": "download_documents", "label": "Документы", "count": counts.get("download_documents", 0)},
        {"key": "extract_estimate", "label": "Сметы", "count": counts.get("extract_estimate", 0)},
        {"key": "find_market_prices", "label": "Цены", "count": counts.get("find_market_prices", 0)},
        {"key": "build_comparison", "label": "Сравнения", "count": counts.get("build_comparison", 0)},
        {"key": "review", "label": "Готово", "count": counts.get("review", 0)},
    ]

    def _filter_counts(key: str, empty_label: str = "") -> list[dict]:
        values: dict[str, int] = {}
        for row in visible_items:
            raw = str(row.get(key) or "").strip()
            value = raw or "__empty__"
            values[value] = values.get(value, 0) + 1
        result = []
        for value, count in sorted(values.items(), key=lambda pair: (pair[0] == "__empty__", pair[0].casefold())):
            result.append({
                "value": value,
                "label": empty_label if value == "__empty__" else value,
                "count": count,
            })
        return result

    law_filters = _filter_counts("law", "Закон не указан")
    stage_filters = _filter_counts("eis_stage", "Статус не указан")
    method_filters = _filter_counts("purchase_method", "Способ не указан")
    region_filters = _filter_counts("region", "Регион не указан")
    return render_template(
        "tenders.html",
        items=visible_items,
        filters=filters,
        selected_action=selected_action,
        overview=_tenders_overview(items, counts),
        law_filters=law_filters,
        stage_filters=stage_filters,
        method_filters=method_filters,
        region_filters=region_filters,
    )


SIMPLE_INDEX_TEMPLATE = """
<!doctype html>
<html lang="ru">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <link rel="icon" href="/favicon.svg" type="image/svg+xml" />
  <title>Тендеры</title>
  <style>
    :root { color-scheme: light; }
    body { margin:0; font-family: Segoe UI, Arial, sans-serif; background:linear-gradient(180deg,#ffffff 0,#f4f7fb 100%); color:#172235; }
    .wrap { max-width: 1180px; margin: 0 auto; padding: 24px; }
    .top { display:flex; justify-content:space-between; gap:16px; align-items:flex-start; margin-bottom: 18px; }
    .card { background:#ffffff; border:1px solid #d9e3ef; border-radius:16px; padding:18px; box-shadow: 0 10px 30px rgba(28,49,84,.08); }
    .muted { color:#62748b; }
    .stats { display:flex; flex-wrap:wrap; gap:12px; }
    .stat { min-width:140px; }
    .stat b { display:block; font-size:22px; margin-top:4px; }
    .group { margin-top: 18px; }
    .group h2 { margin:0 0 10px; font-size:18px; }
    .grid { display:grid; grid-template-columns: repeat(auto-fill, minmax(320px, 1fr)); gap:14px; }
    .tender h3 { margin:0 0 10px; font-size:16px; line-height:1.35; }
    .meta { display:grid; gap:6px; margin-bottom:12px; font-size:14px; }
    .progress { margin: 12px 0; }
    .progress-row { display:flex; justify-content:space-between; gap:8px; font-size:13px; margin-bottom:6px; }
    .track { width:100%; height:10px; background:#edf3fa; border-radius:999px; overflow:hidden; border:1px solid #d6e0ee; }
    .fill { height:100%; background:linear-gradient(90deg, #4f8cff, #63d1ff); }
    .tags { display:flex; flex-wrap:wrap; gap:8px; margin:10px 0 14px; }
    .tag { font-size:12px; padding:5px 9px; border-radius:999px; background:#f4f8fd; border:1px solid #cfd9e8; color:#35506f; }
    .tag.ok { background:#e9f8ef; border-color:#bfe5cc; color:#257347; }
    .tag.warn { background:#fff8e8; border-color:#f0deb1; color:#a06b18; }
    .tag.bad { background:#fff1f1; border-color:#f0c5c5; color:#b04e4e; }
    .actions { display:flex; flex-wrap:wrap; gap:8px; }
    .btn { border:0; border-radius:10px; padding:10px 14px; cursor:pointer; font-size:14px; background:linear-gradient(180deg,#2e80e8,#1f72dc); color:#fff; }
    .btn.secondary { background:#f4f8fd; color:#35506f; border:1px solid #cfd9e8; }
    .btn[disabled] { opacity:.55; cursor:not-allowed; }
    a.btn { text-decoration:none; display:inline-block; }
    .empty { padding:18px; text-align:center; }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <div>
        <h1 style="margin:0 0 8px;">Тендеры</h1>
        <div class="muted">Список закупок, прогресс по поиску цен и быстрые действия.</div>
      </div>
      <div class="card stats">
        <div class="stat"><span class="muted">Всего тендеров</span><b>{{ tender_count }}</b></div>
        <div class="stat"><span class="muted">В показе</span><b>{{ visible_count }}</b></div>
        <div class="stat"><span class="muted">Карточек</span><b>{{ display_report_count }}/{{ report_count }}</b></div>
      </div>
    </div>

    {% if grouped %}
      {% for region, items in grouped %}
      <section class="group">
        <h2>{{ region }}</h2>
        <div class="grid">
          {% for t in items %}
          <article class="card tender">
            <h3>{{ t.display_title }}</h3>
            <div class="meta">
              <div><span class="muted">Тендер:</span> <code>{{ t.tender_id }}</code></div>
              <div><span class="muted">Публикация:</span> {{ t.publish_date or "не указана" }}</div>
              <div><span class="muted">Этап:</span> {{ t.stage_display }}</div>
              <div><span class="muted">Смета:</span> {% if t.has_estimate %}{{ t.estimate_rows }} строк{% else %}ещё не собрана{% endif %}</div>
            </div>

            {% if t.market_progress_total > 0 %}
            <div class="progress">
              <div class="progress-row">
                <span>Поиск цен</span>
                <span>{{ t.market_progress_done }}/{{ t.market_progress_total }}</span>
              </div>
              <div class="track"><div class="fill" style="width: {{ t.market_progress_percent }}%;"></div></div>
            </div>
            {% endif %}

            <div class="tags">
              {% if t.has_merge_report %}
              <span class="tag ok">Карточка готова</span>
              {% elif t.has_market_partial %}
              <span class="tag warn">Есть частичные цены</span>
              {% elif t.has_estimate %}
              <span class="tag ok">Смета готова</span>
              {% else %}
              <span class="tag bad">Нет сметы</span>
              {% endif %}
              <span class="tag">{{ t.stage_display }}</span>
            </div>

            <div class="actions">
              {% if t.has_merge_report %}
              <a class="btn" href="/merge-report/{{ t.tender_id }}/">Открыть карточку</a>
              {% endif %}
              {% if t.has_estimate %}
              <button class="btn secondary" type="button" onclick="runAction('/api/generate-merge-site-one', '{{ t.tender_id }}', 'Запускаю поиск цен…')">Запустить поиск цен</button>
              <button class="btn secondary" type="button" onclick="runAction('/api/generate-merge-site-one-rerun-market', '{{ t.tender_id }}', 'Перезапускаю поиск…')">Перезапустить</button>
              <button class="btn secondary" type="button" onclick="runAction('/api/rebuild-report', '{{ t.tender_id }}', 'Пересобираю карточку…')">Пересобрать карточку</button>
              {% else %}
              <button class="btn secondary" type="button" disabled>Сначала нужна смета</button>
              {% endif %}
              <a class="btn secondary" href="{{ t.eis_url }}" target="_blank" rel="noopener noreferrer">ЕИС</a>
            </div>
          </article>
          {% endfor %}
        </div>
      </section>
      {% endfor %}
    {% else %}
      <div class="card empty">
        <div>Сейчас список пуст.</div>
        <div class="muted" style="margin-top:8px;">Либо ещё нет тендеров, либо включён фильтр, который всё скрывает.</div>
      </div>
    {% endif %}
  </div>

  <script>
    function primeTenderMarketProgress(tenderId, startMessage) {
      const panel = document.getElementById("mergeSitePanel");
      const fill = document.getElementById("mergeBarFill");
      const ptext = document.getElementById("mergePercentText");
      const det = document.getElementById("mergeSiteDetail");
      const logs = document.getElementById("mergeSiteLogs");
      if (panel) panel.hidden = false;
      if (fill) fill.style.width = "3%";
      if (ptext) ptext.textContent = "0% ? ??????? 0 / 1" + (tenderId ? (" ? ??????: " + tenderId) : "");
      if (det) det.textContent = startMessage || "???????? ????? ??? ? ??????? ???????? ??????????";
      if (logs) logs.textContent = (tenderId ? ("????? ?? ??????? " + tenderId) : "????? ???????");
    }

    async function runAction(url, tenderId, startMessage) {
      try {
        const body = tenderId ? { tender_id: tenderId } : {};
        const isMarketRun = String(url || "").includes("generate-merge-site-one");
        if (isMarketRun) {
          primeTenderMarketProgress(tenderId, startMessage);
          if (typeof refreshStatus === "function") refreshStatus();
        } else if (startMessage) {
          alert(startMessage);
        }
        const resp = await fetch(url, {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify(body)
        });
        const data = await resp.json().catch(() => ({}));
        if (!resp.ok) throw new Error(data.error || data.message || ("HTTP " + resp.status));
        if (isMarketRun) {
          if (typeof refreshStatus === "function") refreshStatus();
          if (typeof refreshCoverage === "function") refreshCoverage();
          return;
        }
        alert(data.message || "??????? ??????????.");
        location.reload();
      } catch (err) {
        alert("??????: " + (err.message || err));
      }
    }
  </script>
/body>
</html>
"""


def _render_tenders_index(*, embed_mode: bool = False):
    sidebar_items, tender_count, report_count, display_report_count = collect_sidebar_tenders()
    meta_by_id = load_tender_metadata()
    show_all = (request.args.get("all", "") or "").strip().lower() in ("1", "true", "yes", "on")
    sort_mode = "publish_desc"
    region_options = sorted({str(x.get("region") or "Без региона") for x in sidebar_items})
    selected_region = (request.args.get("region", "") or "").strip()
    if selected_region not in region_options:
        selected_region = ""
    only_submission = not show_all  # True = только «Подача заявок» (режим по умолчанию)
    visible_items = [x for x in sidebar_items if (x.get("stage_open") if only_submission else True)]
    if selected_region:
        visible_items = [x for x in visible_items if str(x.get("region") or "Без региона") == selected_region]
    newest_first = sort_mode == "publish_desc"
    visible_items.sort(
        key=lambda x: (
            _publish_date_sort_key(str(x.get("publish_date") or ""), newest_first=newest_first),
            str(x.get("display_title") or ""),
            str(x.get("tender_id") or ""),
        )
    )
    for item in visible_items:
        meta_row = meta_by_id.get(str(item.get("tender_id") or ""), {}) or {}
        price_rub = _float_or_none(meta_row.get("price_rub"))
        item["price_fmt"] = _fmt_money(price_rub) if price_rub else "—"
        item["deadline_date"] = _tender_deadline_text(meta_row) or "не указано"
    visible_count = len(visible_items)
    rebuild_options = [
        {"tender_id": x["tender_id"], "display_title": x["display_title"]} for x in sidebar_items
    ]
    coverage = _compute_reports_coverage()
    return render_template_string(
        INDEX_TEMPLATE,
        items=visible_items,
        rebuild_options=rebuild_options,
        tender_count=tender_count,
        report_count=report_count,
        display_report_count=display_report_count,
        coverage=coverage,
        show_all=show_all,
        sort_mode=sort_mode,
        visible_count=visible_count,
        region_options=region_options,
        selected_region=selected_region,
        embed_mode=embed_mode,
    )


@app.route("/")
def root_index():
    return redirect(url_for("index"))


@app.route("/tenders")
def index():
    return _render_tenders_board()


@app.route("/tenders/<tender_id>")
def tender_detail_page(tender_id: str):
    tid = str(tender_id or "").strip()
    if not re.fullmatch(r"\d{8,25}", tid):
        abort(404)
    metadata = load_tender_metadata()
    meta = dict(metadata.get(tid) or {})
    estimate_path = REPORTS_DIR / f"ОТЧЕТ_ПО_СМЕТАМ_{tid}.xlsx"
    if not meta and not estimate_path.is_file():
        abort(404)

    workflow_items, _ = _tenders_items()
    workflow = next((dict(item) for item in workflow_items if str(item.get("tender_id") or "") == tid), {})
    if not workflow:
        eis_url = eis_notice_url(tid, meta.get("url"))
        law, method = _tender_law_and_method(tid, eis_url, str(meta.get("law") or ""), str(meta.get("purchase_method") or ""))
        workflow = {
            "tender_id": tid,
            "law": law,
            "purchase_method": method,
            "law_method_label": " · ".join(value for value in (law, method) if value),
            "eis_url": eis_url,
            "eis_stage": str(meta.get("stage") or "").strip(),
            "status_label": "Проверить данные тендера",
            "status_detail": "Продолжите обработку, чтобы получить смету и проверенные цены.",
        }
    tender = build_tender_detail(tid, meta, workflow)
    active_tab = str(request.args.get("tab") or "overview").strip().casefold()
    tender["active_tab"] = "files" if active_tab == "files" else "overview"
    tender["documents"] = list_tender_source_files(tid)
    response = make_response(render_template("tender_detail.html", tender=tender))
    response.headers["Cache-Control"] = "no-store, max-age=0"
    return response


@app.route("/market-audit")
def market_audit_view():
    """Read-only viewer for immutable market evidence captured during verification."""
    record_value = str(request.args.get("record") or "").strip()
    records_root = (REPO_ROOT / "data" / "market_index" / "audit" / "records").resolve()
    try:
        record_path = (REPO_ROOT / record_value).resolve()
        if not record_path.is_relative_to(records_root) or record_path.suffix.casefold() != ".json":
            abort(404)
        payload = json.loads(record_path.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        abort(404)
    if not isinstance(payload, dict):
        abort(404)

    snapshot_value = str(payload.get("snapshot_path") or "").strip()
    snapshot_html = ""
    if snapshot_value:
        blobs_root = (REPO_ROOT / "data" / "market_index" / "audit" / "blobs").resolve()
        try:
            snapshot_path = (REPO_ROOT / snapshot_value).resolve()
            if not snapshot_path.is_relative_to(blobs_root) or snapshot_path.suffix.casefold() != ".gz":
                raise ValueError("invalid audit snapshot path")
            with gzip.open(snapshot_path, "rt", encoding="utf-8", errors="replace") as stream:
                snapshot_html = stream.read(400_000)
        except (OSError, ValueError):
            snapshot_html = ""
    if str(request.args.get("download") or "") == "1" and snapshot_html:
        response = make_response(
            send_file(
                io.BytesIO(snapshot_html.encode("utf-8")),
                mimetype="text/html; charset=utf-8",
                as_attachment=True,
                download_name=f"market-source-{record_path.stem}.html",
                max_age=0,
            )
        )
        response.headers["X-Content-Type-Options"] = "nosniff"
        return response

    response = make_response(render_template_string(
        """<!doctype html><html lang="ru"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
        <title>Аудиторский снимок · AutoBot</title><style>
        body{margin:0;background:#f4f7fb;color:#20334f;font:14px/1.5 Inter,Arial,sans-serif}.page{max-width:1180px;margin:auto;padding:28px}
        header,.card{background:#fff;border:1px solid #dfe7f1;border-radius:14px;padding:20px;box-shadow:0 10px 30px rgba(38,59,88,.06)}
        header{display:flex;justify-content:space-between;gap:20px;align-items:center}h1{margin:4px 0 0;font-size:22px}small{color:#71829a}
        .actions{display:flex;gap:9px}.btn{padding:10px 14px;border-radius:9px;text-decoration:none;color:#fff;background:#1769d2;font-weight:700}.btn.alt{color:#31506f;background:#eef4fb}
        .grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin:16px 0}.grid div{background:#fff;border:1px solid #dfe7f1;border-radius:10px;padding:13px}.grid small,.grid b{display:block}
        pre{margin:0;white-space:pre-wrap;overflow-wrap:anywhere;font:12px/1.45 ui-monospace,Consolas,monospace;color:#344861}.card{max-height:68vh;overflow:auto}
        @media(max-width:720px){header{align-items:flex-start;flex-direction:column}.grid{grid-template-columns:1fr}.page{padding:14px}}
        </style></head><body><main class="page"><header><div><small>AutoBot · доказательство цены</small><h1>{{ title }}</h1></div><div class="actions">{% if has_snapshot %}<a class="btn alt" href="?record={{ record|urlencode }}&download=1">Скачать HTML</a>{% endif %}<a class="btn" href="{{ url }}" target="_blank" rel="noopener noreferrer">Оригинал ↗</a></div></header>
        <section class="grid"><div><small>Цена</small><b>{{ price }} ₽</b></div><div><small>Зафиксировано</small><b>{{ captured }}</b></div><div><small>SHA-256 снимка</small><b>{{ sha or 'нет HTML' }}</b></div></section>
        <section class="card"><small>Сохранённый HTML-код страницы</small><pre>{{ snapshot if snapshot else 'HTML-снимок отсутствует; сохранены URL, время, цена и метаданные проверки.' }}</pre></section></main></body></html>""",
        title=str(payload.get("title") or "Источник цены"),
        price=payload.get("price") or "—",
        captured=str(payload.get("captured_at") or payload.get("observed_at") or "—"),
        sha=str(payload.get("snapshot_sha256") or ""),
        url=str(payload.get("url") or "#"),
        record=record_value,
        snapshot=snapshot_html,
        has_snapshot=bool(snapshot_html),
    ))
    response.headers["Cache-Control"] = "private, no-store, max-age=0"
    response.headers["Content-Security-Policy"] = "default-src 'none'; style-src 'unsafe-inline'; img-src data:; base-uri 'none'; form-action 'none'"
    return response


def _source_file_view_model(tender_id: str, token: str, path: Path) -> dict:
    inventory = list_tender_source_files(tender_id)
    item = next((dict(row) for row in inventory.get("files", []) if row.get("token") == token), None)
    if item is not None:
        return item
    stat = path.stat()
    return {
        "token": token,
        "name": repair_filename(path.name),
        "extension": path.suffix.lstrip(".").upper() or "ФАЙЛ",
        "kind": "other",
        "type_label": "Файл",
        "size_fmt": format_file_size(stat.st_size),
        "updated": datetime.fromtimestamp(stat.st_mtime).strftime("%d.%m.%Y %H:%M"),
    }


@app.route("/tenders/<tender_id>/source-files/<token>/download")
def tender_source_file_download(tender_id: str, token: str):
    try:
        path = resolve_tender_source_file(tender_id, token)
    except (ValueError, FileNotFoundError, OSError):
        abort(404)
    file_model = _source_file_view_model(tender_id, token, path)
    response = make_response(
        send_file(
            path,
            as_attachment=True,
            download_name=file_model["name"],
            conditional=True,
            max_age=0,
        )
    )
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


@app.route("/tenders/<tender_id>/source-files/<token>/preview")
def tender_source_file_preview(tender_id: str, token: str):
    try:
        path = resolve_tender_source_file(tender_id, token)
    except (ValueError, FileNotFoundError, OSError):
        abort(404)
    file_model = _source_file_view_model(tender_id, token, path)
    extension = path.suffix.casefold()
    if extension == ".pdf" or extension in {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}:
        response = make_response(
            send_file(
                path,
                mimetype=mimetypes.guess_type(path.name)[0],
                as_attachment=False,
                download_name=file_model["name"],
                conditional=True,
                max_age=0,
            )
        )
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["Content-Security-Policy"] = "sandbox; default-src 'none'; img-src 'self' data: blob:"
        return response
    try:
        preview = build_source_file_preview(path)
    except Exception as exc:
        preview = {
            "kind": "unavailable",
            "message": f"Не удалось открыть предпросмотр ({type(exc).__name__}). Файл можно скачать без изменений.",
        }
    response = make_response(
        render_template(
            "source_file_preview.html",
            tender_id=str(tender_id),
            file=file_model,
            preview=preview,
            back_url=f"/tenders/{tender_id}?tab=files",
            download_url=f"/tenders/{tender_id}/source-files/{token}/download",
            archive_source_token=token,
        )
    )
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Content-Security-Policy"] = "default-src 'self'; style-src 'self'; img-src 'self' data:; frame-ancestors 'self'"
    return response


@app.route("/tenders/<tender_id>/source-files/<token>/members/<member_token>/download")
def tender_archive_member_download(tender_id: str, token: str, member_token: str):
    try:
        path = resolve_tender_source_file(tender_id, token)
        member = read_archive_member(path, member_token)
    except (ValueError, FileNotFoundError, OSError):
        abort(404)
    response = make_response(
        send_file(
            io.BytesIO(member["data"]),
            mimetype=mimetypes.guess_type(member["name"])[0],
            as_attachment=True,
            download_name=member["name"],
            conditional=True,
            max_age=0,
        )
    )
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Cache-Control"] = "no-store, max-age=0"
    return response


@app.route("/tenders/<tender_id>/source-files/<token>/members/<member_token>/preview")
def tender_archive_member_preview(tender_id: str, token: str, member_token: str):
    try:
        path = resolve_tender_source_file(tender_id, token)
        member = read_archive_member(path, member_token)
    except (ValueError, FileNotFoundError, OSError):
        abort(404)

    extension = Path(member["name"]).suffix.casefold()
    download_url = f"/tenders/{tender_id}/source-files/{token}/members/{member_token}/download"
    if extension == ".pdf" or extension in {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}:
        response = make_response(
            send_file(
                io.BytesIO(member["data"]),
                mimetype=mimetypes.guess_type(member["name"])[0],
                as_attachment=False,
                download_name=member["name"],
                conditional=True,
                max_age=0,
            )
        )
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["Cache-Control"] = "no-store, max-age=0"
        response.headers["Content-Security-Policy"] = "sandbox; default-src 'none'; img-src 'self' data: blob:"
        return response

    try:
        preview = build_source_bytes_preview(member["data"], member["name"], member["chain"])
    except Exception as exc:
        preview = {
            "kind": "unavailable",
            "message": f"Не удалось открыть предпросмотр ({type(exc).__name__}). Файл можно скачать без изменений.",
        }
    response = make_response(
        render_template(
            "source_file_preview.html",
            tender_id=str(tender_id),
            file=member,
            preview=preview,
            back_url=f"/tenders/{tender_id}/source-files/{token}/preview",
            download_url=download_url,
            archive_source_token=token,
        )
    )
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Content-Security-Policy"] = "default-src 'self'; style-src 'self'; img-src 'self' data:; frame-ancestors 'self'"
    return response


@app.route("/tenders/old")
def tenders_old():
    query = request.query_string.decode("utf-8", errors="ignore").strip()
    iframe_src = "/tenders/content"
    if query:
        iframe_src = f"{iframe_src}?{query}"
    return render_template_string(TENDERS_SHELL_TEMPLATE, iframe_src=iframe_src)


@app.route("/tenders/content")
def tenders_content():
    return _render_tenders_index(embed_mode=True)


@app.route("/favicon.svg")
def favicon_svg():
    resp = make_response(FAVICON_SVG)
    resp.headers["Content-Type"] = "image/svg+xml"
    resp.headers["Cache-Control"] = "public, max-age=86400"
    return resp


USER_ESTIMATES_DIR = REPO_ROOT / "data" / "user_estimates"
USER_ESTIMATES_INDEX = USER_ESTIMATES_DIR / "index.json"


def _estimate_upload_allowed(filename: str) -> bool:
    return Path(filename or "").suffix.lower() in (".xlsx", ".xls", ".xlsm", ".pdf")


def _safe_upload_filename(filename: str) -> str:
    raw = Path(filename or "estimate.xlsx").name
    stem = Path(raw).stem
    suffix = Path(raw).suffix.lower()
    stem = re.sub(r"[^0-9A-Za-zА-Яа-я_. -]+", "_", stem).strip(" ._")[:80] or "estimate"
    if suffix not in (".xlsx", ".xls", ".xlsm", ".pdf"):
        suffix = ".xlsx"
    return f"{stem}{suffix}"


def _read_estimates_index() -> list[dict]:
    if not USER_ESTIMATES_INDEX.is_file():
        return []
    try:
        data = json.loads(USER_ESTIMATES_INDEX.read_text(encoding="utf-8"))
    except Exception:
        return []
    return data if isinstance(data, list) else []


def _write_estimates_index(items: list[dict]) -> None:
    USER_ESTIMATES_DIR.mkdir(parents=True, exist_ok=True)
    USER_ESTIMATES_INDEX.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")


def _estimate_meta_path(estimate_id: str) -> Path:
    return USER_ESTIMATES_DIR / estimate_id / "meta.json"


def _estimate_rows_path(estimate_id: str) -> Path:
    return USER_ESTIMATES_DIR / estimate_id / "rows.json"


def _estimate_market_raw_path(estimate_id: str) -> Path:
    return USER_ESTIMATES_DIR / estimate_id / "market_sources.xlsx"


def _estimate_market_merged_path(estimate_id: str) -> Path:
    return USER_ESTIMATES_DIR / estimate_id / "market_compare.xlsx"


def _estimate_dir_path(estimate_id: str) -> Path:
    return USER_ESTIMATES_DIR / estimate_id


def _estimate_market_progress_for_card(estimate_id: str, rows: list[dict] | None = None) -> tuple[int, int]:
    estimate_id = re.sub(r"[^0-9a-fA-F-]", "", estimate_id or "")[:40]
    rows = list(rows or [])
    total = len(rows)
    if total <= 0:
        return 0, 0

    saved_done = 0
    market_path = _estimate_market_raw_path(estimate_id)
    if not market_path.is_file():
        market_path = _estimate_market_merged_path(estimate_id)
    if market_path.is_file():
        try:
            df = pd.read_excel(market_path)
            saved_done = int(len(df.index))
        except Exception:
            saved_done = 0

    with estimate_market_lock:
        job = dict(estimate_market_jobs.get(estimate_id) or {})
    live_total = int(job.get("total") or 0)
    live_done = int(job.get("done") or 0)
    if job.get("running") and live_total > 0:
        return max(0, min(live_done, live_total)), max(0, live_total)
    return max(0, min(saved_done, total)), total


def _load_estimate_meta(estimate_id: str) -> dict | None:
    p = _estimate_meta_path(estimate_id)
    if not p.is_file():
        return None
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None
    return data if isinstance(data, dict) else None


def _load_estimate_rows(estimate_id: str) -> list[dict]:
    p = _estimate_rows_path(estimate_id)
    if not p.is_file():
        return []
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return []
    return data if isinstance(data, list) else []


def _json_num(v) -> float | None:
    try:
        if v is None or pd.isna(v):
            return None
    except Exception:
        if v is None:
            return None
    try:
        f = float(v)
    except Exception:
        return None
    return f if f == f else None


def _estimate_code_type(value: str) -> tuple[str, str] | None:
    text = str(value or "")
    if re.search(r"(?<![\w\u0400-\u04ff])(?:ФСБЦ|FSBC)\s*[-\d]", text, flags=re.IGNORECASE):
        return "material", "Материал"
    if re.search(r"(?<![\w\u0400-\u04ff])(?:ГЭСН|GESN)\s*[A-ZА-Я]?\s*\d", text, flags=re.IGNORECASE):
        return "work", "Работа"
    return None


def _position_type(name: str, unit: str = "", basis_code: str = "") -> tuple[str, str]:
    code_type = _estimate_code_type(f"{basis_code} {name} {unit}")
    if code_type:
        return code_type
    text = f"{name} {unit}".casefold().replace("ё", "е")
    forced_material_keys = (
        "видеокамер",
        "камера ip",
        "камеры видеонаблюден",
        "trassir",
    )
    forced_work_keys = (
        "погруз",
        "перевозк",
        "автосамосвал",
        "комплекс работ",
        "обращен",
        "строительных отход",
        "строительными отход",
    )
    if any(k in text for k in forced_material_keys):
        return "material", "Материал"
    if any(k in text for k in forced_work_keys):
        return "work", "Работа"
    material_keys = (
        "бетон", "раствор", "смесь", "цемент", "песок", "щебень", "грунт", "краска", "эмаль",
        "плитк", "кирпич", "труба", "кабель", "провод", "арматур", "битум", "мастик", "лист",
        "профил", "доска", "брус", "изоляц", "линолеум", "ламинат", "керамзит",
    )
    product_keys = (
        "насос", "шкаф", "щит", "светильник", "радиатор", "кран", "задвижк", "клапан", "вентил",
        "люк", "двер", "окно", "блок", "прибор", "оборудован", "издели", "унитаз", "раковин",
        "смесител", "тройник", "угольник", "муфт", "фланец",
    )
    service_keys = ("аренда", "перевозка", "доставка", "вывоз", "погруз", "разгруз", "обслуживание", "испытание", "пусконалад")
    work_keys = (
        "устройство", "установка", "монтаж", "демонтаж", "разборка", "снятие", "прокладка", "окраска",
        "ремонт", "очистка", "расчистка", "штукатур", "облицов", "сверление", "засыпка", "разработка",
        "укладка", "изоляция", "испытание",
    )
    if any(k in text for k in service_keys):
        return "work", "Работа"
    if any(k in text for k in work_keys):
        return "work", "Работа"
    if any(k in text for k in material_keys):
        return "material", "Материал"
    if any(k in text for k in product_keys):
        return "material", "Материал"
    return "other", "Другое"


def _fmt_money(v: float | None) -> str:
    if v is None:
        return "—"
    return f"{float(v):,.2f}".replace(",", " ").replace(".", ",") + " ₽"


def _fmt_qty(v: float | None) -> str:
    if v is None:
        return "—"
    s = f"{float(v):,.4f}".replace(",", " ").rstrip("0").rstrip(".")
    return s if s else "0"


def _normalize_section_title(value: str) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if not text:
        return ""
    parts = text.split(" ")
    if len(parts) >= 4 and len(parts) % 2 == 0:
        half = len(parts) // 2
        left = " ".join(parts[:half]).strip()
        right = " ".join(parts[half:]).strip()
        if left == right:
            return left
    return text


def _summarize_estimate_rows(rows: list[dict]) -> dict:
    total_sum = 0.0
    has_sum = False
    qty_by_unit: dict[str, float] = {}
    prices: list[float] = []
    type_counts: dict[str, int] = {}
    for r in rows:
        tkey = str(r.get("type") or "other")
        type_counts[tkey] = type_counts.get(tkey, 0) + 1
        sm = _json_num(r.get("total"))
        if sm is not None:
            total_sum += sm
            has_sum = True
        qty = _json_num(r.get("qty"))
        unit = str(r.get("unit") or "без ед.").strip() or "без ед."
        if qty is not None:
            qty_by_unit[unit] = qty_by_unit.get(unit, 0.0) + qty
        up = _json_num(r.get("unit_price"))
        if up is not None and up > 0:
            prices.append(up)
    qty_parts = [f"{_fmt_qty(v)} {u}" for u, v in sorted(qty_by_unit.items(), key=lambda x: x[0])]
    return {
        "row_count": len(rows),
        "total_sum": total_sum if has_sum else None,
        "qty_by_unit": qty_by_unit,
        "qty_text": "; ".join(qty_parts) if qty_parts else "—",
        "avg_price": (sum(prices) / len(prices)) if prices else None,
        "type_counts": type_counts,
    }


def _normalize_selected_estimate_types(raw_values: list[str] | tuple[str, ...] | None) -> list[str]:
    allowed = {"work", "service", "product", "material", "other"}
    out: list[str] = []
    for value in raw_values or []:
        for part in str(value or "").split(","):
            key = part.strip()
            if key and key in allowed and key not in out:
                out.append(key)
    return out


def _filter_estimate_rows(rows_all: list[dict], *, q: str = "", selected_types: list[str] | None = None) -> list[dict]:
    rows = list(rows_all)
    if q:
        q_low = str(q).casefold()
        rows = [r for r in rows if q_low in str(r.get("name") or "").casefold()]
    selected_types = _normalize_selected_estimate_types(selected_types)
    if selected_types:
        allowed = set(selected_types)
        rows = [r for r in rows if str(r.get("type") or "") in allowed]
    return rows


def _estimate_row_to_dict(row) -> dict:
    basis_code = str(getattr(row, "basis_code", "") or "").strip()
    type_key, type_label = _position_type(row.name, row.unit, basis_code)
    return {
        "idx": int(row.idx),
        "name": row.name,
        "unit": row.unit,
        "qty": _json_num(row.qty),
        "unit_price": _json_num(row.unit_price),
        "total": _json_num(row.total),
        "item_no": row.item_no,
        "basis_code": basis_code,
        "code": basis_code,
        "sheet": row.sheet,
        "excel_row": row.excel_row,
        "section": row.section,
        "source": row.source,
        "type": type_key,
        "type_label": type_label,
    }


def _estimate_rows_to_report_df(rows: list[dict]) -> pd.DataFrame:
    from autobot.market_analytics import COL_ITEM, COL_NAME, COL_QTY, COL_SUM, COL_UNIT, COL_UNIT_PRICE

    data: list[dict] = []
    for r in rows:
        data.append(
            {
                COL_ITEM: str(r.get("item_no") or ""),
                COL_NAME: str(r.get("name") or ""),
                COL_UNIT: str(r.get("unit") or ""),
                COL_QTY: _json_num(r.get("qty")),
                COL_UNIT_PRICE: _json_num(r.get("unit_price")),
                COL_SUM: _json_num(r.get("total")),
                "Лист": str(r.get("sheet") or ""),
                "Строка Excel": r.get("excel_row"),
                "Раздел": str(r.get("section") or ""),
                "Тип": str(r.get("type_label") or ""),
            }
        )
    return pd.DataFrame(data)


def _merge_uploaded_estimate_market_df(est_df: pd.DataFrame, market_df: pd.DataFrame) -> pd.DataFrame:
    from autobot.market_analytics import COL_NAME, extract_ruble_amounts, recalc_estimate_qty_price_from_unit
    from autobot.merge_estimate_market import _agg_text, _norm_key, _normalize_market_columns

    est = recalc_estimate_qty_price_from_unit(est_df.copy())
    ali = _normalize_market_columns(market_df.copy())
    if COL_NAME not in est.columns or COL_NAME not in ali.columns:
        return est

    est["__merge_key"] = est[COL_NAME].map(_norm_key)
    ali["__merge_key"] = ali[COL_NAME].map(_norm_key)
    ali = ali.drop(columns=[COL_NAME], errors="ignore")
    agg_cols = [c for c in ali.columns if c != "__merge_key"]
    if agg_cols:
        ali = ali.groupby("__merge_key", as_index=False).agg({c: _agg_text for c in agg_cols})
    merged = est.merge(ali, on="__merge_key", how="left").drop(columns=["__merge_key"], errors="ignore")
    merged = recalc_estimate_qty_price_from_unit(merged)

    def _rub_line(txt) -> str:
        if txt is None or (isinstance(txt, float) and pd.isna(txt)):
            return ""
        amounts = extract_ruble_amounts(str(txt))
        if not amounts:
            return ""
        uniq = sorted({round(x, 2) for x in amounts})[:12]
        return "; ".join(f"{v:,.0f}".replace(",", " ") for v in uniq)

    if "Рыночные источники" in merged.columns:
        merged["Суммы из текста ответа (авто)"] = merged["Рыночные источники"].map(_rub_line)
    if "Цены за ед. (рынок, руб)" in merged.columns:
        strict_prices = merged["Цены за ед. (рынок, руб)"].fillna("").astype(str).str.strip()
        fallback = merged.get("Суммы из текста ответа (авто)", pd.Series([""] * len(merged), index=merged.index))
        merged["Рынок цены за ед. (итог)"] = strict_prices.where(strict_prices != "", fallback)
    elif "Суммы из текста ответа (авто)" in merged.columns:
        merged["Рынок цены за ед. (итог)"] = merged["Суммы из текста ответа (авто)"]
    return merged


def _estimate_market_sections(estimate_id: str, rows_filtered: list[dict], selected_types: list[str] | None = None) -> list[dict]:
    from autobot.market_analytics import COL_NAME
    from autobot.merge_estimate_market import _norm_key

    path = _estimate_market_merged_path(estimate_id)
    if not path.is_file() or not rows_filtered:
        return []
    try:
        df = pd.read_excel(path)
    except Exception:
        return []
    if df.empty or COL_NAME not in df.columns:
        return []

    by_key: dict[str, dict] = {}
    for _, row in df.iterrows():
        key = _norm_key(str(row.get(COL_NAME) or ""))
        if key and key not in by_key:
            by_key[key] = {str(k): row.get(k) for k in df.columns}

    labels_full = {"work": "Работы", "service": "Услуги", "product": "Товары/изделия", "material": "Материалы", "other": "Другое"}
    groups: dict[str, list[dict]] = {}
    for src in rows_filtered:
        key = _norm_key(str(src.get("name") or ""))
        merged = by_key.get(key)
        if not merged:
            continue
        offers_raw = merged.get("Цена-сайт-телефон (json)")
        offers: list[dict] = []
        if isinstance(offers_raw, str) and offers_raw.strip():
            try:
                parsed = json.loads(offers_raw)
                if isinstance(parsed, list):
                    for item in parsed[:5]:
                        if not isinstance(item, dict):
                            continue
                        price_num = _json_num(item.get("price"))
                        offers.append(
                            {
                                "source": str(item.get("source") or "Интернет"),
                                "title": str(item.get("title") or "Источник"),
                                "price": price_num,
                                "price_fmt": _fmt_money(price_num) if price_num else "—",
                                "url": str(item.get("url") or ""),
                                "snippet": str(item.get("snippet") or "")[:320],
                            }
                        )
            except Exception:
                offers = []
        if not offers:
            for i in range(1, 6):
                title = str(merged.get(f"Название объявления {i}") or "").strip()
                url = str(merged.get(f"Ссылка объявления {i}") or "").strip()
                if not title and not url:
                    continue
                price_num = _json_num(merged.get(f"Цена объявления {i}"))
                offers.append(
                    {
                        "source": str(merged.get(f"Источник {i}") or "Интернет"),
                        "title": title or "Источник",
                        "price": price_num,
                        "price_fmt": _fmt_money(price_num) if price_num else "—",
                        "url": url,
                        "snippet": "",
                    }
                )
        if not offers and not str(merged.get("Ошибка / статус") or "").strip():
            continue
        type_key = str(src.get("type") or "other")
        groups.setdefault(type_key, []).append(
            {
                "name": str(src.get("name") or ""),
                "type": type_key,
                "type_label": str(src.get("type_label") or labels_full.get(type_key, type_key)),
                "unit": str(src.get("unit") or ""),
                "qty_fmt": _fmt_qty(_json_num(src.get("qty"))),
                "estimate_price_fmt": _fmt_money(_json_num(src.get("unit_price"))),
                "estimate_total_fmt": _fmt_money(_json_num(src.get("total"))),
                "market_prices": str(merged.get("Рынок цены за ед. (итог)") or merged.get("Цены за ед. (рынок, руб)") or "").strip(),
                "status": str(merged.get("Ошибка / статус") or "").strip(),
                "offers": offers,
            }
        )

    order = selected_types or ["work", "service", "product", "material", "other"]
    sections: list[dict] = []
    for key in order:
        items = groups.get(key) or []
        if not items:
            continue
        sections.append({"key": key, "label": labels_full.get(key, key), "count": len(items), "items": items})
    if sections:
        return sections
    for key, items in groups.items():
        sections.append({"key": key, "label": labels_full.get(key, key), "count": len(items), "items": items})
    return sections


def _estimate_market_links(estimate_id: str, market_sections: list[dict], *, q: str = "", selected_types: list[str] | None = None) -> list[dict]:
    selected_types = _normalize_selected_estimate_types(selected_types)
    out: list[dict] = []
    for sec in market_sections:
        params: list[tuple[str, str]] = []
        if q:
            params.append(("q", q))
        for t in selected_types:
            params.append(("types", t))
        params.append(("market_type", str(sec.get("key") or "")))
        out.append(
            {
                "key": str(sec.get("key") or ""),
                "label": str(sec.get("label") or ""),
                "count": int(sec.get("count") or 0),
                "href": f"/estimates/{estimate_id}/market-view?{urlencode(params, doseq=True)}",
            }
        )
    return out


def _estimate_market_df_for_rows(path: Path, rows_filtered: list[dict]) -> pd.DataFrame:
    from autobot.market_analytics import COL_NAME
    from autobot.merge_estimate_market import _norm_key, _normalize_market_columns

    if not path.is_file():
        return pd.DataFrame()
    try:
        df = pd.read_excel(path)
    except Exception:
        return pd.DataFrame()
    if getattr(df, "empty", True):
        return pd.DataFrame()
    df = _normalize_market_columns(df)
    if not rows_filtered or COL_NAME not in df.columns:
        return df
    allowed_keys = {_norm_key(str(r.get("name") or "")) for r in rows_filtered if str(r.get("name") or "").strip()}
    if not allowed_keys:
        return df
    try:
        filtered = df[df[COL_NAME].fillna("").astype(str).map(_norm_key).isin(allowed_keys)].copy()
    except Exception:
        filtered = df.copy()
    return filtered


def _table_cell_text(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "—"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        fv = float(value)
        if not math.isfinite(fv):
            return "—"
        if abs(fv - round(fv)) < 1e-9:
            return f"{int(round(fv)):,}".replace(",", " ")
        return f"{fv:,.2f}".replace(",", " ").replace(".", ",")
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return "—"
    if len(text) > 900:
        text = text[:900].rstrip() + "…"
    return text


def _build_table_view_from_df(
    df: pd.DataFrame,
    *,
    preferred_columns: list[str],
    fallback_limit: int = 10,
    max_rows: int = 300,
) -> dict:
    if getattr(df, "empty", True):
        return {"available": False, "columns": [], "rows": [], "truncated": False}
    columns = [c for c in preferred_columns if c in df.columns]
    if not columns:
        columns = [str(c) for c in list(df.columns)[:fallback_limit]]
    rows: list[list[str]] = []
    for _, row in df[columns].head(max_rows).iterrows():
        rows.append([_table_cell_text(row.get(col)) for col in columns])
    return {
        "available": bool(rows),
        "columns": columns,
        "rows": rows,
        "truncated": len(df.index) > len(rows),
    }


def _estimate_table_views(estimate_id: str, rows_filtered: list[dict]) -> dict[str, dict]:
    from autobot.market_analytics import COL_ITEM, COL_NAME, COL_QTY, COL_SUM, COL_UNIT, COL_UNIT_PRICE

    compare_df = _estimate_market_df_for_rows(_estimate_market_merged_path(estimate_id), rows_filtered)
    raw_df = _estimate_market_df_for_rows(_estimate_market_raw_path(estimate_id), rows_filtered)

    compare_view = _build_table_view_from_df(
        compare_df,
        preferred_columns=[
            COL_ITEM,
            "Тип",
            COL_NAME,
            COL_UNIT,
            COL_QTY,
            COL_UNIT_PRICE,
            COL_SUM,
            "Рынок цены за ед. (итог)",
            "Медиана цена за ед. (рынок)",
            "Ошибка / статус",
        ],
    )
    raw_view = _build_table_view_from_df(
        raw_df,
        preferred_columns=[
            COL_ITEM,
            "Тип",
            COL_NAME,
            "Поисковый запрос рынка",
            "Цены за ед. (рынок, руб)",
            "Рыночные источники",
            "Ошибка / статус",
        ],
    )
    return {
        "estimate": {"available": bool(rows_filtered)},
        "compare": compare_view,
        "sources": raw_view,
    }


def _pick_estimate_active_table_view(requested: str, table_views: dict[str, dict]) -> str:
    requested_key = str(requested or "").strip().lower()
    if requested_key in ("estimate", "compare", "sources") and table_views.get(requested_key, {}).get("available"):
        return requested_key
    return "estimate"


def _first_url_from_text(text: object) -> str:
    raw = str(text or "").strip()
    if not raw:
        return ""
    m = re.search(r"https?://[^\s<>'\"]+", raw)
    return m.group(0).strip() if m else ""


def _estimate_compare_rows(rows_filtered: list[dict], compare_df: pd.DataFrame) -> list[dict]:
    from autobot.market_analytics import COL_NAME
    from autobot.merge_estimate_market import _norm_key
    from autobot.tender_viability import _estimate_numeric_for_compare, _market_median_for_row, _rub_col

    by_key: dict[str, dict] = {}
    if not getattr(compare_df, "empty", True) and COL_NAME in compare_df.columns:
        for _, row in compare_df.iterrows():
            key = _norm_key(str(row.get(COL_NAME) or ""))
            if key and key not in by_key:
                by_key[key] = {str(k): row.get(k) for k in compare_df.columns}

    rc = _rub_col(compare_df) if not getattr(compare_df, "empty", True) else None
    out: list[dict] = []
    for src in rows_filtered:
        key = _norm_key(str(src.get("name") or ""))
        merged = by_key.get(key, {})
        section = _normalize_section_title(str(src.get("section") or "")) or "Без раздела"
        market_num = None
        est_num = None
        ratio = None
        if merged and rc:
            row_series = pd.Series(merged)
            est_num = _estimate_numeric_for_compare(row_series)
            market_num = _market_median_for_row(row_series, rc)
            if est_num and market_num and market_num > 0:
                ratio = est_num / market_num
        status = str(merged.get("Ошибка / статус") or "").strip()
        if market_num is None and not status:
            status = "Рынок пока не найден"
        first_url = ""
        if merged:
            first_url = _first_url_from_text(merged.get("Ссылки (строго)") or merged.get("Источники (ссылки/телефоны)") or "")
            if not first_url:
                bundle = merged.get("Цена-сайт-телефон (json)")
                if isinstance(bundle, str) and bundle.strip():
                    try:
                        parsed = json.loads(bundle)
                    except Exception:
                        parsed = []
                    if isinstance(parsed, list):
                        for item in parsed:
                            if isinstance(item, dict) and str(item.get("url") or "").strip():
                                first_url = str(item.get("url") or "").strip()
                                break
            if not first_url:
                for i in range(1, 6):
                    maybe = str(merged.get(f"Ссылка объявления {i}") or "").strip()
                    if maybe:
                        first_url = maybe
                        break
        site = urlparse(first_url).netloc.replace("www.", "") if first_url else ""
        if not site:
            site = str(merged.get("Источник 1") or merged.get("Источник") or "").strip()
        if ratio is None:
            compare_label = "Нет данных"
            compare_class = "muted"
        elif ratio < 0.92:
            compare_label = "Ниже рынка"
            compare_class = "bad"
        elif ratio > 1.08:
            compare_label = "Выше рынка"
            compare_class = "good"
        else:
            compare_label = "Около рынка"
            compare_class = "warn"
        out.append(
            {
                "section": section,
                "type_label": str(src.get("type_label") or ""),
                "name": str(src.get("name") or ""),
                "estimate_price": _fmt_money(est_num) if est_num else _fmt_money(_json_num(src.get("unit_price"))),
                "market_price": _fmt_money(market_num) if market_num else "—",
                "site": site or "—",
                "site_url": first_url,
                "status": status or compare_label,
                "compare_label": compare_label,
                "compare_class": compare_class,
                "ratio": ratio,
                "has_market": market_num is not None,
            }
        )
    return out


def _estimate_source_rows(rows_filtered: list[dict], raw_df: pd.DataFrame) -> list[dict]:
    from autobot.market_analytics import COL_NAME
    from autobot.merge_estimate_market import _norm_key

    by_key: dict[str, dict] = {}
    if not getattr(raw_df, "empty", True) and COL_NAME in raw_df.columns:
        for _, row in raw_df.iterrows():
            key = _norm_key(str(row.get(COL_NAME) or ""))
            if key and key not in by_key:
                by_key[key] = {str(k): row.get(k) for k in raw_df.columns}
    out: list[dict] = []
    for src in rows_filtered:
        key = _norm_key(str(src.get("name") or ""))
        merged = by_key.get(key, {})
        text = str(merged.get("Рыночные источники") or "").strip()
        query = str(merged.get("Поисковый запрос рынка") or "").strip()
        status = str(merged.get("Ошибка / статус") or "").strip() or ("Есть источники" if text else "Нет источников")
        first_url = _first_url_from_text(text) or _first_url_from_text(merged.get("Ссылки (строго)") or "")
        site = urlparse(first_url).netloc.replace("www.", "") if first_url else ""
        out.append(
            {
                "section": _normalize_section_title(str(src.get("section") or "")) or "Без раздела",
                "name": str(src.get("name") or ""),
                "market_price": str(merged.get("Цены за ед. (рынок, руб)") or merged.get("Рынок цены за ед. (итог)") or "—").strip() or "—",
                "site": site or "—",
                "status": status,
                "query": query or "—",
            }
        )
    return out


def _estimate_viability_overview(compare_df: pd.DataFrame, compare_rows: list[dict], scope_info: dict | None = None) -> dict:
    from autobot.tender_viability import build_viability_section_html, compute_viability_stats

    scope_info = scope_info or {}
    if getattr(compare_df, "empty", True):
        if scope_info.get("has_notice"):
            return {
                "available": False,
                "title": "РЫНОК НЕ СОБРАН ДЛЯ ЭТОГО ТИПА",
                "subtitle": str(scope_info.get("text") or ""),
                "tone": "warn",
                "facts": [],
                "groups": [],
                "html": "",
            }
        return {
            "available": False,
            "title": "Недостаточно данных",
            "subtitle": "Сначала нужен поиск рынка хотя бы по части позиций.",
            "tone": "warn",
            "facts": [],
            "groups": [],
            "html": "",
        }
    stats = compute_viability_stats(compare_df)
    if stats.comparable < 3 and scope_info.get("has_notice"):
        title = "РЫНОК НЕ СОБРАН ДЛЯ ЭТОГО ТИПА"
        tone = "warn"
    elif stats.comparable < 3:
        title = "НЕДОСТАТОЧНО ДАННЫХ"
        tone = "warn"
    elif stats.median_ratio is not None and stats.median_ratio > 1.08:
        title = "ВЫГОДНО"
        tone = "good"
    elif stats.median_ratio is not None and stats.median_ratio < 0.92:
        title = "НЕВЫГОДНО"
        tone = "bad"
    else:
        title = "ПОГРАНИЧНО"
        tone = "warn"

    types_seen = []
    for row in compare_rows:
        label = str(row.get("type_label") or "").strip()
        if row.get("has_market") and label and label not in types_seen:
            types_seen.append(label)
    comparable_types = ", ".join(types_seen) if types_seen else "пока без уверенного покрытия"
    facts = [
        {"label": "Сравнимых позиций", "value": str(stats.comparable)},
        {"label": "Без рынка", "value": str(stats.no_market)},
        {"label": "Смета / рынок", "value": (f"{stats.median_ratio:.2f}".replace(".", ",") if stats.median_ratio is not None else "—")},
        {"label": "По сумме", "value": (_fmt_money(stats.comparable_gap_total) if stats.comparable_gap_total is not None else "—")},
    ]
    group_map: dict[str, dict] = {}
    for row in compare_rows:
        section = str(row.get("section") or "Без раздела")
        g = group_map.setdefault(section, {"title": section, "good": 0, "warn": 0, "bad": 0, "none": 0})
        cls = str(row.get("compare_class") or "muted")
        if cls == "good":
            g["good"] += 1
        elif cls == "bad":
            g["bad"] += 1
        elif cls == "warn":
            g["warn"] += 1
        else:
            g["none"] += 1
    groups = list(group_map.values())[:18]
    subtitle = str(scope_info.get("text") or f"Анализ построен по уже найденным данным. Сейчас покрыты: {comparable_types}.")
    return {
        "available": True,
        "title": title,
        "subtitle": subtitle,
        "tone": tone,
        "facts": facts,
        "groups": groups,
        "html": build_viability_section_html(stats, "estimate"),
    }


def _estimate_market_scope_info(meta: dict | None, current_selected_types: list[str] | None) -> dict:
    labels_full = {"work": "Работы", "service": "Услуги", "product": "Товары/изделия", "material": "Материалы", "other": "Другое"}
    analyzed_types = _normalize_selected_estimate_types((meta or {}).get("market_selected_types") or [])
    current_types = _normalize_selected_estimate_types(current_selected_types)
    analyzed_set = set(analyzed_types)
    current_set = set(current_types)
    if not analyzed_types:
        return {
            "has_notice": False,
            "tone": "warn",
            "title": "",
            "text": "",
        }
    if not current_types:
        current_set = analyzed_set
    only_analyzed = current_set.issubset(analyzed_set)
    if only_analyzed:
        return {
            "has_notice": False,
            "tone": "warn",
            "title": "",
            "text": "",
        }
    analyzed_labels = ", ".join(labels_full.get(x, x) for x in analyzed_types)
    missing_labels = ", ".join(labels_full.get(x, x) for x in current_types if x not in analyzed_set)
    return {
        "has_notice": True,
        "tone": "warn",
        "title": "Рынок собран не для всех выбранных типов",
        "text": f"Сейчас в файле рынка есть только: {analyzed_labels}. Для этих типов ещё не собраны цены: {missing_labels}. Поэтому вывод ниже не может честно посчитать их как проанализированные.",
    }


def _simple_compare_export_df(compare_rows: list[dict]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Раздел": row["section"],
                "Наименование": row["name"],
                "Цена сметы": row["estimate_price"],
                "Цена рынка": row["market_price"],
                "Сайт": row["site"],
                "Статус": row["status"],
            }
            for row in compare_rows
        ]
    )


def _simple_sources_export_df(source_rows: list[dict]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Раздел": row["section"],
                "Наименование": row["name"],
                "Цена рынка": row["market_price"],
                "Сайт": row["site"],
                "Статус": row["status"],
            }
            for row in source_rows
        ]
    )


def _estimate_upload_log_append(job: dict, line: str) -> None:
    logs = list(job.get("log_lines") or [])
    stamp = datetime.now().strftime("%H:%M:%S")
    logs.append(f"{stamp} · {line}")
    job["log_lines"] = logs[-20:]


def _estimate_market_log_append(job: dict, line: str) -> None:
    logs = list(job.get("log_lines") or [])
    stamp = datetime.now().strftime("%H:%M:%S")
    logs.append(f"{stamp} · {line}")
    job["log_lines"] = logs[-30:]


def _estimate_upload_set(job_id: str, **updates) -> None:
    with estimate_upload_lock:
        job = estimate_upload_jobs.get(job_id)
        if not job:
            return
        for key, value in updates.items():
            job[key] = value


def _estimate_market_set(estimate_id: str, **updates) -> None:
    with estimate_market_lock:
        job = estimate_market_jobs.get(estimate_id)
        if not job:
            return
        for key, value in updates.items():
            job[key] = value


def _estimate_market_cleanup(max_jobs: int = 16) -> None:
    with estimate_market_lock:
        items = sorted(
            estimate_market_jobs.items(),
            key=lambda kv: str(kv[1].get("started_at") or ""),
            reverse=True,
        )
        keep = dict(items[:max_jobs])
        estimate_market_jobs.clear()
        estimate_market_jobs.update(keep)


def _research_queries_from_text(raw: str, *, limit: int = 8) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for line in str(raw or "").splitlines():
        q = re.sub(r"\s+", " ", line).strip(" \t,;")
        if len(q) < 2:
            continue
        key = q.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(q)
        if len(out) >= limit:
            break
    return out


def _crm_login(session, base: str) -> None:
    creds = _crm_credentials()
    if not creds:
        raise RuntimeError("В .env auto_bot нужно задать PMBI_CRM_LOGIN и PMBI_CRM_PASSWORD.")
    response = session.post(
        f"{base}/api/auth/login",
        json={"login": creds[0], "password": creds[1]},
        timeout=15,
    )
    if response.status_code >= 400:
        raise RuntimeError(f"CRM не приняла логин: HTTP {response.status_code}.")


def _crm_projects(session, base: str) -> list[dict]:
    response = session.get(f"{base}/api/projects", timeout=30)
    if response.status_code >= 400:
        raise RuntimeError(f"CRM не вернула список объектов: HTTP {response.status_code}.")
    rows = response.json().get("projects") or []
    projects: list[dict] = []
    for row in rows:
        try:
            project_id = int(row.get("id") or 0)
        except (TypeError, ValueError):
            continue
        if project_id <= 0:
            continue
        projects.append(
            {
                "id": project_id,
                "title": str(row.get("title") or f"Объект #{project_id}").strip(),
                "contract_no": str(row.get("contract_no") or row.get("contractNo") or "").strip(),
                "address": str(row.get("address") or "").strip(),
            }
        )
    return projects


def crm_projects_for_picker() -> list[dict]:
    import requests

    base = _crm_base_url()
    with requests.Session() as session:
        _crm_login(session, base)
        return _crm_projects(session, base)


def _requested_crm_project_id(value) -> int | None:
    if value in (None, ""):
        return None
    try:
        project_id = int(value)
    except (TypeError, ValueError):
        raise RuntimeError("Некорректный объект CRM.")
    if project_id <= 0:
        raise RuntimeError("Некорректный объект CRM.")
    return project_id


def _import_crm_estimate(
    session,
    base: str,
    project_id: int,
    items: list[dict],
    source: dict,
    *,
    source_label: str,
    source_reference: str = "",
) -> dict:
    if not items:
        raise RuntimeError("В смете нет подходящих строк для добавления в объект.")
    response = session.post(
        f"{base}/api/projects/{project_id}/estimate-import",
        json={
            "items": items,
            "source": source,
            "sourceLabel": source_label,
            "sourceReference": source_reference,
            "replace_source": True,
        },
        timeout=90,
    )
    if response.status_code >= 400:
        raise RuntimeError(f"CRM не импортировала смету: HTTP {response.status_code} {response.text[:300]}")
    return response.json()


def _research_specs_from_text(raw: str, *, limit: int = 5) -> list[dict[str, str]]:
    """Lines use `position | unit`; the unit is required for verified prices."""
    specs: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for line in str(raw or "").splitlines():
        clean = re.sub(r"\s+", " ", line).strip(" \t,;")
        if len(clean) < 2:
            continue
        name, sep, unit = clean.rpartition("|")
        if not sep:
            name, unit = clean, ""
        name = name.strip(" ,;")
        unit = unit.strip(" ,;")[:32]
        if len(name) < 2:
            continue
        key = (name.casefold(), unit.casefold())
        if key in seen:
            continue
        seen.add(key)
        specs.append({"query": name, "unit": unit})
        if len(specs) >= limit:
            break
    return specs


def _estimate_upload_progress_cb(job_id: str):
    def _cb(percent: int, stage: str, detail: str = "") -> None:
        with estimate_upload_lock:
            job = estimate_upload_jobs.get(job_id)
            if not job:
                return
            job["progress"] = max(int(job.get("progress") or 0), int(percent))
            job["stage"] = stage
            job["detail"] = detail
            _estimate_upload_log_append(job, f"{stage}" + (f": {detail}" if detail else ""))
    return _cb


def _estimate_upload_cleanup(max_jobs: int = 16) -> None:
    with estimate_upload_lock:
        items = sorted(
            estimate_upload_jobs.items(),
            key=lambda kv: str(kv[1].get("started_at") or ""),
            reverse=True,
        )
        keep = dict(items[:max_jobs])
        estimate_upload_jobs.clear()
        estimate_upload_jobs.update(keep)


def _run_estimate_upload_worker(job_id: str, *, estimate_id: str, title_raw: str, original_name: str, src_path: Path) -> None:
    try:
        from autobot.estimate_excel_analysis import load_estimate_session

        _estimate_upload_set(job_id, running=True, progress=max(30, int(estimate_upload_jobs.get(job_id, {}).get("progress") or 0)), stage="Файл получен", detail="Запускаю разбор Excel")
        _estimate_upload_set(job_id, updated_at=datetime.now().isoformat(timespec="seconds"))
        session = load_estimate_session(src_path, progress_cb=_estimate_upload_progress_cb(job_id))
        rows = [_estimate_row_to_dict(r) for r in session.rows]
        summary = _summarize_estimate_rows(rows)
        _estimate_upload_set(job_id, progress=97, stage="Сохраняю смету", detail="Записываю карточку и таблицу")
        meta = {
            "id": estimate_id,
            "title": (title_raw or Path(original_name).stem)[:160],
            "original_filename": original_name,
            "created_at": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "row_count": len(rows),
            "total_sum": summary.get("total_sum"),
            "source_path": str(src_path.relative_to(REPO_ROOT)),
        }
        _estimate_rows_path(estimate_id).write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
        _estimate_meta_path(estimate_id).write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
        index_items = [x for x in _read_estimates_index() if str(x.get("id") or "") != estimate_id]
        index_items.insert(0, meta)
        _write_estimates_index(index_items)
        with estimate_upload_lock:
            job = estimate_upload_jobs.get(job_id)
            if job:
                job["running"] = False
                job["ok"] = True
                job["progress"] = 100
                job["stage"] = "Готово"
                job["detail"] = f"Строк: {len(rows)} · смета сохранена"
                job["estimate_id"] = estimate_id
                job["ended_at"] = datetime.now().isoformat(timespec="seconds")
                _estimate_upload_log_append(job, f"Готово: смета сохранена, строк {len(rows)}")
    except Exception as e:
        with estimate_upload_lock:
            job = estimate_upload_jobs.get(job_id)
            if job:
                job["running"] = False
                job["ok"] = False
                job["error"] = str(e)[:500]
                job["stage"] = "Ошибка"
                job["detail"] = "Не удалось распарсить Excel или сохранить смету"
                job["ended_at"] = datetime.now().isoformat(timespec="seconds")
                _estimate_upload_log_append(job, f"Ошибка: {str(e)[:300]}")
    finally:
        _estimate_upload_cleanup()


def _run_estimate_market_worker(estimate_id: str, *, city: str, sources: list[str], selected_types: list[str] | None = None) -> None:
    try:
        from autobot.market_analytics import COL_NAME
        from autobot.market_strategy import build_search_plan
        from autobot.merge_estimate_market import _norm_key
        from autobot.real_market_scraper import (
            AvitoBrowserFetcher,
            _build_output_row,
            _compact_query,
            _dedupe_and_sort,
            _eligible_rows,
            _merge_rows,
            _processed_keys,
            _read_previous,
            _verify_offers,
            search_market,
        )

        meta = _load_estimate_meta(estimate_id) or {}
        rows_json = _load_estimate_rows(estimate_id)
        if not rows_json:
            raise ValueError("У этой сметы нет строк для поиска рынка.")
        selected_types = _normalize_selected_estimate_types(selected_types)
        filtered_rows_json = _filter_estimate_rows(rows_json, selected_types=selected_types)
        if not filtered_rows_json:
            raise ValueError("По выбранным типам позиций нет строк для поиска рынка.")
        est_df = _estimate_rows_to_report_df(filtered_rows_json)
        total_rows = len(_eligible_rows(est_df))
        raw_path = _estimate_market_raw_path(estimate_id)
        merged_path = _estimate_market_merged_path(estimate_id)
        prev = _read_previous(raw_path)
        if prev is not None and not getattr(prev, "empty", True) and COL_NAME in prev.columns:
            allowed_keys = {_norm_key(str(x)) for x in est_df[COL_NAME].fillna("").astype(str).tolist() if str(x).strip()}
            if allowed_keys:
                prev = prev[prev[COL_NAME].fillna("").astype(str).map(_norm_key).isin(allowed_keys)].copy()
        done_keys = _processed_keys(prev)
        eligible = _eligible_rows(est_df)
        total = len(eligible)
        active_sources = list(sources)
        _estimate_market_set(
            estimate_id,
            running=True,
            progress=3,
            stage="Готовлю строки сметы",
            detail=f"К обработке: {total} строк" + (f" · город: {city}" if city else ""),
            selected_types=selected_types,
            updated_at=datetime.now().isoformat(timespec="seconds"),
        )
        use_browser = (os.environ.get("MARKET_AVITO_BROWSER", "1") or "1").strip().lower() not in ("0", "false", "no", "off")
        browser_headless = (os.environ.get("MARKET_AVITO_HEADLESS", "1") or "1").strip().lower() not in ("0", "false", "no", "off")
        max_results = max(1, min(10, int((os.environ.get("MARKET_MAX_RESULTS") or "5").strip() or "5")))
        pause = max(0.0, float((os.environ.get("MARKET_PAUSE_SEC") or "4").strip() or "4"))
        new_rows: list[dict] = []
        with AvitoBrowserFetcher(enabled=use_browser, headless=browser_headless) as browser:
            for seq, (_, row) in enumerate(eligible, start=1):
                with estimate_market_lock:
                    job = estimate_market_jobs.get(estimate_id)
                    stop_requested = bool(job.get("stop_requested")) if job else False
                if stop_requested:
                    with estimate_market_lock:
                        job = estimate_market_jobs.get(estimate_id)
                        if job:
                            job["running"] = False
                            job["ok"] = True
                            job["stage"] = "Остановлено"
                            job["detail"] = f"Остановлено пользователем · обработано {len(done_keys) + len(new_rows)} из {total}"
                            job["ended_at"] = datetime.now().isoformat(timespec="seconds")
                            job["done"] = max(0, len(done_keys) + len(new_rows))
                            job["total"] = total
                            job["has_raw"] = raw_path.is_file()
                            job["has_merged"] = merged_path.is_file()
                            _estimate_market_log_append(job, "Остановлено пользователем")
                    return
                work_name = str(row.get(COL_NAME, "") or "").strip()
                key = _norm_key(work_name)
                if key in done_keys:
                    continue
                plan = build_search_plan(
                    work_name,
                    row.get("Ед. изм.", ""),
                    row.get("basis_code", ""),
                    row.get("Раздел", ""),
                    city,
                )
                queries = list(plan.queries) or [_compact_query(work_name)]
                query = " | ".join(queries)
                _estimate_market_set(
                    estimate_id,
                    progress=max(4, min(96, int(round(((seq - 1) / max(1, total)) * 100)))),
                    stage="Ищу цены",
                    detail=f"{seq}/{total} · {work_name[:140]}",
                    current_item=work_name[:220],
                    done=max(0, len(done_keys) + len(new_rows)),
                    total=total,
                )
                with estimate_market_lock:
                    job = estimate_market_jobs.get(estimate_id)
                    if job:
                        _estimate_market_log_append(job, f"Поиск: {seq}/{total} · {work_name[:160]}")
                offers = []
                error_parts: list[str] = []
                if plan.can_auto_price:
                    primary_sources = [source for source in active_sources if source != "avito"] or active_sources
                    for planned_query in queries[:2]:
                        found, found_error = search_market(
                            planned_query,
                            region="" if plan.queries else city,
                            sources=primary_sources,
                            max_results=max_results,
                            browser_fetcher=browser,
                        )
                        checked = _verify_offers(
                            row,
                            found,
                            plan,
                            browser_fetcher=browser,
                            reference_offers=offers,
                        )
                        offers = _dedupe_and_sort(offers + checked, max_results=max_results)
                        if found_error:
                            error_parts.append(found_error)
                        if sum(1 for offer in offers if offer.verification == "verified") >= min(3, max_results):
                            break
                    if "avito" in active_sources and "web" in active_sources and not any(
                        offer.verification == "verified" for offer in offers
                    ):
                        found, found_error = search_market(
                            queries[0],
                            region="" if plan.queries else city,
                            sources=["avito"],
                            max_results=max_results,
                            browser_fetcher=browser,
                        )
                        checked = _verify_offers(
                            row,
                            found,
                            plan,
                            browser_fetcher=browser,
                            reference_offers=offers,
                        )
                        offers = _dedupe_and_sort(offers + checked, max_results=max_results)
                        if found_error:
                            error_parts.append(found_error)
                err = "; ".join(dict.fromkeys(error_parts))
                err_low = str(err or "").casefold()
                if "avito" in active_sources and ("ограничил доступ" in err_low or "ip/vpn" in err_low or "captcha" in err_low):
                    active_sources = [x for x in active_sources if x != "avito"]
                    with estimate_market_lock:
                        job = estimate_market_jobs.get(estimate_id)
                        if job:
                            _estimate_market_log_append(job, "Авито заблокировал доступ — продолжаю только по интернету")
                new_rows.append(_build_output_row(row, offers=offers, query=query, err=err, plan=plan))
                merged_raw = _merge_rows(prev, new_rows)
                raw_path.parent.mkdir(parents=True, exist_ok=True)
                merged_raw.to_excel(raw_path, index=False)
                merged_df = _merge_uploaded_estimate_market_df(est_df, merged_raw)
                merged_df.to_excel(merged_path, index=False)
                detail = f"{len(offers)} ист." if offers else "ничего не найдено"
                if err and not offers:
                    detail += f" · {err[:120]}"
                with estimate_market_lock:
                    job = estimate_market_jobs.get(estimate_id)
                    if job:
                        _estimate_market_log_append(job, f"Готово: {seq}/{total} · {detail}")
                if pause > 0 and seq < total:
                    time.sleep(pause)
        if new_rows:
            merged_raw = _merge_rows(prev, new_rows)
        else:
            merged_raw = prev
        if merged_raw is None or (hasattr(merged_raw, "empty") and merged_raw.empty and not raw_path.is_file()):
            pd.DataFrame().to_excel(raw_path, index=False)
        else:
            merged_raw.to_excel(raw_path, index=False)
        _merge_uploaded_estimate_market_df(est_df, merged_raw).to_excel(merged_path, index=False)
        meta["market_city"] = city
        meta["market_sources"] = ",".join(sources)
        meta["market_selected_types"] = selected_types
        meta["market_updated_at"] = datetime.now().strftime("%d.%m.%Y %H:%M")
        _estimate_meta_path(estimate_id).write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
        with estimate_market_lock:
            job = estimate_market_jobs.get(estimate_id)
            if job:
                job["running"] = False
                job["ok"] = True
                job["progress"] = 100
                job["stage"] = "Готово"
                job["detail"] = f"Поиск рынка завершён · строк: {total_rows}"
                job["done"] = total
                job["total"] = total
                job["ended_at"] = datetime.now().isoformat(timespec="seconds")
                job["has_raw"] = raw_path.is_file()
                job["has_merged"] = merged_path.is_file()
                _estimate_market_log_append(job, "Готово: файлы рынка сохранены")
    except Exception as e:
        with estimate_market_lock:
            job = estimate_market_jobs.get(estimate_id)
            if job:
                job["running"] = False
                job["ok"] = False
                job["stage"] = "Ошибка"
                job["detail"] = "Не удалось выполнить поиск рынка по этой смете"
                job["error"] = str(e)[:500]
                job["ended_at"] = datetime.now().isoformat(timespec="seconds")
                _estimate_market_log_append(job, f"Ошибка: {str(e)[:300]}")
    finally:
        _estimate_market_cleanup()


ESTIMATES_TEMPLATE = """
<!doctype html>
<html lang="ru">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <link rel="icon" href="/favicon.svg" type="image/svg+xml" />
  <title>Сметы</title>
  <style>
    :root { color-scheme: light; --bg:#f4f7fb; --panel:#ffffff; --panel2:#f7fafe; --border:#d9e3ef; --muted:#62748b; --text:#172235; --accent:#1f72dc; --ok:#2e8b57; }
    body { margin:0; font-family: Segoe UI, Arial, sans-serif; background:linear-gradient(180deg,#ffffff 0,#f4f7fb 100%); color:var(--text); }
    .page { max-width:1220px; margin:0 auto; padding:26px 18px 44px; }
    h1 { margin:0 0 8px; font-size:34px; }
    .sub,.muted { color:var(--muted); }
    .tabs { display:flex; flex-wrap:wrap; gap:8px; margin:14px 0 18px; }
    .tab { display:inline-flex; padding:9px 12px; border-radius:999px; color:#35506f; text-decoration:none; background:#f4f8fd; border:1px solid var(--border); font-weight:700; font-size:13px; }
    .tab.is-active { color:#fff; background:linear-gradient(180deg,#2e80e8,#1f72dc); border-color:#2e80e8; }
    .panel,.card { background:linear-gradient(180deg,#ffffff,#f8fbff); border:1px solid var(--border); border-radius:16px; box-shadow:0 18px 45px rgba(28,49,84,.08); }
    .panel { padding:16px; margin-bottom:16px; }
    .upload-row { display:flex; flex-wrap:wrap; gap:10px; align-items:center; }
    input[type=file], input[type=text], select { background:#fff; border:1px solid #cfd9e8; color:var(--text); border-radius:10px; padding:10px; }
    input[type=text] { min-width:280px; }
    .btn { border:1px solid #2e80e8; background:linear-gradient(180deg,#2e80e8,#1f72dc); color:white; border-radius:10px; padding:10px 14px; font-weight:700; cursor:pointer; text-decoration:none; display:inline-flex; align-items:center; justify-content:center; }
    .btn.secondary { background:#f4f8fd; border-color:#cfd9e8; color:#35506f; }
    .grid { display:grid; grid-template-columns:repeat(auto-fill,minmax(290px,1fr)); gap:12px; }
    .card { padding:14px; text-decoration:none; color:var(--text); }
    .card:hover { border-color:#9ec0ef; transform:translateY(-1px); }
    .card h3 { margin:0 0 9px; font-size:15px; line-height:1.35; }
    .meta { display:grid; grid-template-columns:1fr 1fr; gap:8px; font-size:12px; }
    .meta b { display:block; color:#1b2a41; font-size:15px; margin-top:2px; }
    .table-wrap { overflow-x:auto; overflow-y:visible; border-radius:14px; border:1px solid var(--border); background:#fff; position:relative; }
    table { width:100%; border-collapse:collapse; font-size:12px; }
    th,td { padding:9px 10px; border-bottom:1px solid #e5ecf4; vertical-align:top; }
    th { position:sticky; top:72px; background:#f1f6fc; color:#35506f; text-align:left; z-index:12; }
    tr:hover td { background:#f7faff; }
    .num { text-align:right; font-variant-numeric:tabular-nums; white-space:nowrap; }
    .tag { display:inline-flex; border-radius:999px; padding:3px 8px; border:1px solid #cfd9e8; background:#f4f8fd; color:#35506f; font-size:11px; }
    .summary { display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:10px; margin-top:12px; }
    .summary-item { padding:11px; background:#fff; border:1px solid #dfe7f1; border-radius:12px; }
    .summary-item span { display:block; color:var(--muted); font-size:11px; }
    .summary-item b { display:block; margin-top:4px; font-size:16px; }
    .upload-progress { margin-top:14px; padding:14px; border-radius:14px; border:1px solid #d9e3ef; background:linear-gradient(180deg,#ffffff,#f7fafe); }
    .upload-progress[hidden] { display:none; }
    .upload-progress-head { display:flex; justify-content:space-between; gap:10px; align-items:center; margin-bottom:8px; }
    .upload-progress-title { font-size:14px; font-weight:700; color:#1b2a41; }
    .upload-progress-pct { font-size:13px; color:#2e80e8; font-variant-numeric:tabular-nums; }
    .upload-progress-bar { height:12px; border-radius:999px; overflow:hidden; background:#edf3fa; border:1px solid #d6e0ee; }
    .upload-progress-fill { height:100%; width:0%; background:linear-gradient(90deg, #4f8cff, #5ecf8a); transition:width .28s ease; }
    .upload-progress-stage { margin-top:10px; color:#1f3957; font-size:13px; font-weight:700; }
    .upload-progress-detail { margin-top:5px; color:#9fb0d6; font-size:12px; line-height:1.45; }
    .upload-progress-steps { display:flex; flex-wrap:wrap; gap:8px; margin-top:10px; }
    .upload-step { border:1px solid #cfd9e8; background:#f4f8fd; color:#6d7f96; border-radius:999px; padding:4px 9px; font-size:11px; }
    .upload-step.is-active { color:#fff; border-color:#2e80e8; background:#2e80e8; }
    .upload-step.is-done { color:#257347; border-color:#bfe5cc; background:#e9f8ef; }
    .upload-progress-error { margin-top:10px; color:#b04e4e; font-size:12px; white-space:pre-wrap; }
    .upload-progress-logs { margin-top:10px; border-radius:10px; border:1px solid #dfe7f1; background:#f8fbff; padding:9px; max-height:180px; overflow:auto; font-size:11px; color:#576a84; line-height:1.45; white-space:pre-wrap; }
    .empty { text-align:center; padding:28px; color:var(--muted); }
    @media (max-width:720px){ h1{font-size:28px}.upload-row{align-items:stretch;flex-direction:column}.btn,input[type=text],select{width:100%;box-sizing:border-box}.meta{grid-template-columns:1fr} }
  </style>
</head>
<body>
  <div class="page">
    <h1>Сметы</h1>
    <div class="sub">Загрузите Excel-смету, сохраните её карточкой и смотрите все найденные позиции в таблице.</div>
    <nav class="tabs">
      <a class="tab" href="/tenders">📋 Тендеры</a>
      <a class="tab is-active" href="/estimates">📊 Сметы</a>
      <a class="tab" href="/research">🔎 Поиск по позиции</a>
    </nav>

    <section class="panel">
      <h2 style="margin:0 0 10px;font-size:18px;">Загрузить Excel-смету</h2>
      <form id="estimateUploadForm" class="upload-row">
        <input type="file" name="file" accept=".xlsx,.xls,.xlsm,.pdf,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/pdf" required />
        <input type="text" name="title" placeholder="Название сметы (необязательно)" />
        <button class="btn" type="submit">Загрузить и распарсить</button>
      </form>
      <div id="uploadStatus" class="muted" style="margin-top:10px;"></div>
      <div id="estimateUploadProgress" class="upload-progress" hidden>
        <div class="upload-progress-head">
          <div class="upload-progress-title" id="estimateUploadStage">Подготовка…</div>
          <div class="upload-progress-pct" id="estimateUploadPct">0%</div>
        </div>
        <div class="upload-progress-bar"><div id="estimateUploadFill" class="upload-progress-fill"></div></div>
        <div id="estimateUploadDetail" class="upload-progress-detail"></div>
        <div class="upload-progress-steps" id="estimateUploadSteps">
          <span class="upload-step" data-step="upload">Отправка файла</span>
          <span class="upload-step" data-step="received">Файл получен</span>
          <span class="upload-step" data-step="parse">Разбор Excel</span>
          <span class="upload-step" data-step="catalogue">Каталог позиций</span>
          <span class="upload-step" data-step="save">Сохранение</span>
          <span class="upload-step" data-step="done">Готово</span>
        </div>
        <div id="estimateUploadError" class="upload-progress-error" hidden></div>
        <div id="estimateUploadLogs" class="upload-progress-logs" hidden></div>
      </div>
    </section>

    <section class="panel">
      <h2 style="margin:0 0 10px;font-size:18px;">Список смет</h2>
      {% if estimates %}
      <div class="grid">
        {% for e in estimates %}
        <a class="card" href="/estimates/{{ e.id }}">
          <h3>{{ e.title }}</h3>
          <div class="muted" style="font-size:12px;margin-bottom:10px;">{{ e.original_filename }}</div>
          <div class="meta">
            <div><span class="muted">Строк</span><b>{{ e.row_count }}</b></div>
            <div><span class="muted">Сумма</span><b>{{ e.total_sum_fmt }}</b></div>
            <div><span class="muted">Загружено</span><b style="font-size:12px;">{{ e.created_at }}</b></div>
            <div><span class="muted">Типы</span><b style="font-size:12px;">{{ e.types_text }}</b></div>
          </div>
        </a>
        {% endfor %}
      </div>
      {% else %}
      <div class="empty">Пока нет загруженных смет.</div>
      {% endif %}
    </section>
  </div>
  <script>
    (function() {
      const form = document.getElementById("estimateUploadForm");
      const status = document.getElementById("uploadStatus");
      const panel = document.getElementById("estimateUploadProgress");
      const fill = document.getElementById("estimateUploadFill");
      const pct = document.getElementById("estimateUploadPct");
      const stage = document.getElementById("estimateUploadStage");
      const detail = document.getElementById("estimateUploadDetail");
      const errBox = document.getElementById("estimateUploadError");
      const logs = document.getElementById("estimateUploadLogs");
      const stepNodes = Array.from(document.querySelectorAll("#estimateUploadSteps .upload-step"));
      let activePoll = 0;

      function showProgress() {
        panel.hidden = false;
        logs.hidden = false;
      }

      function markStep(progress, currentStage, done) {
        const stageLow = String(currentStage || "").toLowerCase();
        const currentKey =
          done ? "done"
          : progress < 25 ? "upload"
          : progress < 40 ? "received"
          : stageLow.includes("каталог") ? "catalogue"
          : stageLow.includes("сохраня") ? "save"
          : progress >= 40 ? "parse"
          : "received";
        const order = ["upload", "received", "parse", "catalogue", "save", "done"];
        const currentIndex = order.indexOf(currentKey);
        stepNodes.forEach(function(node) {
          const key = node.getAttribute("data-step");
          const idx = order.indexOf(key);
          node.classList.toggle("is-done", idx >= 0 && idx < currentIndex);
          node.classList.toggle("is-active", key === currentKey);
        });
      }

      function renderProgress(data) {
        const value = Math.max(0, Math.min(100, Number(data.progress || 0)));
        fill.style.width = value + "%";
        pct.textContent = value + "%";
        stage.textContent = data.stage || "Подготовка…";
        detail.textContent = data.detail || "";
        const resultOk = !!data.result_ok;
        status.textContent = data.running ? "Смета обрабатывается…" : (resultOk ? "Смета готова." : (data.error ? "Во время обработки возникла ошибка." : ""));
        if (Array.isArray(data.log_tail) && data.log_tail.length) {
          logs.hidden = false;
          logs.textContent = data.log_tail.join("\\n");
        } else {
          logs.hidden = true;
          logs.textContent = "";
        }
        if (data.error) {
          errBox.hidden = false;
          errBox.textContent = "Ошибка: " + data.error;
        } else {
          errBox.hidden = true;
          errBox.textContent = "";
        }
        markStep(value, data.stage || "", resultOk && !data.running);
      }

      async function pollJob(jobId) {
        const pollId = ++activePoll;
        for (;;) {
          if (pollId !== activePoll) return;
          let resp, data;
          try {
            resp = await fetch("/api/estimates/upload-status/" + encodeURIComponent(jobId), { cache: "no-store" });
            data = await resp.json();
          } catch (e) {
            status.textContent = "Не удалось обновить статус обработки.";
            return;
          }
          if (!resp.ok || !data.ok) {
            status.textContent = (data && data.message) || "Статус обработки недоступен.";
            return;
          }
          renderProgress(data);
          if (!data.running) {
            if (data.result_ok && data.estimate_id) {
              status.textContent = "Готово, открываю смету…";
              setTimeout(function() { location.href = "/estimates/" + data.estimate_id; }, 450);
            }
            return;
          }
          await new Promise(function(resolve) { setTimeout(resolve, 500); });
        }
      }

      form.addEventListener("submit", function(e) {
        e.preventDefault();
        const fd = new FormData(form);
        const xhr = new XMLHttpRequest();
        activePoll += 1;
        showProgress();
        errBox.hidden = true;
        errBox.textContent = "";
        logs.textContent = "";
        logs.hidden = true;
        status.textContent = "Начинаю загрузку файла…";
        renderProgress({ progress: 2, stage: "Отправляю файл", detail: "Загружаю смету на сервер", running: true, result_ok: false, log_tail: [] });
        xhr.open("POST", "/api/estimates/upload");
        xhr.upload.addEventListener("progress", function(ev) {
          if (!ev.lengthComputable) return;
          showProgress();
          const uploadPct = Math.max(2, Math.min(24, Math.round(ev.loaded / ev.total * 24)));
          renderProgress({ progress: uploadPct, stage: "Отправляю файл", detail: "Передано " + ev.loaded + " из " + ev.total + " байт", running: true, result_ok: false, log_tail: [] });
        });
        xhr.onreadystatechange = function() {
          if (xhr.readyState !== 4) return;
          let data = {};
          try { data = JSON.parse(xhr.responseText || "{}"); } catch (e) {}
          if (xhr.status < 200 || xhr.status >= 300 || !data.ok) {
            showProgress();
            renderProgress({
              progress: 100,
              stage: "Ошибка",
              detail: "Загрузка или запуск обработки не удались",
              running: false,
              result_ok: false,
              error: data.message || ("HTTP " + xhr.status),
              log_tail: []
            });
            status.textContent = "Ошибка: " + (data.message || ("HTTP " + xhr.status));
            return;
          }
          showProgress();
          renderProgress({
            progress: Math.max(26, Number(data.progress || 26)),
            stage: data.stage || "Файл получен",
            detail: data.detail || "Сервер принял файл и начал разбор",
            running: true,
            result_ok: false,
            log_tail: data.log_tail || []
          });
          status.textContent = "Файл получен, идёт разбор сметы…";
          pollJob(data.job_id);
        };
        xhr.send(fd);
      });

      document.addEventListener("click", async function(e) {
        const btn = e.target.closest("[data-estimate-delete]");
        if (!btn) return;
        e.preventDefault();
        e.stopPropagation();
        const estimateId = btn.getAttribute("data-estimate-delete") || "";
        const title = btn.getAttribute("data-estimate-title") || "эта смета";
        const ok = confirm(`Удалить смету "${title}"?\n\nБудут удалены карточка сметы, ее строки и сохраненные файлы рынка.`);
        if (!ok) return;
        const initialText = btn.textContent;
        btn.disabled = true;
        btn.textContent = "Удаляем...";
        try {
          const resp = await fetch("/api/estimates/" + encodeURIComponent(estimateId) + "/delete", {
            method: "POST",
            headers: { "Accept": "application/json" },
          });
          let data = {};
          try { data = await resp.json(); } catch (err) {}
          if (!resp.ok || !data.ok) {
            alert(data.message || ("Не удалось удалить смету (HTTP " + resp.status + ")."));
            btn.disabled = false;
            btn.textContent = initialText;
            return;
          }
          const card = btn.closest("[data-estimate-card]");
          if (card) card.remove();
          setTimeout(function() { window.location.reload(); }, 120);
        } catch (err) {
          alert("Не удалось удалить смету: " + err);
          btn.disabled = false;
          btn.textContent = initialText;
        }
      });

      document.addEventListener("click", function(e) {
        if (e.target.closest("[data-estimate-delete]")) return;
        const card = e.target.closest("[data-estimate-open]");
        if (!card) return;
        const href = card.getAttribute("data-estimate-open") || "";
        if (!href) return;
        window.location.href = href;
      });

      const catalogSearch = document.getElementById("estimateCatalogSearch");
      const catalogEmpty = document.getElementById("estimateCatalogEmpty");
      catalogSearch?.addEventListener("input", function() {
        const query = String(catalogSearch.value || "").trim().toLocaleLowerCase("ru");
        const cards = Array.from(document.querySelectorAll("[data-estimate-search]"));
        let visible = 0;
        cards.forEach(function(card) {
          const matches = !query || String(card.getAttribute("data-estimate-search") || "").includes(query);
          card.hidden = !matches;
          if (matches) visible += 1;
        });
        if (catalogEmpty) catalogEmpty.hidden = visible > 0 || !cards.length;
      });
    })();
  </script>
</body>
</html>
"""

ESTIMATES_TEMPLATE_V2 = """
<!doctype html>
<html lang="ru">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <link rel="icon" href="/favicon.svg" type="image/svg+xml" />
  <title>Сметы</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f7f9fc;
      --panel: #ffffff;
      --panel-soft: #f4f8fd;
      --border: #d9e4f0;
      --border-strong: #c8d7e8;
      --text: #182537;
      --muted: #62748b;
      --accent: #1f72dc;
      --ok: #257347;
      --ok-soft: #e8f7ed;
      --warn: #9b6a1b;
      --warn-soft: #fff6df;
      --shadow: 0 18px 44px rgba(33, 63, 110, 0.08);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: "Segoe UI", Arial, sans-serif;
      color: var(--text);
      background:
        radial-gradient(circle at top left, rgba(79, 140, 255, 0.10), transparent 26%),
        radial-gradient(circle at top right, rgba(94, 207, 138, 0.08), transparent 24%),
        linear-gradient(180deg, #ffffff 0%, var(--bg) 38%, #f6f9fc 100%);
    }
    a { color: inherit; }
    .page { max-width: 1280px; margin: 0 auto; padding: 28px 20px 48px; }
    .muted { color: var(--muted); }
    .tabs { display: flex; flex-wrap: wrap; gap: 8px; margin: 0 0 18px; }
    .tab {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 40px;
      padding: 9px 14px;
      border-radius: 999px;
      color: #35506f;
      text-decoration: none;
      background: rgba(255,255,255,0.88);
      border: 1px solid var(--border);
      font-weight: 700;
      font-size: 13px;
      box-shadow: 0 8px 20px rgba(36, 67, 112, 0.05);
    }
    .tab.is-active {
      color: #ffffff;
      background: linear-gradient(180deg, #2e80e8, #1f72dc);
      border-color: #2e80e8;
    }
    .tab:hover { border-color: #a9c4e9; background: #f7fbff; }
    .eyebrow {
      display: inline-flex;
      align-items: center;
      padding: 6px 10px;
      border-radius: 999px;
      border: 1px solid var(--border);
      background: rgba(255,255,255,0.88);
      color: #35506f;
      font-size: 12px;
      font-weight: 700;
      letter-spacing: 0.02em;
      text-transform: uppercase;
    }
    .hero-grid {
      display: grid;
      grid-template-columns: minmax(0, 1.2fr) minmax(360px, 0.8fr);
      gap: 18px;
      margin-bottom: 18px;
      align-items: stretch;
    }
    .hero-card,
    .upload-card,
    .catalog-panel,
    .estimate-card,
    .upload-progress {
      background: linear-gradient(180deg, #ffffff, #f9fbff);
      border: 1px solid var(--border);
      border-radius: 20px;
      box-shadow: var(--shadow);
    }
    .hero-card {
      position: relative;
      overflow: hidden;
      padding: 22px 24px 24px;
      min-height: 184px;
    }
    .hero-card::after {
      content: "";
      position: absolute;
      right: -60px;
      top: -70px;
      width: 220px;
      height: 220px;
      border-radius: 50%;
      background: radial-gradient(circle, rgba(31, 114, 220, 0.12), rgba(31, 114, 220, 0));
      pointer-events: none;
    }
    h1 {
      margin: 12px 0 10px;
      font-size: clamp(34px, 4vw, 48px);
      line-height: 1.02;
      letter-spacing: -0.03em;
    }
    .hero-text {
      max-width: 760px;
      margin: 10px 0 0;
      font-size: 14px;
      line-height: 1.55;
      color: #516984;
    }
    .upload-card {
      padding: 20px;
      display: flex;
      flex-direction: column;
      gap: 14px;
    }
    .panel-title {
      margin: 0;
      font-size: 22px;
      line-height: 1.15;
    }
    .panel-subtitle {
      margin: 6px 0 0;
      font-size: 13px;
      line-height: 1.5;
      color: var(--muted);
    }
    .upload-row { display: grid; gap: 10px; }
    .file-picker {
      display: flex;
      flex-wrap: wrap;
      align-items: center;
      gap: 10px;
    }
    .file-input-native {
      position: absolute;
      width: 1px;
      height: 1px;
      padding: 0;
      margin: -1px;
      overflow: hidden;
      clip: rect(0, 0, 0, 0);
      white-space: nowrap;
      border: 0;
    }
    .file-picker-btn {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 8px;
      min-height: 44px;
      padding: 0 15px;
      border-radius: 12px;
      border: 1px solid #cfd9e8;
      background: linear-gradient(180deg, #ffffff, #f4f8fd);
      color: #35506f;
      font-size: 14px;
      font-weight: 700;
      cursor: pointer;
      transition: border-color .18s ease, background .18s ease, transform .18s ease;
    }
    .file-picker-btn:hover {
      border-color: #a9c4e9;
      background: linear-gradient(180deg, #ffffff, #eef5ff);
      transform: translateY(-1px);
    }
    .file-picker-name {
      color: var(--muted);
      font-size: 13px;
      line-height: 1.4;
      word-break: break-word;
    }
    .file-picker-name[hidden] { display: none; }
    .icon-clip {
      width: 16px;
      height: 16px;
      display: block;
      stroke: currentColor;
      fill: none;
      stroke-width: 1.9;
      stroke-linecap: round;
      stroke-linejoin: round;
      pointer-events: none;
    }
    .upload-hint {
      display: grid;
      gap: 8px;
      padding: 12px 14px;
      border-radius: 16px;
      background: linear-gradient(180deg, #f4f8fd, #eef4fb);
      border: 1px solid var(--border);
      color: #35506f;
      font-size: 13px;
      line-height: 1.5;
    }
    input[type=text] {
      width: 100%;
      min-width: 0;
      background: #ffffff;
      border: 1px solid #cfd9e8;
      color: var(--text);
      border-radius: 12px;
      padding: 12px 13px;
      font-size: 14px;
      outline: none;
    }
    input[type=text]:focus {
      border-color: #8db4eb;
      box-shadow: 0 0 0 3px rgba(46, 128, 232, 0.12);
    }
    .btn {
      border: 1px solid #2e80e8;
      background: linear-gradient(180deg, #2e80e8, #1f72dc);
      color: #ffffff;
      border-radius: 12px;
      padding: 11px 15px;
      font-weight: 700;
      font-size: 14px;
      cursor: pointer;
      text-decoration: none;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 44px;
    }
    .overview-strip {
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      margin-bottom: 18px;
    }
    .overview-pill {
      display: inline-flex;
      align-items: center;
      gap: 10px;
      padding: 10px 14px;
      border-radius: 999px;
      border: 1px solid var(--border);
      background: linear-gradient(180deg, #ffffff, #f7fbff);
      box-shadow: var(--shadow);
      color: #29415f;
      font-size: 13px;
      line-height: 1.2;
      font-weight: 700;
    }
    .overview-pill b {
      color: #172235;
      font-size: 18px;
      line-height: 1;
      letter-spacing: -0.03em;
    }
    .catalog-panel { padding: 18px; }
    .panel-head {
      display: flex;
      flex-wrap: wrap;
      justify-content: space-between;
      align-items: flex-end;
      gap: 12px;
      margin-bottom: 14px;
    }
    .panel-note {
      color: var(--muted);
      font-size: 13px;
      line-height: 1.45;
    }
    .grid {
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(310px, 1fr));
      gap: 14px;
    }
    .estimate-card {
      padding: 16px;
      color: var(--text);
      transition: transform .18s ease, box-shadow .18s ease, border-color .18s ease;
    }
    .estimate-card:hover {
      transform: translateY(-2px);
      border-color: #9ec0ef;
      box-shadow: 0 20px 36px rgba(43, 78, 131, 0.12);
    }
    .estimate-card-body-link {
      display: block;
      color: inherit;
      text-decoration: none;
    }
    .estimate-card-top {
      display: flex;
      justify-content: space-between;
      align-items: flex-start;
      gap: 8px;
      margin-bottom: 10px;
    }
    .estimate-card-meta {
      display: flex;
      align-items: center;
      justify-content: flex-end;
      gap: 8px;
      flex-wrap: wrap;
    }
    .status-pill,
    .date-pill,
    .chip {
      display: inline-flex;
      align-items: center;
      border-radius: 999px;
      border: 1px solid var(--border);
      font-size: 11px;
      font-weight: 700;
      line-height: 1;
    }
    .status-pill { padding: 6px 10px; }
    .date-pill {
      padding: 6px 9px;
      color: var(--muted);
      background: #f8fbff;
    }
    .status-ready { background: var(--ok-soft); border-color: #bfe5cc; color: var(--ok); }
    .status-partial { background: var(--warn-soft); border-color: #f1ddb2; color: var(--warn); }
    .status-idle { background: #fff1f1; border-color: #f0c9c9; color: #bb4545; }
    .estimate-card h3 {
      margin: 0;
      font-size: 17px;
      line-height: 1.35;
      color: #172235;
      display: -webkit-box;
      -webkit-box-orient: vertical;
      -webkit-line-clamp: 2;
      line-clamp: 2;
      overflow: hidden;
      text-overflow: ellipsis;
      min-height: calc(1.35em * 2);
      word-break: break-word;
    }
    .chip-row {
      display: flex;
      flex-wrap: wrap;
      gap: 7px;
      margin-top: 12px;
      min-height: 26px;
    }
    .chip {
      padding: 6px 9px;
      background: var(--panel-soft);
      color: #35506f;
    }
    .estimate-progress {
      margin-top: 14px;
      padding: 11px 12px 12px;
      border-radius: 14px;
      background: linear-gradient(180deg, #ffffff, #f7fafd);
      border: 1px solid #dfe7f1;
    }
    .estimate-progress-head {
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 10px;
      margin-bottom: 8px;
    }
    .estimate-progress-label {
      color: var(--muted);
      font-size: 11px;
      line-height: 1.35;
      text-transform: uppercase;
      letter-spacing: 0.03em;
    }
    .estimate-progress-value {
      color: #1b2a41;
      font-size: 14px;
      font-weight: 800;
      line-height: 1.15;
    }
    .estimate-progress-track {
      position: relative;
      width: 100%;
      height: 7px;
      border-radius: 999px;
      overflow: hidden;
      background: #e8eff7;
    }
    .estimate-progress-fill {
      position: absolute;
      inset: 0 auto 0 0;
      border-radius: inherit;
      background: linear-gradient(90deg, #5b94dd, #84b5f1);
    }
    .estimate-progress-note {
      margin-top: 7px;
      color: var(--muted);
      font-size: 12px;
      line-height: 1.4;
    }
    .metric-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 10px;
      margin-top: 14px;
    }
    .metric-box {
      padding: 11px 12px;
      border-radius: 14px;
      background: linear-gradient(180deg, #ffffff, #f5f9fe);
      border: 1px solid #dfe7f1;
    }
    .metric-box span {
      display: block;
      color: var(--muted);
      font-size: 11px;
      line-height: 1.35;
      text-transform: uppercase;
      letter-spacing: 0.03em;
    }
    .metric-box b {
      display: block;
      margin-top: 5px;
      color: #1b2a41;
      font-size: 16px;
      line-height: 1.25;
    }
    .card-delete-btn {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 34px;
      height: 34px;
      padding: 0;
      border-radius: 999px;
      border: 1px solid #efc7c2;
      background: #fff4f2;
      color: #b4392c;
      cursor: pointer;
      transition: background .18s ease, border-color .18s ease, transform .18s ease;
    }
    .card-delete-btn:hover {
      background: #ffe9e5;
      border-color: #e8a7a0;
      transform: translateY(-1px);
    }
    .card-delete-btn:disabled {
      opacity: .6;
      cursor: wait;
      transform: none;
    }
    .icon-trash {
      width: 16px;
      height: 16px;
      display: block;
      stroke: currentColor;
      fill: none;
      stroke-width: 1.9;
      stroke-linecap: round;
      stroke-linejoin: round;
      pointer-events: none;
    }
    .metric-good {
      color: #20744c !important;
    }
    .metric-bad {
      color: #bb4545 !important;
    }
    .btn-danger-lite:disabled {
      opacity: .6;
      cursor: wait;
      transform: none;
    }
    .open-link {
      color: var(--accent);
      font-weight: 700;
      white-space: nowrap;
    }
    .upload-progress {
      margin-top: 2px;
      padding: 14px;
      border-radius: 18px;
    }
    .upload-progress[hidden] { display:none; }
    .upload-progress-head { display:flex; justify-content:space-between; gap:10px; align-items:center; margin-bottom:8px; }
    .upload-progress-title { font-size:14px; font-weight:700; color:#1b2a41; }
    .upload-progress-pct { font-size:13px; color:#2e80e8; font-variant-numeric:tabular-nums; }
    .upload-progress-bar { height:12px; border-radius:999px; overflow:hidden; background:#edf3fa; border:1px solid #d6e0ee; }
    .upload-progress-fill { height:100%; width:0%; background:linear-gradient(90deg, #4f8cff, #5ecf8a); transition:width .28s ease; }
    .upload-progress-detail { margin-top:5px; color:#7389a9; font-size:12px; line-height:1.45; }
    .upload-progress-steps { display:flex; flex-wrap:wrap; gap:8px; margin-top:10px; }
    .upload-step { border:1px solid #cfd9e8; background:#f4f8fd; color:#6d7f96; border-radius:999px; padding:4px 9px; font-size:11px; }
    .upload-step.is-active { color:#fff; border-color:#2e80e8; background:#2e80e8; }
    .upload-step.is-done { color:#257347; border-color:#bfe5cc; background:#e9f8ef; }
    .upload-progress-error { margin-top:10px; color:#b04e4e; font-size:12px; white-space:pre-wrap; }
    .upload-progress-logs { margin-top:10px; border-radius:10px; border:1px solid #dfe7f1; background:#f8fbff; padding:9px; max-height:180px; overflow:auto; font-size:11px; color:#576a84; line-height:1.45; white-space:pre-wrap; }
    .empty {
      text-align: center;
      padding: 34px 20px;
      border-radius: 18px;
      border: 1px dashed var(--border-strong);
      background: linear-gradient(180deg, #fbfdff, #f5f9fd);
      color: var(--muted);
    }
    @media (max-width: 1080px) {
      .hero-grid { grid-template-columns: 1fr; }
    }
    @media (max-width: 720px) {
      .page { padding: 20px 12px 34px; }
      h1 { font-size: 30px; }
      .metric-grid,
      .grid { grid-template-columns: 1fr; }
      .tabs,
      .chip-row { gap: 6px; }
      .panel-head,
      .estimate-card-top,
      .estimate-progress-head { flex-direction: column; align-items: flex-start; }
      .estimate-card-meta { justify-content: flex-start; }
    }
  </style>
  <link rel="stylesheet" href="/static/autobot-ui.css?v=20260821" />
</head>
<body class="autobot-page estimates-index-page">
  <header class="topbar autobot-section-bar">
    <a class="brand" href="/tenders">
      <span class="brand-mark" aria-hidden="true"><i></i></span>
      <span class="brand-copy"><strong>AutoBot</strong><small>Закупки без рутины</small></span>
    </a>
    <nav class="topnav" aria-label="Разделы AutoBot">
      <a href="/tenders">Тендеры</a>
      <a class="is-active" href="/estimates">Сметы</a>
      <a href="/research">Поиск позиции</a>
    </nav>
  </header>
  <div class="page">
    <section class="hero-grid">
      <div class="hero-card">
        <span class="eyebrow">Сметы</span>
        <h1>Загрузка и список смет</h1>
        <p class="hero-text">Загрузите Excel или PDF. AutoBot выделит позиции, посчитает суммы и подготовит их к сравнению с рынком.</p>
      </div>

      <section class="upload-card">
        <div>
          <h2 class="panel-title">Загрузить смету (Excel/PDF)</h2>
        </div>
        <form id="estimateUploadForm" class="upload-row">
          <div class="file-picker">
            <input class="file-input-native" id="estimateUploadFile" type="file" name="file" accept=".xlsx,.xls,.xlsm,.pdf,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/pdf" required />
            <label class="file-picker-btn" for="estimateUploadFile">
              <svg class="icon-clip" viewBox="0 0 24 24" aria-hidden="true">
                <path d="M21.44 11.05l-8.49 8.49a6 6 0 0 1-8.49-8.49l9.19-9.19a4 4 0 0 1 5.66 5.66l-9.2 9.19a2 2 0 0 1-2.82-2.83l8.49-8.48"></path>
              </svg>
              <span>Прикрепить файл</span>
            </label>
            <span class="file-picker-name" id="estimateUploadFileName" hidden></span>
          </div>
          <input type="text" name="title" placeholder="Название сметы, если нужно переименовать карточку" />
          <button class="btn" type="submit">Загрузить и распарсить</button>
        </form>
        <div id="uploadStatus" class="muted"></div>
        <div id="estimateUploadProgress" class="upload-progress" hidden>
          <div class="upload-progress-head">
            <div class="upload-progress-title" id="estimateUploadStage">Подготовка…</div>
            <div class="upload-progress-pct" id="estimateUploadPct">0%</div>
          </div>
          <div class="upload-progress-bar"><div id="estimateUploadFill" class="upload-progress-fill"></div></div>
          <div id="estimateUploadDetail" class="upload-progress-detail"></div>
          <div class="upload-progress-steps" id="estimateUploadSteps">
            <span class="upload-step" data-step="upload">Отправка файла</span>
            <span class="upload-step" data-step="received">Файл получен</span>
            <span class="upload-step" data-step="parse">Разбор файла</span>
            <span class="upload-step" data-step="catalogue">Каталог позиций</span>
            <span class="upload-step" data-step="save">Сохранение</span>
            <span class="upload-step" data-step="done">Готово</span>
          </div>
          <div id="estimateUploadError" class="upload-progress-error" hidden></div>
          <div id="estimateUploadLogs" class="upload-progress-logs" hidden></div>
        </div>
      </section>
    </section>

    <section class="overview-strip" aria-label="Сводка по сметам">
      <div class="overview-pill">Всего смет <b>{{ overview.total_count }}</b></div>
      <div class="overview-pill">С анализом рынка <b>{{ overview.with_compare }}</b></div>
    </section>

    <section class="catalog-panel">
      <div class="panel-head">
        <label class="catalog-search"><span aria-hidden="true"></span><input type="search" id="estimateCatalogSearch" placeholder="Найти смету" autocomplete="off" /></label>
      </div>
      {% if estimates %}
      <div class="grid">
        {% for e in estimates %}
        <article class="estimate-card" data-estimate-card="{{ e.id }}" data-estimate-search="{{ (e.title ~ ' ' ~ e.original_filename ~ ' ' ~ e.types_short)|lower }}">
          <div class="estimate-card-top">
            <span class="status-pill {{ e.market_status_class }}">{{ e.market_status_label }}</span>
            <div class="estimate-card-meta">
              <span class="date-pill">{{ e.created_at }}</span>
              <button class="card-delete-btn" type="button" data-estimate-delete="{{ e.id }}" data-estimate-title="{{ e.title }}" onclick="deleteEstimateCard(event, this)" title="Удалить смету" aria-label="Удалить смету">
                <svg class="icon-trash" viewBox="0 0 24 24" aria-hidden="true">
                  <path d="M4 7h16"></path>
                  <path d="M10 11v6"></path>
                  <path d="M14 11v6"></path>
                  <path d="M6 7l1 12a2 2 0 0 0 2 2h6a2 2 0 0 0 2-2l1-12"></path>
                  <path d="M9 7V4a1 1 0 0 1 1-1h4a1 1 0 0 1 1 1v3"></path>
                </svg>
              </button>
            </div>
          </div>
          <a class="estimate-card-body-link" href="/estimates/{{ e.id }}" title="{{ e.title }}">
            <h3 title="{{ e.title }}">{{ e.title }}</h3>

            <div class="chip-row">
              {% for badge in e.type_badges %}
              <span class="chip">{{ badge.label }} · {{ badge.count }}</span>
              {% endfor %}
              {% if not e.type_badges %}
              <span class="chip">Типы не определены</span>
              {% endif %}
            </div>

            <div class="estimate-progress">
              <div class="estimate-progress-head">
                <span class="estimate-progress-label">Проанализировано</span>
                <span class="estimate-progress-value">{{ e.market_progress_done }} из {{ e.market_progress_total }}</span>
              </div>
              <div class="estimate-progress-track">
                <div class="estimate-progress-fill" style="width: {{ e.market_progress_percent }}%;"></div>
              </div>
              <div class="estimate-progress-note">{{ e.market_progress_note }}</div>
            </div>

            <div class="metric-grid">
              <div class="metric-box">
                <span>Строк</span>
                <b>{{ e.row_count }}</b>
              </div>
              <div class="metric-box">
                <span>Сумма</span>
                <b>{{ e.total_sum_fmt }}</b>
              </div>
              <div class="metric-box">
                <span>Состояние рынка</span>
                <b class="{{ e.market_summary_class }}" style="font-size:14px;">{{ e.market_summary }}</b>
              </div>
              <div class="metric-box">
                <span>Типы</span>
                <b style="font-size:14px;">{{ e.types_short }}</b>
              </div>
            </div>
          </a>
        </article>
        {% endfor %}
      </div>
      <div class="empty" id="estimateCatalogEmpty" hidden>По этому запросу смет нет. Очистите поиск и попробуйте снова.</div>
      {% else %}
      <div class="empty">Смет пока нет.</div>
      {% endif %}
    </section>
  </div>
  <script>
    (function() {
      const form = document.getElementById("estimateUploadForm");
      const status = document.getElementById("uploadStatus");
      const panel = document.getElementById("estimateUploadProgress");
      const fill = document.getElementById("estimateUploadFill");
      const pct = document.getElementById("estimateUploadPct");
      const stage = document.getElementById("estimateUploadStage");
      const detail = document.getElementById("estimateUploadDetail");
      const fileInput = document.getElementById("estimateUploadFile");
      const fileName = document.getElementById("estimateUploadFileName");
      const errBox = document.getElementById("estimateUploadError");
      const logs = document.getElementById("estimateUploadLogs");
      const stepNodes = Array.from(document.querySelectorAll("#estimateUploadSteps .upload-step"));
      let activePoll = 0;

      function syncChosenFile() {
        if (!fileInput || !fileName) return;
        const name = fileInput.files && fileInput.files[0] ? fileInput.files[0].name : "";
        fileName.textContent = name;
        fileName.hidden = !name;
      }

      if (fileInput) {
        fileInput.addEventListener("change", syncChosenFile);
      }

      function showProgress() {
        panel.hidden = false;
        logs.hidden = false;
      }

      function markStep(progress, currentStage, done) {
        const stageLow = String(currentStage || "").toLowerCase();
        const currentKey =
          done ? "done"
          : progress < 25 ? "upload"
          : progress < 40 ? "received"
          : stageLow.includes("каталог") ? "catalogue"
          : stageLow.includes("сохраня") ? "save"
          : progress >= 40 ? "parse"
          : "received";
        const order = ["upload", "received", "parse", "catalogue", "save", "done"];
        const currentIndex = order.indexOf(currentKey);
        stepNodes.forEach(function(node) {
          const key = node.getAttribute("data-step");
          const idx = order.indexOf(key);
          node.classList.toggle("is-done", idx >= 0 && idx < currentIndex);
          node.classList.toggle("is-active", key === currentKey);
        });
      }

      function renderProgress(data) {
        const value = Math.max(0, Math.min(100, Number(data.progress || 0)));
        fill.style.width = value + "%";
        pct.textContent = value + "%";
        stage.textContent = data.stage || "Подготовка…";
        detail.textContent = data.detail || "";
        const resultOk = !!data.result_ok;
        status.textContent = data.running ? "Смета обрабатывается…" : (resultOk ? "Смета готова." : (data.error ? "Во время обработки возникла ошибка." : ""));
        if (Array.isArray(data.log_tail) && data.log_tail.length) {
          logs.hidden = false;
          logs.textContent = data.log_tail.join("\\n");
        } else {
          logs.hidden = true;
          logs.textContent = "";
        }
        if (data.error) {
          errBox.hidden = false;
          errBox.textContent = "Ошибка: " + data.error;
        } else {
          errBox.hidden = true;
          errBox.textContent = "";
        }
        markStep(value, data.stage || "", resultOk && !data.running);
      }

      async function pollJob(jobId) {
        const pollId = ++activePoll;
        for (;;) {
          if (pollId !== activePoll) return;
          let resp, data;
          try {
            resp = await fetch("/api/estimates/upload-status/" + encodeURIComponent(jobId), { cache: "no-store" });
            data = await resp.json();
          } catch (e) {
            status.textContent = "Не удалось обновить статус обработки.";
            return;
          }
          if (!resp.ok || !data.ok) {
            status.textContent = (data && data.message) || "Статус обработки недоступен.";
            return;
          }
          renderProgress(data);
          if (!data.running) {
            if (data.result_ok && data.estimate_id) {
              status.textContent = "Готово, открываю смету…";
              setTimeout(function() { location.href = "/estimates/" + data.estimate_id; }, 450);
            }
            return;
          }
          await new Promise(function(resolve) { setTimeout(resolve, 500); });
        }
      }

      form.addEventListener("submit", function(e) {
        e.preventDefault();
        if (!fileInput || !fileInput.files || !fileInput.files.length) {
          status.textContent = "Сначала прикрепите Excel-файл.";
          return;
        }
        const fd = new FormData(form);
        const xhr = new XMLHttpRequest();
        activePoll += 1;
        showProgress();
        errBox.hidden = true;
        errBox.textContent = "";
        logs.textContent = "";
        logs.hidden = true;
        status.textContent = "Начинаю загрузку файла…";
        renderProgress({ progress: 2, stage: "Отправляю файл", detail: "Загружаю Excel на сервер", running: true, result_ok: false, log_tail: [] });
        xhr.open("POST", "/api/estimates/upload");
        xhr.upload.addEventListener("progress", function(ev) {
          if (!ev.lengthComputable) return;
          showProgress();
          const uploadPct = Math.max(2, Math.min(24, Math.round(ev.loaded / ev.total * 24)));
          renderProgress({ progress: uploadPct, stage: "Отправляю файл", detail: "Передано " + ev.loaded + " из " + ev.total + " байт", running: true, result_ok: false, log_tail: [] });
        });
        xhr.onreadystatechange = function() {
          if (xhr.readyState !== 4) return;
          let data = {};
          try { data = JSON.parse(xhr.responseText || "{}"); } catch (e) {}
          if (xhr.status < 200 || xhr.status >= 300 || !data.ok) {
            showProgress();
            renderProgress({
              progress: 100,
              stage: "Ошибка",
              detail: "Загрузка или запуск обработки не удались",
              running: false,
              result_ok: false,
              error: data.message || ("HTTP " + xhr.status),
              log_tail: []
            });
            status.textContent = "Ошибка: " + (data.message || ("HTTP " + xhr.status));
            return;
          }
          showProgress();
          renderProgress({
            progress: Math.max(26, Number(data.progress || 26)),
            stage: data.stage || "Файл получен",
            detail: data.detail || "Сервер принял файл и начал разбор",
            running: true,
            result_ok: false,
            log_tail: data.log_tail || []
          });
          status.textContent = "Файл получен, идёт разбор сметы…";
          pollJob(data.job_id);
        };
        xhr.send(fd);
      });

      window.deleteEstimateCard = async function(event, btn) {
        if (event) {
          event.preventDefault();
          event.stopPropagation();
        }
        if (!btn || btn.disabled) return;
        const estimateId = btn.getAttribute("data-estimate-delete") || "";
        const title = btn.getAttribute("data-estimate-title") || "эта смета";
        const ok = confirm(`Удалить смету "${title}"?\n\nБудут удалены карточка сметы, ее строки и сохраненные файлы рынка.`);
        if (!ok) return;
        const initialHtml = btn.innerHTML;
        btn.disabled = true;
        btn.innerHTML = "...";
        try {
          const resp = await fetch("/api/estimates/" + encodeURIComponent(estimateId) + "/delete", {
            method: "POST",
            headers: { "Accept": "application/json" },
          });
          let data = {};
          try { data = await resp.json(); } catch (err) {}
          if (!resp.ok || !data.ok) {
            alert(data.message || ("Не удалось удалить смету (HTTP " + resp.status + ")."));
            btn.disabled = false;
            btn.innerHTML = initialHtml;
            return;
          }
          const card = btn.closest("[data-estimate-card]");
          if (card) {
            card.remove();
          } else {
            window.location.reload();
            return;
          }
          setTimeout(function() { window.location.reload(); }, 120);
        } catch (err) {
          alert("Не удалось удалить смету: " + err);
          btn.disabled = false;
          btn.innerHTML = initialHtml;
        }
      };

      document.addEventListener("click", function(e) {
        if (e.target.closest("[data-estimate-delete]")) return;
        const card = e.target.closest("[data-estimate-open]");
        if (!card) return;
        if (e.target.closest("a, button, input, textarea, select, label")) return;
        const href = card.getAttribute("data-estimate-open") || "";
        if (!href) return;
        window.location.href = href;
      });
    })();
  </script>
</body>
</html>
"""


RESEARCH_TEMPLATE = """
<!doctype html>
<html lang="ru">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <link rel="icon" href="/favicon.svg" type="image/svg+xml" />
  <title>Поиск по позиции</title>
  <style>
    :root { color-scheme: light; --bg:#f4f7fb; --panel:#ffffff; --panel2:#f7fafe; --border:#d9e3ef; --muted:#62748b; --text:#172235; --accent:#1f72dc; }
    body { margin:0; font-family: Segoe UI, Arial, sans-serif; background:linear-gradient(180deg,#ffffff 0,#f4f7fb 100%); color:var(--text); }
    .page { max-width:1220px; margin:0 auto; padding:26px 18px 44px; }
    h1 { margin:0 0 8px; font-size:34px; }
    .sub,.muted { color:var(--muted); }
    .tabs { display:flex; flex-wrap:wrap; gap:8px; margin:14px 0 18px; }
    .tab { display:inline-flex; padding:9px 12px; border-radius:999px; color:#35506f; text-decoration:none; background:#f4f8fd; border:1px solid var(--border); font-weight:700; font-size:13px; }
    .tab.is-active { color:#fff; background:linear-gradient(180deg,#2e80e8,#1f72dc); border-color:#2e80e8; }
    .panel,.card { background:linear-gradient(180deg,#ffffff,#f8fbff); border:1px solid var(--border); border-radius:16px; box-shadow:0 18px 45px rgba(28,49,84,.08); }
    .panel { padding:16px; margin-bottom:16px; }
    .grid { display:grid; gap:12px; }
    .form-grid { display:grid; grid-template-columns:minmax(320px,1.4fr) minmax(220px,.8fr); gap:12px; align-items:start; }
    label { display:grid; gap:6px; color:var(--muted); font-size:12px; }
    textarea,input { background:#fff; border:1px solid #cfd9e8; color:var(--text); border-radius:12px; padding:12px; font:inherit; }
    textarea { min-height:180px; resize:vertical; }
    .btn-row { display:flex; flex-wrap:wrap; gap:10px; margin-top:12px; }
    .btn { border:1px solid #2e80e8; background:linear-gradient(180deg,#2e80e8,#1f72dc); color:white; border-radius:10px; padding:10px 14px; font-weight:700; cursor:pointer; text-decoration:none; }
    .btn.secondary { background:#f4f8fd; border-color:#cfd9e8; color:#35506f; }
    .btn[disabled] { opacity:.6; cursor:not-allowed; }
    .results { display:grid; gap:12px; }
    .result-card { padding:14px; background:#fff; border:1px solid #dfe7f1; border-radius:14px; }
    .result-card h3 { margin:0 0 6px; font-size:18px; }
    .meta { color:#62748b; font-size:13px; margin-bottom:10px; }
    .offers { display:grid; gap:10px; }
    .offer { padding:11px 12px; border-radius:12px; background:#f8fbff; border:1px solid #dfe7f1; }
    .offer-top { display:flex; justify-content:space-between; gap:12px; margin-bottom:6px; }
    .offer-source { display:inline-flex; padding:4px 8px; border-radius:999px; background:#edf4fd; border:1px solid #cfd9e8; font-size:12px; color:#35506f; }
    .offer-price { font-weight:800; color:#2e8b57; white-space:nowrap; }
    .offer-price.is-candidate { color:#a06b18; }
    .offer.is-verified { border-color:#b9dfc7; background:#f4fcf7; }
    .offer.is-candidate { border-color:#ecd8aa; background:#fffaf0; }
    .verification { margin-top:8px; padding:7px 9px; border-radius:8px; font-size:12px; line-height:1.4; background:#fff; border:1px solid #dfe7f1; color:#50627a; }
    .verification b { color:#1f6f43; }.offer.is-candidate .verification b { color:#986315; }
    .strategy-note { margin:10px 0 0; padding:10px 12px; border-radius:10px; background:#eef5ff; color:#4d6480; font-size:12px; line-height:1.45; }
    .offer a { color:#1f72dc; font-weight:700; text-decoration:none; }
    .offer-snippet { margin-top:6px; color:#62748b; font-size:13px; }
    .empty { padding:16px; text-align:center; color:#62748b; }
    @media (max-width:760px){ .form-grid{grid-template-columns:1fr} .btn-row{flex-direction:column} .btn{width:100%;box-sizing:border-box} }
  </style>
  <link rel="stylesheet" href="/static/autobot-ui.css?v=20260821" />
</head>
<body class="autobot-page research-page">
  <header class="topbar autobot-section-bar">
    <a class="brand" href="/tenders">
      <span class="brand-mark" aria-hidden="true"><i></i></span>
      <span class="brand-copy"><strong>AutoBot</strong><small>Закупки без рутины</small></span>
    </a>
    <nav class="topnav" aria-label="Разделы AutoBot">
      <a href="/tenders">Тендеры</a>
      <a href="/estimates">Сметы</a>
      <a class="is-active" href="/research">Поиск позиции</a>
    </nav>
  </header>
  <div class="page">
    <header class="research-hero">
      <span class="eyebrow">Быстрая проверка рынка</span>
      <h1>Найти цену по позиции</h1>
      <p class="sub">AutoBot найдёт кандидатов, откроет прямые страницы и отдельно покажет проверенные цены и отклонённые источники.</p>
    </header>

    <section class="panel research-form-panel">
      <div class="form-grid">
        <label>Позиции для поиска
          <textarea id="researchQueries" placeholder="Название | единица&#10;Кабель ВВГнг 3х2,5 | м&#10;Укладка тротуарной плитки | м2">{{ default_query }}</textarea>
        </label>
        <div class="grid">
          <label>Город
            <input id="researchCity" type="text" placeholder="Например: Челябинск" value="{{ default_city }}" />
          </label>
          <div class="research-hint"><strong>Единица обязательна для проверки</strong><span>Формат строки: <b>название | единица</b>. Например: «Кабель ВВГнг 3х2,5 | м». Без единицы найденные цены останутся кандидатами. За раз — до 5 позиций.</span></div>
          <div class="btn-row">
            <button class="btn" id="researchRunBtn" type="button" onclick="runResearch()"><span class="research-button-icon" aria-hidden="true"></span><span data-research-button-label>Найти цены</span></button>
            <button class="btn secondary" type="button" onclick="fillExample()">Подставить пример</button>
          </div>
        </div>
      </div>
    </section>

    <section class="panel research-results-panel">
      <div class="research-status"><span aria-hidden="true"></span><div id="researchStatus" class="muted" aria-live="polite">Пока ничего не искали.</div></div>
      <div id="researchResults" class="results" style="margin-top:12px;"></div>
    </section>
  </div>

  <script>
    function money(v) {
      const num = Number(v || 0);
      if (!Number.isFinite(num) || num <= 0) return "цена не указана";
      return new Intl.NumberFormat("ru-RU", { maximumFractionDigits: 0 }).format(num) + " ₽";
    }

    function fillExample() {
      const q = document.getElementById("researchQueries");
      const c = document.getElementById("researchCity");
      if (q) q.value = "Кабель ВВГнг 3х2,5 | м\nУкладка тротуарной плитки | м2";
    }

    function renderResearch(data) {
      const root = document.getElementById("researchResults");
      const status = document.getElementById("researchStatus");
      if (!root || !status) return;
      root.replaceChildren();
      const items = Array.isArray(data.results) ? data.results : [];
      status.textContent = data.message || (items.length ? "Готово." : "Ничего не найдено.");
      if (!items.length) {
        const empty = document.createElement("div");
        empty.className = "empty";
        empty.textContent = "Нет результатов.";
        root.appendChild(empty);
        return;
      }
      for (const item of items) {
        const card = document.createElement("div");
        card.className = "result-card";
        const h = document.createElement("h3");
        h.textContent = String(item.query || "");
        const meta = document.createElement("div");
        meta.className = "meta";
        const prices = Array.isArray(item.offers) ? item.offers.filter(x => x.verified).map(x => Number(x.price || 0)).filter(x => Number.isFinite(x) && x > 0) : [];
        const verifiedCount = Array.isArray(item.offers) ? item.offers.filter(x => x.verified).length : 0;
        const candidateCount = Array.isArray(item.offers) ? item.offers.filter(x => !x.verified).length : 0;
        const cityText = item.region ? (" · город: " + item.region) : "";
        meta.textContent = (item.position_label ? item.position_label + (item.unit ? " · ед.: " + item.unit : "") + " · " : "") + "проверено: " + verifiedCount + " · кандидатов: " + candidateCount + cityText + (prices.length ? (" · диапазон: " + money(Math.min(...prices)) + " — " + money(Math.max(...prices))) : "");
        card.appendChild(h);
        card.appendChild(meta);
        if (item.strategy || item.warning) {
          const strategy = document.createElement("div");
          strategy.className = "strategy-note";
          strategy.textContent = [item.strategy, item.warning].filter(Boolean).join(" · ");
          card.appendChild(strategy);
        }
        const offersWrap = document.createElement("div");
        offersWrap.className = "offers";
        const offers = Array.isArray(item.offers) ? item.offers : [];
        if (!offers.length) {
          const empty = document.createElement("div");
          empty.className = "offer";
          empty.textContent = item.errors || "Ничего не найдено.";
          offersWrap.appendChild(empty);
        } else {
          for (const offer of offers) {
            const box = document.createElement("div");
            box.className = "offer " + (offer.verified ? "is-verified" : "is-candidate");
            const top = document.createElement("div");
            top.className = "offer-top";
            const source = document.createElement("span");
            source.className = "offer-source";
            source.textContent = (offer.verified ? "✓ Проверен · " : "? Кандидат · ") + String(offer.source || "Источник");
            const price = document.createElement("span");
            price.className = "offer-price" + (offer.verified ? "" : " is-candidate");
            price.textContent = offer.verified ? money(offer.price) : "не принято в расчёт";
            top.appendChild(source);
            top.appendChild(price);
            const link = document.createElement("a");
            link.href = String(offer.url || "#");
            link.target = "_blank";
            link.rel = "noopener noreferrer";
            link.textContent = String(offer.title || offer.url || "Открыть источник");
            box.appendChild(top);
            box.appendChild(link);
            if (offer.snippet) {
              const sn = document.createElement("div");
              sn.className = "offer-snippet";
              sn.textContent = String(offer.snippet);
              box.appendChild(sn);
            }
            const verification = document.createElement("div");
            verification.className = "verification";
            const adapter = offer.adapter ? (" · адаптер: " + offer.adapter) : "";
            const unit = offer.matched_unit ? (" · единица: " + offer.matched_unit) : "";
            verification.textContent = (offer.verified ? "Цена подтверждена" : "Источник отклонён") + " · " + String(offer.reason || "Нет доказательства на прямой странице") + adapter + unit;
            box.appendChild(verification);
            offersWrap.appendChild(box);
          }
        }
        if (item.errors && offers.length) {
          const warn = document.createElement("div");
          warn.className = "meta";
          warn.textContent = "Ограничения поиска: " + item.errors;
          card.appendChild(warn);
        }
        card.appendChild(offersWrap);
        root.appendChild(card);
      }
    }

    async function runResearch() {
      const btn = document.getElementById("researchRunBtn");
      const status = document.getElementById("researchStatus");
      const queries = document.getElementById("researchQueries");
      const city = document.getElementById("researchCity");
      if (!queries) return;
      if (!String(queries.value || "").trim()) {
        if (status) status.textContent = "Добавьте хотя бы одну позицию для поиска.";
        queries.focus();
        return;
      }
      if (btn) btn.disabled = true;
      document.body.classList.add("research-is-running");
      const buttonLabel = btn ? btn.querySelector("[data-research-button-label]") : null;
      if (buttonLabel) buttonLabel.textContent = "Ищем предложения…";
      if (status) status.textContent = "Ищу кандидатов и проверяю цены на прямых страницах…";
      try {
        const resp = await fetch("/api/research-items", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({
            queries: String(queries.value || ""),
            city: city ? String(city.value || "").trim() : ""
          })
        });
        const data = await resp.json();
        if (!resp.ok || !data.ok) {
          if (status) status.textContent = data.message || "Не удалось выполнить поиск.";
          return;
        }
        renderResearch(data);
      } catch (e) {
        if (status) status.textContent = "Не удалось выполнить поиск.";
      } finally {
        if (btn) btn.disabled = false;
        document.body.classList.remove("research-is-running");
        if (buttonLabel) buttonLabel.textContent = "Найти цены";
      }
    }

    document.getElementById("researchQueries")?.addEventListener("keydown", function(event) {
      if ((event.ctrlKey || event.metaKey) && event.key === "Enter") {
        event.preventDefault();
        runResearch();
      }
    });
  </script>
</body>
</html>
"""


ESTIMATE_DETAIL_TEMPLATE = """
<!doctype html>
<html lang="ru">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <link rel="icon" href="/favicon.svg" type="image/svg+xml" />
  <title>{{ meta.title }} · Смета</title>
  <style>
    :root { color-scheme: light; --bg:#f4f7fb; --panel:#ffffff; --border:#d9e3ef; --muted:#62748b; --text:#172235; --accent:#1f72dc; }
    body { margin:0; font-family: Segoe UI, Arial, sans-serif; background:linear-gradient(180deg,#ffffff 0,#f4f7fb 100%); color:var(--text); }
    .page { max-width:1240px; margin:0 auto; padding:18px 14px 34px; }
    a { color:#1f72dc; }
    .page-head { display:flex; flex-wrap:wrap; justify-content:space-between; gap:10px; align-items:center; margin-bottom:12px; }
    .page-head-actions { display:flex; flex-wrap:wrap; gap:8px; align-items:center; }
    .top-back {
      display:inline-flex;
      align-items:center;
      gap:8px;
      padding:9px 13px;
      border-radius:999px;
      border:1px solid #cfd9e8;
      background:linear-gradient(180deg,#ffffff,#f4f8fd);
      color:#35506f;
      text-decoration:none;
      font-size:13px;
      font-weight:700;
      box-shadow:0 10px 24px rgba(43, 78, 131, 0.08);
    }
    .top-back:hover { background:#eef5fd; border-color:#9ec0ef; color:#173a65; }
    h1 { margin:0 0 6px; font-size:24px; }
    .muted { color:var(--muted); }
    .panel { background:linear-gradient(180deg,#ffffff,#f8fbff); border:1px solid var(--border); border-radius:16px; padding:12px; margin-bottom:12px; box-shadow:0 16px 34px rgba(28,49,84,.08); }
    .filters { display:flex; flex-wrap:wrap; gap:8px; align-items:end; }
    label { display:grid; gap:4px; color:var(--muted); font-size:11px; }
    input,select,textarea { background:#fff; border:1px solid #cfd9e8; color:var(--text); border-radius:10px; padding:8px 10px; min-width:190px; }
    textarea { min-height:96px; resize:vertical; font-family:inherit; }
    .btn { border:1px solid #2e80e8; background:linear-gradient(180deg,#2e80e8,#1f72dc); color:white; border-radius:10px; padding:8px 12px; font-weight:700; cursor:pointer; text-decoration:none; font-size:12px; display:inline-flex; align-items:center; justify-content:center; }
    .btn.secondary { background:#f4f8fd; border-color:#cfd9e8; color:#35506f; }
    .btn.danger-soft { background:linear-gradient(180deg,#fff4f4,#fdeaea); border-color:#efcaca; color:#a24c4c; }
    .btn.danger-soft:hover { background:linear-gradient(180deg,#feecec,#fbdede); border-color:#e5b2b2; color:#933f3f; }
    .btn.icon-only { width:38px; height:38px; padding:0; border-radius:999px; }
    .btn.icon-only:disabled { opacity:.6; cursor:wait; }
    .icon-trash {
      width: 16px;
      height: 16px;
      display: block;
      stroke: currentColor;
      fill: none;
      stroke-width: 1.9;
      stroke-linecap: round;
      stroke-linejoin: round;
      pointer-events: none;
    }
    .btn.is-stop { background:linear-gradient(180deg,#f7e2e2,#efcdcd); border-color:#ddb1b1; color:#8a3f3f; }
    .btn.is-stop:hover { background:linear-gradient(180deg,#f4d6d6,#ebc2c2); border-color:#d39c9c; color:#7a3232; }
    .actions-row { display:flex; flex-wrap:wrap; gap:8px; margin-top:12px; }
    .crm-callout { display:flex; flex-wrap:wrap; gap:10px; align-items:center; margin-top:12px; padding:10px 12px; border-radius:14px; background:linear-gradient(180deg,#eef5ff,#f7fbff); border:1px solid #d7e5f6; }
    .crm-callout-note { color:#45627f; font-size:12px; line-height:1.45; }
    .crm-callout-note strong { color:#1b2a41; }
    .type-picker { margin-top:10px; padding:10px; background:#f8fbff; border:1px solid #dfe7f1; border-radius:14px; }
    .type-picker-title { margin:0 0 8px; font-size:13px; font-weight:700; color:#1b2a41; }
    .type-checks { display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:8px; }
    .type-check { display:flex; align-items:flex-start; gap:8px; padding:8px 10px; background:#fff; border:1px solid #dfe7f1; border-radius:12px; min-height:48px; }
    .type-check input { min-width:18px; width:18px; height:18px; margin-top:2px; accent-color:#2e80e8; }
    .type-check strong { display:block; font-size:13px; color:#1b2a41; }
    .type-check span { display:block; color:#62748b; font-size:11px; margin-top:2px; }
    .summary { display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:10px; }
    .summary-item { padding:9px 10px; background:#fff; border:1px solid #dfe7f1; border-radius:12px; }
    .summary-item span { display:block; color:var(--muted); font-size:11px; }
    .summary-item b { display:block; margin-top:4px; font-size:15px; }
    .download-box { margin-top:10px; padding:10px; background:#f8fbff; border:1px solid #dfe7f1; border-radius:12px; }
    .download-title { margin:0 0 8px; font-size:13px; font-weight:700; color:#1b2a41; }
    .download-links { display:flex; flex-wrap:wrap; gap:8px; }
    .download-note { margin-top:8px; color:#62748b; font-size:11px; line-height:1.35; }
    .table-switch { display:flex; flex-wrap:wrap; gap:8px; align-items:center; }
    .table-switch .btn.is-active { background:linear-gradient(180deg,#2e80e8,#1f72dc); border-color:#2e80e8; color:#fff; }
    .table-switch .btn[disabled] { opacity:.45; cursor:not-allowed; }
    .table-panel[hidden] { display:none; }
    .table-panel-head { display:flex; flex-wrap:wrap; justify-content:space-between; gap:8px; align-items:center; margin-bottom:8px; }
    .table-panel-title { margin:0; font-size:15px; color:#1b2a41; }
    .viability-box { margin-top:10px; padding:12px; border-radius:14px; border:1px solid #c6d8f0; background:linear-gradient(180deg,#f7fbff,#eef5fd); }
    .viability-box.good { border-color:#bfe5cc; background:linear-gradient(180deg,#f3fcf6,#e8f8ee); }
    .viability-box.warn { border-color:#f0deb1; background:linear-gradient(180deg,#fffaf0,#fff4dd); }
    .viability-box.bad { border-color:#f0c5c5; background:linear-gradient(180deg,#fff7f7,#fff0f0); }
    .viability-title { margin:0; font-size:24px; font-weight:800; letter-spacing:.04em; }
    .viability-sub { margin-top:6px; color:#445870; font-size:13px; line-height:1.45; }
    .viability-facts { display:grid; grid-template-columns:repeat(auto-fit,minmax(160px,1fr)); gap:8px; margin-top:10px; }
    .viability-fact { padding:9px 10px; border-radius:12px; background:#fff; border:1px solid #dfe7f1; }
    .viability-fact span { display:block; font-size:10px; color:#62748b; text-transform:uppercase; letter-spacing:.06em; }
    .viability-fact b { display:block; margin-top:4px; font-size:16px; color:#1b2a41; }
    .viability-groups { display:grid; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); gap:8px; margin-top:10px; }
    .viability-group { padding:10px; border-radius:12px; background:#fff; border:1px solid #dfe7f1; }
    .viability-group-title { font-size:13px; font-weight:700; color:#1b2a41; margin-bottom:6px; }
    .viability-group-line { font-size:11px; color:#4e6582; line-height:1.4; }
    .tone-good { color:#2e8b57; }
    .tone-warn { color:#a06b18; }
    .tone-bad { color:#c05757; }
    .table-wrap { overflow:hidden; border-radius:14px; border:1px solid var(--border); background:#fff; position:relative; }
    .table-scroll { overflow:auto; max-height:calc(100vh - 190px); border-radius:14px; }
    table { width:100%; border-collapse:collapse; font-size:11px; min-width:900px; }
    th,td { padding:6px 7px; border-bottom:1px solid #e5ecf4; vertical-align:top; }
    th { position:sticky; top:0; background:#f1f6fc; color:#35506f; text-align:left; z-index:2; }
    tr:hover td { background:#f7faff; }
    tr.section-row td { background:#edf4fd; color:#1b2a41; font-weight:700; border-bottom-color:#d4e0ef; }
    tr.section-row:hover td { background:#edf4fd; }
    tr.sheet-total-row td { background:#fff3f3; color:#a94444; font-weight:700; border-top:1px solid #f0c5c5; border-bottom:1px solid #f0c5c5; }
    tr.sheet-break-row td { background:#f9dde0; border-bottom:0; height:14px; padding:0; }
    .num { text-align:right; font-variant-numeric:tabular-nums; white-space:nowrap; }
    .name { min-width:300px; }
    .tag { display:inline-flex; border-radius:999px; padding:2px 7px; border:1px solid #cfd9e8; background:#f4f8fd; color:#35506f; font-size:10px; white-space:nowrap; }
    .where { color:#62748b; font-size:10px; line-height:1.3; }
    .market-box { display:grid; gap:10px; }
    .market-links { display:flex; flex-wrap:wrap; gap:10px; }
    .market-link-chip { display:inline-flex; align-items:center; gap:8px; text-decoration:none; border:1px solid #cfd9e8; background:#fff; color:#1f3957; border-radius:12px; padding:10px 14px; font-weight:700; }
    .market-link-chip:hover { border-color:#9ec0ef; color:#173a65; background:#f5f9ff; }
    .market-links-note { font-size:12px; line-height:1.45; }
    .status-box { background:#f8fbff; border:1px solid #dfe7f1; border-radius:12px; padding:12px; }
    .status-line { color:var(--muted); font-size:12px; }
    .logs { margin:0; background:#fff; border:1px solid #dfe7f1; border-radius:12px; padding:10px; max-height:200px; overflow:auto; white-space:pre-wrap; font-size:12px; color:#576a84; }
    .crm-drawer[hidden] { display:none; }
    .crm-drawer { position:fixed; inset:0; z-index:90; }
    .crm-drawer-backdrop {
      position:absolute;
      inset:0;
      background:rgba(17, 32, 53, 0.34);
      backdrop-filter:blur(2px);
      opacity:0;
      transition:opacity .26s ease;
    }
    .crm-drawer-panel {
      position:absolute;
      top:0;
      right:0;
      width:min(50vw, 760px);
      height:100%;
      background:linear-gradient(180deg,#ffffff,#f8fbff);
      border-left:1px solid #dfe7f1;
      box-shadow:-16px 0 40px rgba(33, 63, 110, 0.14);
      display:flex;
      flex-direction:column;
      transform:translateX(100%);
      transition:transform .3s cubic-bezier(.2,.8,.2,1);
    }
    .crm-drawer.is-open .crm-drawer-backdrop { opacity:1; }
    .crm-drawer.is-open .crm-drawer-panel { transform:translateX(0); }
    .crm-drawer-head { display:flex; align-items:flex-start; justify-content:space-between; gap:10px; padding:18px 18px 12px; border-bottom:1px solid #e6edf6; }
    .crm-drawer-title { margin:0; font-size:20px; color:#1b2a41; }
    .crm-drawer-sub { margin:6px 0 0; color:#62748b; font-size:12px; line-height:1.45; }
    .crm-drawer-close { border:1px solid #d8e2ef; background:#f7fbff; color:#506985; border-radius:10px; padding:8px 10px; font-size:12px; font-weight:700; cursor:pointer; }
    .crm-drawer-body { padding:14px 18px 18px; overflow:auto; display:grid; gap:12px; }
    .crm-source-card { padding:12px; border-radius:14px; border:1px solid #dfe7f1; background:linear-gradient(180deg,#eef5ff,#f8fbff); }
    .crm-source-card strong { display:block; color:#1b2a41; font-size:14px; }
    .crm-source-card span { display:block; margin-top:4px; color:#62748b; font-size:12px; line-height:1.45; }
    .crm-form-grid { display:grid; gap:10px; }
    .crm-form-grid label { font-size:12px; }
    .crm-form-grid label small { display:block; margin-top:5px; color:#718198; line-height:1.35; }
    .crm-status { min-height:18px; color:#62748b; font-size:12px; line-height:1.4; }
    .crm-status.is-error { color:#b14b4b; }
    .crm-status.is-success { color:#257347; }
    .crm-drawer-foot { display:flex; flex-wrap:wrap; gap:8px; justify-content:flex-end; padding-top:4px; }
    @media (max-width:760px){
      .filters{align-items:stretch;flex-direction:column}
      .btn,input,select,textarea{width:100%;box-sizing:border-box}
      .crm-drawer-panel { width:100vw; }
    }
  </style>
  <link rel="stylesheet" href="/static/autobot-ui.css?v=20260821" />
</head>
<body class="autobot-page estimate-detail-page">
  <header class="topbar autobot-section-bar">
    <a class="brand" href="/tenders">
      <span class="brand-mark" aria-hidden="true"><i></i></span>
      <span class="brand-copy"><strong>AutoBot</strong><small>Закупки без рутины</small></span>
    </a>
    <nav class="topnav" aria-label="Разделы AutoBot">
      <a href="/tenders">Тендеры</a>
      <a class="is-active" href="/estimates">Сметы</a>
      <a href="/research">Поиск позиции</a>
    </nav>
  </header>
  <div class="page">
    <div class="page-head">
      <a class="top-back" href="/estimates">← Все сметы</a>
      <div class="page-head-actions">
        <button class="btn danger-soft icon-only" type="button" onclick="deleteEstimate(this)" title="Удалить смету" aria-label="Удалить смету">
          <svg class="icon-trash" viewBox="0 0 24 24" aria-hidden="true">
            <path d="M4 7h16"></path>
            <path d="M10 11v6"></path>
            <path d="M14 11v6"></path>
            <path d="M6 7l1 12a2 2 0 0 0 2 2h6a2 2 0 0 0 2-2l1-12"></path>
            <path d="M9 7V4a1 1 0 0 1 1-1h4a1 1 0 0 1 1 1v3"></path>
          </svg>
        </button>
      </div>
    </div>
    <h1>{{ meta.title }}</h1>
    <div class="muted">{{ meta.original_filename }} · загружено {{ meta.created_at }} · всего строк {{ meta.row_count }}</div>

    <section class="panel">
      <form class="filters" method="get" action="/estimates/{{ meta.id }}" id="estimateFilterForm">
        <input type="hidden" name="table_view" id="estimateTableViewInput" value="{{ active_table_view }}" />
        <label>Поиск по наименованию
          <input type="text" name="q" value="{{ q }}" placeholder="например: бетон, демонтаж, труба" />
        </label>
        <button class="btn" type="submit">Применить</button>
        <a class="btn secondary" href="/estimates/{{ meta.id }}">Сбросить</a>
      </form>
      <div class="type-picker">
        <div class="type-picker-title">Фильтр по типам позиций</div>
        <div class="type-checks">
          {% for opt in type_options %}
          <label class="type-check">
            <input type="checkbox" name="types" value="{{ opt.key }}" form="estimateFilterForm" {% if opt.key in selected_types %}checked{% endif %} />
            <span>
              <strong>{{ opt.label }}</strong>
              <span>Строк: {{ opt.count }}</span>
            </span>
          </label>
          {% endfor %}
        </div>
      </div>
      <div class="summary" style="margin-top:12px;">
        <div class="summary-item"><span>Строк после фильтра</span><b>{{ summary.row_count }}</b></div>
        <div class="summary-item"><span>Общее количество / объём</span><b>{{ summary.qty_text }}</b></div>
        <div class="summary-item"><span>Общая сумма</span><b>{{ summary.total_sum_fmt }}</b></div>
        <div class="summary-item"><span>Средняя цена</span><b>{{ summary.avg_price_fmt }}</b></div>
      </div>
      <div class="crm-callout">
        <button class="btn" id="estimateCrmOpenBtn" type="button">Добавить в объекты</button>
        <div class="crm-callout-note">
          На основе сметы <strong>{{ crm_prefill.estimate_title }}</strong>. Поля объекта можно открыть и отредактировать перед созданием.
        </div>
      </div>
      <div class="download-box">
        <div class="download-title">Таблица ниже</div>
        <div class="table-switch">
          <button class="btn secondary{% if active_table_view == 'estimate' %} is-active{% endif %}" type="button" data-estimate-view-btn="estimate" data-download-href="/estimates/{{ meta.id }}/download.xlsx?{{ filter_query }}" data-download-label="Смета">Смета</button>
          <button class="btn secondary{% if active_table_view == 'compare' %} is-active{% endif %}" type="button" data-estimate-view-btn="compare" data-download-href="/estimates/{{ meta.id }}/market-compare.xlsx?{{ filter_query }}" data-download-label="Смета vs Рынок" {% if not compare_table.available %}disabled{% endif %}>Смета vs Рынок</button>
          <button class="btn secondary{% if active_table_view == 'sources' %} is-active{% endif %}" type="button" data-estimate-view-btn="sources" data-download-href="/estimates/{{ meta.id }}/market-sources.xlsx?{{ filter_query }}" data-download-label="Источники цен" {% if not sources_table.available %}disabled{% endif %}>Источники цен</button>
          <a class="btn" id="activeTableDownloadBtn" href="/estimates/{{ meta.id }}/download.xlsx?{{ filter_query }}">Скачать Excel</a>
        </div>
        <div class="download-note">Сначала выберите нужную таблицу, потом нажмите скачать Excel для текущей вкладки.</div>
      </div>
      {% if viability.title %}
      <div class="viability-box {{ viability.tone }}">
        <h2 class="viability-title">{{ viability.title }}</h2>
        <div class="viability-sub">{{ viability.subtitle }}</div>
        {% if viability.facts %}
        <div class="viability-facts">
          {% for fact in viability.facts %}
          <div class="viability-fact">
            <span>{{ fact.label }}</span>
            <b>{{ fact.value }}</b>
          </div>
          {% endfor %}
        </div>
        {% endif %}
        {% if viability.groups %}
        <div class="viability-groups">
          {% for group in viability.groups %}
          <div class="viability-group">
            <div class="viability-group-title">{{ group.title }}</div>
            <div class="viability-group-line"><span class="tone-good">Выше рынка: {{ group.good }}</span></div>
            <div class="viability-group-line"><span class="tone-warn">Около рынка: {{ group.warn }}</span></div>
            <div class="viability-group-line"><span class="tone-bad">Ниже рынка: {{ group.bad }}</span></div>
            <div class="viability-group-line">Без данных: {{ group.none }}</div>
          </div>
          {% endfor %}
        </div>
        {% endif %}
      </div>
      {% endif %}
    </section>

    <section class="panel">
      <div style="display:grid;gap:10px;">
        <div>
          <h2 style="margin:0 0 8px;">Сравнение по найденным ценам</h2>
          <div class="muted market-links-note">Откройте нужный тип позиций, чтобы посмотреть сайты, цены и найденные источники.</div>
        </div>
        {% if market_links %}
        <div class="market-links">
          {% for link in market_links %}
          <a class="market-link-chip" href="{{ link.href }}">{{ link.label }} · {{ link.count }}</a>
          {% endfor %}
        </div>
        {% else %}
        <div class="status-box">
          <div>Пока нет сохранённых сравнений по текущему фильтру.</div>
          <div class="status-line">Запустите поиск цен, затем обновите страницу — здесь появятся ссылки на отдельные страницы сравнения.</div>
        </div>
        {% endif %}
      </div>
    </section>

    <section class="panel">
      <div class="market-box">
        <div>
          <h2 style="margin:0 0 8px;">Поиск цен по этой смете</h2>
          <div class="muted">Ищем цены по интернету и на Авито для строк этой сметы. Можно указать город, чтобы сузить выдачу.</div>
        </div>
        <div class="filters">
          <label>Город для поиска
            <input type="text" id="marketCityInput" value="{{ market_city }}" placeholder="например: Челябинск" />
          </label>
          <button class="btn" type="button" id="marketStartBtn" onclick="toggleEstimateMarket()">Найти цены</button>
          <a class="btn secondary" id="marketMergedBtn" href="/estimates/{{ meta.id }}/market-compare.xlsx" {% if not has_market_merged %}hidden{% endif %}>Excel: смета vs рынок</a>
          <a class="btn secondary" id="marketRawBtn" href="/estimates/{{ meta.id }}/market-sources.xlsx" {% if not has_market_raw %}hidden{% endif %}>Excel: источники рынка</a>
        </div>
        <div class="status-box">
          <div id="marketStatusMain">Пока поиск рынка не запускался.</div>
          <div class="status-line" id="marketStatusDetail"></div>
        </div>
        <pre class="logs" id="marketLogs">—</pre>
      </div>
    </section>

    <section class="panel">
      <div class="table-panel" data-estimate-view-panel="estimate" {% if active_table_view != 'estimate' %}hidden{% endif %}>
        <div class="table-panel-head">
          <h2 class="table-panel-title">Смета</h2>
        </div>
      <div class="table-wrap">
        <div class="table-scroll">
        <table>
          <thead>
            <tr>
              <th>№</th>
              <th>Тип</th>
              <th class="name">Наименование</th>
              <th>Ед.</th>
              <th class="num">Кол-во</th>
              <th class="num">Цена за ед.</th>
              <th class="num">Сумма</th>
              <th>Где найдено</th>
            </tr>
          </thead>
          <tbody>
            {% for r in rows %}
            {% if r._is_section %}
            <tr class="section-row">
              <td colspan="8">{{ r.section_title }}</td>
            </tr>
            {% elif r._is_sheet_total %}
            <tr class="sheet-total-row">
              <td colspan="6">Итого по листу: {{ r.sheet_title or "без названия" }}</td>
              <td class="num">{{ r.sheet_total_fmt }}</td>
              <td></td>
            </tr>
            {% elif r._is_sheet_break %}
            <tr class="sheet-break-row">
              <td colspan="8"></td>
            </tr>
            {% else %}
            <tr>
              <td>{{ r.display_no }}</td>
              <td><span class="tag">{{ r.type_label }}</span></td>
              <td class="name">{{ r.name }}</td>
              <td>{{ r.unit or "—" }}</td>
              <td class="num">{{ r.qty_fmt }}</td>
              <td class="num">{{ r.unit_price_fmt }}</td>
              <td class="num">{{ r.total_fmt }}</td>
              <td class="where">
                {% if r.sheet %}лист: {{ r.sheet }}<br>{% endif %}
                {% if r.excel_row %}строка Excel: {{ r.excel_row }}<br>{% endif %}
              </td>
            </tr>
            {% endif %}
            {% endfor %}
            {% if not rows %}
            <tr><td colspan="8" style="text-align:center;color:#9fb0d6;padding:24px;">По фильтру ничего не найдено.</td></tr>
            {% endif %}
          </tbody>
        </table>
        </div>
      </div>
      </div>
      <div class="table-panel" data-estimate-view-panel="compare" {% if active_table_view != 'compare' %}hidden{% endif %}>
        <div class="table-panel-head">
          <h2 class="table-panel-title">Смета vs Рынок</h2>
        </div>
        {% if compare_table.available %}
        <div class="table-wrap">
          <div class="table-scroll">
          <table>
            <thead>
              <tr>
                <th>Раздел</th>
                <th>Тип</th>
                <th class="name">Наименование</th>
                <th class="num">Цена сметы</th>
                <th class="num">Цена рынка</th>
                <th>Сайт</th>
                <th>Вывод</th>
              </tr>
            </thead>
            <tbody>
              {% for row in compare_rows %}
              <tr>
                <td class="where">{{ row.section }}</td>
                <td><span class="tag">{{ row.type_label }}</span></td>
                <td class="name">{{ row.name }}</td>
                <td class="num">{{ row.estimate_price }}</td>
                <td class="num">{{ row.market_price }}</td>
                <td class="where">
                  {% if row.site_url %}
                  <a href="{{ row.site_url }}" target="_blank" rel="noopener noreferrer">{{ row.site }}</a>
                  {% else %}
                  {{ row.site }}
                  {% endif %}
                </td>
                <td class="where">
                  {% if row.compare_class == 'good' %}
                  <span class="tone-good">{{ row.status }}</span>
                  {% elif row.compare_class == 'bad' %}
                  <span class="tone-bad">{{ row.status }}</span>
                  {% else %}
                  <span class="tone-warn">{{ row.status }}</span>
                  {% endif %}
                </td>
              </tr>
              {% endfor %}
            </tbody>
          </table>
          </div>
        </div>
        {% else %}
        <div class="status-box">Сравнение рынка пока не готово. Сначала выполните поиск цен по этой смете.</div>
        {% endif %}
      </div>
      <div class="table-panel" data-estimate-view-panel="sources" {% if active_table_view != 'sources' %}hidden{% endif %}>
        <div class="table-panel-head">
          <h2 class="table-panel-title">Источники цен</h2>
        </div>
        {% if sources_table.available %}
        <div class="table-wrap">
          <div class="table-scroll">
          <table>
            <thead>
              <tr>
                <th>Раздел</th>
                <th class="name">Наименование</th>
                <th>Цена рынка</th>
                <th>Сайт</th>
                <th>Статус</th>
                <th>Запрос</th>
              </tr>
            </thead>
            <tbody>
              {% for row in source_rows %}
              <tr>
                <td class="where">{{ row.section }}</td>
                <td class="name">{{ row.name }}</td>
                <td class="where">{{ row.market_price }}</td>
                <td class="where">{{ row.site }}</td>
                <td class="where">{{ row.status }}</td>
                <td class="where">{{ row.query }}</td>
              </tr>
              {% endfor %}
            </tbody>
          </table>
          </div>
        </div>
        {% else %}
        <div class="status-box">Источники цен пока не готовы. Сначала выполните поиск цен по этой смете.</div>
        {% endif %}
      </div>
    </section>
  </div>
  <div class="crm-drawer" id="estimateCrmDrawer" hidden>
    <div class="crm-drawer-backdrop" id="estimateCrmBackdrop"></div>
    <aside class="crm-drawer-panel" role="dialog" aria-modal="true" aria-labelledby="estimateCrmDrawerTitle">
      <div class="crm-drawer-head">
        <div>
          <h2 class="crm-drawer-title" id="estimateCrmDrawerTitle">Добавить смету в объект</h2>
          <div class="crm-drawer-sub">Выберите существующий объект или оставьте создание нового. Одна и та же смета при повторном импорте обновится без дублей.</div>
        </div>
        <button class="crm-drawer-close" id="estimateCrmCloseBtn" type="button">Закрыть</button>
      </div>
      <div class="crm-drawer-body">
        <div class="crm-source-card">
          <strong>На основе сметы: {{ crm_prefill.estimate_title }}</strong>
          <span>{{ crm_prefill.original_filename }} · строк: {{ crm_prefill.row_count }} · сумма: {{ crm_prefill.total_sum_fmt }}</span>
        </div>
        <form id="estimateCrmForm" class="crm-form-grid">
          <label>Объект CRM
            <select id="estimateCrmProject" name="project_id">
              <option value="">Создать новый объект</option>
            </select>
            <small id="estimateCrmProjectHint">Загружаю доступные объекты…</small>
          </label>
          <label>Название объекта
            <input type="text" id="estimateCrmTitle" name="title" required />
          </label>
          <label>Клиент / заказчик
            <input type="text" id="estimateCrmClient" name="client_name" />
          </label>
          <label>Адрес
            <input type="text" id="estimateCrmAddress" name="address" />
          </label>
          <label>Регион
            <input type="text" id="estimateCrmRegion" name="region" />
          </label>
          <label>Код / договор
            <input type="text" id="estimateCrmContractNo" name="contract_no" />
          </label>
          <label>Бюджет
            <input type="text" id="estimateCrmBudget" name="budget" inputmode="decimal" />
          </label>
          <label>Описание
            <textarea id="estimateCrmDescription" name="description"></textarea>
          </label>
          <div class="crm-status" id="estimateCrmStatus"></div>
          <div class="crm-drawer-foot">
            <button class="btn secondary" id="estimateCrmCancelBtn" type="button">Отмена</button>
            <button class="btn" type="submit" id="estimateCrmSubmitBtn">Создать объект</button>
          </div>
        </form>
      </div>
    </aside>
  </div>
  <script>
    let estimateMarketRenderFresh = {{ 'true' if (compare_table.available or sources_table.available) else 'false' }};
    let estimateMarketReloadPending = false;
    let estimateCrmDrawerTimer = null;
    let estimateCrmProjectsLoaded = false;
    const estimateCrmPrefill = {{ crm_prefill|tojson }};

    function setEstimateCrmStatus(message, tone) {
      const box = document.getElementById("estimateCrmStatus");
      if (!box) return;
      box.textContent = message || "";
      box.classList.toggle("is-error", tone === "error");
      box.classList.toggle("is-success", tone === "success");
    }

    function fillEstimateCrmForm(data) {
      const project = data && data.project ? data.project : {};
      const budget = project.budget == null ? "" : String(project.budget);
      const map = {
        estimateCrmTitle: project.title || "",
        estimateCrmClient: project.client_name || "",
        estimateCrmAddress: project.address || "",
        estimateCrmRegion: project.region || "",
        estimateCrmContractNo: project.contract_no || "",
        estimateCrmBudget: budget,
        estimateCrmDescription: project.description || "",
      };
      Object.entries(map).forEach(([id, value]) => {
        const node = document.getElementById(id);
        if (node) node.value = value;
      });
    }

    function getEstimateCrmFieldValue(id) {
      const node = document.getElementById(id);
      return node ? (node.value || "") : "";
    }

    function syncEstimateCrmMode() {
      const select = document.getElementById("estimateCrmProject");
      const submit = document.getElementById("estimateCrmSubmitBtn");
      const hint = document.getElementById("estimateCrmProjectHint");
      const adding = Boolean(select && select.value);
      if (submit) submit.textContent = adding ? "Добавить смету" : "Создать объект";
      if (hint && estimateCrmProjectsLoaded) {
        hint.textContent = adding
          ? "Смета будет добавлена отдельным файлом к выбранному объекту."
          : "Будет создан новый объект с данными из этой сметы.";
      }
    }

    async function loadEstimateCrmProjects() {
      if (estimateCrmProjectsLoaded) return;
      const select = document.getElementById("estimateCrmProject");
      const hint = document.getElementById("estimateCrmProjectHint");
      if (!select) return;
      try {
        const response = await fetch("/api/crm/projects", { headers: { "Accept": "application/json" }, cache: "no-store" });
        const data = await response.json().catch(function() { return {}; });
        if (!response.ok || !data.ok) throw new Error(data.message || ("HTTP " + response.status));
        (data.projects || []).forEach(function(project) {
          const option = document.createElement("option");
          option.value = String(project.id);
          option.textContent = "#" + project.id + " · " + (project.title || "Без названия") + (project.contract_no ? " · " + project.contract_no : "");
          select.appendChild(option);
        });
        estimateCrmProjectsLoaded = true;
        if (hint) hint.textContent = (data.projects || []).length ? "Выберите объект или создайте новый." : "Доступных объектов пока нет — будет создан новый.";
        syncEstimateCrmMode();
      } catch (error) {
        if (hint) hint.textContent = "Не удалось загрузить объекты: " + (error.message || error);
      }
    }

    function navigateEstimateCrmProject(url) {
      if (!url) return;
      try {
        if (window.parent && window.parent !== window) {
          window.parent.postMessage({ type: "pmbi:navigate", href: url }, "*");
          return;
        }
      } catch (e) {}
      window.location.href = url;
    }

    window.openEstimateCrmDrawer = function() {
      const drawer = document.getElementById("estimateCrmDrawer");
      if (!drawer) return;
      if (estimateCrmDrawerTimer) {
        clearTimeout(estimateCrmDrawerTimer);
        estimateCrmDrawerTimer = null;
      }
      fillEstimateCrmForm(estimateCrmPrefill);
      setEstimateCrmStatus("На основе сметы «" + (estimateCrmPrefill.estimate_title || "") + "».", "");
      drawer.hidden = false;
      requestAnimationFrame(() => drawer.classList.add("is-open"));
      document.body.style.overflow = "hidden";
      loadEstimateCrmProjects();
    };

    window.closeEstimateCrmDrawer = function() {
      const drawer = document.getElementById("estimateCrmDrawer");
      if (!drawer) return;
      drawer.classList.remove("is-open");
      if (estimateCrmDrawerTimer) clearTimeout(estimateCrmDrawerTimer);
      estimateCrmDrawerTimer = setTimeout(() => {
        drawer.hidden = true;
        estimateCrmDrawerTimer = null;
      }, 320);
      document.body.style.overflow = "";
    };

    window.submitEstimateCrmForm = async function(event) {
      event.preventDefault();
      const submitBtn = document.getElementById("estimateCrmSubmitBtn");
      if (submitBtn) submitBtn.disabled = true;
      const payload = {
        project_id: getEstimateCrmFieldValue("estimateCrmProject") || null,
        title: getEstimateCrmFieldValue("estimateCrmTitle"),
        client_name: getEstimateCrmFieldValue("estimateCrmClient"),
        address: getEstimateCrmFieldValue("estimateCrmAddress"),
        region: getEstimateCrmFieldValue("estimateCrmRegion"),
        contract_no: getEstimateCrmFieldValue("estimateCrmContractNo"),
        budget: getEstimateCrmFieldValue("estimateCrmBudget"),
        description: getEstimateCrmFieldValue("estimateCrmDescription"),
      };
      const addingToExisting = Boolean(payload.project_id);
      setEstimateCrmStatus(addingToExisting ? "Добавляю смету в выбранный объект…" : "Создаю объект в CRM…", "");
      try {
        const resp = await fetch("/api/estimates/{{ meta.id }}/export-to-crm", {
          method: "POST",
          headers: { "Content-Type": "application/json", "Accept": "application/json" },
          body: JSON.stringify(payload),
        });
        let data = {};
        try { data = await resp.json(); } catch (e) {}
        if (!resp.ok || !data.ok) {
          setEstimateCrmStatus(data.message || ("Не удалось добавить смету (HTTP " + resp.status + ")."), "error");
          return;
        }
        if (data.added_to_existing) {
          setEstimateCrmStatus("Готово: смета добавлена в объект #" + data.project_id + ". Строк обновлено: " + (data.materials_sent || 0) + ".", "success");
          if (data.project_url) navigateEstimateCrmProject(data.project_url);
          return;
        }
        const summary = data.summary || {};
        setEstimateCrmStatus("Готово: объект #" + data.project_id + " создан. Материалов отправлено: " + (data.materials_sent || 0) + ".", "success");
        if (data.project_url) navigateEstimateCrmProject(data.project_url);
      } catch (e) {
        setEstimateCrmStatus("Не удалось отправить данные в CRM: " + e, "error");
      } finally {
        if (submitBtn) submitBtn.disabled = false;
      }
    };

    const estimateCrmOpenBtn = document.getElementById("estimateCrmOpenBtn");
    if (estimateCrmOpenBtn) estimateCrmOpenBtn.addEventListener("click", window.openEstimateCrmDrawer);
    const estimateCrmBackdrop = document.getElementById("estimateCrmBackdrop");
    if (estimateCrmBackdrop) estimateCrmBackdrop.addEventListener("click", window.closeEstimateCrmDrawer);
    const estimateCrmCloseBtn = document.getElementById("estimateCrmCloseBtn");
    if (estimateCrmCloseBtn) estimateCrmCloseBtn.addEventListener("click", window.closeEstimateCrmDrawer);
    const estimateCrmCancelBtn = document.getElementById("estimateCrmCancelBtn");
    if (estimateCrmCancelBtn) estimateCrmCancelBtn.addEventListener("click", window.closeEstimateCrmDrawer);
    const estimateCrmForm = document.getElementById("estimateCrmForm");
    if (estimateCrmForm) estimateCrmForm.addEventListener("submit", window.submitEstimateCrmForm);
    const estimateCrmProject = document.getElementById("estimateCrmProject");
    if (estimateCrmProject) estimateCrmProject.addEventListener("change", syncEstimateCrmMode);

    async function deleteEstimate(btn) {
      const title = String({{ meta.title|tojson }});
      const ok = confirm(`Удалить смету "${title}"?\n\nБудут удалены карточка сметы, её строки и все сохранённые файлы рынка по этой смете.`);
      if (!ok) return;
      const initialHtml = btn ? btn.innerHTML : "";
      if (btn) {
        btn.disabled = true;
        btn.innerHTML = "...";
      }
      try {
        const resp = await fetch("/api/estimates/{{ meta.id }}/delete", {
          method: "POST",
          headers: { "Accept": "application/json" },
        });
        let data = {};
        try { data = await resp.json(); } catch (e) {}
        if (!resp.ok || !data.ok) {
          alert(data.message || ("Не удалось удалить смету (HTTP " + resp.status + ")."));
          if (btn) {
            btn.disabled = false;
            btn.innerHTML = initialHtml;
          }
          return;
        }
        window.location.href = "/estimates";
      } catch (e) {
        alert("Не удалось удалить смету: " + e);
        if (btn) {
          btn.disabled = false;
          btn.innerHTML = initialHtml;
        }
      }
    }

    function setEstimateTableView(viewKey) {
      const buttons = Array.from(document.querySelectorAll("[data-estimate-view-btn]"));
      const panels = Array.from(document.querySelectorAll("[data-estimate-view-panel]"));
      const activeBtn = buttons.find((btn) => btn.getAttribute("data-estimate-view-btn") === viewKey && !btn.disabled) || buttons.find((btn) => !btn.disabled);
      const nextKey = activeBtn ? activeBtn.getAttribute("data-estimate-view-btn") : "estimate";
      buttons.forEach((btn) => btn.classList.toggle("is-active", btn === activeBtn));
      panels.forEach((panel) => {
        panel.hidden = panel.getAttribute("data-estimate-view-panel") !== nextKey;
      });
      const hiddenInput = document.getElementById("estimateTableViewInput");
      if (hiddenInput) hiddenInput.value = nextKey;
      const downloadBtn = document.getElementById("activeTableDownloadBtn");
      if (downloadBtn && activeBtn) {
        downloadBtn.href = activeBtn.getAttribute("data-download-href") || "#";
        const label = activeBtn.getAttribute("data-download-label") || "";
        downloadBtn.textContent = label ? ("Скачать Excel: " + label) : "Скачать Excel";
      }
    }

    document.querySelectorAll("[data-estimate-view-btn]").forEach((btn) => {
      btn.addEventListener("click", function() {
        if (btn.disabled) return;
        setEstimateTableView(btn.getAttribute("data-estimate-view-btn") || "estimate");
      });
    });

    async function refreshEstimateMarketStatus() {
      try {
        const resp = await fetch("/api/estimates/{{ meta.id }}/market-status");
        if (!resp.ok) return;
        const data = await resp.json();
        const main = document.getElementById("marketStatusMain");
        const detail = document.getElementById("marketStatusDetail");
        const logs = document.getElementById("marketLogs");
        const startBtn = document.getElementById("marketStartBtn");
        const mergedBtn = document.getElementById("marketMergedBtn");
        const rawBtn = document.getElementById("marketRawBtn");
        const compareBtn = document.querySelector('[data-estimate-view-btn="compare"]');
        const sourcesBtn = document.querySelector('[data-estimate-view-btn="sources"]');
        if (startBtn) {
          startBtn.dataset.running = data.running ? "1" : "0";
          startBtn.disabled = startBtn.dataset.busy === "1";
          startBtn.textContent = data.running ? "Остановить поиск" : "Найти цены";
          startBtn.classList.toggle("is-stop", !!data.running);
        }
        if (mergedBtn) mergedBtn.hidden = !data.has_merged;
        if (rawBtn) rawBtn.hidden = !data.has_raw;
        if (!data.running && !estimateMarketRenderFresh && (data.has_merged || data.has_raw) && !estimateMarketReloadPending) {
          estimateMarketRenderFresh = true;
          estimateMarketReloadPending = true;
          if (compareBtn && data.has_merged) compareBtn.disabled = false;
          if (sourcesBtn && data.has_raw) sourcesBtn.disabled = false;
          const nextUrl = new URL(window.location.href);
          nextUrl.searchParams.set("table_view", data.has_merged ? "compare" : "sources");
          window.location.replace(nextUrl.toString());
          return;
        }
        if (main) {
          if (data.running) {
            main.textContent = "Идёт поиск цен: " + (data.done || 0) + " / " + (data.total || 0);
          } else if (data.result_ok) {
            main.textContent = "Поиск рынка завершён.";
          } else if (data.error) {
            main.textContent = "Поиск завершился с ошибкой.";
          } else {
            main.textContent = "Пока поиск рынка не запускался.";
          }
        }
        if (detail) {
          const bits = [];
          if (data.stage) bits.push(data.stage);
          if (data.detail) bits.push(data.detail);
          if (data.city) bits.push("город: " + data.city);
          detail.textContent = bits.join(" · ");
        }
        if (logs) {
          const arr = Array.isArray(data.log_tail) ? data.log_tail : [];
          logs.textContent = arr.length ? arr.join("\\n") : "—";
          logs.scrollTop = logs.scrollHeight;
        }
      } catch (e) {}
    }

    async function toggleEstimateMarket() {
      const btn = document.getElementById("marketStartBtn");
      const isRunning = btn && btn.dataset.running === "1";
      if (isRunning) {
        await stopEstimateMarket();
      } else {
        await startEstimateMarket();
      }
    }

    async function startEstimateMarket() {
      const cityInput = document.getElementById("marketCityInput");
      const city = cityInput ? String(cityInput.value || "").trim() : "";
      const selectedTypes = Array.from(document.querySelectorAll('input[name="types"]:checked')).map(x => String(x.value || ""));
      const btn = document.getElementById("marketStartBtn");
      if (btn) {
        btn.dataset.busy = "1";
        btn.disabled = true;
      }
      try {
        const resp = await fetch("/api/estimates/{{ meta.id }}/market-start", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ city, selected_types: selectedTypes })
        });
        const data = await resp.json();
        if (!resp.ok || !data.ok) {
          alert(data.message || "Не удалось запустить поиск рынка");
        }
      } catch (e) {
        alert("Не удалось запустить поиск рынка");
      } finally {
        if (btn) btn.dataset.busy = "0";
        refreshEstimateMarketStatus();
      }
    }

    async function stopEstimateMarket() {
      const btn = document.getElementById("marketStartBtn");
      if (btn) {
        btn.dataset.busy = "1";
        btn.disabled = true;
      }
      try {
        const resp = await fetch("/api/estimates/{{ meta.id }}/market-stop", { method: "POST" });
        const data = await resp.json();
        if (!resp.ok || !data.ok) {
          alert(data.message || "Не удалось остановить поиск");
        }
      } catch (e) {
        alert("Не удалось остановить поиск");
      } finally {
        if (btn) btn.dataset.busy = "0";
        refreshEstimateMarketStatus();
      }
    }

    setEstimateTableView("{{ active_table_view }}");
    refreshEstimateMarketStatus();
    setInterval(refreshEstimateMarketStatus, 3000);
  </script>
</body>
</html>
"""


ESTIMATE_MARKET_VIEW_TEMPLATE = """
<!doctype html>
<html lang="ru">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <link rel="icon" href="/favicon.svg" type="image/svg+xml" />
  <title>{{ meta.title }} · Сравнение цен</title>
  <style>
    :root { color-scheme: light; --bg:#f4f7fb; --panel:#ffffff; --border:#d9e3ef; --muted:#62748b; --text:#172235; }
    body { margin:0; font-family: Segoe UI, Arial, sans-serif; background:linear-gradient(180deg,#ffffff 0,#f4f7fb 100%); color:var(--text); }
    .page { max-width:1040px; margin:0 auto; padding:14px 12px 24px; }
    .panel { background:linear-gradient(180deg,#ffffff,#f8fbff); border:1px solid var(--border); border-radius:14px; box-shadow:0 14px 34px rgba(28,49,84,.08); padding:10px; margin-bottom:10px; }
    .muted,.where { color:var(--muted); }
    .chips { display:flex; flex-wrap:wrap; gap:6px; }
    .chip { display:inline-flex; align-items:center; gap:5px; text-decoration:none; border:1px solid #cfd9e8; background:#fff; color:#35506f; border-radius:10px; padding:6px 10px; font-weight:700; font-size:11px; }
    .chip.is-active { background:linear-gradient(180deg,#2e80e8,#1f72dc); border-color:#2e80e8; color:#fff; }
    .items { display:grid; gap:0; }
    .item { position:relative; border:1px solid #d9e3ef; border-radius:12px; background:linear-gradient(180deg,#ffffff,#f8fbff); padding:8px 9px; box-shadow:0 8px 18px rgba(28,49,84,.06); }
    .item + .item { margin-top:8px; }
    .item + .item::after { content:""; position:absolute; left:10px; right:10px; top:-12px; height:2px; background:linear-gradient(90deg, rgba(110,168,255,0), rgba(110,168,255,.55), rgba(46,139,87,.75), rgba(110,168,255,.55), rgba(110,168,255,0)); box-shadow:0 0 8px rgba(110,168,255,.16); }
    .item-head { display:flex; justify-content:space-between; gap:8px; align-items:flex-start; margin-bottom:6px; padding-bottom:6px; border-bottom:1px solid #e5ecf4; }
    .item-title { font-weight:700; font-size:12px; line-height:1.28; }
    .item-index { display:inline-flex; align-items:center; justify-content:center; min-width:22px; height:22px; padding:0 6px; border-radius:999px; background:linear-gradient(180deg,#2e80e8,#1f72dc); color:#fff; font-weight:800; font-size:10px; border:1px solid #2e80e8; box-shadow:0 4px 12px rgba(46,128,232,.16); }
    .tag { display:inline-flex; border-radius:999px; padding:1px 6px; border:1px solid #cfd9e8; background:#f4f8fd; color:#35506f; font-size:9px; white-space:nowrap; }
    .meta { display:flex; flex-wrap:wrap; gap:5px; margin-bottom:6px; }
    .offers { display:grid; gap:6px; margin-top:6px; }
    .offer { border:1px solid #dfe7f1; border-radius:9px; background:#fff; padding:6px; }
    .offer-top { display:flex; justify-content:space-between; gap:8px; align-items:flex-start; }
    .offer-title a { color:#1f72dc; text-decoration:none; }
    .offer-title a:hover { text-decoration:underline; }
    .offer-snippet { color:#62748b; font-size:10px; line-height:1.25; margin-top:4px; white-space:pre-wrap; }
    .status-note { color:#a06b18; font-size:10px; margin-top:5px; }
    .num { text-align:right; font-variant-numeric:tabular-nums; white-space:nowrap; }
  </style>
  <link rel="stylesheet" href="/static/autobot-ui.css?v=20260821" />
</head>
<body class="autobot-page market-view-page">
  <header class="topbar autobot-section-bar">
    <a class="brand" href="/tenders">
      <span class="brand-mark" aria-hidden="true"><i></i></span>
      <span class="brand-copy"><strong>AutoBot</strong><small>Закупки без рутины</small></span>
    </a>
    <nav class="topnav" aria-label="Разделы AutoBot">
      <a href="/tenders">Тендеры</a>
      <a class="is-active" href="/estimates">Сметы</a>
      <a href="/research">Поиск позиции</a>
    </nav>
  </header>
  <div class="page">
    <p style="margin:0 0 10px;"><a href="/estimates/{{ meta.id }}?{{ back_query }}">← Назад к смете</a> · <a href="/estimates">Все сметы</a> · <a href="/research">Поиск по позиции</a></p>
    <h1 style="margin:0 0 8px;">{{ meta.title }} · Сравнение цен</h1>
    <div class="muted">Показываю найденные сайты и цены по выбранному типу позиций.</div>

    <section class="panel">
      {% if market_links %}
      <div class="chips">
        {% for link in market_links %}
        <a class="chip{% if link.key == active_market_type %} is-active{% endif %}" href="{{ link.href }}">{{ link.label }} · {{ link.count }}</a>
        {% endfor %}
      </div>
      {% else %}
      <div>Пока нет сохранённых сравнений по текущему фильтру.</div>
      {% endif %}
    </section>

    <section class="panel">
      {% if active_section %}
      <div class="items">
        {% for item in active_section["items"] %}
        <article class="item">
          <div class="item-head">
            <div style="display:flex; gap:12px; align-items:flex-start;">
              <div class="item-index">{{ loop.index }}</div>
              <div class="item-title">{{ item.name }}</div>
            </div>
            <span class="tag">{{ item.type_label }}</span>
          </div>
          <div class="meta">
            <span class="tag">Кол-во: {{ item.qty_fmt }}</span>
            <span class="tag">Ед.: {{ item.unit or "—" }}</span>
            <span class="tag">Смета за ед.: {{ item.estimate_price_fmt }}</span>
            <span class="tag">Смета всего: {{ item.estimate_total_fmt }}</span>
            <span class="tag">Рынок: {{ item.market_prices or "—" }}</span>
          </div>
          {% if item["offers"] %}
          <div class="offers">
            {% for offer in item["offers"] %}
            <div class="offer">
              <div class="offer-top">
                <div class="offer-title">
                  {% if offer.url %}
                  <a href="{{ offer.url }}" target="_blank" rel="noopener noreferrer">{{ offer.title }}</a>
                  {% else %}
                  {{ offer.title }}
                  {% endif %}
                </div>
                <div class="num">{{ offer.price_fmt }}</div>
              </div>
              {% if offer.source and offer.source != "Интернет" %}
              <div class="where">{{ offer.source }}</div>
              {% endif %}
              {% if offer.snippet %}
              <div class="offer-snippet">{{ offer.snippet }}</div>
              {% endif %}
            </div>
            {% endfor %}
          </div>
          {% endif %}
          {% if item.status %}
          <div class="status-note">{{ item.status }}</div>
          {% endif %}
        </article>
        {% endfor %}
      </div>
      {% else %}
      <div>По выбранному типу пока нет сохранённых данных.</div>
      {% endif %}
    </section>
  </div>
</body>
</html>
"""

def _render_estimates_page_v2():
    def _sort_created_key(raw: object) -> tuple[int, float]:
        text = str(raw or "").strip()
        if not text:
            return (1, 0.0)
        try:
            dt = datetime.strptime(text, "%d.%m.%Y %H:%M")
            return (0, -dt.timestamp())
        except Exception:
            return (0, 0.0)

    type_order = ["material", "work", "service", "product", "other"]
    type_labels = {
        "material": "Материалы",
        "work": "Работы",
        "service": "Услуги",
        "product": "Товары",
        "other": "Другое",
    }
    cards = []
    total_rows = 0
    total_sum = 0.0
    total_sum_known = False
    with_compare = 0

    for meta in _read_estimates_index():
        if not isinstance(meta, dict):
            continue
        estimate_id = str(meta.get("id") or "")
        rows = _load_estimate_rows(estimate_id)
        summary = _summarize_estimate_rows(rows)
        type_counts = summary.get("type_counts") or {}
        type_badges = [{"key": key, "label": type_labels.get(key, key), "count": int(type_counts.get(key) or 0)} for key in type_order if int(type_counts.get(key) or 0) > 0]
        types_short = ", ".join(b["label"] for b in type_badges[:3]) if type_badges else "Без типов"
        has_market_compare = _estimate_market_merged_path(estimate_id).is_file()
        has_market_sources = _estimate_market_raw_path(estimate_id).is_file()
        market_done, market_total = _estimate_market_progress_for_card(estimate_id, rows)
        market_total = max(market_total, int(summary.get("row_count") or 0))
        market_done = max(0, min(market_done, market_total)) if market_total > 0 else 0
        market_pct = int(min(100, max(0, round(100.0 * market_done / market_total)))) if market_total > 0 else 0
        if has_market_compare:
            market_status_label = "Смета vs рынок"
            market_status_class = "status-ready"
            market_summary = "Сравнение уже готово"
            market_summary_class = "metric-good"
            market_progress_note = "Сравнение собрано, можно открывать смету и смотреть разбор."
            with_compare += 1
        elif has_market_sources:
            market_status_label = "Есть источники"
            market_status_class = "status-partial"
            market_summary = "Собраны источники рынка"
            market_summary_class = ""
            market_progress_note = f"Найдено цен: {market_done} из {market_total}. Можно дособрать сравнение внутри сметы."
        else:
            market_status_label = "Нет рынка"
            market_status_class = "status-idle"
            market_summary = "Рынок ещё не анализировали"
            market_summary_class = "metric-bad"
            market_progress_note = "Поиск цен ещё не запускался."

        item = dict(meta)
        item["total_sum_fmt"] = _fmt_money(summary.get("total_sum"))
        item["type_badges"] = type_badges
        item["types_short"] = types_short
        item["market_status_label"] = market_status_label
        item["market_status_class"] = market_status_class
        item["market_summary"] = market_summary
        item["market_summary_class"] = market_summary_class
        item["market_progress_done"] = market_done
        item["market_progress_total"] = market_total
        item["market_progress_percent"] = market_pct
        item["market_progress_note"] = market_progress_note
        cards.append(item)

        total_rows += int(summary.get("row_count") or 0)
        sum_value = summary.get("total_sum")
        if sum_value is not None:
            total_sum += float(sum_value)
            total_sum_known = True

    cards.sort(key=lambda x: _sort_created_key(x.get("created_at")))
    overview = {
        "total_count": len(cards),
        "total_rows": total_rows,
        "with_compare": with_compare,
        "total_sum_fmt": _fmt_money(total_sum) if total_sum_known else "—",
    }
    return render_template_string(ESTIMATES_TEMPLATE_V2, estimates=cards, overview=overview, crm_prefill={})


@app.route("/estimates")
def estimates_page():
    return _render_estimates_page_v2()
    cards = []
    for meta in _read_estimates_index():
        if not isinstance(meta, dict):
            continue
        rows = _load_estimate_rows(str(meta.get("id") or ""))
        summary = _summarize_estimate_rows(rows)
        type_counts = summary.get("type_counts") or {}
        labels = {"work": "работ", "service": "услуг", "product": "товаров", "material": "материалов", "other": "другое"}
        types_text = ", ".join(f"{labels.get(k, k)}: {v}" for k, v in type_counts.items()) or "—"
        item = dict(meta)
        item["total_sum_fmt"] = _fmt_money(summary.get("total_sum"))
        item["types_text"] = types_text
        cards.append(item)
    cards.sort(key=lambda x: str(x.get("created_at") or ""), reverse=True)
    return render_template_string(ESTIMATES_TEMPLATE, estimates=cards)


@app.route("/research")
def research_page():
    return render_template_string(
        RESEARCH_TEMPLATE,
        default_query=(request.args.get("q", "") or "").strip(),
        default_city=(request.args.get("city", "") or "").strip(),
    )


@app.route("/api/research-items", methods=["POST"])
def api_research_items():
    data = request.get_json(silent=True) or {}
    specs = _research_specs_from_text(str(data.get("queries") or data.get("query") or ""))
    city = re.sub(r"\s+", " ", str(data.get("city") or "").strip())[:120]
    if not specs:
        return jsonify({"ok": False, "message": "Нужна хотя бы одна позиция для поиска."}), 400

    from autobot.item_research import parse_sources, research_item

    sources = parse_sources(
        os.environ.get("MARKET_SUMMARY_SOURCES")
        or os.environ.get("MARKET_SOURCES")
        or "web,avito"
    )

    results: list[dict] = []
    for spec in specs:
        query = spec["query"]
        unit = spec["unit"]
        try:
            item = research_item(query, unit=unit, region=city, sources=sources, max_results=3)
            offers = []
            for offer in item.offers[:5]:
                offers.append(
                    {
                        "source": str(offer.source or ""),
                        "title": str(offer.title or ""),
                        "price": float(offer.price or 0) if offer.price else 0,
                        "url": str(offer.url or ""),
                        "snippet": str(offer.snippet or "")[:500],
                        "verification": str(offer.verification or "candidate"),
                        "verified": str(offer.verification or "") == "verified",
                        "confidence": float(offer.confidence or 0),
                        "reason": str(offer.verification_reason or offer.page_error or "")[:500],
                        "matched_unit": str(offer.matched_unit or ""),
                        "page_checked": bool(offer.page_checked),
                        "adapter": str(offer.adapter or ""),
                        "price_scope": str(offer.price_scope or ""),
                    }
                )
            results.append(
                {
                    "query": item.query,
                    "unit": item.unit,
                    "region": item.region,
                    "position_type": item.position_type,
                    "position_label": item.position_label,
                    "strategy": item.strategy,
                    "warning": item.warning,
                    "offers": offers,
                    "errors": str(item.errors or ""),
                }
            )
        except Exception as e:
            results.append(
                {
                    "query": query,
                    "unit": unit,
                    "region": city,
                    "offers": [],
                    "errors": str(e)[:400],
                }
            )
    return jsonify(
        {
            "ok": True,
            "message": f"Готово. Проверено позиций: {len(results)}.",
            "results": results,
        }
    )


@app.route("/estimates/<estimate_id>")
def estimate_detail_page(estimate_id: str):
    estimate_id = re.sub(r"[^0-9a-fA-F-]", "", estimate_id or "")[:40]
    meta = _load_estimate_meta(estimate_id)
    if not meta:
        abort(404)
    rows_all = _load_estimate_rows(estimate_id)
    q = (request.args.get("q", "") or "").strip()
    selected_types = _normalize_selected_estimate_types(request.args.getlist("types"))
    if not selected_types:
        legacy_type = (request.args.get("type", "") or "").strip()
        if legacy_type:
            selected_types = _normalize_selected_estimate_types([legacy_type])
        elif str(request.args.get("hide_work", "") or "").strip() in {"1", "true", "yes", "on"}:
            selected_types = [x for x in ["service", "product", "material", "other"]]
    rows = _filter_estimate_rows(rows_all, q=q, selected_types=selected_types)
    summary = _summarize_estimate_rows(rows)
    summary["total_sum_fmt"] = _fmt_money(summary.get("total_sum"))
    summary["avg_price_fmt"] = _fmt_money(summary.get("avg_price"))
    type_counts = _summarize_estimate_rows(rows_all).get("type_counts") or {}
    labels_full = {"work": "Работы", "service": "Услуги", "product": "Товары/изделия", "material": "Материалы", "other": "Другое"}
    order = ["work", "service", "product", "material", "other"]
    type_options = [
        {"key": k, "label": labels_full.get(k, k), "count": int(type_counts.get(k, 0))}
        for k in order
        if int(type_counts.get(k, 0)) > 0
    ]
    filter_query = urlencode([("q", q)] + [("types", t) for t in selected_types], doseq=True)
    rows_view = []
    display_no = 0
    last_section = None
    current_sheet = None
    current_sheet_total = 0.0
    current_sheet_has_sum = False
    for r in rows:
        rr = dict(r)
        sheet = str(rr.get("sheet") or "").strip()
        if current_sheet is None:
            current_sheet = sheet
        elif sheet != current_sheet:
            rows_view.append(
                {
                    "_is_sheet_total": True,
                    "sheet_title": current_sheet,
                    "sheet_total_fmt": _fmt_money(current_sheet_total if current_sheet_has_sum else None),
                }
            )
            rows_view.append({"_is_sheet_break": True})
            current_sheet = sheet
            current_sheet_total = 0.0
            current_sheet_has_sum = False
            last_section = None
        section = _normalize_section_title(str(rr.get("section") or ""))
        rr["section"] = section
        if section and section != last_section:
            rows_view.append(
                {
                    "_is_section": True,
                    "section_title": section,
                }
            )
            last_section = section
        display_no += 1
        rr["_is_section"] = False
        rr["display_no"] = display_no
        rr["qty_fmt"] = _fmt_qty(_json_num(rr.get("qty")))
        rr["unit_price_fmt"] = _fmt_money(_json_num(rr.get("unit_price")))
        rr["total_fmt"] = _fmt_money(_json_num(rr.get("total")))
        total_num = _json_num(rr.get("total"))
        if total_num is not None:
            current_sheet_total += total_num
            current_sheet_has_sum = True
        rows_view.append(rr)
    if current_sheet is not None and rows_view:
        rows_view.append(
            {
                "_is_sheet_total": True,
                "sheet_title": current_sheet,
                "sheet_total_fmt": _fmt_money(current_sheet_total if current_sheet_has_sum else None),
            }
        )
    compare_df = _estimate_market_df_for_rows(_estimate_market_merged_path(estimate_id), rows)
    raw_df = _estimate_market_df_for_rows(_estimate_market_raw_path(estimate_id), rows)
    compare_rows = _estimate_compare_rows(rows, compare_df)
    source_rows = _estimate_source_rows(rows, raw_df)
    scope_info = _estimate_market_scope_info(meta, selected_types)
    table_views = {
        "estimate": {"available": bool(rows)},
        "compare": {"available": bool(compare_rows)},
        "sources": {"available": bool(source_rows)},
    }
    active_table_view = _pick_estimate_active_table_view(request.args.get("table_view", ""), table_views)
    viability = _estimate_viability_overview(compare_df, compare_rows, scope_info)
    market_sections = _estimate_market_sections(estimate_id, rows, selected_types=selected_types)
    market_links = _estimate_market_links(estimate_id, market_sections, q=q, selected_types=selected_types)
    crm_prefill = _estimate_crm_prefill(estimate_id)
    return render_template_string(
        ESTIMATE_DETAIL_TEMPLATE,
        meta=meta,
        rows=rows_view,
        q=q,
        selected_types=selected_types,
        filter_query=filter_query,
        market_city=str(meta.get("market_city") or ""),
        has_market_raw=_estimate_market_raw_path(estimate_id).is_file(),
        has_market_merged=_estimate_market_merged_path(estimate_id).is_file(),
        active_table_view=active_table_view,
        compare_table=table_views["compare"],
        compare_rows=compare_rows,
        sources_table=table_views["sources"],
        source_rows=source_rows,
        viability=viability,
        market_links=market_links,
        type_options=type_options,
        summary=summary,
        crm_prefill=crm_prefill,
    )


@app.route("/estimates/<estimate_id>/market-view")
def estimate_market_view_page(estimate_id: str):
    estimate_id = re.sub(r"[^0-9a-fA-F-]", "", estimate_id or "")[:40]
    meta = _load_estimate_meta(estimate_id)
    if not meta:
        abort(404)
    rows_all = _load_estimate_rows(estimate_id)
    q = (request.args.get("q", "") or "").strip()
    selected_types = _normalize_selected_estimate_types(request.args.getlist("types"))
    if not selected_types:
        legacy_type = (request.args.get("type", "") or "").strip()
        if legacy_type:
            selected_types = _normalize_selected_estimate_types([legacy_type])
        elif str(request.args.get("hide_work", "") or "").strip() in {"1", "true", "yes", "on"}:
            selected_types = [x for x in ["service", "product", "material", "other"]]
    rows = _filter_estimate_rows(rows_all, q=q, selected_types=selected_types)
    market_sections = _estimate_market_sections(estimate_id, rows, selected_types=selected_types)
    market_links = _estimate_market_links(estimate_id, market_sections, q=q, selected_types=selected_types)
    active_market_type = (request.args.get("market_type", "") or "").strip()
    active_section = None
    if market_sections:
        active_section = next((sec for sec in market_sections if str(sec.get("key") or "") == active_market_type), None) or market_sections[0]
        active_market_type = str(active_section.get("key") or "")
    back_query = urlencode([("q", q)] + [("types", t) for t in selected_types], doseq=True)
    return render_template_string(
        ESTIMATE_MARKET_VIEW_TEMPLATE,
        meta=meta,
        market_links=market_links,
        active_market_type=active_market_type,
        active_section=active_section,
        back_query=back_query,
    )


@app.route("/estimates/<estimate_id>/download.xlsx")
def estimate_detail_download_xlsx(estimate_id: str):
    estimate_id = re.sub(r"[^0-9a-fA-F-]", "", estimate_id or "")[:40]
    meta = _load_estimate_meta(estimate_id)
    if not meta:
        abort(404)
    rows_all = _load_estimate_rows(estimate_id)
    q = (request.args.get("q", "") or "").strip()
    selected_types = _normalize_selected_estimate_types(request.args.getlist("types"))
    if not selected_types:
        legacy_type = (request.args.get("type", "") or "").strip()
        if legacy_type:
            selected_types = _normalize_selected_estimate_types([legacy_type])
        elif str(request.args.get("hide_work", "") or "").strip() in {"1", "true", "yes", "on"}:
            selected_types = [x for x in ["service", "product", "material", "other"]]
    rows = _filter_estimate_rows(rows_all, q=q, selected_types=selected_types)
    from openpyxl.styles import Font, PatternFill

    columns = ["№", "Тип", "Наименование", "Ед.", "Кол-во", "Цена за ед.", "Сумма", "Лист", "Строка Excel", "Раздел"]
    export_rows: list[dict] = []
    total_row_excel_numbers: list[int] = []
    break_row_excel_numbers: list[int] = []
    display_no = 0
    current_sheet = None
    current_sheet_total = 0.0
    current_sheet_has_sum = False

    def _append_sheet_total(sheet_name: str | None) -> None:
        export_rows.append(
            {
                "№": "",
                "Тип": "",
                "Наименование": f"Итого по листу: {sheet_name or 'без названия'}",
                "Ед.": "",
                "Кол-во": None,
                "Цена за ед.": None,
                "Сумма": current_sheet_total if current_sheet_has_sum else None,
                "Лист": sheet_name or "",
                "Строка Excel": None,
                "Раздел": "",
            }
        )
        total_row_excel_numbers.append(len(export_rows) + 1)

    def _append_break_row() -> None:
        export_rows.append({c: "" for c in columns})
        break_row_excel_numbers.append(len(export_rows) + 1)

    for r in rows:
        sheet = str(r.get("sheet") or "").strip()
        if current_sheet is None:
            current_sheet = sheet
        elif sheet != current_sheet:
            _append_sheet_total(current_sheet)
            _append_break_row()
            current_sheet = sheet
            current_sheet_total = 0.0
            current_sheet_has_sum = False

        display_no += 1
        total_num = _json_num(r.get("total"))
        if total_num is not None:
            current_sheet_total += total_num
            current_sheet_has_sum = True
        export_rows.append(
            {
                "№": display_no,
                "Тип": str(r.get("type_label") or ""),
                "Наименование": str(r.get("name") or ""),
                "Ед.": str(r.get("unit") or ""),
                "Кол-во": _json_num(r.get("qty")),
                "Цена за ед.": _json_num(r.get("unit_price")),
                "Сумма": total_num,
                "Лист": sheet,
                "Строка Excel": r.get("excel_row"),
                "Раздел": _normalize_section_title(str(r.get("section") or "")),
            }
        )

    if current_sheet is not None and rows:
        _append_sheet_total(current_sheet)

    df = pd.DataFrame(export_rows, columns=columns)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Смета")
        ws = writer.book["Смета"]
        total_fill = PatternFill(fill_type="solid", fgColor="2F4F6F")
        total_font = Font(color="FFD7D7", bold=True)
        break_fill = PatternFill(fill_type="solid", fgColor="6B2331")
        for row_no in total_row_excel_numbers:
            for cell in ws[row_no]:
                cell.fill = total_fill
                cell.font = total_font
        for row_no in break_row_excel_numbers:
            for cell in ws[row_no]:
                cell.fill = break_fill
            ws.row_dimensions[row_no].height = 12
    buf.seek(0)
    safe_stem = re.sub(r"[^0-9A-Za-zА-Яа-яЁё._ -]+", "_", str(meta.get("title") or estimate_id)).strip(" ._") or estimate_id
    filename = f"{safe_stem}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
        max_age=0,
    )


@app.route("/estimates/<estimate_id>/market-sources.xlsx")
def estimate_market_sources_download_xlsx(estimate_id: str):
    estimate_id = re.sub(r"[^0-9a-fA-F-]", "", estimate_id or "")[:40]
    meta = _load_estimate_meta(estimate_id) or {}
    rows_all = _load_estimate_rows(estimate_id)
    if not rows_all:
        abort(404)
    q = (request.args.get("q", "") or "").strip()
    selected_types = _normalize_selected_estimate_types(request.args.getlist("types"))
    rows = _filter_estimate_rows(rows_all, q=q, selected_types=selected_types)
    raw_df = _estimate_market_df_for_rows(_estimate_market_raw_path(estimate_id), rows)
    source_rows = _estimate_source_rows(rows, raw_df)
    if not source_rows:
        abort(404)
    filename = _safe_download_stem(str(meta.get("title") or estimate_id), estimate_id)
    export_df = _simple_sources_export_df(source_rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Источники")
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"{filename} - источники рынка.xlsx", max_age=0)


@app.route("/estimates/<estimate_id>/market-compare.xlsx")
def estimate_market_compare_download_xlsx(estimate_id: str):
    estimate_id = re.sub(r"[^0-9a-fA-F-]", "", estimate_id or "")[:40]
    meta = _load_estimate_meta(estimate_id) or {}
    rows_all = _load_estimate_rows(estimate_id)
    if not rows_all:
        abort(404)
    q = (request.args.get("q", "") or "").strip()
    selected_types = _normalize_selected_estimate_types(request.args.getlist("types"))
    rows = _filter_estimate_rows(rows_all, q=q, selected_types=selected_types)
    compare_df = _estimate_market_df_for_rows(_estimate_market_merged_path(estimate_id), rows)
    compare_rows = _estimate_compare_rows(rows, compare_df)
    if not compare_rows:
        abort(404)
    filename = _safe_download_stem(str(meta.get("title") or estimate_id), estimate_id)
    export_df = _simple_compare_export_df(compare_rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Сравнение")
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"{filename} - сравнение рынка.xlsx", max_age=0)


@app.route("/tenders/<tender_id>/estimate.xlsx")
def tender_estimate_download_xlsx(tender_id: str):
    tid = (tender_id or "").strip()
    if not tid or "/" in tid or ".." in tid:
        abort(404)
    path = REPORTS_DIR / f"ОТЧЕТ_ПО_СМЕТАМ_{tid}.xlsx"
    if not path.is_file():
        abort(404)
    meta = load_tender_metadata().get(tid) or {}
    filename = _safe_download_stem(meta.get("title") or tid, tid)
    return send_file(path, as_attachment=True, download_name=f"{filename} - смета.xlsx", max_age=0)


@app.route("/tenders/<tender_id>/market-sources.xlsx")
def tender_market_sources_download_xlsx(tender_id: str):
    tid = (tender_id or "").strip()
    if not tid or "/" in tid or ".." in tid:
        abort(404)
    path = _price_output_path_for_tender(tid)
    if not path.is_file():
        abort(404)
    meta = load_tender_metadata().get(tid) or {}
    filename = _safe_download_stem(meta.get("title") or tid, tid)
    return send_file(path, as_attachment=True, download_name=f"{filename} - источники рынка.xlsx", max_age=0)


@app.route("/tenders/<tender_id>/svodka.xlsx")
def tender_svodka_download_xlsx(tender_id: str):
    tid = (tender_id or "").strip()
    if not tid or "/" in tid or ".." in tid:
        abort(404)
    from autobot.merge_estimate_market import OUT_PREFIX

    path = REPORTS_DIR / f"{OUT_PREFIX}{tid}.xlsx"
    if not path.is_file():
        abort(404)
    meta = load_tender_metadata().get(tid) or {}
    filename = _safe_download_stem(meta.get("title") or tid, tid)
    return send_file(path, as_attachment=True, download_name=f"{filename} - выгодность.xlsx", max_age=0)


@app.route("/api/estimates/<estimate_id>/market-status")
def api_estimate_market_status(estimate_id: str):
    estimate_id = re.sub(r"[^0-9a-fA-F-]", "", estimate_id or "")[:40]
    meta = _load_estimate_meta(estimate_id) or {}
    with estimate_market_lock:
        job = dict(estimate_market_jobs.get(estimate_id) or {})
    payload = {
        "ok": True,
        "running": bool(job.get("running")),
        "result_ok": bool(job.get("ok")),
        "progress": int(job.get("progress") or 0),
        "stage": str(job.get("stage") or ""),
        "detail": str(job.get("detail") or ""),
        "error": str(job.get("error") or ""),
        "done": int(job.get("done") or 0),
        "total": int(job.get("total") or 0),
        "city": str(job.get("city") or meta.get("market_city") or ""),
        "selected_types": _normalize_selected_estimate_types(job.get("selected_types") or meta.get("market_selected_types") or []),
        "log_tail": list(job.get("log_lines") or []),
        "started_at": job.get("started_at"),
        "ended_at": job.get("ended_at"),
        "has_raw": _estimate_market_raw_path(estimate_id).is_file(),
        "has_merged": _estimate_market_merged_path(estimate_id).is_file(),
    }
    return jsonify(payload)


@app.route("/api/estimates/<estimate_id>/market-start", methods=["POST"])
def api_estimate_market_start(estimate_id: str):
    estimate_id = re.sub(r"[^0-9a-fA-F-]", "", estimate_id or "")[:40]
    meta = _load_estimate_meta(estimate_id)
    if not meta:
        return jsonify({"ok": False, "message": "Смета не найдена."}), 404
    rows = _load_estimate_rows(estimate_id)
    if not rows:
        return jsonify({"ok": False, "message": "У этой сметы нет строк для поиска рынка."}), 400
    with estimate_market_lock:
        cur = estimate_market_jobs.get(estimate_id)
        if cur and cur.get("running"):
            return jsonify({"ok": False, "message": "Поиск рынка уже выполняется."}), 409
    data = request.get_json(silent=True) or {}
    city = re.sub(r"\s+", " ", str(data.get("city") or "").strip())[:120]
    selected_types = _normalize_selected_estimate_types(data.get("selected_types") or [])
    filtered_rows = _filter_estimate_rows(rows, selected_types=selected_types)
    if not filtered_rows:
        return jsonify({"ok": False, "message": "По выбранным типам позиций нет строк для поиска цен."}), 400
    sources = ["avito", "web"]
    with estimate_market_lock:
        estimate_market_jobs[estimate_id] = {
            "running": True,
            "ok": False,
            "progress": 1,
            "stage": "Старт",
            "detail": "Запускаю поиск цен по строкам сметы",
            "error": "",
            "city": city,
            "sources": ",".join(sources),
            "selected_types": selected_types,
            "stop_requested": False,
            "done": 0,
            "total": len(filtered_rows),
            "started_at": datetime.now().isoformat(timespec="seconds"),
            "ended_at": None,
            "log_lines": [f"{datetime.now().strftime('%H:%M:%S')} · Старт поиска рынка" + (f" · город: {city}" if city else "")],
        }
    threading.Thread(
        target=_run_estimate_market_worker,
        kwargs={"estimate_id": estimate_id, "city": city, "sources": sources, "selected_types": selected_types},
        daemon=True,
    ).start()
    return jsonify({"ok": True, "message": "Поиск рынка запущен."})


@app.route("/api/estimates/<estimate_id>/market-stop", methods=["POST"])
def api_estimate_market_stop(estimate_id: str):
    estimate_id = re.sub(r"[^0-9a-fA-F-]", "", estimate_id or "")[:40]
    with estimate_market_lock:
        cur = estimate_market_jobs.get(estimate_id)
        if not cur:
            return jsonify({"ok": False, "message": "Активный поиск для этой сметы не найден."}), 404
        if not cur.get("running"):
            return jsonify({"ok": True, "message": "Поиск уже остановлен."})
        cur["stop_requested"] = True
        cur["stage"] = "Останавливаю"
        cur["detail"] = "Жду завершения текущей позиции и сохраняю уже найденные результаты"
        _estimate_market_log_append(cur, "Запрошена остановка поиска")
    return jsonify({"ok": True, "message": "Остановка поиска запрошена."})


@app.route("/api/estimates/upload", methods=["POST"])
def api_estimates_upload():
    f = request.files.get("file")
    if not f or not getattr(f, "filename", None):
        return jsonify({"ok": False, "message": "Выберите Excel-файл со сметой."}), 400
    if not _estimate_upload_allowed(f.filename):
        return jsonify({"ok": False, "message": "Нужен файл сметы: .xlsx, .xls, .xlsm или .pdf."}), 400
    estimate_id = uuid.uuid4().hex[:16]
    job_id = uuid.uuid4().hex[:16]
    est_dir = USER_ESTIMATES_DIR / estimate_id
    est_dir.mkdir(parents=True, exist_ok=True)
    original_name = _safe_upload_filename(f.filename)
    src_path = est_dir / original_name
    f.save(src_path)
    title_raw = (request.form.get("title", "") or "").strip()
    with estimate_upload_lock:
        estimate_upload_jobs[job_id] = {
            "job_id": job_id,
            "estimate_id": None,
            "running": True,
            "ok": False,
            "progress": 26,
            "stage": "Файл получен",
            "detail": "Сохраняю Excel и запускаю разбор",
            "error": "",
            "started_at": datetime.now().isoformat(timespec="seconds"),
            "ended_at": None,
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "log_lines": [f"{datetime.now().strftime('%H:%M:%S')} · Файл получен: {original_name}"],
        }
    threading.Thread(
        target=_run_estimate_upload_worker,
        kwargs={
            "job_id": job_id,
            "estimate_id": estimate_id,
            "title_raw": title_raw,
            "original_name": original_name,
            "src_path": src_path,
        },
        daemon=True,
    ).start()
    return jsonify(
        {
            "ok": True,
            "job_id": job_id,
            "progress": 26,
            "stage": "Файл получен",
            "detail": "Сервер принял файл и начал разбор",
            "message": "Смета загружена на сервер.",
        }
    )


@app.route("/api/estimates/upload-status/<job_id>")
def api_estimates_upload_status(job_id: str):
    with estimate_upload_lock:
        job = dict(estimate_upload_jobs.get(job_id) or {})
    if not job:
        return jsonify({"ok": False, "message": "Статус загрузки не найден."}), 404
    return jsonify(
        {
            "ok": True,
            "job_id": job_id,
            "estimate_id": job.get("estimate_id"),
            "running": bool(job.get("running")),
            "result_ok": bool(job.get("ok")),
            "progress": int(job.get("progress") or 0),
            "stage": job.get("stage") or "",
            "detail": job.get("detail") or "",
            "error": job.get("error") or "",
            "started_at": job.get("started_at"),
            "ended_at": job.get("ended_at"),
            "log_tail": list(job.get("log_lines") or [])[-12:],
        }
    )


@app.route("/reports/<path:filename>")
def report_file(filename: str):
    target = REPORTS_DIR / filename
    if not target.exists() or not target.is_file():
        abort(404)
    return send_from_directory(REPORTS_DIR, filename)


MERGE_REPORTS_SITE_DIR = REPO_ROOT / "data" / "reports_site"
NMCK_PREVIEW_DIR = REPO_ROOT / "data" / "nmck_previews"


NMCK_PREVIEW_PAGE = """
<!doctype html>
<html lang="ru">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <link rel="icon" href="/favicon.svg" type="image/svg+xml" />
  <title>{{ title|e }} — таблица НМЦК</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f4f7fb;
      --panel: #ffffff;
      --border: #d9e3ef;
      --text: #172235;
      --muted: #62748b;
      --accent: #1f72dc;
    }
    html, body { margin: 0; min-height: 100%; background: linear-gradient(180deg, #ffffff 0%, var(--bg) 100%); color: var(--text); font-family: Segoe UI, Arial, sans-serif; }
    .page { max-width: 100%; padding: 18px 16px 32px; box-sizing: border-box; }
    .head { max-width: 1400px; margin: 0 auto 14px; }
    h1 { font-size: 1.2rem; font-weight: 700; margin: 0 0 6px 0; letter-spacing: -0.02em; line-height: 1.35; word-break: break-word; }
    .sub { font-size: 13px; color: var(--muted); margin: 0 0 12px 0; }
    a.back { color: #1f72dc; font-size: 13px; text-decoration: none; }
    a.back:hover { text-decoration: underline; }
    .table-shell {
      max-width: 1400px;
      margin: 0 auto;
      border: 1px solid var(--border);
      border-radius: 12px;
      background: linear-gradient(180deg, var(--panel), #f8fbff);
      overflow: hidden;
      box-shadow: 0 10px 30px rgba(28, 49, 84, 0.08);
    }
    .table-scroll { overflow-x: auto; -webkit-overflow-scrolling: touch; }
    table { border-collapse: collapse; width: 100%; font-size: 12px; }
    th, td {
      border: 1px solid #e5ecf4;
      padding: 8px 10px;
      text-align: left;
      vertical-align: top;
      line-height: 1.35;
    }
    th {
      background: #f1f6fc;
      color: #35506f;
      font-weight: 600;
      max-width: 28em;
      word-break: break-word;
    }
    tbody tr:nth-child(even) { background: #fafcff; }
    tbody tr:hover { background: #f3f8ff; }
    td { word-break: break-word; max-width: 36em; }
    td.num { font-variant-numeric: tabular-nums; white-space: nowrap; max-width: none; }
    .foot { max-width: 1400px; margin: 14px auto 0; font-size: 11px; color: #62748b; }
  </style>
</head>
<body>
  <div class="page">
    <div class="head">
      <p style="margin:0 0 8px 0;"><a class="back" href="/">← На главную</a></p>
      <h1>{{ title|e }}</h1>
      <p class="sub">{{ subtitle|e }}</p>
    </div>
    <div class="table-shell">
      <div class="table-scroll">
        <table>
          <thead>
            <tr>
              {% for col in columns %}
              <th>{{ col }}</th>
              {% endfor %}
            </tr>
          </thead>
          <tbody>
            {% for row in rows %}
            <tr>
              {% for col in columns %}
              {% set v = row.get(col) %}
              <td class="{% if v is number %}num{% endif %}">{{ v if v is not none else '' }}</td>
              {% endfor %}
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
    <p class="foot">Колонки и строки как в загруженном Excel (обоснование НМЦК).</p>
  </div>
</body>
</html>
"""


MISSING_MERGE_PAGE = """
<!doctype html>
<html lang="ru">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <link rel="icon" href="/favicon.svg" type="image/svg+xml" />
  <title>Сравнение цен ещё не готово</title>
  <style>
    body { font-family: Segoe UI, sans-serif; background:linear-gradient(180deg,#ffffff 0,#f4f7fb 100%); color:#172235; margin:0; padding:24px; line-height:1.5; }
    a { color:#1f72dc; }
    .box { max-width:580px; margin:0 auto; border:1px solid #d9e3ef; border-radius:12px; padding:20px; background:#ffffff; box-shadow:0 12px 28px rgba(28,49,84,.08); }
    h1 { font-size:1.15rem; margin-top:0; }
    .btn { border:1px solid #2e80e8; background:linear-gradient(180deg,#2e80e8,#1f72dc); color:#ffffff; border-radius:8px; padding:10px 16px; cursor:pointer; font-size:14px; margin-top:12px; margin-right:8px; }
    .btn:disabled { opacity:.5; cursor:not-allowed; }
    .merge-bar-wrap { height:12px; background:#edf3fa; border-radius:8px; overflow:hidden; margin-top:12px; border:1px solid #d6e0ee; }
    .merge-bar-fill { height:100%; background:linear-gradient(90deg,#3d5290,#5ecf8a); transition:width .35s ease; }
    .logs { margin-top:10px; max-height:140px; overflow:auto; font-family:Consolas,monospace; font-size:11px; white-space:pre-wrap; background:#f8fbff; padding:8px; border-radius:8px; border:1px solid #dfe7f1; color:#576a84; }
    .hint { font-size:13px; color:#62748b; }
  </style>
</head>
<body>
  <div class="box">
    <h1>Сравнение цен ещё не готово</h1>
    <p>Закупка <strong>№ {{ tender_id }}</strong>. Чтобы увидеть результат, программе нужно извлечь смету, найти рыночные цены и собрать страницу сравнения.</p>
    <p class="hint">
    {% if not has_svodka_for_tid %}Для этой закупки рыночные цены ещё не найдены. Нажмите первую кнопку ниже.{% else %}Рыночные цены уже найдены, осталось обновить страницу результата.{% endif %}
    </p>
    <p><a href="/">← На главную</a> · <a id="retryLink" href="#">Обновить страницу</a></p>
    <button type="button" class="btn" id="genOneBtn">Подготовить сравнение для этой закупки</button>
    <button type="button" class="btn" id="genBtn" style="background:#283247;border-color:#4a567e;">Подготовить сравнения для всех</button>
    <div id="panel" style="margin-top:16px;display:none;">
      <div class="merge-bar-wrap"><div id="bar" class="merge-bar-fill" style="width:0%"></div></div>
      <p id="pct" style="margin:8px 0 0;font-size:14px;">0%</p>
      <div class="logs" id="logs"></div>
    </div>
    <p id="idleLine" class="hint" style="margin-top:12px;"></p>
  </div>
  <script>
    const REQUESTED_TID = {{ tender_id|tojson }};
    document.getElementById("retryLink").addEventListener("click", function(e) {
      e.preventDefault();
      location.reload();
    });
    let prevRun = false;
    async function tick() {
      try {
        const r = await fetch("/api/merge-site-status");
        const m = await r.json();
        const run = !!m.running;
        const panel = document.getElementById("panel");
        const idle = document.getElementById("idleLine");
        if (run || (m.log_tail && m.log_tail.length)) panel.style.display = "block";
        const pct = typeof m.percent === "number" ? m.percent : 0;
        document.getElementById("bar").style.width = Math.min(100, Math.max(0, pct)) + "%";
        document.getElementById("pct").textContent = pct + "% · " + (m.done||0) + " / " + (m.total||0) + (m.current_tid ? " · " + m.current_tid : "");
        document.getElementById("logs").textContent = (m.log_tail||[]).join("\\n");
        if (!run && m.last_ended_at) idle.textContent = "Последняя обработка: " + m.last_ended_at + " — " + (m.last_summary||"");
        else if (run) idle.textContent = "";
        document.getElementById("genBtn").disabled = run;
        document.getElementById("genOneBtn").disabled = run;
        if (prevRun && !run && m.total > 0) {
          try {
            const chk = await fetch("/merge-report/" + encodeURIComponent(REQUESTED_TID) + "/?t=" + Date.now(), { method: "GET", cache: "no-store" });
            if (chk.ok) location.reload();
          } catch (e) {}
        }
        prevRun = run;
      } catch (e) {}
    }
    document.getElementById("genBtn").addEventListener("click", async function() {
      try {
        const r = await fetch("/api/generate-merge-site-all", { method: "POST", headers: { "Content-Type": "application/json" } });
        const d = await r.json();
        if (!d.ok) alert(d.message || "Ошибка");
      } catch (e) { alert("Сеть"); }
      tick();
    });
    document.getElementById("genOneBtn").addEventListener("click", async function() {
      try {
        const r = await fetch("/api/generate-merge-site-one", {
          method: "POST",
          headers: { "Content-Type": "application/json", "Accept": "application/json" },
          body: JSON.stringify({ tender_id: REQUESTED_TID }),
        });
        const d = await r.json();
        if (!d.ok) alert(d.message || "Ошибка");
      } catch (e) { alert("Сеть"); }
      tick();
    });
    setInterval(tick, 1000);
    tick();
  </script>
</body>
</html>
"""


def _svodka_xlsx_tender_ids() -> list[str]:
    from autobot.merge_estimate_market import OUT_PREFIX

    if not REPORTS_DIR.is_dir():
        return []
    out: list[str] = []
    for p in REPORTS_DIR.glob(f"{OUT_PREFIX}*.xlsx"):
        tid = p.stem[len(OUT_PREFIX) :]
        if tid:
            out.append(tid)
    return sorted(out)


def _estimate_xlsx_tender_ids() -> list[str]:
    prefix = "ОТЧЕТ_ПО_СМЕТАМ_"
    if not REPORTS_DIR.is_dir():
        return []
    out: list[str] = []
    for p in REPORTS_DIR.glob(f"{prefix}*.xlsx"):
        if "ОБЩИЙ" in p.name:
            continue
        tid = p.stem[len(prefix) :]
        if tid:
            out.append(tid)
    return sorted(set(out))


def _compute_reports_coverage() -> dict[str, int]:
    """Тендеры из tenders.json vs готовые веб-сводки и Excel СВОДКА_РЫНОК."""
    merge_root = REPO_ROOT / "data" / "reports_site"
    from autobot.merge_estimate_market import OUT_PREFIX

    meta = load_tender_metadata()
    tender_ids = list(meta.keys())
    n_t = len(tender_ids)
    n_merge = sum(1 for tid in tender_ids if (merge_root / tid / "index.html").is_file())
    n_svodka = len(_svodka_xlsx_tender_ids())
    n_estimate = len(_estimate_xlsx_tender_ids())
    no_est = 0
    no_svodka = 0
    no_html = 0
    for tid in tender_ids:
        est_ok = (REPORTS_DIR / f"ОТЧЕТ_ПО_СМЕТАМ_{tid}.xlsx").is_file()
        sv_ok = (REPORTS_DIR / f"{OUT_PREFIX}{tid}.xlsx").is_file()
        html_ok = (merge_root / tid / "index.html").is_file()
        if not est_ok:
            no_est += 1
        if est_ok and not sv_ok:
            no_svodka += 1
        if sv_ok and not html_ok:
            no_html += 1
    return {
        "tender_count": n_t,
        "merge_html_among_tenders": n_merge,
        "tenders_missing_merge_html": max(0, n_t - n_merge),
        "svodka_xlsx_count": n_svodka,
        "estimate_xlsx_count": n_estimate,
        "missing_no_estimate": no_est,
        "missing_no_svodka": no_svodka,
        "missing_no_html": no_html,
    }


def _merge_site_busy() -> bool:
    with merge_site_lock:
        return bool(merge_site_state["running"])


def _missing_or_error_tender_ids() -> list[str]:
    """Только тендеры из tenders.json, где нет HTML-сводки или нет/битая цепочка до неё."""
    merge_root = REPO_ROOT / "data" / "reports_site"
    from autobot.merge_estimate_market import OUT_PREFIX

    out: list[str] = []
    for tid in sorted(load_tender_metadata().keys()):
        est_ok = (REPORTS_DIR / f"ОТЧЕТ_ПО_СМЕТАМ_{tid}.xlsx").is_file()
        sv_ok = (REPORTS_DIR / f"{OUT_PREFIX}{tid}.xlsx").is_file()
        html_ok = (merge_root / tid / "index.html").is_file()
        if (not html_ok) or (est_ok and not sv_ok):
            out.append(tid)
    return out


def _extract_tender_id(raw: str) -> str:
    """
    Извлекает regNumber / tender_id из ссылки zakupki.gov.ru или прямого ввода.
    Поддерживает:
    - полный URL с regNumber=...
    - номер закупки 223-ФЗ (11 цифр) или 44-ФЗ (19 цифр)
    - любой текст, где встречается такой номер
    """
    s = (raw or "").strip()
    if not s:
        return ""
    if s.isdigit() and len(s) in {11, 19}:
        return s
    try:
        p = urlparse(s)
        q = parse_qs(p.query)
        for key in ("regNumber", "purchaseNoticeNumber"):
            if key not in q or not q[key]:
                continue
            cand = (q[key][0] or "").strip()
            if cand.isdigit() and len(cand) in {11, 19}:
                return cand
    except Exception:
        pass
    import re

    m = re.search(r"(?<!\d)(\d{19}|\d{11})(?!\d)", s)
    if m:
        return m.group(1)
    return ""


def _truthy_env(name: str, default: str = "0") -> bool:
    v = (os.environ.get(name, default) or "").strip().lower()
    return v in ("1", "true", "yes", "on")


def _run_market_for_tender(
    tid: str,
    *,
    force_no_resume: bool = False,
    max_rows_override: int | None = None,
    rerun_selected: bool = False,
    sources_override: str | None = None,
    only_without_verified: bool = False,
    avito_collect_only: bool = False,
) -> tuple[int, str]:
    rows_map = _estimate_rows_by_tender_id()
    max_rows_arg: str | None = None
    max_rows_raw = str(max_rows_override or "").strip()
    if not max_rows_raw:
        max_rows_raw = (os.environ.get("MARKET_MAX_ROWS") or os.environ.get("MARKET_MAX_ROWS") or "").strip()
    if max_rows_raw:
        try:
            cap_rows = int(max_rows_raw)
        except ValueError:
            cap_rows = 0
        if cap_rows > 0:
            cap_rows = min(cap_rows, 5000)
            est_n = int(rows_map.get(tid, 0) or 0)
            use = cap_rows if est_n <= 0 else min(est_n, cap_rows)
            max_rows_arg = str(max(1, use))
    pause = "0" if avito_collect_only else ((os.environ.get("MARKET_PAUSE_SEC") or os.environ.get("MARKET_PAUSE_SEC") or "4").strip() or "4")
    sources = (sources_override or os.environ.get("MARKET_SOURCES") or "web,avito").strip() or "web,avito"
    max_results = "3" if avito_collect_only else ((os.environ.get("MARKET_MAX_RESULTS") or "5").strip() or "5")
    cmd = [
        sys.executable,
        str(_TOOLS_RUN_MODULE),
        "autobot.real_market_scraper",
        "--tender-id",
        tid,
        "--pause",
        pause,
        "--sources",
        sources,
        "--max-results-per-row",
        max_results,
    ]
    if max_rows_arg:
        cmd.extend(["--max-rows", max_rows_arg])
    if force_no_resume or _truthy_env("MARKET_NO_RESUME") or _truthy_env("MARKET_NO_RESUME"):
        cmd.append("--no-resume")
    elif rerun_selected:
        cmd.append("--rerun-selected")
    if only_without_verified:
        cmd.append("--only-without-verified")
    if avito_collect_only:
        cmd.extend(["--avito-collect-only", "--avito-safe-interval-sec", "35"])
    timeout_raw = (os.environ.get("MARKET_TIMEOUT_SEC") or os.environ.get("MARKET_TIMEOUT_SEC") or "21600").strip() or "21600"
    try:
        timeout_sec = max(60, int(timeout_raw))
    except ValueError:
        timeout_sec = 21600
    try:
        r = subprocess.run(cmd, cwd=str(REPO_ROOT), timeout=timeout_sec)
        return int(r.returncode), " ".join(cmd)
    except subprocess.TimeoutExpired:
        return 124, " ".join(cmd) + f" [timeout={timeout_sec}s]"


def _run_main_fetch_for_tender(tid: str, tender_url: str) -> tuple[int, str, str]:
    cmd = [
        sys.executable,
        str(_TOOLS_RUN_MODULE),
        "autobot.main",
        "--from-tender-id",
        tid,
        "--from-tender-url",
        tender_url,
    ]
    r = subprocess.run(cmd, cwd=str(REPO_ROOT), capture_output=True, text=True, errors="replace")
    out = (r.stdout or "") + ("\n" + r.stderr if r.stderr else "")
    return int(r.returncode), " ".join(cmd), out.strip()


def _main_failure_reason(output: str) -> str:
    """Короткая причина из вывода main.py для Telegram."""
    if not output:
        return "main.py не вернул подробный вывод."
    lines = [x.strip() for x in str(output).splitlines() if x.strip()]
    if not lines:
        return "main.py не вернул подробный вывод."
    hints = [
        "Документы не скачались",
        "не скачались",
        "Не удалось",
        "RAR",
        "excel",
        "xlsx",
        "Error",
        "Traceback",
        "Timeout",
        "captcha",
        "403",
        "404",
    ]
    picked: list[str] = []
    for line in lines:
        if any(h.lower() in line.lower() for h in hints):
            picked.append(line)
    tail = picked[-2:] if picked else lines[-2:]
    txt = " | ".join(tail)
    if len(txt) > 320:
        txt = txt[:317] + "..."
    return txt


def _estimate_diagnostics(tid: str, *, report_exists: bool, report_rows: int | None = None) -> str:
    """
    Короткая диагностика, почему не получилась смета/ЛСР для тендера.
    Возвращает строку для лога/Telegram (1-2 фразы).
    """
    tid_s = str(tid or "").strip()
    dl_dir = REPO_ROOT / "data" / "downloads" / tid_s
    ex_dir = REPO_ROOT / "data" / "extracted" / tid_s

    dl_files = [p for p in dl_dir.rglob("*") if p.is_file()] if dl_dir.is_dir() else []
    ex_files = [p for p in ex_dir.rglob("*") if p.is_file()] if ex_dir.is_dir() else []
    all_files = dl_files + ex_files

    excel_cnt = sum(1 for p in all_files if p.suffix.lower() in (".xlsx", ".xls"))
    rar_cnt = sum(1 for p in all_files if p.suffix.lower() == ".rar")
    zip_cnt = sum(1 for p in all_files if p.suffix.lower() == ".zip")

    ok_dl = 0
    failed_dl = 0
    log_path = dl_dir / "download_log.json"
    if log_path.is_file():
        try:
            payload = json.loads(log_path.read_text(encoding="utf-8"))
            if isinstance(payload, list):
                for row in payload:
                    st = str((row or {}).get("status") or "").strip().lower()
                    if st == "ok":
                        ok_dl += 1
                    elif st:
                        failed_dl += 1
        except Exception:
            pass

    if not dl_dir.is_dir():
        return "нет папки downloads по тендеру (документы не скачаны)."
    if ok_dl == 0 and failed_dl > 0:
        return f"скачивание документов неуспешно: ok=0, failed={failed_dl}."
    if ok_dl == 0 and not dl_files:
        return "скачивание не дало файлов (возможны капча/403/блокировка)."

    if not report_exists:
        if excel_cnt == 0:
            if rar_cnt > 0:
                return (
                    f"отчёт сметы не создан: Excel не найден, хотя есть архивы "
                    f"(RAR={rar_cnt}, ZIP={zip_cnt}); вероятно, не распаковано."
                )
            return "отчёт сметы не создан: среди скачанных файлов нет Excel."
        return f"отчёт сметы не создан при наличии Excel (найдено {excel_cnt})."

    if report_rows is not None and report_rows <= 0:
        if excel_cnt == 0:
            return "смета пустая (0 позиций): Excel-источники не обнаружены."
        if rar_cnt > 0 and ex_dir.is_dir() and not any(p.suffix.lower() in (".xlsx", ".xls") for p in ex_files):
            return (
                f"смета пустая (0 позиций): есть RAR={rar_cnt}, но из extracted нет Excel; "
                "проверьте 7-Zip/UnRAR."
            )
        return (
            f"смета пустая (0 позиций): Excel найдено {excel_cnt}, "
            "но структура ЛСР в них не распознана (возможен нестандартный формат)."
        )

    return "диагностика: смета и ЛСР найдены."


def _run_merge_site_all_worker(
    *,
    only_missing: bool = False,
    ids_override: list[str] | None = None,
    tender_url_by_id: dict[str, str] | None = None,
    force_market_no_resume: bool = False,
    market_max_rows: int | None = None,
    market_rerun_selected: bool = False,
    market_sources_override: str | None = None,
    market_only_without_verified: bool = False,
    market_avito_collect_only: bool = False,
) -> None:
    errors: list[str] = []
    ok_html = 0
    ok_full = 0
    ids: list[str] = []
    reason_counts = {"no_estimate": 0, "market_failed": 0, "merge_failed": 0, "html_failed": 0}
    try:
        from autobot.merge_estimate_market import merge_estimate_and_market
        from autobot.report_merge_html import write_tender_report_site

        if ids_override is not None:
            ids = [x for x in ids_override if str(x).strip()]
        else:
            ids = _missing_or_error_tender_ids() if only_missing else _estimate_xlsx_tender_ids()
        cap = 250
        with merge_site_lock:
            merge_site_state["running"] = True
            merge_site_state["total"] = len(ids)
            merge_site_state["done"] = 0
            merge_site_state["current_tid"] = ""
            merge_site_state["market_done"] = 0
            merge_site_state["market_total"] = 0
            merge_site_state["last_market_chat_done"] = 0
            merge_site_state["started_at"] = datetime.now().isoformat(timespec="seconds")
            merge_site_state["ended_at"] = None
            merge_site_state["error_ids"] = []
            merge_site_state["chat_events"] = []
            mode = (
                f"Авито: до {market_max_rows or 10} позиций без подтверждённой цены"
                if market_avito_collect_only else (
                    f"контрольная выборка по {market_max_rows} позиций"
                    if market_max_rows else ("только отсутствующие/ошибки" if only_missing else "все сметы")
                )
            )
            merge_site_state["log_lines"] = [f"Режим: {mode}. К обработке: {len(ids)}"]
            if not ids:
                merge_site_state["log_lines"].append(
                    "Нечего обрабатывать."
                )
                merge_site_state["running"] = False
                merge_site_state["ended_at"] = datetime.now().isoformat(timespec="seconds")
                merge_site_state["last_ended_at"] = merge_site_state["ended_at"]
                merge_site_state["last_summary"] = "Нет смет для обработки"
                merge_site_state["last_reason_counts"] = reason_counts
                return

        _merge_chat_add("start", f"Старт подготовки сравнений. Режим: {mode}. К обработке: {len(ids)}")
        for i, tid in enumerate(ids):
            pref = f"📊 <b>{i + 1}/{len(ids)}</b> · <code>{tid}</code>"
            with merge_site_lock:
                merge_site_state["current_tid"] = tid
                merge_site_state["market_done"] = 0
                merge_site_state["market_total"] = 0
                merge_site_state["last_market_chat_done"] = 0
                merge_site_state["log_lines"].append(f"[{i + 1}/{len(ids)}] {tid}…")
                merge_site_state["log_lines"] = merge_site_state["log_lines"][-cap:]
            _merge_chat_add("tender", f"📊 {i + 1}/{len(ids)} · тендер {tid}: старт", tender_id=tid)
            _tg_send(f"{pref}\n🟡 Старт")
            try:
                est_path = REPORTS_DIR / f"ОТЧЕТ_ПО_СМЕТАМ_{tid}.xlsx"
                cout = ""
                explicit_turl = ""
                if tender_url_by_id and tid in tender_url_by_id:
                    explicit_turl = (tender_url_by_id.get(tid) or "").strip()
                if not est_path.is_file() or explicit_turl:
                    turl = explicit_turl
                    if not turl:
                        # Пробуем URL из metadata для автодозагрузки.
                        md = load_tender_metadata().get(tid, {})
                        turl = str(md.get("url") or "").strip()
                    if turl:
                        _tg_send(f"{pref}\n🟡 Скачиваю смету…")
                        c, ccmd, cout = _run_main_fetch_for_tender(tid, turl)
                        with merge_site_lock:
                            merge_site_state["log_lines"].append(f"  main: {ccmd}")
                            merge_site_state["log_lines"].append(f"  main code: {c}")
                            if cout:
                                for line in cout.splitlines()[-8:]:
                                    merge_site_state["log_lines"].append("    " + line[:300])
                        if c != 0:
                            reason = _main_failure_reason(cout)
                            diag = _estimate_diagnostics(tid, report_exists=False)
                            _tg_send(
                                f"{pref}\n⚠️ main.py код <code>{c}</code>\n"
                                f"<code>{html_mod.escape(reason)}</code>\n<code>{html_mod.escape(diag)}</code>"
                            )
                        est_path = REPORTS_DIR / f"ОТЧЕТ_ПО_СМЕТАМ_{tid}.xlsx"
                    if not est_path.is_file():
                        reason_counts["no_estimate"] += 1
                        with merge_site_lock:
                            merge_site_state["log_lines"].append("  → пропуск: нет ОТЧЕТ_ПО_СМЕТАМ")
                        reason = _main_failure_reason(cout)
                        diag = _estimate_diagnostics(tid, report_exists=False)
                        _tg_send(
                            f"{pref}\n⚠️ Нет <code>ОТЧЕТ_ПО_СМЕТАМ_{tid}.xlsx</code>\n"
                            f"<code>{html_mod.escape(reason)}</code>\n<code>{html_mod.escape(diag)}</code>\n"
                            "Если ссылка уже была, обычно это временная недоступность ЕИС (таймаут/капча/блокировка); повторите позже."
                        )
                        errors.append(tid)
                        continue
                try:
                    _cnt = int(len(pd.read_excel(est_path, usecols=[0])))
                except Exception:
                    _cnt = 0
                if _cnt > 0:
                    _merge_chat_add("estimate", f"Смета: {_cnt} позиций. Запускаю поиск рынка…", tender_id=tid)
                    _tg_send(f"{pref}\n🟡 Смета: <b>{_cnt}</b> поз.")
                    _tg_flush_spool()
                    # Старт поиска — до тяжёлых проверок и до subprocess, иначе при spool сообщение
                    # может уехать в конец и появиться после всех строк прогресса.
                    _tg_send(f"{pref}\n🟡 Поиск рынка…")
                    _tg_flush_spool()
                else:
                    reason_counts["no_estimate"] += 1
                    diag = _estimate_diagnostics(tid, report_exists=True, report_rows=_cnt)
                    with merge_site_lock:
                        merge_site_state["log_lines"].append(f"  → пропуск: смета пустая (0 позиций) [{diag}]")
                    _tg_send(
                        f"{pref}\n⚠️ Смета пустая (0 поз.)\n<code>{html_mod.escape(diag)}</code>"
                    )
                    errors.append(tid)
                    continue

                if market_max_rows:
                    done_before, total_works = 0, min(int(market_max_rows), _cnt)
                else:
                    done_before, total_works = _market_progress_for_tender(tid)
                rem_before = max(0, total_works - done_before)
                with merge_site_lock:
                    merge_site_state["market_done"] = int(done_before)
                    merge_site_state["market_total"] = int(total_works)
                if total_works > 0:
                    if rem_before > 0:
                        _merge_chat_add("market", f"Рынок: обработано строк сметы {done_before}/{total_works}, осталось {rem_before}", tender_id=tid, seq=done_before, total=total_works)
                        _tg_send(f"{pref}\n🟡 Рынок: обработано строк сметы <b>{done_before}/{total_works}</b>…")
                    else:
                        _merge_chat_add("market", f"Рынок уже обработал строки сметы: {done_before}/{total_works}", tender_id=tid, seq=done_before, total=total_works)
                        _tg_send(f"{pref}\n🟢 Рынок уже обработал строки сметы: <b>{done_before}/{total_works}</b>")
                    _tg_flush_spool()
                _tg_flush_spool()
                market_code, market_cmd = _run_market_for_tender(
                    tid,
                    force_no_resume=force_market_no_resume,
                    max_rows_override=market_max_rows,
                    rerun_selected=market_rerun_selected,
                    sources_override=market_sources_override,
                    only_without_verified=market_only_without_verified,
                    avito_collect_only=market_avito_collect_only,
                )
                with merge_site_lock:
                    merge_site_state["log_lines"].append(f"  market: {market_cmd}")
                if market_code != 0:
                    reason_counts["market_failed"] += 1
                    with merge_site_lock:
                        if market_code == 124:
                            merge_site_state["log_lines"].append("  → поиск рынка завис и остановлен по таймауту")
                        else:
                            merge_site_state["log_lines"].append(f"  → поиск рынка код {market_code}")
                    if market_code == 124:
                        _merge_chat_add("error", "⚠️ Поиск рынка завис и остановлен по таймауту", tender_id=tid)
                        _tg_send(f"{pref}\n⚠️ Поиск рынка завис и остановлен по таймауту")
                    else:
                        _merge_chat_add("error", f"⚠️ Поиск рынка завершился с кодом {market_code}", tender_id=tid)
                        _tg_send(f"{pref}\n⚠️ Рынок код <code>{market_code}</code>")
                    errors.append(tid)
                    continue

                if market_max_rows:
                    done_after, total_after = min(int(market_max_rows), _cnt), min(int(market_max_rows), _cnt)
                else:
                    done_after, total_after = _market_progress_for_tender(tid)
                rem_after = max(0, total_after - done_after)
                with merge_site_lock:
                    merge_site_state["market_done"] = int(done_after)
                    merge_site_state["market_total"] = int(total_after)
                if total_after > 0:
                    if rem_after > 0:
                        _merge_chat_add("market", f"Рынок: обработано строк сметы {done_after}/{total_after}, осталось {rem_after}", tender_id=tid, seq=done_after, total=total_after)
                        _tg_send(f"{pref}\n🟡 Рынок: обработано строк сметы <b>{done_after}/{total_after}</b>, осталось <b>{rem_after}</b>")
                    else:
                        _merge_chat_add("market", f"🟢 Рынок обработал строки сметы: {done_after}/{total_after}", tender_id=tid, seq=done_after, total=total_after)
                        _tg_send(f"{pref}\n🟢 Рынок: строки сметы обработаны <b>{done_after}/{total_after}</b>")

                out = merge_estimate_and_market(tid)
                _merge_chat_add("merge", "Собираю СВОДКА_РЫНОК и страницу сравнения…", tender_id=tid)
                _tg_send(f"{pref}\n🟡 Merge…")
                if not out or not out.is_file():
                    reason_counts["merge_failed"] += 1
                    _merge_chat_add("error", f"⚠️ Не собрался СВОДКА_РЫНОК_{tid}.xlsx", tender_id=tid)
                    with merge_site_lock:
                        merge_site_state["log_lines"].append("  → merge не собрал СВОДКА_РЫНОК")
                    _tg_send(f"{pref}\n⚠️ Нет <code>СВОДКА_РЫНОК_{tid}.xlsx</code>")
                    errors.append(tid)
                    continue
                p = write_tender_report_site(tid)
                if p and p.is_file():
                    ok_html += 1
                    ok_full += 1
                    _merge_chat_add("done", f"✅ Тендер {tid}: сравнение готово", tender_id=tid)
                    site_url = get_report_site_public_base()
                    link = f"{site_url}/tenders/{tid}" if site_url else ""
                    with merge_site_lock:
                        merge_site_state["log_lines"].append("  → OK (рынок + merge + HTML)")
                    if link:
                        safe_link = html_mod.escape(link, quote=True)
                        _tg_send(
                            f"{pref}\n✅ Готово\n🔗 <a href=\"{safe_link}\">Отчёт</a>"
                        )
                    else:
                        _tg_send(
                            f"{pref}\n✅ Готово\n<code>data/reports_site/{tid}/index.html</code>"
                        )
                    try:
                        from autobot.tender_viability import format_viability_for_telegram

                        vmsg = format_viability_for_telegram(tid)
                        if vmsg:
                            _tg_flush_spool()
                            _tg_send(vmsg)
                            _tg_flush_spool()
                    except Exception:
                        pass
                else:
                    reason_counts["html_failed"] += 1
                    with merge_site_lock:
                        merge_site_state["log_lines"].append("  → HTML не создан")
                    _tg_send(f"{pref}\n⚠️ HTML не создан")
                    errors.append(tid)
            except Exception as e:
                reason_counts["merge_failed"] += 1
                with merge_site_lock:
                    merge_site_state["log_lines"].append(f"  → ошибка: {e}")
                _tg_send(f"{pref}\n⚠️ Ошибка: <code>{html_mod.escape(str(e)[:280])}</code>")
                errors.append(tid)
            with merge_site_lock:
                merge_site_state["done"] = i + 1
                merge_site_state["market_done"] = 0
                merge_site_state["market_total"] = 0
                merge_site_state["last_market_chat_done"] = 0
                merge_site_state["log_lines"] = merge_site_state["log_lines"][-cap:]

        ended = datetime.now().isoformat(timespec="seconds")
        with merge_site_lock:
            merge_site_state["done"] = len(ids)
            merge_site_state["current_tid"] = ""
            merge_site_state["market_done"] = 0
            merge_site_state["market_total"] = 0
            merge_site_state["last_market_chat_done"] = 0
            merge_site_state["running"] = False
            merge_site_state["ended_at"] = ended
            merge_site_state["error_ids"] = errors
            merge_site_state["last_ended_at"] = ended
            merge_site_state["last_summary"] = (
                f"Готово сравнений: {ok_full} из {len(ids)} "
                f"(не удалось обработать: {len(errors)})"
            )
            merge_site_state["last_reason_counts"] = reason_counts
            merge_site_state["log_lines"].append("--- Готово. Можно обновить страницу тендера. ---")
            merge_site_state["log_lines"] = merge_site_state["log_lines"][-cap:]
    except Exception as e:
        t = datetime.now().isoformat(timespec="seconds")
        with merge_site_lock:
            merge_site_state.setdefault("log_lines", []).append(f"Критическая ошибка: {e}")
            merge_site_state["last_summary"] = str(e)[:240]
            merge_site_state["last_ended_at"] = t
            merge_site_state["ended_at"] = t
            merge_site_state["last_reason_counts"] = reason_counts
    finally:
        with merge_site_lock:
            if merge_site_state["running"]:
                merge_site_state["running"] = False
                merge_site_state["current_tid"] = ""
                merge_site_state["market_done"] = 0
                merge_site_state["market_total"] = 0
                merge_site_state["last_market_chat_done"] = 0
                t = datetime.now().isoformat(timespec="seconds")
                merge_site_state["ended_at"] = t
                if not merge_site_state.get("last_ended_at"):
                    merge_site_state["last_ended_at"] = t
                merge_site_state.setdefault("last_summary", "Прервано / ошибка")


@app.route("/merge-report/<tender_id>/")
@app.route("/merge-report/<tender_id>/index.html")
def merge_report_site(tender_id: str):
    """Сводка смета + рынок (report_merge_html → data/reports_site/<id>/index.html)."""
    tid = (tender_id or "").strip()
    if not tid or "/" in tid or ".." in tid:
        abort(404)
    folder = MERGE_REPORTS_SITE_DIR / tid
    target = folder / "index.html"
    from autobot.merge_estimate_market import OUT_PREFIX

    svodka = REPORTS_DIR / f"{OUT_PREFIX}{tid}.xlsx"
    # Всегда пересобираем HTML из сводки — иначе остаётся старый index.html без карточек/скриптов.
    try:
        from autobot.report_merge_html import write_tender_report_site

        out_path = write_tender_report_site(tid)
        print(f"[merge-report] HTML attempt tender={tid} -> {out_path}", file=sys.stderr, flush=True)
    except Exception as e:
        print(f"[merge-report] HTML FAILED tender={tid}: {e}", file=sys.stderr, flush=True)
        traceback.print_exc(file=sys.stderr)
    if target.is_file():
        resp = make_response(send_from_directory(folder, "index.html"))
        resp.headers["Cache-Control"] = "no-store, max-age=0"
        return resp

    n = len(_svodka_xlsx_tender_ids())
    has_svodka = svodka.is_file()
    return render_template_string(
        MISSING_MERGE_PAGE,
        tender_id=tid,
        svodka_count=n,
        has_svodka_for_tid=has_svodka,
    )


@app.route("/nmck-preview/<preview_id>/")
def nmck_preview_table(preview_id: str):
    """Таблица из последнего разбора Excel обоснования НМЦК (payload в data/nmck_previews/)."""
    pid = (preview_id or "").strip().lower()
    if not re.fullmatch(r"[a-f0-9]{32}", pid):
        abort(404)
    path = NMCK_PREVIEW_DIR / pid / "payload.json"
    if not path.is_file():
        abort(404)
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        abort(404)
    columns = data.get("columns") or []
    rows = data.get("rows") or []
    meta = data.get("meta") or {}
    title = (meta.get("filename") or "Обоснование НМЦК").strip() or "Обоснование НМЦК"
    subtitle = (
        f"{meta.get('row_count', '?')} поз. · лист «{meta.get('sheet', '')}» · "
        f"{meta.get('column_count', '?')} колонок"
    )
    resp = make_response(
        render_template_string(
            NMCK_PREVIEW_PAGE,
            title=title,
            subtitle=subtitle,
            columns=columns,
            rows=rows,
        )
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    return resp


def _parse_env() -> dict[str, str]:
    env = os.environ.copy()
    env.setdefault("PYTHONUTF8", "1")
    env.setdefault("PYTHONUNBUFFERED", "1")
    return env


def _cmd_display(cmd: list[str]) -> str:
    return " ".join(repr(x) if any(c in x for c in " \t\"") else x for x in cmd)


def _stream_main_py(cli_args: list[str], *, log_cap: int = 400) -> int:
    """Один запуск main.py; дописывает строки в parse_state['log_lines']. Возвращает код выхода."""
    cmd = [sys.executable, "-u", str(_TOOLS_RUN_MODULE), "autobot.main"] + cli_args
    try:
        proc = subprocess.Popen(
            cmd,
            cwd=str(REPO_ROOT),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            errors="replace",
            env=_parse_env(),
            bufsize=1,
        )
    except OSError as e:
        with parse_lock:
            parse_state["log_lines"].append(f"Ошибка запуска процесса: {e}")
            parse_state["log_lines"] = parse_state["log_lines"][-log_cap:]
        return -1

    assert proc.stdout is not None
    for line in proc.stdout:
        line = line.rstrip("\n")
        with parse_lock:
            parse_state["log_lines"].append(line)
            parse_state["log_lines"] = parse_state["log_lines"][-log_cap:]
    return proc.wait()


def _run_main_worker(cli_args: list[str], task: str) -> None:
    """Запуск main.py тем же Python, что и веб-сервер (не py.exe из PATH)."""
    cmd = [sys.executable, "-u", str(_TOOLS_RUN_MODULE), "autobot.main"] + cli_args
    cmd_display = _cmd_display(cmd)
    with parse_lock:
        parse_state["running"] = True
        parse_state["task"] = task
        parse_state["command"] = cmd_display
        parse_state["started_at"] = datetime.now().isoformat(timespec="seconds")
        parse_state["ended_at"] = None
        parse_state["exit_code"] = None
        parse_state["log_lines"] = [f">>> {cmd_display}"]
    exit_code = -1
    try:
        exit_code = _stream_main_py(cli_args, log_cap=300)
    except Exception:
        err = traceback.format_exc(limit=8)
        with parse_lock:
            parse_state["log_lines"].append("!!! Внутренняя ошибка фонового парсинга:")
            parse_state["log_lines"].extend(err.rstrip().splitlines())
            parse_state["log_lines"] = parse_state["log_lines"][-300:]
    finally:
        with parse_lock:
            parse_state["running"] = False
            parse_state["task"] = ""
            parse_state["command"] = ""
            parse_state["ended_at"] = datetime.now().isoformat(timespec="seconds")
            parse_state["exit_code"] = exit_code
            if exit_code == 0:
                parse_state["log_lines"].append(
                    "--- Готово. Обновите страницу (F5), чтобы подтянуть список тендеров. ---"
                )


def _run_rebuild_all_worker() -> None:
    """Последовательно пересобирает Excel/HTML для каждого тендера из tenders.json (как --from-downloaded-tender-id)."""
    meta = load_tender_metadata()
    ids = sorted(meta.keys(), key=lambda x: x)
    started = datetime.now().isoformat(timespec="seconds")
    with parse_lock:
        parse_state["running"] = True
        parse_state["task"] = "пересбор всех отчётов"
        parse_state["command"] = f"{len(ids)} тендеров"
        parse_state["started_at"] = started
        parse_state["ended_at"] = None
        parse_state["exit_code"] = None
        if not ids:
            parse_state["log_lines"] = ["В tenders.json нет тендеров — нечего пересобирать."]
            parse_state["running"] = False
            parse_state["task"] = ""
            parse_state["ended_at"] = datetime.now().isoformat(timespec="seconds")
            parse_state["exit_code"] = 0
            return
        parse_state["log_lines"] = [
            f">>> Пересбор отчётов для {len(ids)} тендеров из tenders.json (по очереди, тот же алгоритм, что «Пересобрать отчёт»)."
        ]

    failed: list[tuple[str, int]] = []
    cap = 600
    for i, tid in enumerate(ids, 1):
        with parse_lock:
            parse_state["task"] = f"пересбор {i}/{len(ids)}: {tid}"
            parse_state["command"] = _cmd_display(
                [
                    sys.executable,
                    str(_TOOLS_RUN_MODULE),
                    "autobot.main",
                    "--from-downloaded-tender-id",
                    tid,
                ]
            )
            parse_state["log_lines"].append(f"--- [{i}/{len(ids)}] {tid} ---")
            parse_state["log_lines"] = parse_state["log_lines"][-cap:]

        code = _stream_main_py(["--from-downloaded-tender-id", tid], log_cap=cap)
        if code != 0:
            failed.append((tid, code))
        with parse_lock:
            parse_state["log_lines"].append(f"--- конец {tid}, код {code} ---")
            parse_state["log_lines"] = parse_state["log_lines"][-cap:]

    with parse_lock:
        parse_state["running"] = False
        parse_state["task"] = ""
        parse_state["command"] = ""
        parse_state["ended_at"] = datetime.now().isoformat(timespec="seconds")
        parse_state["exit_code"] = 0 if not failed else 1
        parse_state["log_lines"].append(
            "--- Пересбор всех отчётов завершён. Обновите страницу (F5). ---"
        )
        if failed:
            parse_state["log_lines"].append(
                "Тендеры с ненулевым кодом выхода: " + ", ".join(f"{t} ({c})" for t, c in failed)
            )
        parse_state["log_lines"] = parse_state["log_lines"][-cap:]


def _nmck_upload_allowed(filename: str) -> bool:
    fn = (filename or "").lower().strip()
    return fn.endswith((".xlsx", ".xls", ".xlsm"))


@app.route("/api/parse-nmck-justification", methods=["POST"])
def api_parse_nmck_justification():
    """Excel «Обоснование НМЦК» (приложение №2) → JSON (позиции и все колонки таблицы)."""
    f = request.files.get("file")
    if not f or not getattr(f, "filename", None):
        return jsonify({"ok": False, "message": "Выберите файл в поле «Обоснование НМЦК»."}), 400
    if not _nmck_upload_allowed(f.filename):
        return jsonify({"ok": False, "message": "Нужен файл Excel: .xlsx, .xls или .xlsm."}), 400
    try:
        raw = f.read()
    except Exception as e:
        return jsonify({"ok": False, "message": f"Не удалось прочитать файл: {e}"}), 400
    if not raw:
        return jsonify({"ok": False, "message": "Пустой файл."}), 400
    try:
        from autobot.nmck_justification_parse import parse_nmck_justification_excel

        out = parse_nmck_justification_excel(raw, original_name=f.filename)
    except ValueError as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": f"Ошибка разбора Excel: {e}"}), 400
    preview_id = uuid.uuid4().hex
    folder = NMCK_PREVIEW_DIR / preview_id
    try:
        folder.mkdir(parents=True, exist_ok=True)
        payload = {"columns": out["columns"], "rows": out["rows"], "meta": out["meta"]}
        (folder / "payload.json").write_text(
            json.dumps(payload, ensure_ascii=False),
            encoding="utf-8",
        )
    except OSError as e:
        return jsonify({"ok": False, "message": f"Не удалось сохранить превью: {e}"}), 500
    out["preview_id"] = preview_id
    out["preview_url"] = f"/nmck-preview/{preview_id}/"
    return jsonify({"ok": True, **out})


@app.route("/api/start-parse", methods=["POST"])
def api_start_parse():
    if _merge_site_busy():
        return jsonify({"ok": False, "message": "Сначала дождитесь окончания подготовки сравнений цен."}), 409
    with parse_lock:
        if parse_state["running"]:
            return jsonify({"ok": False, "message": "Уже выполняется задание"}), 409
    data = request.get_json(silent=True) or {}
    try:
        max_pages = int(data.get("max_pages", 2))
        max_tenders = int(data.get("max_tenders", 15))
        days_back = int(data.get("days_back", 60))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "message": "Некорректные числа (max_pages, max_tenders, days_back)"}), 400
    max_pages = max(1, min(max_pages, 20))
    max_tenders = max(1, min(max_tenders, 100))
    days_back = max(1, min(days_back, 365))
    args = [
        "--max-pages",
        str(max_pages),
        "--max-tenders",
        str(max_tenders),
        "--days-back",
        str(days_back),
    ]
    catalog_only_raw = data.get("catalog_only", True)
    catalog_only = catalog_only_raw not in (False, 0, "0", "false", "False", "no", "off")
    if catalog_only:
        args.append("--catalog-only")
    worker = threading.Thread(
        target=_run_main_worker,
        kwargs={"cli_args": args, "task": "поиск закупок для каталога" if catalog_only else "поиск новых закупок"},
        daemon=True,
    )
    with parse_lock:
        if parse_state["running"]:
            return jsonify({"ok": False, "message": "Уже выполняется задание"}), 409
        parse_state["running"] = True
        parse_state["task"] = "поиск закупок для каталога" if catalog_only else "запуск поиска новых закупок"
        parse_state["command"] = ""
        parse_state["started_at"] = datetime.now().isoformat(timespec="seconds")
        parse_state["ended_at"] = None
        parse_state["exit_code"] = None
        parse_state["log_lines"] = ["Подготавливаем запуск поиска…"]
    try:
        worker.start()
    except RuntimeError as e:
        with parse_lock:
            parse_state["running"] = False
            parse_state["task"] = ""
            parse_state["ended_at"] = datetime.now().isoformat(timespec="seconds")
            parse_state["exit_code"] = -1
            parse_state["log_lines"].append(f"Не удалось запустить фоновую задачу: {e}")
        return jsonify({"ok": False, "message": "Не удалось запустить фоновую задачу поиска."}), 500
    return jsonify({"ok": True})


@app.route("/api/rebuild-report", methods=["POST"])
def api_rebuild_report():
    if _merge_site_busy():
        return jsonify({"ok": False, "message": "Сначала дождитесь окончания подготовки сравнений цен."}), 409
    with parse_lock:
        if parse_state["running"]:
            return jsonify({"ok": False, "message": "Уже выполняется задание"}), 409
    data = request.get_json(silent=True) or {}
    tid = str(data.get("tender_id", "")).strip()
    if not tid:
        return jsonify({"ok": False, "message": "Укажите tender_id"}), 400
    if not _AUTOBOT_MAIN_FILE.is_file():
        return jsonify({"ok": False, "message": f"Не найден {_AUTOBOT_MAIN_FILE}"}), 500
    threading.Thread(
        target=_run_main_worker,
        kwargs={"cli_args": ["--from-downloaded-tender-id", tid], "task": f"повторное извлечение сметы {tid}"},
        daemon=True,
    ).start()
    return jsonify({"ok": True})


@app.route("/api/rebuild-all-reports", methods=["POST"])
def api_rebuild_all_reports():
    if _merge_site_busy():
        return jsonify({"ok": False, "message": "Сначала дождитесь окончания подготовки сравнений цен."}), 409
    with parse_lock:
        if parse_state["running"]:
            return jsonify({"ok": False, "message": "Уже выполняется задание"}), 409
    if not _AUTOBOT_MAIN_FILE.is_file():
        return jsonify({"ok": False, "message": f"Не найден {_AUTOBOT_MAIN_FILE}"}), 500
    if not TENDERS_JSON.is_file():
        return jsonify({"ok": False, "message": "Нет файла tenders.json"}), 400
    meta = load_tender_metadata()
    if not meta:
        return jsonify({"ok": False, "message": "В tenders.json нет тендеров"}), 400
    threading.Thread(target=_run_rebuild_all_worker, daemon=True).start()
    return jsonify({"ok": True, "count": len(meta)})


@app.route("/api/parse-status")
def api_parse_status():
    with parse_lock:
        payload = {
            "running": parse_state["running"],
            "task": parse_state["task"],
            "command": parse_state["command"],
            "started_at": parse_state["started_at"],
            "ended_at": parse_state["ended_at"],
            "exit_code": parse_state["exit_code"],
            "log_lines_count": len(parse_state["log_lines"]),
            "log_tail": parse_state["log_lines"][-80:],
        }
    return jsonify(payload)


@app.route("/api/tenders")
def api_tenders():
    if not TENDERS_JSON.exists():
        return jsonify({"items": []})
    try:
        data = json.loads(TENDERS_JSON.read_text(encoding="utf-8"))
    except Exception:
        return jsonify({"items": []})
    items = []
    for row in data:
        tid = str(row.get("tender_id", "") or "").strip()
        stage_raw = (str(row.get("stage") or "")).strip()
        stage_open = stage_raw == STAGE_SUBMISSION
        stage_display = stage_raw if stage_raw else "—"
        items.append(
            {
                "tender_id": row.get("tender_id"),
                "region": row.get("region"),
                "title": row.get("title"),
                "price_rub": row.get("price_rub"),
                "eis_url": eis_notice_url(tid, row.get("url")),
                "stage_display": stage_display,
                "stage_open": stage_open,
                "publish_date": (row.get("publish_date") or ""),
                "updated_date": (row.get("updated_date") or ""),
                "customer_name": (row.get("customer_name") or ""),
                "law": (row.get("law") or ""),
                "purchase_method": (row.get("purchase_method") or ""),
            }
        )
    items.sort(key=lambda x: (x.get("region") or "", str(x.get("tender_id") or "")))
    return jsonify({"items": items[:200]})


@app.route("/api/tenders/<tender_id>/delete", methods=["POST"])
def api_delete_tender(tender_id: str):
    tid = str(tender_id or "").strip()
    if not re.fullmatch(r"\d{8,25}", tid):
        return jsonify({"ok": False, "message": "Некорректный номер тендера."}), 400
    data = request.get_json(silent=True) or {}
    if str(data.get("confirm_tender_id") or "").strip() != tid:
        return jsonify({"ok": False, "message": "Удаление не подтверждено номером тендера."}), 400
    with parse_lock:
        parse_running = bool(parse_state.get("running"))
    with merge_site_lock:
        merge_running = bool(merge_site_state.get("running"))
    if parse_running or merge_running:
        return jsonify({"ok": False, "message": "Дождитесь завершения текущей задачи AutoBot и повторите удаление."}), 409
    try:
        with tender_delete_lock:
            result = delete_tender_data(tid)
    except FileNotFoundError:
        return jsonify({"ok": False, "message": "Тендер уже удалён или не найден."}), 404
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)[:500]}), 400
    except Exception as exc:
        return jsonify({"ok": False, "message": f"Не удалось удалить тендер: {str(exc)[:500]}"}), 500
    return jsonify({"ok": True, **result})


@app.route("/api/crm/projects")
def api_crm_projects():
    try:
        projects = crm_projects_for_picker()
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)[:700]}), 502
    return jsonify({"ok": True, "projects": projects})


@app.route("/api/export-to-crm", methods=["POST"])
def api_export_to_crm():
    data = request.get_json(silent=True) or {}
    tid = str(data.get("tender_id", "")).strip()
    if not tid:
        return jsonify({"ok": False, "message": "Нужен tender_id"}), 400
    if tid not in load_tender_metadata():
        return jsonify({"ok": False, "message": "Такой тендер не найден в tenders.json"}), 404
    try:
        result = export_tender_to_crm(tid, project_id=data.get("project_id", data.get("projectId")))
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)[:700]}), 500
    return jsonify({"ok": True, "tender_id": tid, **result})


@app.route("/api/estimates/<estimate_id>/export-to-crm", methods=["POST"])
def api_export_estimate_to_crm(estimate_id: str):
    estimate_id = re.sub(r"[^0-9a-fA-F-]", "", estimate_id or "")[:40]
    meta = _load_estimate_meta(estimate_id)
    if not meta:
        return jsonify({"ok": False, "message": "Смета не найдена."}), 404
    data = request.get_json(silent=True) or {}
    try:
        result = export_estimate_to_crm(
            estimate_id,
            overrides=data,
            project_id=data.get("project_id", data.get("projectId")),
        )
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)[:700]}), 500
    return jsonify({"ok": True, "estimate_id": estimate_id, **result})


@app.route("/api/estimates/<estimate_id>/delete", methods=["POST"])
def api_delete_estimate(estimate_id: str):
    estimate_id = re.sub(r"[^0-9a-fA-F-]", "", estimate_id or "")[:40]
    if not estimate_id:
        return jsonify({"ok": False, "message": "Нужен estimate_id."}), 400
    try:
        delete_estimate(estimate_id)
    except Exception as e:
        message = str(e)[:700]
        status = 409 if "идёт поиск рынка" in message else 500
        if "не найдена" in message.casefold():
            status = 404
        return jsonify({"ok": False, "message": message}), status
    return jsonify({"ok": True, "estimate_id": estimate_id})


@app.route("/api/merge-site-status")
def api_merge_site_status():
    with merge_site_lock:
        total = int(merge_site_state["total"] or 0)
        done = int(merge_site_state["done"] or 0)
        running = bool(merge_site_state["running"])
        current_tid = merge_site_state.get("current_tid") or ""
        market_done = int(merge_site_state.get("market_done") or 0)
        market_total = int(merge_site_state.get("market_total") or 0)

    if running and current_tid:
        live_market_done, live_market_total = _market_progress_for_tender(current_tid)
        if live_market_total > 0:
            if live_market_total == market_total:
                market_done = max(market_done, int(live_market_done))
            else:
                market_done = int(live_market_done)
            market_total = int(live_market_total)
            with merge_site_lock:
                if merge_site_state.get("current_tid") == current_tid:
                    merge_site_state["market_done"] = market_done
                    merge_site_state["market_total"] = market_total

    market_percent = int(min(100, max(0, round(100.0 * market_done / market_total)))) if market_total > 0 else 0
    current_fraction = (market_done / market_total) if (running and market_total > 0) else 0.0
    if running and total > 0:
        pct = int(min(99, max(0, round(100.0 * (done + current_fraction) / total))))
        if market_done > 0 and done < total:
            pct = max(1, pct)
    elif not running and total > 0 and done >= total:
        pct = 100
    else:
        pct = 0 if total == 0 else int(min(100, max(0, round(100.0 * done / total))))

    market_events = _read_market_web_events(current_tid) if current_tid else []
    market_event_max_seq = max((int(e.get("seq") or 0) for e in market_events), default=0)
    if running and current_tid and market_total > 0 and market_done > 0 and market_done > market_event_max_seq:
        should_add_fallback = False
        with merge_site_lock:
            last_chat_done = int(merge_site_state.get("last_market_chat_done") or 0)
            if market_done > last_chat_done:
                merge_site_state["last_market_chat_done"] = market_done
                should_add_fallback = True
        if should_add_fallback:
            _merge_chat_add("done", f"✅ {market_done}/{market_total} · готово.", tender_id=current_tid, seq=market_done, total=market_total)
            if market_done < market_total:
                _merge_chat_add("begin", f"Работа {market_done + 1} из {market_total} началась.", tender_id=current_tid, seq=market_done + 1, total=market_total)

    with merge_site_lock:
        web_events = list(merge_site_state.get("chat_events") or [])
        chat_events = sorted(
            web_events + market_events,
            key=lambda e: (str(e.get("ts") or ""), str(e.get("source") or ""), int(e.get("seq") or 0)),
        )[-140:]
        payload = {
            "running": bool(merge_site_state["running"]),
            "total": total,
            "done": done,
            "percent": pct,
            "current_tid": merge_site_state.get("current_tid") or "",
            "market_done": market_done,
            "market_total": market_total,
            "market_left": max(0, market_total - market_done),
            "market_percent": market_percent,
            "started_at": merge_site_state.get("started_at"),
            "ended_at": merge_site_state.get("ended_at"),
            "error_ids": list(merge_site_state.get("error_ids") or []),
            "log_tail": (merge_site_state.get("log_lines") or [])[-60:],
            "chat_events": chat_events,
            "last_ended_at": merge_site_state.get("last_ended_at"),
            "last_summary": merge_site_state.get("last_summary") or "",
            "last_reason_counts": merge_site_state.get("last_reason_counts") or {},
        }
    return jsonify(payload)


@app.route("/api/avito-status")
def api_avito_status():
    """Read the persistent Avito pause without opening Avito."""
    try:
        from autobot.real_market_scraper import avito_guard_status

        status = avito_guard_status()
    except Exception as exc:
        return jsonify({"ok": False, "message": f"Не удалось прочитать состояние Авито: {str(exc)[:300]}"}), 500
    blocked_until = float(status.get("blocked_until") or 0)
    status["blocked_until_iso"] = (
        datetime.fromtimestamp(blocked_until, timezone.utc).isoformat(timespec="seconds") if blocked_until > 0 else ""
    )
    return jsonify({"ok": True, **status})


_AGENT_MARKET_SOURCE_HINTS = (
    {
        "regions": ("яросл",),
        "markers": ("бетон в15", "бетон м200", "бст в15"),
        "urls": (
            "https://beton-yrs.ru/price/",
            "https://yaroslavl.gamma-beton.ru/price",
            "https://yar-beton.ru/",
        ),
    },
    {
        "regions": ("яросл",),
        "markers": ("песок строительный", "песок карьерный", "песок мелкий"),
        "urls": (
            "https://yaroslavl.scheben-rf.ru/pesok_karerniy/",
            "https://xn--90ahb6al8czar.xn--p1ai/karernyj-pesok/",
            "https://pesok-yaroslavl.ru/kariernyy-pesok",
            "https://postavka76.ru/pages/pesok.htm",
        ),
    },
    {
        "regions": ("яросл",),
        "markers": ("щебень",),
        "urls": (
            "https://yaroslavl.scheben-rf.ru/scheben_granitniy/",
            "https://yaroslavl.scheben-rf.ru/scheben_graviyniy/",
            "https://xn--90ahb6al8czar.xn--p1ai/shcheben-20-40/",
        ),
    },
)


def _agent_market_start_urls(name: object, queries: object, region: object) -> list[str]:
    region_text = str(region or "").casefold().replace("ё", "е")
    query_text = " ".join([str(name or ""), *(str(item or "") for item in list(queries or []))]).casefold().replace("ё", "е")
    for hint in _AGENT_MARKET_SOURCE_HINTS:
        if not any(marker in region_text for marker in hint["regions"]):
            continue
        if any(marker.casefold().replace("ё", "е") in query_text for marker in hint["markers"]):
            return list(hint["urls"])
    return []


@app.route("/api/tenders/<tender_id>/agent-market/jobs", methods=["GET", "POST"])
def api_tender_agent_market_jobs(tender_id: str):
    """Create browser-agent jobs from real estimate rows or show their current state."""
    tid = str(tender_id or "").strip()
    if not re.fullmatch(r"\d{8,25}", tid):
        return jsonify({"ok": False, "message": "Некорректный номер тендера"}), 400
    from autobot.agent_market_queue import enqueue_jobs, job_progress, job_summary, list_jobs

    requested_mode = str(request.args.get("mode") or "web").strip().casefold()
    job_mode = "avito" if requested_mode == "avito" else "web"

    if request.method == "GET":
        jobs = list_jobs(tid, mode=job_mode)
        latest_jobs: dict[str, dict] = {}
        for job in jobs:
            key = str(job.get("position_key") or "").strip()
            if key and key not in latest_jobs:
                latest_jobs[key] = job

        public_results: list[dict] = []
        result_totals = {"found": 0, "verified": 0, "candidate": 0, "rejected": 0}
        for job in latest_jobs.values():
            result = job.get("result") or {}
            imported = result.get("import") or {}
            outcomes = {
                str(item.get("url") or "").strip(): item
                for item in list(imported.get("offer_outcomes") or [])
                if isinstance(item, dict) and str(item.get("url") or "").strip()
            }
            raw_offers = [item for item in list(result.get("offers") or [])[:10] if isinstance(item, dict)]
            for raw_offer in raw_offers:
                url = str(raw_offer.get("url") or "").strip()
                outcome = outcomes.get(url) or {}
                verification = str(outcome.get("verification") or "").strip().casefold()
                if not verification:
                    if imported.get("offer_outcomes") is not None:
                        verification = "rejected"
                    else:
                        verification = "verified" if len(raw_offers) == 1 and int(imported.get("verified") or 0) else "candidate"
                if verification not in {"verified", "candidate"}:
                    verification = "rejected"
                result_totals["found"] += 1
                result_totals[verification] += 1
                public_results.append(
                    {
                        "job_id": job.get("id"),
                        "position_key": job.get("position_key"),
                        "position_name": job.get("position_name"),
                        "title": str(raw_offer.get("title") or "Источник цены")[:500],
                        "price": raw_offer.get("price"),
                        "unit": str(raw_offer.get("unit") or outcome.get("matched_unit") or "")[:80],
                        "url": url,
                        "evidence": str(raw_offer.get("evidence") or "")[:800],
                        "verification": verification,
                        "reason": str(outcome.get("verification_reason") or imported.get("message") or "")[:800],
                        "completed_at": job.get("completed_at"),
                    }
                )
        public_jobs = [
            {
                "id": job.get("id"),
                "position_key": job.get("position_key"),
                "position_name": job.get("position_name"),
                "job_mode": job.get("job_mode") or "web",
                "status": job.get("status"),
                "attempts": job.get("attempts"),
                "worker_id": job.get("worker_id"),
                "error": job.get("error"),
                "created_at": job.get("created_at"),
                "updated_at": job.get("updated_at"),
                "completed_at": job.get("completed_at"),
                "offers_found": len((job.get("result") or {}).get("offers") or []),
                "notes": str((job.get("result") or {}).get("notes") or "")[:500],
                "import": (job.get("result") or {}).get("import") or {},
            }
            for job in jobs
        ]
        return jsonify(
            {
                "ok": True,
                "enabled": bool(_agent_market_token()),
                "mode": job_mode,
                "summary": job_summary(tid, mode=job_mode),
                "progress": job_progress(tid, mode=job_mode),
                "jobs": public_jobs,
                "results": public_results[:60],
                "result_totals": result_totals,
            }
        )

    estimate_path = REPORTS_DIR / f"ОТЧЕТ_ПО_СМЕТАМ_{tid}.xlsx"
    if not estimate_path.is_file():
        return jsonify({"ok": False, "message": "Для тендера ещё нет распознанной сметы"}), 404
    data = request.get_json(silent=True) or {}
    requested_mode = str(data.get("mode") or requested_mode).strip().casefold()
    job_mode = "avito" if requested_mode == "avito" else "web"
    requested_keys = {
        str(value or "").strip()
        for value in list(data.get("position_keys") or [])[:100]
        if str(value or "").strip()
    }
    try:
        default_limit = 5 if job_mode == "avito" else 20
        limit = max(1, min(int(data.get("limit") or default_limit), 50))
    except (TypeError, ValueError):
        limit = 5 if job_mode == "avito" else 20
    metadata = load_tender_metadata()
    meta = dict(metadata.get(tid) or {})
    workflow_items, _ = _tenders_items()
    workflow = next((dict(item) for item in workflow_items if str(item.get("tender_id") or "") == tid), {})
    tender = build_tender_detail(tid, meta, workflow)
    eligible_positions: list[tuple[int, int, dict, list[str]]] = []
    skipped_ineligible: list[dict[str, str]] = []
    type_priority = {"material": 0, "product": 0, "service": 1, "work": 1}
    for row_index, position in enumerate(tender.get("positions") or []):
        key = str(position.get("position_key") or "")
        if requested_keys and key not in requested_keys:
            continue
        if not requested_keys and position.get("verified_count"):
            continue
        queries = [str(query or "").strip() for query in list(position.get("queries") or []) if str(query or "").strip()]
        position_type = str(position.get("type_slug") or "").strip().casefold()
        unit = str(position.get("unit") or "").strip()
        can_auto_price = position.get("can_auto_price")
        if can_auto_price is None:
            can_auto_price = bool(queries and unit and unit != "—")
        reason = ""
        if not can_auto_price:
            reason = str(position.get("warning") or "Позиция не готова к автоматическому сравнению")
        elif not queries:
            reason = "Нет безопасного рыночного запроса"
        elif position_type in {"aggregate", "other"}:
            reason = "Сводную или неоднозначную строку сначала нужно разложить"
        elif not unit or unit == "—":
            reason = "Нет единицы измерения"
        if reason:
            skipped_ineligible.append({"position_key": key, "name": str(position.get("name") or ""), "reason": reason})
            continue
        eligible_positions.append((type_priority.get(position_type, 2), row_index, position, queries))

    selected: list[dict] = []
    for search_rank, (_, row_index, position, queries) in enumerate(sorted(eligible_positions, key=lambda item: (item[0], item[1]))):
        key = str(position.get("position_key") or "")
        primary_query = queries[0]
        region = str(tender.get("region") or "").strip()
        position_type = str(position.get("type_slug") or "").strip().casefold()
        start_urls = _agent_market_start_urls(position.get("name"), queries, region)
        payload = {
            "schema_version": 2,
            "tender_id": tid,
            "position_key": key,
            "item_no": position.get("item_no"),
            "name": position.get("name"),
            "unit": position.get("unit"),
            "quantity": position.get("quantity"),
            "section": position.get("section"),
            "source_file": position.get("source_file"),
            "basis_code": position.get("basis_code"),
            "position_type": position_type,
            "region": region,
            "estimate_unit_price": position.get("estimate_unit"),
            "queries": queries,
            "job_mode": job_mode,
            "max_offers": 3,
            "max_sources": 3,
            "max_turns": 16,
            "max_seconds": 180,
            "max_attempts": 1,
            "retry_policy": "network_only",
            "queue_priority": (60 if position_type in {"material", "product"} else 70) + search_rank,
            "start_urls": start_urls,
            "result_schema": {
                "schema_version": 2,
                "position_key": key,
                "offers": [
                    {
                        "title": "",
                        "price": 0,
                        "currency": "RUB",
                        "unit": "",
                        "url": "",
                        "evidence": "",
                        "observed_at": "ISO-8601",
                        "published_at": "",
                        "location": "",
                        "confidence": 0.0,
                    }
                ],
                "notes": "",
            },
        }
        if job_mode == "avito":
            avito_query = re.sub(r"\s+", " ", f"{primary_query} {region}".strip())
            avito_url = "https://www.avito.ru/all?" + urlencode({"q": avito_query})
            payload.update(
                {
                    "search_mode": "avito_agent",
                    "max_offers": 3,
                    "max_sources": 3,
                    "max_turns": 24,
                    "max_seconds": 360,
                    "allowed_domains": ["avito.ru"],
                    "start_urls": [avito_url],
                    "task": (
                        "Ищи цену только на Авито через браузерную сессию Mac mini. "
                        f"Начни с готовой страницы поиска: {avito_url}. "
                        "Открой не более 3 подходящих объявлений и верни только прямые ссылки вида avito.ru/..._123456789. "
                        "Для каждого предложения запиши точное название, цену, единицу, город и короткий видимый фрагмент страницы в evidence. "
                        "В price пиши число ровно как на странице, а в unit — его знаменатель (например, 650 и м или 65000 и 100 м); не пересчитывай сам. "
                        "Не используй сниппеты поисковиков, другие домены, цену доставки, кредита или похожего товара. "
                        "Не обходи CAPTCHA и ограничения доступа, не перезагружай заблокированную страницу многократно. "
                        "Если Авито показал CAPTCHA или ограничение IP, сразу верни пустой offers и укажи причину в notes. "
                        "Остановись после 3 валидных объявлений и верни только JSON по схеме result_schema."
                    ),
                }
            )
        else:
            source_instruction = (
                "Сначала открой эти прямые источники по порядку: " + ", ".join(start_urls) + ". "
                if start_urls
                else ""
            )
            payload.update(
                {
                    "search_mode": "fast_web",
                    "excluded_domains": ["avito.ru"],
                    "task": (
                    "Быстрый поиск только по обычным сайтам поставщиков, производителей и подрядчиков. "
                    + source_instruction
                    + f"Используй запросы из queries по порядку; товар/работа уже определены как {position_type}. "
                    "Не используй Авито. Открой не более 3 наиболее перспективных прямых страниц, "
                    "не делай искусственных пауз и не обходи CAPTCHA или ограничения сайта. "
                    "Остановись сразу после 2 валидных цен. Если 3 страницы не дали цену, сразу верни результат. "
                    "Не используй цену доставки, кредита или похожего товара. Верни только JSON по схеме result_schema."
                    " В price верни цену ровно как она видна на странице, а в unit — точную единицу/блок; сам цену не умножай. Evidence должен содержать видимые название, цену и единицу."
                ),
                }
            )
        selected.append(payload)
        if len(selected) >= limit:
            break
    if not selected:
        return jsonify({"ok": False, "message": "Нет позиций, которые можно безопасно сравнить с рынком", "skipped_ineligible": skipped_ineligible}), 400
    outcome = enqueue_jobs(tid, selected, priority=80 if job_mode == "avito" else 100)
    return jsonify(
        {
            "ok": True,
            "mode": job_mode,
            "created": len(outcome["created"]),
            "skipped_active": len(outcome["skipped_active"]),
            "skipped_ineligible": skipped_ineligible,
            "jobs": outcome["created"],
            "summary": job_summary(tid, mode=job_mode),
        }
    )


@app.route("/api/tenders/<tender_id>/agent-market/jobs/<job_id>/cancel", methods=["POST"])
def api_tender_agent_market_cancel(tender_id: str, job_id: str):
    from autobot.agent_market_queue import cancel_job

    if not cancel_job(str(job_id or ""), str(tender_id or "")):
        return jsonify({"ok": False, "message": "Активное задание не найдено"}), 404
    return jsonify({"ok": True})


@app.route("/api/agent-market/v1/status")
def api_agent_market_worker_status():
    auth_error = _require_agent_market_token()
    if auth_error:
        return auth_error
    return jsonify({"ok": True, "schema_version": 1, "service": "autobot-agent-market"})


@app.route("/api/agent-market/v1/claim", methods=["POST"])
def api_agent_market_claim():
    auth_error = _require_agent_market_token()
    if auth_error:
        return auth_error
    from autobot.agent_market_queue import claim_job

    data = request.get_json(silent=True) or {}
    worker_id = str(data.get("worker_id") or "").strip()
    if not worker_id:
        return jsonify({"ok": False, "message": "Нужен worker_id"}), 400
    try:
        lease_seconds = int(data.get("lease_seconds") or 600)
        job = claim_job(worker_id, lease_seconds=lease_seconds)
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    if not job:
        return jsonify({"ok": True, "job": None}), 200
    return jsonify({"ok": True, "job": job})


@app.route("/api/agent-market/v1/jobs/<job_id>/heartbeat", methods=["POST"])
def api_agent_market_heartbeat(job_id: str):
    auth_error = _require_agent_market_token()
    if auth_error:
        return auth_error
    from autobot.agent_market_queue import heartbeat_job

    data = request.get_json(silent=True) or {}
    worker_id = str(data.get("worker_id") or "").strip()
    if not heartbeat_job(job_id, worker_id, lease_seconds=int(data.get("lease_seconds") or 600)):
        return jsonify({"ok": False, "message": "Задание не принадлежит этому агенту"}), 409
    return jsonify({"ok": True})


@app.route("/api/agent-market/v1/jobs/<job_id>/complete", methods=["POST"])
def api_agent_market_complete(job_id: str):
    auth_error = _require_agent_market_token()
    if auth_error:
        return auth_error
    from autobot.agent_market_queue import complete_job, get_job
    from autobot.real_market_scraper import import_agent_market_result

    data = request.get_json(silent=True) or {}
    worker_id = str(data.get("worker_id") or "").strip()
    job = get_job(job_id)
    if not job or job.get("status") != "leased" or job.get("worker_id") != worker_id:
        return jsonify({"ok": False, "message": "Задание не принадлежит этому агенту"}), 409
    try:
        validated = _validate_agent_market_result(data.get("result"), str(job.get("position_key") or ""))
        with agent_market_import_lock:
            imported = import_agent_market_result(str(job.get("tender_id") or ""), job.get("payload") or {}, validated)
        validated["import"] = imported
    except (OSError, ValueError, TypeError) as exc:
        return jsonify({"ok": False, "message": f"Результат не принят: {str(exc)[:500]}"}), 422
    completed = complete_job(job_id, worker_id, validated)
    if not completed:
        return jsonify({"ok": False, "message": "Не удалось завершить задание"}), 409
    return jsonify({"ok": True, "job_id": job_id, "import": imported})


@app.route("/api/agent-market/v1/jobs/<job_id>/fail", methods=["POST"])
def api_agent_market_fail(job_id: str):
    auth_error = _require_agent_market_token()
    if auth_error:
        return auth_error
    from autobot.agent_market_queue import complete_job, fail_job, get_job
    from autobot.real_market_scraper import import_agent_market_result, probe_agent_market_start_urls

    data = request.get_json(silent=True) or {}
    worker_id = str(data.get("worker_id") or "").strip()
    job = get_job(job_id)
    if job and job.get("status") == "leased" and job.get("worker_id") == worker_id:
        payload = job.get("payload") or {}
        if str(job.get("job_mode") or "web") == "web" and payload.get("start_urls"):
            try:
                recovered = probe_agent_market_start_urls(str(job.get("tender_id") or ""), payload, max_sources=1)
                if recovered.get("offers"):
                    validated = _validate_agent_market_result(recovered, str(job.get("position_key") or ""))
                    validated["_autobot_direct_probe"] = True
                    with agent_market_import_lock:
                        imported = import_agent_market_result(str(job.get("tender_id") or ""), payload, validated)
                    validated["import"] = imported
                    completed = complete_job(job_id, worker_id, validated)
                    if completed:
                        return jsonify({"ok": True, "recovered": True, "job_id": job_id, "import": imported})
            except (OSError, ValueError, TypeError):
                # The original worker failure remains the authoritative result
                # when the bounded direct-source fallback cannot be verified.
                pass
    try:
        ok = fail_job(
            job_id,
            worker_id,
            str(data.get("error") or "Ошибка агента"),
            retry=bool(data.get("retry")),
        )
    except (TypeError, ValueError) as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    if not ok:
        return jsonify({"ok": False, "message": "Задание не принадлежит этому агенту"}), 409
    return jsonify({"ok": True})


@app.route("/api/reports-coverage")
def api_reports_coverage():
    return jsonify(_compute_reports_coverage())


@app.route("/api/workflow-overview")
def api_workflow_overview():
    include_storage = str(request.args.get("storage") or "").strip().lower() in {"1", "true", "yes"}
    return jsonify(build_workflow_payload(include_storage=include_storage))


@app.route("/api/storage-overview")
def api_storage_overview():
    return jsonify({"storage": [item.to_dict() for item in build_storage_overview()]})


@app.route("/api/push-state")
def api_push_state():
    cov = _compute_reports_coverage()
    with parse_lock:
        pr_running = bool(parse_state.get("running"))
        pr_exit = parse_state.get("exit_code")
        pr_end = parse_state.get("ended_at")
    with merge_site_lock:
        mr_running = bool(merge_site_state.get("running"))
        mr_last_end = merge_site_state.get("last_ended_at")
        mr_summary = str(merge_site_state.get("last_summary") or "")
    return jsonify(
        {
            "parse_running": pr_running,
            "parse_exit_code": pr_exit,
            "parse_ended_at": pr_end,
            "merge_running": mr_running,
            "merge_last_ended_at": mr_last_end,
            "merge_last_summary": mr_summary,
            "coverage_merge_html": int(cov.get("merge_html_among_tenders", 0) or 0),
        }
    )


@app.route("/api/generate-merge-site-all", methods=["POST"])
def api_generate_merge_site_all():
    if _merge_site_busy():
        return jsonify({"ok": False, "message": "Сравнения цен уже подготавливаются."}), 409
    with parse_lock:
        if parse_state["running"]:
            return jsonify({"ok": False, "message": "Сначала дождитесь окончания текущей работы с документами."}), 409
    threading.Thread(target=_run_merge_site_all_worker, kwargs={"only_missing": False}, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/generate-merge-site-missing", methods=["POST"])
def api_generate_merge_site_missing():
    if _merge_site_busy():
        return jsonify({"ok": False, "message": "Сравнения цен уже подготавливаются."}), 409
    with parse_lock:
        if parse_state["running"]:
            return jsonify({"ok": False, "message": "Сначала дождитесь окончания текущей работы с документами."}), 409
    threading.Thread(target=_run_merge_site_all_worker, kwargs={"only_missing": True}, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/generate-merge-site-selected", methods=["POST"])
def api_generate_merge_site_selected():
    if _merge_site_busy():
        return jsonify({"ok": False, "message": "Сравнения цен уже подготавливаются."}), 409
    with parse_lock:
        if parse_state["running"]:
            return jsonify({"ok": False, "message": "Сначала дождитесь окончания текущей работы с документами."}), 409
    data = request.get_json(silent=True) or {}
    raw_ids = data.get("tender_ids")
    if not isinstance(raw_ids, list):
        return jsonify({"ok": False, "message": "Передайте список tender_ids."}), 400
    ids: list[str] = []
    seen: set[str] = set()
    known_ids = set(load_tender_metadata().keys())
    for raw in raw_ids:
        tid = str(raw or "").strip()
        if not re.fullmatch(r"\d{8,25}", tid) or tid in seen or tid not in known_ids:
            continue
        seen.add(tid)
        ids.append(tid)
        if len(ids) >= 100:
            break
    if not ids:
        return jsonify({"ok": False, "message": "Выберите хотя бы один тендер из каталога."}), 400
    threading.Thread(
        target=_run_merge_site_all_worker,
        kwargs={"ids_override": ids},
        daemon=True,
    ).start()
    return jsonify({"ok": True, "tender_ids": ids, "count": len(ids)})


@app.route("/api/generate-merge-site-one", methods=["POST"])
def api_generate_merge_site_one():
    if _merge_site_busy():
        return jsonify({"ok": False, "message": "Сравнения цен уже подготавливаются."}), 409
    with parse_lock:
        if parse_state["running"]:
            return jsonify({"ok": False, "message": "Сначала дождитесь окончания текущей работы с документами."}), 409
    data = request.get_json(silent=True) or {}
    tid = str(data.get("tender_id", "")).strip()
    if not tid:
        return jsonify({"ok": False, "message": "Нужен tender_id"}), 400
    threading.Thread(target=_run_merge_site_all_worker, kwargs={"ids_override": [tid]}, daemon=True).start()
    return jsonify({"ok": True, "tender_id": tid})


@app.route("/api/generate-merge-site-one-rerun-market", methods=["POST"])
def api_generate_merge_site_one_rerun_market():
    if _merge_site_busy():
        return jsonify({"ok": False, "message": "Сравнения цен уже подготавливаются."}), 409
    with parse_lock:
        if parse_state["running"]:
            return jsonify({"ok": False, "message": "Сначала дождитесь окончания текущей работы с документами."}), 409
    data = request.get_json(silent=True) or {}
    tid = str(data.get("tender_id", "")).strip()
    if not tid:
        return jsonify({"ok": False, "message": "Нужен tender_id"}), 400
    threading.Thread(
        target=_run_merge_site_all_worker,
        kwargs={"ids_override": [tid], "force_market_no_resume": True},
        daemon=True,
    ).start()
    return jsonify({"ok": True, "tender_id": tid, "mode": "rerun_market_no_resume"})


@app.route("/api/generate-merge-site-one-sample-market", methods=["POST"])
def api_generate_merge_site_one_sample_market():
    """Recheck a small batch without deleting the remaining market report."""
    if _merge_site_busy():
        return jsonify({"ok": False, "message": "Сравнения цен уже подготавливаются."}), 409
    with parse_lock:
        if parse_state["running"]:
            return jsonify({"ok": False, "message": "Сначала дождитесь окончания текущей работы с документами."}), 409
    data = request.get_json(silent=True) or {}
    tid = str(data.get("tender_id", "")).strip()
    if not tid:
        return jsonify({"ok": False, "message": "Нужен tender_id"}), 400
    try:
        sample_rows = int(data.get("sample_rows") or 10)
    except (TypeError, ValueError):
        sample_rows = 10
    sample_rows = max(1, min(20, sample_rows))
    threading.Thread(
        target=_run_merge_site_all_worker,
        kwargs={
            "ids_override": [tid],
            "market_max_rows": sample_rows,
            "market_rerun_selected": True,
        },
        daemon=True,
    ).start()
    return jsonify({"ok": True, "tender_id": tid, "mode": "sample_market", "sample_rows": sample_rows})


@app.route("/api/generate-avito-safe-sample", methods=["POST"])
def api_generate_avito_safe_sample():
    """Collect a small fresh Avito sample after the persistent cooldown expires."""
    if _merge_site_busy():
        return jsonify({"ok": False, "message": "Сравнения цен уже подготавливаются."}), 409
    with parse_lock:
        if parse_state["running"]:
            return jsonify({"ok": False, "message": "Сначала дождитесь окончания текущей работы с документами."}), 409
    data = request.get_json(silent=True) or {}
    tid = str(data.get("tender_id", "")).strip()
    if not re.fullmatch(r"\d{8,25}", tid):
        return jsonify({"ok": False, "message": "Нужен корректный tender_id."}), 400
    if tid not in set(_estimate_xlsx_tender_ids()):
        return jsonify({"ok": False, "message": "Для этого тендера ещё нет готовой сметы."}), 404
    try:
        from autobot.real_market_scraper import avito_guard_status

        avito_status = avito_guard_status()
    except Exception as exc:
        return jsonify({"ok": False, "message": f"Не удалось прочитать состояние Авито: {str(exc)[:300]}"}), 500
    if bool(avito_status.get("blocked")):
        remaining_minutes = max(1, int((int(avito_status.get("remaining_seconds") or 0) + 59) // 60))
        remaining_hours, remaining_tail = divmod(remaining_minutes, 60)
        remaining_text = (
            f"{remaining_hours} ч {remaining_tail} мин"
            if remaining_hours and remaining_tail
            else (f"{remaining_hours} ч" if remaining_hours else f"{remaining_tail} мин")
        )
        return jsonify(
            {
                "ok": False,
                "message": f"Авито пока на паузе — осталось примерно {remaining_text}. AutoBot не будет обращаться к сайту раньше.",
                "avito": avito_status,
            }
        ), 429
    try:
        sample_rows = int(data.get("sample_rows") or 2)
    except (TypeError, ValueError):
        sample_rows = 2
    # Одна строка сметы = одна полноценная поисковая навигация Авито. Для
    # домашнего IP безопасная проба намеренно мала; большой объём берём из кэша
    # и локального индекса, а не повторными открытиями сайта.
    sample_rows = max(1, min(2, sample_rows))
    threading.Thread(
        target=_run_merge_site_all_worker,
        kwargs={
            "ids_override": [tid],
            "market_max_rows": sample_rows,
            "market_rerun_selected": True,
            "market_sources_override": "avito",
            "market_only_without_verified": True,
            "market_avito_collect_only": True,
        },
        daemon=True,
    ).start()
    return jsonify(
        {
            "ok": True,
            "tender_id": tid,
            "mode": "avito_safe_collect",
            "sample_rows": sample_rows,
            "message": "Бережный сбор Авито запущен.",
            "safety": {
                "max_rows": 2,
                "daily_remaining": avito_status.get("daily_remaining"),
                "next_request_in_seconds": avito_status.get("next_request_in_seconds"),
            },
        }
    )


@app.route("/api/generate-merge-site-by-link", methods=["POST"])
def api_generate_merge_site_by_link():
    if _merge_site_busy():
        return jsonify({"ok": False, "message": "Сравнения цен уже подготавливаются."}), 409
    with parse_lock:
        if parse_state["running"]:
            return jsonify({"ok": False, "message": "Сначала дождитесь окончания текущей работы с документами."}), 409
    data = request.get_json(silent=True) or {}
    raw = str(data.get("tender_link", "")).strip()
    tid = _extract_tender_id(raw)
    if not tid:
        return jsonify({"ok": False, "message": "Не удалось извлечь номер тендера из ссылки/текста."}), 400
    turl = raw if raw.startswith("http://") or raw.startswith("https://") else ""
    kwargs = {"ids_override": [tid]}
    if turl:
        kwargs["tender_url_by_id"] = {tid: turl}
    threading.Thread(target=_run_merge_site_all_worker, kwargs=kwargs, daemon=True).start()
    return jsonify({"ok": True, "tender_id": tid})


@app.route("/api/tender-viability-refresh", methods=["POST"])
def api_tender_viability_refresh():
    """Пересборка HTML отчёта с блоком оценки + опционально Telegram (по готовой СВОДКА_РЫНОК)."""
    data = request.get_json(silent=True) or {}
    tid = str(data.get("tender_id", "")).strip()
    if not tid:
        return jsonify({"ok": False, "message": "Нужен tender_id"}), 400
    from autobot.merge_estimate_market import OUT_PREFIX

    sv = REPORTS_DIR / f"{OUT_PREFIX}{tid}.xlsx"
    if not sv.is_file():
        return jsonify(
            {
                "ok": False,
                "message": "Для этой закупки ещё нет готового сравнения цен. Сначала найдите рыночные цены.",
            }
        ), 400
    try:
        from autobot.report_merge_html import write_tender_report_site

        out_path = write_tender_report_site(tid)
        if not out_path or not out_path.is_file():
            return jsonify({"ok": False, "message": "Не удалось записать index.html в data/reports_site."}), 500
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)[:500]}), 500

    telegram_sent = False
    try:
        from autobot.tender_viability import format_viability_for_telegram

        vmsg = format_viability_for_telegram(tid)
        if vmsg and _telegram_cfg():
            _tg_flush_spool()
            _tg_send(vmsg)
            _tg_flush_spool()
            telegram_sent = True
    except Exception:
        pass

    base = (get_report_site_public_base() or "").strip().rstrip("/")
    report_url = f"{base}/tenders/{tid}" if base else ""
    return jsonify(
        {
            "ok": True,
            "tender_id": tid,
            "message": "Страница отчёта и блок «Оценка по сравнению» обновлены.",
            "report_url": report_url,
            "telegram_sent": telegram_sent,
        }
    )


if __name__ == "__main__":
    # Для ссылки из Telegram с телефона в той же Wi‑Fi: WEB_UI_HOST=0.0.0.0 и в .env
    # REPORT_SITE_PUBLIC_BASE_URL=http://<IP_ПК>:8765
    _host = (os.environ.get("WEB_UI_HOST") or "127.0.0.1").strip() or "127.0.0.1"
    _port = int((os.environ.get("WEB_UI_PORT") or "8765").strip() or "8765")
    app.run(host=_host, port=_port, debug=False)
